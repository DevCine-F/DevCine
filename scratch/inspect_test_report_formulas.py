import sys
import io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx', data_only=False)
wb_val = openpyxl.load_workbook(r'C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx', data_only=True)

ws_tr = wb_val['Test Report']
ws_tr_raw = wb['Test Report']

print("=== ALL ROWS IN TEST REPORT SHEET ===")
for r in range(1, ws_tr.max_row + 1):
    vals = [str(ws_tr.cell(r, c).value or '').strip() for c in range(1, ws_tr.max_column + 1)]
    raws = [str(ws_tr_raw.cell(r, c).value or '').strip() for c in range(1, ws_tr_raw.max_column + 1)]
    if any(vals):
        print(f"Row {r:2d} (Val):", ' | '.join(vals[:8]))
        if any('=' in str(x) for x in raws):
            print(f"       (Raw):", ' | '.join(raws[:8]))
