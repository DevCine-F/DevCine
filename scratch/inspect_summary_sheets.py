import sys
import io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx', data_only=False)
wb_val = openpyxl.load_workbook(r'C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx', data_only=True)

print("=== CHECK COVER, TEST CASE LIST, TEST REPORT, FUNCTION ===")

ws_tr = wb_val['Test Report']
print(f"Test Report rows: {ws_tr.max_row}")
for r in range(1, ws_tr.max_row + 1):
    vals = [str(ws_tr.cell(r, c).value or '').strip() for c in range(1, 9)]
    if any(vals):
        print(f"TR Row {r:2d}:", ' | '.join(vals))

ws_tcl = wb_val['Test case List']
print(f"\nTest case List rows: {ws_tcl.max_row}")
for r in range(1, ws_tcl.max_row + 1):
    vals = [str(ws_tcl.cell(r, c).value or '').strip() for c in range(1, 7)]
    if any(vals):
        print(f"TCL Row {r:2d}:", ' | '.join(vals))

ws_func = wb_val['FUNCTION']
print(f"\nFUNCTION rows: {ws_func.max_row}")
for r in range(1, ws_func.max_row + 1):
    vals = [str(ws_func.cell(r, c).value or '').strip() for c in range(1, 7)]
    if any(vals):
        print(f"FUNC Row {r:2d}:", ' | '.join(vals))
