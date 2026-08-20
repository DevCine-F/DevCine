import sys
import io
import openpyxl
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx', data_only=True)

test_case_sheets = [s for s in wb.sheetnames if s not in ['Cover (Tổng quan)', 'Test case List', 'Test Report', 'FUNCTION']]

kho_keywords = ['kho', 'tồn kho', 'định mức', 'nguyên vật liệu', 'nhập kho', 'xuất kho', 'hoàn kho', 'trừ tồn kho']
shift_keywords = ['ca làm', 'chia ca', 'giao ca', 'kết ca', 'đổi ca', 'phân ca', 'bàn giao ca']

total_tcs = 0
sheet_stats = {}
all_tc_details = {}

for s in test_case_sheets:
    ws = wb[s]
    header_row = 10
    for r in range(1, 15):
        vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        if any('Mã Test Case' in v or 'Test Case ID' in v or 'Tiêu đề' in v for v in vals):
            header_row = r
            break
            
    tcs = []
    kho_matches = []
    shift_matches = []
    vague_data = []
    cloned_actual = 0
    title_thanh_cong_issues = []
    
    for r in range(header_row + 1, ws.max_row + 1):
        tc_id = str(ws.cell(r, 1).value or '').strip()
        title = str(ws.cell(r, 2).value or '').strip()
        desc = str(ws.cell(r, 3).value or '').strip()
        steps = str(ws.cell(r, 4).value or '').strip()
        data = str(ws.cell(r, 5).value or '').strip()
        expected = str(ws.cell(r, 6).value or '').strip()
        actual = str(ws.cell(r, 7).value or '').strip()
        status = str(ws.cell(r, 9).value or '').strip()
        
        if not tc_id or 'KIỂM TRA' in tc_id or 'Section' in tc_id:
            continue
            
        tc_item = {
            'row': r, 'id': tc_id, 'title': title, 'desc': desc, 
            'steps': steps, 'data': data, 'expected': expected, 
            'actual': actual, 'status': status
        }
        tcs.append(tc_item)
        
        full_text = f"{title} {desc} {steps} {data} {expected} {actual}".lower()
        for kw in kho_keywords:
            if kw in full_text:
                kho_matches.append((tc_id, kw, title, desc, expected))
                break
        for kw in shift_keywords:
            if kw in full_text:
                shift_matches.append((tc_id, kw, title, desc, expected))
                break
                
        if data in ['N/A', '', 'None'] and any(w in desc.lower() or w in title.lower() for w in ['nhập', 'bỏ trống', 'sai', 'vượt', 'ký tự', 'số']):
            vague_data.append((tc_id, title, desc))
            
        if expected == actual and expected != '':
            cloned_actual += 1
            
        if 'thất bại' in title.lower() and 'thành công' in title.lower():
            title_thanh_cong_issues.append((tc_id, title))
        elif any(w in title.lower() for w in ['sai', 'không hợp lệ', 'bỏ trống', 'trùng', 'chặn', 'không tìm thấy', 'thất bại', 'quá hạn']) and title.endswith('thành công'):
            title_thanh_cong_issues.append((tc_id, title))
            
    total_tcs += len(tcs)
    tot = len(tcs)
    all_tc_details[s] = tcs
    sheet_stats[s] = {
        'total': tot,
        'kho_matches': kho_matches,
        'shift_matches': shift_matches,
        'vague_data': vague_data,
        'cloned_actual': cloned_actual,
        'title_issues': title_thanh_cong_issues,
        'sample_tc': tcs[0] if tcs else None
    }

print(f"Total Test Cases in all sheets: {total_tcs}\n")

print("================================================================================")
print("1. PHÂN HỆ / TÍNH NĂNG NGOÀI PHẠM VI (CA LÀM / CHIA CA, KHO / ĐỊNH MỨC)")
print("================================================================================")
for s, st in sheet_stats.items():
    if st['kho_matches'] or st['shift_matches'] or s in ['Ca làm việc', 'POS Kết ca & Bàn giao']:
        tot = st['total']
        print(f"\n[!] Sheet [{s}] (Tổng {tot} Test cases):")
        if st['kho_matches']:
            print(f"  * Dính từ khóa KHO / ĐỊNH MỨC / TỒN KHO ({len(st['kho_matches'])} ca):")
            for m in st['kho_matches']:
                print(f"    - [{m[0]}] ({m[1]}): {m[2]}")
                print(f"      Desc: {m[3]}")
                print(f"      Expected: {m[4]}")
        if st['shift_matches']:
            print(f"  * Dính từ khóa CA LÀM / CHIA CA ({len(st['shift_matches'])} ca):")
            for m in st['shift_matches']:
                print(f"    - [{m[0]}] ({m[1]}): {m[2]}")
                print(f"      Desc: {m[3]}")
                print(f"      Expected: {m[4]}")

print("\n================================================================================")
print("2. ĐÁNH GIÁ CÁC SHEET QUẢN LÝ MASTER DATA RIÊNG BIỆT (ĐẠO DIỄN, DIỄN VIÊN, V.V.)")
print("================================================================================")
for s in ['Đạo diễn', 'Diễn viên', 'Thể loại phim', 'Định dạng chiếu', 'Banner quảng cáo', 'Cài đặt hệ thống', 'Điểm thưởng Loyalty']:
    if s in sheet_stats:
        print(f"\nSheet [{s}] ({sheet_stats[s]['total']} TCs):")
        ws = wb[s]
        req = ws.cell(2, 2).value
        print(f"  Yêu cầu test: {req}")
        for tc in all_tc_details[s][:3]:
            print(f"  - [{tc['id']}] {tc['title']}")
            print(f"    Data: {tc['data']} | Exp: {tc['expected']}")

print("\n================================================================================")
print("3. PHÂN TÍCH LỖI DIỄN ĐẠT, THIẾU RÕ RÀNG (VAGUE DATA, CLONED ACTUAL, TIÊU ĐỀ GƯỢNG)")
print("================================================================================")
for s, st in sheet_stats.items():
    issues_found = []
    if len(st['vague_data']) > 0:
        issues_found.append(f"{len(st['vague_data'])} ca thiếu test data cụ thể")
    if len(st['title_issues']) > 0:
        issues_found.append(f"{len(st['title_issues'])} ca tiêu đề phủ định nhưng đuôi 'thành công'")
    if st['cloned_actual'] == st['total'] and st['total'] > 0:
        issues_found.append("100% Actual Result copy paste nguyên văn Expected Result")
        
    if issues_found:
        print(f"Sheet [{s:28s}]: {', '.join(issues_found)}")
        if st['title_issues']:
            print(f"   Ví dụ tiêu đề gượng: {st['title_issues'][0][0]} - {st['title_issues'][0][1]}")
        if st['vague_data']:
            print(f"   Ví dụ thiếu data: {st['vague_data'][0][0]} - {st['vague_data'][0][1]}")
