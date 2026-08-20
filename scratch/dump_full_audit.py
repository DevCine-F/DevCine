import sys
import io
import openpyxl
import json

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx', data_only=True)

test_case_sheets = [s for s in wb.sheetnames if s not in ['Cover (Tổng quan)', 'Test case List', 'Test Report', 'FUNCTION']]

results = {}

for s in test_case_sheets:
    ws = wb[s]
    header_row = 10
    for r in range(1, 15):
        vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)]
        if any('Mã Test Case' in v or 'Test Case ID' in v or 'Tiêu đề' in v for v in vals):
            header_row = r
            break
            
    module_code = str(ws.cell(1, 2).value or '').strip()
    test_req = str(ws.cell(2, 2).value or '').strip()
    tester = str(ws.cell(3, 2).value or '').strip()
    
    tcs = []
    for r in range(header_row + 1, ws.max_row + 1):
        c1 = str(ws.cell(r, 1).value or '').strip()
        c2 = str(ws.cell(r, 2).value or '').strip()
        c3 = str(ws.cell(r, 3).value or '').strip()
        c4 = str(ws.cell(r, 4).value or '').strip()
        c5 = str(ws.cell(r, 5).value or '').strip()
        c6 = str(ws.cell(r, 6).value or '').strip()
        c7 = str(ws.cell(r, 7).value or '').strip()
        c9 = str(ws.cell(r, 9).value or '').strip()
        
        if not c1 or 'KIỂM TRA' in c1 or 'Section' in c1 or 'KỸ THUẬT' in c1:
            continue
            
        tcs.append({
            'row': r,
            'id': c1,
            'title': c2,
            'desc': c3,
            'steps': c4,
            'data': c5,
            'expected': c6,
            'actual': c7,
            'status': c9
        })
        
    results[s] = {
        'module_code': module_code,
        'test_requirement': test_req,
        'tester': tester,
        'total_tcs': len(tcs),
        'tcs': tcs
    }

print(f"Parsed {len(results)} sheets successfully.")
with open('scratch/full_audit_data.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print("Saved to scratch/full_audit_data.json")
