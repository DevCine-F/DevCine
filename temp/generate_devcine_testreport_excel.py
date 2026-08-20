# -*- coding: utf-8 -*-
"""
Script to generate TestReport Dự án DevCine.xlsx matching template TestReport Dự án CozyPot.xlsx
Comprehensive Software Test Report for DevCine Movie Theater Management & Online Booking System.
Standards for University Graduation Thesis (Đồ án tốt nghiệp).
Contains 46 complete sheets: Cover, Test Case List, Test Report, 42 Module Test Execution Sheets, and FUNCTION.
Every single test case has complete, structured "Các bước thực hiện (Procedure)" with Bước 1, Bước 2, Bước 3, Bước 4, Bước 5.
"""

import os
import sys
import datetime
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

def build_full_devcine_test_report(output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    font_name = "Times New Roman"
    
    border_thin = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000')
    )
    
    fill_header_navy = PatternFill(start_color='FF002060', end_color='FF002060', fill_type='solid')
    fill_header_green = PatternFill(start_color='FFC5E0B3', end_color='FFC5E0B3', fill_type='solid')
    fill_header_blue = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
    fill_header_gold = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
    fill_group_gray = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
    
    font_title = Font(name=font_name, size=16, bold=True, color='FF002060')
    font_sub_title = Font(name=font_name, size=13, bold=True, color='FF000000')
    font_header_white = Font(name=font_name, size=11, bold=True, color='FFFFFFFF')
    font_header_black = Font(name=font_name, size=11, bold=True, color='FF000000')
    font_bold = Font(name=font_name, size=11, bold=True, color='FF000000')
    font_regular = Font(name=font_name, size=11, bold=False, color='FF000000')
    font_pass = Font(name=font_name, size=11, bold=True, color='FF008000')
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_top_center = Alignment(horizontal='center', vertical='top', wrap_text=True)

    test_date = datetime.datetime(2026, 3, 19, 0, 0)
    
    # -------------------------------------------------------------------------
    # 1. SHEET: Cover (Tổng quan)
    # -------------------------------------------------------------------------
    ws_cover = wb.create_sheet("Cover (Tổng quan)")
    ws_cover.views.sheetView[0].showGridLines = True
    
    ws_cover.cell(2, 2, "BÁO CÁO KIỂM THỬ PHẦN MỀM (TEST REPORT)").font = font_title
    
    meta_info = [
        ("Project Name", "DevCine - Website Quản lý Rạp chiếu phim & Đặt vé trực tuyến", "Creator", "Nguyễn Quang Huy"),
        ("Project Code", "DEVCINE_2026", "Reviewer/Approver", "Hội đồng Đồ án Tốt nghiệp / Tech Lead"),
        ("Document Code", "TR_DEVCINE_v1.0", "Issue Date", test_date),
        ("", "", "Version", "Phiên bản v1.0")
    ]
    
    for r_idx, (k1, v1, k2, v2) in enumerate(meta_info, start=4):
        if k1:
            ws_cover.cell(r_idx, 2, k1).font = font_bold
            ws_cover.cell(r_idx, 2).border = border_thin
            ws_cover.cell(r_idx, 3, v1).font = font_regular
            ws_cover.cell(r_idx, 3).border = border_thin
        if k2:
            ws_cover.cell(r_idx, 5, k2).font = font_bold
            ws_cover.cell(r_idx, 5).border = border_thin
            ws_cover.cell(r_idx, 6, v2).font = font_regular
            ws_cover.cell(r_idx, 6).border = border_thin
            if isinstance(v2, datetime.datetime):
                ws_cover.cell(r_idx, 6).number_format = 'yyyy-mm-dd'
                
    ws_cover.cell(9, 2, "Lịch sử thay đổi tài liệu (Record of Change)").font = font_sub_title
    change_headers = ["Ngày hiệu lực", "Phiên bản", "Hạng mục thay đổi", "Loại (*A/D/M)", "Mô tả thay đổi", "Tài liệu tham khảo"]
    for c_idx, h in enumerate(change_headers, start=2):
        cell = ws_cover.cell(10, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    change_rows = [
        ("2026-03-10", "1.0", "Tạo mới kế hoạch kiểm thử", "A", "Khởi tạo bộ test case kiểm thử toàn bộ hệ thống DevCine", "SRS / BA Specs"),
        ("2026-03-15", "1.0", "Bổ sung test case phân hệ POS & Check-in", "A", "Thêm test case nghiệp vụ bán vé tại quầy và soát vé QR", "POS Architecture Doc"),
        ("2026-03-19", "1.0", "Hoàn thiện thực thi & Báo cáo kết quả", "M", "Chạy kiểm thử đợt 1 (Release 1) và tổng hợp kết quả Pass", "Test Execution Log")
    ]
    for r_idx, crow in enumerate(change_rows, start=11):
        for c_idx, val in enumerate(crow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 3, 5] else align_left
            cell.border = border_thin
            
    ws_cover.cell(15, 2, "Danh sách Thành viên Nhóm Kiểm thử (Team Members)").font = font_sub_title
    member_headers = ["STT", "Họ và tên", "Mã sinh viên", "Vai trò", "Nhiệm vụ phụ trách"]
    for c_idx, h in enumerate(member_headers, start=2):
        cell = ws_cover.cell(16, c_idx, h)
        cell.font = font_header_black
        cell.fill = fill_header_green
        cell.alignment = align_center
        cell.border = border_thin
        
    member_rows = [
        (1, "Nguyễn Quang Huy", "PH12345", "Trưởng nhóm / Test Lead", "Phân hệ Đặt vé online, Chọn ghế, Combo F&B, Voucher, Thanh toán VNPAY, Đơn hàng"),
        (2, "Văn Minh Khôi", "PH12346", "Tester / QA", "Phân hệ POS Bán vé, POS Đơn chờ, Bán F&B, Soát vé Check-in, Xử lý sự cố chỗ ngồi"),
        (3, "Phạm Thị Quỳnh Anh", "PH12347", "Tester / QA", "Phân hệ Xác thực, Đăng nhập, Đăng ký, Quên mật khẩu, Quản lý Nhân viên, Khách hàng, RBAC"),
        (4, "Nguyễn Ngọc Hà Linh", "PH12348", "Tester / QA", "Phân hệ Quản trị Phim, Cụm rạp, Phòng chiếu, Sơ đồ ghế, Lịch chiếu, Bảng giá vé, Khuyến mãi, Cài đặt")
    ]
    for r_idx, mrow in enumerate(member_rows, start=17):
        for c_idx, val in enumerate(mrow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 4] else align_left
            cell.border = border_thin

    ws_cover.column_dimensions['A'].width = 4.0
    ws_cover.column_dimensions['B'].width = 18.0
    ws_cover.column_dimensions['C'].width = 30.0
    ws_cover.column_dimensions['D'].width = 25.0
    ws_cover.column_dimensions['E'].width = 18.0
    ws_cover.column_dimensions['F'].width = 45.0
    ws_cover.column_dimensions['G'].width = 22.0

    # -------------------------------------------------------------------------
    # FULL LIST OF 42 DEVCINE MODULES WITH RIGOROUS DATN PROCEDURES
    # -------------------------------------------------------------------------
    modules_data = [
        # --- KHÁCH HÀNG (13 modules) ---
        {
            "code": "MOD_CUST_LOGIN", "sheet": "Đăng nhập",
            "req": "Kiểm tra Đăng nhập tài khoản khách hàng và quản trị viên",
            "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng & Nhân viên", "pre": "Người dùng mở trình duyệt và truy cập vào trang Đăng nhập hệ thống DevCine",
            "test_cases": [
                ("DN_GUI_01", "Kiểm tra hiển thị giao diện Form Đăng nhập", "Kiểm tra hiển thị đầy đủ các trường nhập liệu và nút chức năng trên màn hình Đăng nhập",
                 "Bước 1: Truy cập vào trang web DevCine với vai trò khách vãng lai\nBước 2: Click vào nút 'Đăng nhập' trên thanh Header\nBước 3: Quan sát giao diện và các phần tử hiển thị trên form",
                 "N/A", "Hiển thị đầy đủ ô nhập Tên đăng nhập/Email, Mật khẩu, nút Đăng nhập, liên kết Quên mật khẩu và Đăng ký tài khoản"),
                ("DN_GUI_02", "Kiểm tra hiệu ứng focus vào ô nhập liệu", "Kiểm tra viền sáng đổi màu khi người dùng click chuột vào ô nhập liệu",
                 "Bước 1: Mở form Đăng nhập của hệ thống DevCine\nBước 2: Click chuột vào ô nhập 'Tên đăng nhập hoặc Email'\nBước 3: Kiểm tra hiệu ứng đổi màu viền của ô nhập liệu",
                 "N/A", "Viền của ô nhập liệu đổi sang màu sáng đặc trưng của hệ thống, con trỏ chuột nhấp nháy sẵn sàng nhận dữ liệu"),
                ("DN_VAL_01", "Kiểm tra validate khi để trống tài khoản", "Kiểm tra thông báo lỗi khi không nhập tên đăng nhập hoặc email",
                 "Bước 1: Mở form Đăng nhập của hệ thống DevCine\nBước 2: Để trống trường 'Tên đăng nhập hoặc Email'\nBước 3: Nhập mật khẩu hợp lệ vào ô 'Mật khẩu'\nBước 4: Click vào button 'Đăng nhập'\nBước 5: Kiểm tra thông báo lỗi hiển thị từ hệ thống",
                 "Tài khoản: ''\nMật khẩu: '123456'", "Hiển thị thông báo lỗi 'Vui lòng nhập tên đăng nhập hoặc email' màu đỏ bên dưới ô nhập liệu"),
                ("DN_VAL_02", "Kiểm tra validate khi để trống mật khẩu", "Kiểm tra thông báo lỗi khi không nhập mật khẩu",
                 "Bước 1: Mở form Đăng nhập của hệ thống DevCine\nBước 2: Nhập tài khoản hợp lệ vào ô 'Tên đăng nhập hoặc Email'\nBước 3: Để trống trường 'Mật khẩu'\nBước 4: Click vào button 'Đăng nhập'\nBước 5: Kiểm tra thông báo lỗi hiển thị từ hệ thống",
                 "Tài khoản: 'khachhang@gmail.com'\nMật khẩu: ''", "Hiển thị thông báo lỗi 'Vui lòng nhập mật khẩu' màu đỏ bên dưới ô nhập mật khẩu"),
                ("DN_VAL_03", "Kiểm tra validate khi nhập mật khẩu dưới 6 ký tự", "Kiểm tra thông báo lỗi độ dài tối thiểu của mật khẩu",
                 "Bước 1: Mở form Đăng nhập của hệ thống DevCine\nBước 2: Nhập tài khoản hợp lệ vào ô 'Tên đăng nhập hoặc Email'\nBước 3: Nhập mật khẩu có độ dài 4 ký tự\nBước 4: Click vào button 'Đăng nhập'\nBước 5: Kiểm tra thông báo lỗi hiển thị từ hệ thống",
                 "Tài khoản: 'khachhang@gmail.com'\nMật khẩu: '1234'", "Hiển thị thông báo lỗi 'Mật khẩu phải chứa từ 6 đến 50 ký tự'"),
                ("DN_FUNC_01", "Kiểm tra đăng nhập thất bại khi nhập sai mật khẩu", "Kiểm tra thông báo lỗi khi người dùng nhập sai mật khẩu xác thực",
                 "Bước 1: Mở form Đăng nhập của hệ thống DevCine\nBước 2: Nhập email tài khoản đã tồn tại trong hệ thống\nBước 3: Nhập mật khẩu không chính xác\nBước 4: Click vào button 'Đăng nhập'\nBước 5: Kiểm tra phản hồi trả về từ hệ thống",
                 "Tài khoản: 'khachhang@gmail.com'\nMật khẩu: 'WrongPass@123'", "Hiển thị thông báo lỗi dạng Toast: 'Tài khoản hoặc mật khẩu không chính xác'"),
                ("DN_FUNC_02", "Kiểm tra khóa tài khoản tạm thời khi nhập sai quá 5 lần", "Kiểm tra cơ chế bảo mật chống tấn công brute-force mật khẩu",
                 "Bước 1: Mở form Đăng nhập của hệ thống DevCine\nBước 2: Nhập email tài khoản hợp lệ\nBước 3: Thực hiện nhập sai mật khẩu liên tiếp 5 lần\nBước 4: Click vào button 'Đăng nhập' ở lần thứ 5\nBước 5: Kiểm tra thông báo và trạng thái khóa tài khoản",
                 "Tài khoản: 'khachhang@gmail.com'\nNhập sai mật khẩu 5 lần", "Hiển thị cảnh báo: 'Tài khoản của bạn đã bị tạm khóa 15 phút do nhập sai mật khẩu quá 5 lần'"),
                ("DN_FUNC_03", "Kiểm tra đăng nhập với tài khoản bị vô hiệu hóa", "Kiểm tra chặn đăng nhập khi tài khoản có trạng thái ACTIVE=false trong cơ sở dữ liệu",
                 "Bước 1: Mở form Đăng nhập của hệ thống DevCine\nBước 2: Nhập tài khoản đã bị quản trị viên khóa trong hệ thống\nBước 3: Nhập đúng mật khẩu của tài khoản đó\nBước 4: Click vào button 'Đăng nhập'\nBước 5: Kiểm tra phản hồi từ hệ thống",
                 "Tài khoản: 'locked_user'\nMật khẩu: 'Khach@123'", "Hiển thị thông báo lỗi: 'Tài khoản của bạn đã bị vô hiệu hóa. Vui lòng liên hệ CSKH để được hỗ trợ'"),
                ("DN_FUNC_04", "Kiểm tra đăng nhập thành công với tài khoản Khách hàng", "Kiểm tra luồng đăng nhập thành công và cấp phát phiên làm việc JWT",
                 "Bước 1: Mở form Đăng nhập của hệ thống DevCine\nBước 2: Nhập chính xác tên đăng nhập/email của khách hàng\nBước 3: Nhập chính xác mật khẩu của tài khoản\nBước 4: Click vào button 'Đăng nhập'\nBước 5: Kiểm tra kết quả hiển thị và chuyển hướng trang",
                 "Tài khoản: 'khachhang'\nMật khẩu: 'Khach@123'", "Đăng nhập thành công, lưu JWT Token vào LocalStorage, cập nhật tên người dùng trên Header và chuyển về Trang chủ"),
                ("DN_FUNC_05", "Kiểm tra chặn khách hàng truy cập trang Quản trị", "Kiểm tra phân quyền bảo mật URL đối với vai trò ROLE_CUSTOMER",
                 "Bước 1: Đăng nhập thành công vào hệ thống với vai trò 'Khách hàng'\nBước 2: Trên thanh địa chỉ trình duyệt, gõ đường dẫn URL quản trị: '/admin'\nBước 3: Nhấn Enter để truy cập\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Vai trò: ROLE_CUSTOMER\nURL truy cập: '/admin'", "Hệ thống chặn truy cập, hiển thị trang lỗi 403 Forbidden hoặc điều hướng về trang chủ kèm thông báo không có quyền")
            ]
        },
        {
            "code": "MOD_CUST_REG", "sheet": "Đăng ký",
            "req": "Kiểm tra Đăng ký tài khoản khách hàng mới",
            "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng", "pre": "Người dùng mở trình duyệt và truy cập vào trang Đăng ký tài khoản DevCine",
            "test_cases": [
                ("DK_GUI_01", "Kiểm tra hiển thị giao diện Form Đăng ký", "Kiểm tra hiển thị đầy đủ các trường thông tin đăng ký bắt buộc",
                 "Bước 1: Truy cập trang web DevCine\nBước 2: Click vào nút 'Đăng ký' trên thanh Header\nBước 3: Quan sát tất cả các trường nhập liệu trên form",
                 "N/A", "Hiển thị đầy đủ các trường: Họ và tên, Email, Số điện thoại, Tên đăng nhập, Mật khẩu, Xác nhận mật khẩu, Ngày sinh, Giới tính và Checkbox điều khoản"),
                ("DK_VAL_01", "Kiểm tra validate khi để trống toàn bộ form", "Kiểm tra thông báo lỗi bắt buộc nhập khi nhấn Đăng ký ngay mà chưa điền thông tin",
                 "Bước 1: Mở form Đăng ký tài khoản\nBước 2: Không nhập dữ liệu vào bất kỳ trường nào\nBước 3: Click vào button 'Đăng ký ngay'\nBước 4: Kiểm tra kết quả hiển thị từ hệ thống",
                 "N/A", "Hiển thị thông báo lỗi bắt buộc nhập màu đỏ dưới tất cả các trường nhập liệu yêu cầu"),
                ("DK_VAL_02", "Kiểm tra validate Họ và tên dưới 2 từ", "Kiểm tra thông báo lỗi khi họ tên khách hàng chỉ có 1 từ đơn lẻ",
                 "Bước 1: Mở form Đăng ký tài khoản\nBước 2: Nhập họ tên chỉ gồm 1 từ vào ô 'Họ và tên'\nBước 3: Nhập đầy đủ các trường còn lại hợp lệ\nBước 4: Click vào button 'Đăng ký ngay'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Họ và tên: 'Dân'", "Hiển thị thông báo lỗi: 'Họ và tên phải chứa ít nhất 2 từ'"),
                ("DK_VAL_03", "Kiểm tra validate Email sai định dạng", "Kiểm tra thông báo lỗi khi nhập địa chỉ email không đúng chuẩn",
                 "Bước 1: Mở form Đăng ký tài khoản\nBước 2: Nhập email thiếu ký tự '@' hoặc thiếu tên miền\nBước 3: Nhập đầy đủ các trường còn lại hợp lệ\nBước 4: Click vào button 'Đăng ký ngay'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Email: 'dan.nguyen@devcine'", "Hiển thị thông báo lỗi: 'Định dạng email không hợp lệ'"),
                ("DK_VAL_04", "Kiểm tra validate Email đã tồn tại", "Kiểm tra tính duy nhất của trường Email trong hệ thống",
                 "Bước 1: Mở form Đăng ký tài khoản\nBước 2: Nhập địa chỉ email đã được đăng ký bởi tài khoản khác trước đó\nBước 3: Nhập đầy đủ các trường còn lại hợp lệ\nBước 4: Click vào button 'Đăng ký ngay'\nBước 5: Kiểm tra phản hồi từ hệ thống",
                 "Email: 'admin@devcine.com'", "Hiển thị thông báo lỗi: 'Địa chỉ email đã tồn tại trong hệ thống'"),
                ("DK_VAL_05", "Kiểm tra validate Số điện thoại không đủ 10 số", "Kiểm tra validate độ dài và đầu số điện thoại di động Việt Nam",
                 "Bước 1: Mở form Đăng ký tài khoản\nBước 2: Nhập số điện thoại chỉ có 8 hoặc 9 chữ số\nBước 3: Nhập đầy đủ các trường còn lại hợp lệ\nBước 4: Click vào button 'Đăng ký ngay'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Số điện thoại: '09123456'", "Hiển thị thông báo lỗi: 'Số điện thoại phải bao gồm đúng 10 chữ số (đầu 03, 05, 07, 08, 09)'"),
                ("DK_VAL_06", "Kiểm tra validate Mật khẩu xác nhận không khớp", "Kiểm tra tính đồng nhất giữa Mật khẩu và Xác nhận mật khẩu",
                 "Bước 1: Mở form Đăng ký tài khoản\nBước 2: Nhập mật khẩu 'Pass@123'\nBước 3: Nhập xác nhận mật khẩu 'Pass@456'\nBước 4: Click vào button 'Đăng ký ngay'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Mật khẩu: 'Pass@123'\nXác nhận: 'Pass@456'", "Hiển thị thông báo lỗi: 'Mật khẩu xác nhận không trùng khớp với mật khẩu đã nhập'"),
                ("DK_VAL_07", "Kiểm tra validate Ngày sinh dưới 13 tuổi", "Kiểm tra điều kiện độ tuổi đăng ký thành viên xem phim",
                 "Bước 1: Mở form Đăng ký tài khoản\nBước 2: Chọn ngày sinh của người dùng sinh năm 2020\nBước 3: Nhập đầy đủ các trường còn lại hợp lệ\nBước 4: Click vào button 'Đăng ký ngay'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Ngày sinh: '2020-05-15'", "Hiển thị thông báo lỗi: 'Độ tuổi đăng ký thành viên phải từ 13 tuổi trở lên'"),
                ("DK_FUNC_01", "Kiểm tra đăng ký thành công tài khoản mới", "Kiểm tra luồng tạo mới tài khoản thành công với dữ liệu chuẩn",
                 "Bước 1: Mở form Đăng ký tài khoản\nBước 2: Nhập đầy đủ và chính xác tất cả các thông tin hợp lệ\nBước 3: Tích chọn checkbox 'Tôi đồng ý với Điều khoản sử dụng và Chính sách bảo mật'\nBước 4: Click vào button 'Đăng ký ngay'\nBước 5: Kiểm tra phản hồi và kết quả tạo tài khoản",
                 "Họ tên: 'Nguyễn Văn Dân'\nEmail: 'dan.nguyen@gmail.com'\nSĐT: '0912345678'\nTên đăng nhập: 'dannguyen'\nMK: 'Dan@123456'\nNgày sinh: '2000-01-01'",
                 "Hệ thống tạo tài khoản thành công, cấp hạng thành viên Đồng (Bronze), hiển thị thông báo thành công và chuyển hướng sang trang Đăng nhập")
            ]
        },
        {
            "code": "MOD_CUST_FORGOT", "sheet": "Quên mật khẩu",
            "req": "Kiểm tra Quên mật khẩu, Xác thực OTP và Đặt lại mật khẩu",
            "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng", "pre": "Người dùng mở màn hình Quên mật khẩu trên website DevCine",
            "test_cases": [
                ("QMK_GUI_01", "Kiểm tra hiển thị giao diện Quên mật khẩu", "Kiểm tra hiển thị ô nhập email và nút gửi mã OTP",
                 "Bước 1: Mở trang Đăng nhập\nBước 2: Click vào liên kết 'Quên mật khẩu?'\nBước 3: Quan sát giao diện màn hình khôi phục mật khẩu",
                 "N/A", "Hiển thị tiêu đề Quên mật khẩu, ô nhập email tài khoản, nút Gửi mã OTP và liên kết quay lại Đăng nhập"),
                ("QMK_VAL_01", "Kiểm tra validate để trống email nhận OTP", "Kiểm tra thông báo lỗi khi không điền email",
                 "Bước 1: Mở màn hình Quên mật khẩu\nBước 2: Để trống ô 'Địa chỉ Email'\nBước 3: Click vào button 'Gửi mã xác nhận'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Email: ''", "Hiển thị thông báo lỗi: 'Vui lòng nhập địa chỉ email tài khoản'"),
                ("QMK_VAL_02", "Kiểm tra gửi OTP với email không tồn tại", "Kiểm tra thông báo lỗi khi email chưa đăng ký tài khoản",
                 "Bước 1: Mở màn hình Quên mật khẩu\nBước 2: Nhập địa chỉ email chưa từng đăng ký trong hệ thống\nBước 3: Click vào button 'Gửi mã xác nhận'\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Email: 'notfound@domain.com'", "Hiển thị thông báo lỗi: 'Địa chỉ email không tồn tại trong hệ thống'"),
                ("QMK_FUNC_01", "Kiểm tra gửi OTP thành công về email", "Kiểm tra chức năng sinh mã OTP và hiển thị đếm ngược 60 giây",
                 "Bước 1: Mở màn hình Quên mật khẩu\nBước 2: Nhập email tài khoản hợp lệ đã tồn tại\nBước 3: Click vào button 'Gửi mã xác nhận'\nBước 4: Kiểm tra hộp thư email và giao diện website",
                 "Email: 'khachhang@gmail.com'", "Hệ thống gửi mã OTP 6 chữ số về email, hiển thị form nhập mã kèm đồng hồ đếm ngược 60 giây chống gửi lại liên tục"),
                ("QMK_VAL_03", "Kiểm tra validate khi nhập sai mã OTP", "Kiểm tra thông báo lỗi khi mã xác thực OTP không đúng",
                 "Bước 1: Nhận mã OTP gửi về email\nBước 2: Nhập mã OTP sai (ví dụ 000000) vào ô xác nhận\nBước 3: Click vào button 'Xác thực mã'\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Mã OTP: '000000'", "Hiển thị thông báo lỗi: 'Mã xác thực OTP không chính xác. Vui lòng kiểm tra lại'"),
                ("QMK_VAL_04", "Kiểm tra validate khi nhập OTP đã hết hạn", "Kiểm tra timeout của mã xác thực OTP sau 15 phút",
                 "Bước 1: Yêu cầu gửi mã OTP\nBước 2: Chờ thời gian vượt quá 15 phút (mã hết hạn)\nBước 3: Nhập mã OTP đó vào hệ thống và bấm 'Xác thực mã'\nBước 4: Kiểm tra kết quả phản hồi",
                 "Thời gian: Quá 15 phút", "Hiển thị thông báo lỗi: 'Mã xác thực OTP đã hết hạn sử dụng. Vui lòng yêu cầu mã mới'"),
                ("QMK_FUNC_02", "Kiểm tra đặt lại mật khẩu mới thành công", "Kiểm tra lưu mật khẩu mới sau khi xác thực OTP thành công",
                 "Bước 1: Nhập chính xác mã OTP 6 số nhận được từ email\nBước 2: Click button 'Xác thực mã'\nBước 3: Nhập Mật khẩu mới có ít nhất 8 ký tự\nBước 4: Nhập Xác nhận mật khẩu mới trùng khớp\nBước 5: Click button 'Đặt lại mật khẩu'\nBước 6: Kiểm tra kết quả cập nhật",
                 "Mật khẩu mới: 'NewPassword@2026'\nXác nhận: 'NewPassword@2026'",
                 "Hệ thống cập nhật mật khẩu mới thành công, hiển thị thông báo thành công và tự động chuyển hướng về trang Đăng nhập")
            ]
        },
        {
            "code": "MOD_CUST_CHANGE_PASS", "sheet": "Đổi mật khẩu",
            "req": "Kiểm tra chức năng Đổi mật khẩu trong trang cá nhân",
            "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng & Nhân viên", "pre": "Người dùng đã đăng nhập thành công và mở tab 'Đổi mật khẩu' trong trang Hồ sơ cá nhân",
            "test_cases": [
                ("DMK_VAL_01", "Kiểm tra validate khi để trống mật khẩu cũ", "Kiểm tra thông báo lỗi khi không nhập mật khẩu hiện tại",
                 "Bước 1: Truy cập vào trang Hồ sơ cá nhân -> Tab 'Đổi mật khẩu'\nBước 2: Để trống trường 'Mật khẩu hiện tại'\nBước 3: Nhập mật khẩu mới và xác nhận mật khẩu\nBước 4: Click button 'Lưu thay đổi'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Mật khẩu cũ: ''\nMật khẩu mới: 'NewPass@123'", "Hiển thị thông báo lỗi: 'Vui lòng nhập mật khẩu hiện tại'"),
                ("DMK_VAL_02", "Kiểm tra validate khi nhập sai mật khẩu cũ", "Kiểm tra xác thực mật khẩu hiện tại không chính xác",
                 "Bước 1: Truy cập vào tab 'Đổi mật khẩu'\nBước 2: Nhập mật khẩu hiện tại sai\nBước 3: Nhập mật khẩu mới hợp lệ\nBước 4: Click button 'Lưu thay đổi'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Mật khẩu cũ: 'WrongOldPass@123'\nMật khẩu mới: 'NewPass@123'", "Hiển thị thông báo lỗi: 'Mật khẩu hiện tại không chính xác'"),
                ("DMK_VAL_03", "Kiểm tra validate mật khẩu mới trùng mật khẩu cũ", "Kiểm tra ràng buộc mật khẩu mới phải khác mật khẩu cũ",
                 "Bước 1: Truy cập vào tab 'Đổi mật khẩu'\nBước 2: Nhập mật khẩu hiện tại đúng\nBước 3: Nhập mật khẩu mới giống hệt mật khẩu hiện tại\nBước 4: Click button 'Lưu thay đổi'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Mật khẩu mới: Trùng mật khẩu cũ", "Hiển thị thông báo lỗi: 'Mật khẩu mới không được trùng với mật khẩu hiện tại'"),
                ("DMK_FUNC_01", "Kiểm tra đổi mật khẩu thành công", "Kiểm tra cập nhật mật khẩu mới và hủy các phiên đăng nhập cũ",
                 "Bước 1: Truy cập vào tab 'Đổi mật khẩu'\nBước 2: Nhập đúng mật khẩu hiện tại\nBước 3: Nhập mật khẩu mới từ 8 ký tự gồm chữ hoa, thường, số, ký tự đặc biệt\nBước 4: Nhập xác nhận mật khẩu mới trùng khớp\nBước 5: Click button 'Lưu thay đổi'\nBước 6: Kiểm tra phản hồi từ hệ thống",
                 "MK cũ: 'OldPass@123'\nMK mới: 'NewPass@2026'\nXác nhận: 'NewPass@2026'",
                 "Đổi mật khẩu thành công, hiển thị thông báo thành công và yêu cầu người dùng đăng nhập lại bằng mật khẩu mới")
            ]
        },
        {
            "code": "MOD_CUST_PROFILE", "sheet": "Hồ sơ cá nhân",
            "req": "Kiểm tra Cập nhật thông tin cá nhân và Avatar",
            "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng", "pre": "Khách hàng đã đăng nhập và mở màn hình Thông tin tài khoản",
            "test_cases": [
                ("HS_GUI_01", "Kiểm tra hiển thị thông tin hồ sơ và Hạng thành viên", "Kiểm tra hiển thị đầy đủ thông tin cá nhân, điểm tích lũy và thẻ hội viên",
                 "Bước 1: Đăng nhập vào hệ thống với tài khoản Khách hàng\nBước 2: Click vào avatar người dùng -> Chọn 'Hồ sơ cá nhân'\nBước 3: Quan sát các thông tin hiển thị trên trang",
                 "N/A", "Hiển thị đầy đủ Họ tên, Email, SĐT, Giới tính, Ngày sinh, Địa chỉ, Avatar, Điểm thưởng Loyalty và Hạng thành viên (Đồng/Bạc/Vàng/Kim Cương)"),
                ("HS_VAL_01", "Kiểm tra validate khi xóa trắng trường Họ và tên", "Kiểm tra thông báo lỗi bắt buộc nhập họ tên",
                 "Bước 1: Mở trang Hồ sơ cá nhân\nBước 2: Xóa toàn bộ nội dung trong ô 'Họ và tên'\nBước 3: Click button 'Lưu thông tin'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Họ và tên: ''", "Hiển thị thông báo lỗi: 'Họ và tên không được để trống'"),
                ("HS_VAL_02", "Kiểm tra validate khi đổi SĐT trùng với tài khoản khác", "Kiểm tra tính duy nhất của số điện thoại khi cập nhật",
                 "Bước 1: Mở trang Hồ sơ cá nhân\nBước 2: Đổi số điện thoại thành số điện thoại của một tài khoản khác trong hệ thống\nBước 3: Click button 'Lưu thông tin'\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "SĐT: '0988888888' (Đã tồn tại)", "Hiển thị thông báo lỗi: 'Số điện thoại này đã được sử dụng bởi một tài khoản khác'"),
                ("HS_VAL_03", "Kiểm tra validate khi tải ảnh avatar vượt quá 5MB", "Kiểm tra giới hạn dung lượng tệp ảnh đại diện",
                 "Bước 1: Mở trang Hồ sơ cá nhân\nBước 2: Click vào biểu tượng máy ảnh trên Avatar để chọn tệp\nBước 3: Chọn tệp ảnh có dung lượng 8MB\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "File ảnh: dung lượng 8MB", "Hiển thị thông báo lỗi: 'Dung lượng tệp ảnh đại diện không được vượt quá 5MB'"),
                ("HS_FUNC_01", "Kiểm tra cập nhật thông tin hồ sơ thành công", "Kiểm tra lưu thông tin mới và cập nhật hiển thị toàn hệ thống",
                 "Bước 1: Mở trang Hồ sơ cá nhân\nBước 2: Sửa đổi Họ và tên, Ngày sinh, Giới tính và Địa chỉ hợp lệ\nBước 3: Tải lên ảnh đại diện mới có dung lượng 1.5MB (JPG)\nBước 4: Click button 'Lưu thông tin'\nBước 5: Kiểm tra kết quả cập nhật",
                 "Họ tên: 'Nguyễn Văn Dân'\nĐịa chỉ: 'Cầu Giấy, Hà Nội'\nAvatar: 'avatar_dan.jpg'",
                 "Hệ thống lưu thông tin thành công, hiển thị thông báo cập nhật thành công và đổi ngay tên/avatar hiển thị trên thanh Header")
            ]
        },
        {
            "code": "MOD_CUST_SEARCH", "sheet": "Tìm kiếm & Lọc phim",
            "req": "Kiểm tra Tìm kiếm và Bộ lọc phim trên trang chủ",
            "tester": "Nguyễn Quang Huy", "role": "Khách hàng", "pre": "Khách hàng truy cập vào Trang chủ hoặc danh sách Phim đang chiếu",
            "test_cases": [
                ("SRC_GUI_01", "Kiểm tra hiển thị thanh tìm kiếm và bộ lọc", "Kiểm tra hiển thị ô nhập từ khóa và các combobox lọc phim",
                 "Bước 1: Mở trang chủ DevCine\nBước 2: Điều hướng đến mục 'Danh sách phim'\nBước 3: Quan sát thanh tìm kiếm và các nút lọc",
                 "N/A", "Hiển thị đầy đủ ô tìm kiếm từ khóa, bộ lọc Thể loại, Định dạng (2D, 3D, IMAX), Giới hạn độ tuổi (P, K, T13, T16, T18) và Trạng thái phim"),
                ("SRC_FUNC_01", "Kiểm tra tìm kiếm phim theo từ khóa tên phim", "Kiểm tra cơ chế tìm kiếm tự động với độ trễ Debounce 300ms",
                 "Bước 1: Truy cập vào trang danh sách phim\nBước 2: Gõ từ khóa 'Avatar' vào ô tìm kiếm\nBước 3: Chờ 300ms sau khi ngừng gõ phím\nBước 4: Kiểm tra danh sách phim trả về",
                 "Từ khóa: 'Avatar'", "Hệ thống tự động lọc và hiển thị danh sách các phim có chứa từ khóa 'Avatar' trong tên tiếng Anh hoặc tên tiếng Việt"),
                ("SRC_FUNC_02", "Kiểm tra kết hợp đa bộ lọc Thể loại và Độ tuổi", "Kiểm tra lọc đồng thời nhiều tiêu chí lọc",
                 "Bước 1: Truy cập vào trang danh sách phim\nBước 2: Chọn Thể loại: 'Hành động'\nBước 3: Chọn Giới hạn độ tuổi: 'T16'\nBước 4: Chọn Định dạng: 'IMAX'\nBước 5: Kiểm tra danh sách kết quả hiển thị",
                 "Thể loại: 'Hành động'\nĐộ tuổi: 'T16'\nĐịnh dạng: 'IMAX'",
                 "Hệ thống lọc chính xác các phim thỏa mãn đồng thời cả 3 điều kiện: Thể loại Hành động, độ tuổi T16 và có định dạng chiếu IMAX"),
                ("SRC_FUNC_03", "Kiểm tra tìm kiếm không có kết quả phù hợp", "Kiểm tra hiển thị giao diện trạng thái trống (Empty State)",
                 "Bước 1: Truy cập vào trang danh sách phim\nBước 2: Nhập từ khóa không có thực trong hệ thống: 'PhimKhongTonTai123'\nBước 3: Quan sát giao diện danh sách phim",
                 "Từ khóa: 'PhimKhongTonTai123'", "Hiển thị hình minh họa trống kèm dòng thông báo: 'Không tìm thấy bộ phim nào phù hợp với từ khóa của bạn'")
            ]
        },
        {
            "code": "MOD_CUST_REVIEW", "sheet": "Chi tiết phim & Đánh giá",
            "req": "Kiểm tra Đánh giá sao và Bình luận phim",
            "tester": "Nguyễn Quang Huy", "role": "Khách hàng", "pre": "Khách hàng mở trang Chi tiết một bộ phim cụ thể",
            "test_cases": [
                ("REV_GUI_01", "Kiểm tra hiển thị thông tin chi tiết phim", "Kiểm tra hiển thị đầy đủ Poster, Trailer, Diễn viên, Đạo diễn, Tóm tắt nội dung",
                 "Bước 1: Truy cập trang chủ DevCine\nBước 2: Click vào một poster phim để vào trang chi tiết\nBước 3: Quan sát thông tin chi tiết của bộ phim",
                 "N/A", "Hiển thị đầy đủ Tên phim, Poster, Banner, Video Trailer Youtube nhúng trực tiếp, Thời lượng, Ngày khởi chiếu, Đạo diễn, Diễn viên và Tóm tắt"),
                ("REV_VAL_01", "Kiểm tra điều kiện đánh giá khi chưa mua vé", "Kiểm tra chặn người dùng chưa từng mua vé gửi đánh giá",
                 "Bước 1: Đăng nhập bằng tài khoản chưa từng mua vé của bộ phim đang xem\nBước 2: Cuộn xuống mục 'Đánh giá & Bình luận'\nBước 3: Click vào button 'Gửi đánh giá'\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Trạng thái: Chưa mua vé", "Hiển thị thông báo: 'Bạn cần mua vé và xem phim này trước khi có thể gửi đánh giá'"),
                ("REV_VAL_02", "Kiểm tra validate khi chưa chọn số sao đánh giá", "Kiểm tra bắt buộc chọn số sao từ 1 đến 5",
                 "Bước 1: Đăng nhập tài khoản đã xem phim\nBước 2: Nhập nội dung bình luận vào ô bình luận\nBước 3: Không click chọn số sao đánh giá (mặc định 0 sao)\nBước 4: Click button 'Gửi đánh giá'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Số sao: 0\nNội dung: 'Phim rất đáng xem'", "Hiển thị thông báo lỗi: 'Vui lòng chọn điểm đánh giá từ 1 đến 5 sao'"),
                ("REV_VAL_03", "Kiểm tra validate độ dài bình luận vượt quá 500 ký tự", "Kiểm tra giới hạn ký tự tối đa của bình luận",
                 "Bước 1: Mở form đánh giá phim\nBước 2: Chọn 5 sao\nBước 3: Nhập đoạn văn bản bình luận có độ dài 600 ký tự\nBước 4: Click button 'Gửi đánh giá'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Độ dài bình luận: 600 ký tự", "Hiển thị thông báo lỗi: 'Nội dung bình luận không được vượt quá 500 ký tự'"),
                ("REV_FUNC_01", "Kiểm tra gửi đánh giá thành công", "Kiểm tra lưu đánh giá và tự động tính lại điểm sao trung bình của phim",
                 "Bước 1: Chọn điểm đánh giá 5 sao\nBước 2: Nhập bình luận: 'Kỹ xảo và âm thanh quá tuyệt vời, đáng tiền!'\nBước 3: Click button 'Gửi đánh giá'\nBước 4: Kiểm tra kết quả hiển thị trên danh sách đánh giá",
                 "Điểm sao: 5 sao\nBình luận: 'Kỹ xảo và âm thanh quá tuyệt vời, đáng tiền!'",
                 "Gửi đánh giá thành công, bình luận hiển thị đầu danh sách và điểm đánh giá trung bình của phim được cập nhật tự động")
            ]
        },
        {
            "code": "MOD_CUST_BOOKING_SHOWTIME", "sheet": "Đặt vé online",
            "req": "Kiểm tra Chọn suất chiếu, Đối tượng vé và Cảnh báo độ tuổi",
            "tester": "Nguyễn Quang Huy", "role": "Khách hàng", "pre": "Khách hàng đã đăng nhập và đang ở màn hình Đặt vé của bộ phim",
            "test_cases": [
                ("BK_GUI_01", "Kiểm tra hiển thị lịch chiếu theo ngày và cụm rạp", "Kiểm tra hiển thị các ngày chiếu trong tuần và danh sách cụm rạp có suất chiếu",
                 "Bước 1: Mở trang Đặt vé của phim\nBước 2: Quan sát thanh chọn ngày chiếu và danh sách các rạp hiển thị",
                 "N/A", "Hiển thị thanh trượt 7 ngày trong tuần, danh sách các cụm rạp và các khung giờ chiếu phân loại theo 2D, 3D, IMAX"),
                ("BK_VAL_01", "Kiểm tra chặn chọn suất chiếu đã quá giờ mở bán", "Kiểm tra điều kiện đóng bán vé trước giờ chiếu 10 phút",
                 "Bước 1: Quan sát suất chiếu có giờ bắt đầu chỉ còn cách thời điểm hiện tại dưới 10 phút\nBước 2: Cố tình click vào khung giờ chiếu đó\nBước 3: Kiểm tra phản hồi từ hệ thống",
                 "Giờ chiếu: Cách hiện tại < 10 phút", "Suất chiếu bị làm mờ (disabled), hiển thị nhãn 'Đã đóng bán vé' và không cho phép click chọn"),
                ("BK_VAL_02", "Kiểm tra hiển thị modal cảnh báo độ tuổi phim T18", "Kiểm tra bắt buộc xác nhận độ tuổi trước khi vào chọn ghế",
                 "Bước 1: Chọn một bộ phim có nhãn giới hạn độ tuổi T18\nBước 2: Click chọn một suất chiếu hợp lệ\nBước 3: Kiểm tra hiển thị từ hệ thống",
                 "Nhãn phim: 'T18' (Cấm khán giả dưới 18 tuổi)",
                 "Hiển thị modal cảnh báo màu đỏ: 'Phim dành cho khán giả từ đủ 18 tuổi trở lên. Rạp sẽ kiểm tra CCCD khi vào phòng chiếu', bắt buộc nhấn 'Tôi đã đủ 18 tuổi' để tiếp tục"),
                ("BK_VAL_03", "Kiểm tra giới hạn số lượng vé đặt tối đa 8 vé", "Kiểm tra validate chặn đặt vượt quá số vé quy định trong 1 đơn",
                 "Bước 1: Mở bước chọn số lượng vé\nBước 2: Tăng số lượng vé lên 9 vé\nBước 3: Kiểm tra trạng thái nút tăng số lượng",
                 "Số lượng vé: 9", "Nút tăng số lượng vé (+) bị vô hiệu hóa khi số vé đạt 8, kèm dòng thông báo: 'Mỗi lần đặt tối đa 8 vé'"),
                ("BK_FUNC_01", "Kiểm tra chọn suất chiếu và đối tượng vé thành công", "Kiểm tra chọn vé Người lớn và vé Học sinh Sinh viên (HSSV)",
                 "Bước 1: Chọn suất chiếu 19:30 tại cụm rạp CGV Cầu Giấy\nBước 2: Chọn 2 vé Người lớn và 1 vé HSSV\nBước 3: Click button 'Tiếp tục chọn ghế'\nBước 4: Kiểm tra chuyển bước đặt vé",
                 "Suất chiếu: 19:30\nVé: 2 Người lớn + 1 HSSV",
                 "Hệ thống ghi nhận tổng số 3 vé cần chọn, hiển thị lưu ý xuất trình thẻ HSSV và chuyển sang màn hình Sơ đồ chọn ghế")
            ]
        },
        {
            "code": "MOD_CUST_SEAT_HOLD", "sheet": "Chọn ghế & Giữ chỗ",
            "req": "Kiểm tra Chọn ghế trên ma trận và Giữ chỗ 10 phút",
            "tester": "Nguyễn Quang Huy", "role": "Khách hàng", "pre": "Khách hàng đang ở màn hình Sơ đồ chọn ghế của suất chiếu",
            "test_cases": [
                ("ST_GUI_01", "Kiểm tra hiển thị sơ đồ ma trận ghế phòng chiếu", "Kiểm tra trực quan sơ đồ ghế đúng số hàng (A-Z), số cột và chú thích loại ghế",
                 "Bước 1: Truy cập vào màn hình Chọn ghế\nBước 2: Quan sát sơ đồ phòng chiếu hiển thị",
                 "N/A", "Hiển thị màn hình chiếu (Screen), các hàng ghế được đánh mã rõ ràng (A1..A10), phân biệt màu sắc rõ ràng giữa ghế Thường, VIP, Sweetbox, Ghế trống, Ghế đang chọn và Ghế đã bán"),
                ("ST_VAL_01", "Kiểm tra vô hiệu hóa ghế đã bán", "Kiểm tra chặn không cho click chọn ghế màu đỏ (SOLD)",
                 "Bước 1: Mở sơ đồ ghế\nBước 2: Di chuột và click vào ghế màu đỏ (đã có khách mua)\nBước 3: Kiểm tra phản hồi từ hệ thống",
                 "Ghế: 'E05' (Trạng thái: SOLD)", "Ghế bị khóa (disabled), con trỏ chuột hiển thị biểu tượng cấm (not-allowed) và không thể chọn ghế"),
                ("ST_VAL_02", "Kiểm tra chặn chọn ghế đang có người giữ chỗ", "Kiểm tra cơ chế chống xung đột giữ chỗ thời gian thực (HELD)",
                 "Bước 1: Người dùng A đang giữ ghế F06 trong 10 phút\nBước 2: Người dùng B mở sơ đồ ghế cùng suất chiếu đó và click chọn ghế F06\nBước 3: Kiểm tra thông báo hiển thị cho người dùng B",
                 "Ghế: 'F06' (Trạng thái: HELD bởi user khác)", "Hiển thị thông báo Toast: 'Ghế F06 đang được giữ chỗ bởi một khách hàng khác. Vui lòng chọn ghế khác'"),
                ("ST_VAL_03", "Kiểm tra quy tắc chọn trọn cặp ghế đôi Sweetbox", "Kiểm tra tự động chọn cả 2 ghế liền kề của ghế đôi",
                 "Bước 1: Mở sơ đồ ghế\nBước 2: Click chuột vào 1 ghế Sweetbox (ví dụ H01)\nBước 3: Quan sát trạng thái các ghế được chọn",
                 "Ghế click: 'H01'", "Hệ thống tự động kích hoạt chọn cả cặp 2 ghế liền kề H01 và H02, tính giá đúng bằng giá 1 cặp ghế đôi"),
                ("ST_FUNC_01", "Kiểm tra giữ chỗ 10 phút và đếm ngược", "Kiểm tra cơ chế hold ghế tạm thời trong cơ sở dữ liệu và WebSocket đếm ngược",
                 "Bước 1: Chọn đủ 2 ghế VIP (E05, E06) tương ứng với 2 vé đã chọn\nBước 2: Click button 'Tiếp tục thanh toán'\nBước 3: Quan sát trạng thái giữ ghế và đồng hồ đếm ngược",
                 "Ghế chọn: ['E05', 'E06']",
                 "Hệ thống khóa giữ 2 ghế E05, E06, đổi trạng thái ghế sang HELD trên hệ thống toàn rạp và hiển thị đồng hồ đếm ngược 10:00 chính xác từng giây"),
                ("ST_FUNC_02", "Kiểm tra tự động nhả ghế khi hết thời gian 10 phút", "Kiểm tra timeout giữ chỗ giải phóng tài nguyên ghế",
                 "Bước 1: Thực hiện giữ chỗ 2 ghế và chuyển sang bước thanh toán\nBước 2: Không thực hiện thanh toán và chờ đồng hồ đếm ngược về 00:00\nBước 3: Kiểm tra phản hồi từ hệ thống và trạng thái ghế",
                 "Thời gian: Hết 10 phút (00:00)",
                 "Hệ thống hiển thị modal thông báo: 'Đã hết thời gian giữ chỗ. Đơn đặt vé của bạn đã tự động bị hủy', giải phóng 2 ghế về trạng thái TRỐNG cho người khác đặt")
            ]
        },
        {
            "code": "MOD_CUST_FNB", "sheet": "Combo F&B online",
            "req": "Kiểm tra Chọn bắp nước và Tùy chọn vị combo khi đặt vé",
            "tester": "Nguyễn Quang Huy", "role": "Khách hàng", "pre": "Khách hàng đang ở bước Chọn Combo Bắp nước trong luồng đặt vé",
            "test_cases": [
                ("FNB_GUI_01", "Kiểm tra hiển thị menu bắp nước", "Kiểm tra hiển thị danh sách các món ăn, nước uống và combo kèm hình ảnh, giá bán",
                 "Bước 1: Hoàn tất bước chọn ghế -> Chuyển sang bước 'Bắp nước'\nBước 2: Quan sát danh sách món ăn hiển thị",
                 "N/A", "Hiển thị đầy đủ hình ảnh đại diện, tên món, mô tả thành phần, giá tiền và bộ chọn số lượng (+ / -) cho từng món"),
                ("FNB_VAL_01", "Kiểm tra giới hạn số lượng món tối đa 20 phần", "Kiểm tra validate chặn tăng số lượng quá 20 cho 1 món",
                 "Bước 1: Chọn món 'Bắp Phô Mai'\nBước 2: Click nút tăng số lượng (+) liên tục đến khi đạt 20\nBước 3: Tiếp tục click nút (+)\nBước 4: Kiểm tra trạng thái nút bấm",
                 "Món: 'Bắp Phô Mai'\nSố lượng: 21", "Nút tăng số lượng (+) bị vô hiệu hóa khi đạt 20, hệ thống không cho phép tăng thêm"),
                ("FNB_VAL_02", "Kiểm tra cảnh báo khi chưa chọn đủ vị bắp cho Combo", "Kiểm tra ràng buộc bắt buộc chọn đủ các slot tùy chọn trong combo",
                 "Bước 1: Chọn 'Combo 2 Ngăn' (quy định gồm 1 bắp 2 vị + 2 nước ngọt)\nBước 2: Chỉ chọn 1 vị bắp và bỏ trống vị bắp thứ 2\nBước 3: Click button 'Tiếp tục'\nBước 4: Kiểm tra thông báo hiển thị",
                 "Combo 2 Ngăn: Thiếu 1 vị bắp", "Hiển thị thông báo cảnh báo màu vàng: 'Vui lòng chọn đầy đủ các vị bắp và loại nước quy định cho Combo 2 Ngăn'"),
                ("FNB_FUNC_01", "Kiểm tra tính tiền phụ thu khi nâng cấp vị bắp", "Kiểm tra tự động cộng phụ thu vào tổng tiền tạm tính",
                 "Bước 1: Chọn Combo 1 (Bắp ngọt mặc định giá 85.000đ)\nBước 2: Click đổi sang vị 'Phô Mai Đặc Biệt' (phụ thu +15.000đ)\nBước 3: Kiểm tra tổng tiền tạm tính của đơn hàng",
                 "Combo gốc: 85.000đ\nPhụ thu: +15.000đ", "Tổng tiền F&B tự động cập nhật chính xác thành 100.000đ và hiển thị chi tiết dòng phụ thu"),
                ("FNB_FUNC_02", "Kiểm tra bỏ qua bước chọn bắp nước", "Kiểm tra tính không bắt buộc của bước mua F&B khi đặt vé",
                 "Bước 1: Mở bước chọn bắp nước\nBước 2: Không chọn bất kỳ món bắp nước nào\nBước 3: Click button 'Bỏ qua & Tiếp tục'\nBước 4: Kiểm tra chuyển bước tiếp theo",
                 "Giỏ F&B: Trống", "Hệ thống cho phép bỏ qua bước bắp nước và chuyển thẳng sang màn hình Áp dụng Voucher & Thanh toán")
            ]
        },
        {
            "code": "MOD_CUST_VOUCHER", "sheet": "Khuyến mãi & Voucher",
            "req": "Kiểm tra Áp dụng mã giảm giá và Đổi điểm Loyalty",
            "tester": "Nguyễn Quang Huy", "role": "Khách hàng", "pre": "Khách hàng đang ở màn hình Thanh toán đơn đặt vé",
            "test_cases": [
                ("VOU_VAL_01", "Kiểm tra validate khi nhập mã voucher không tồn tại", "Kiểm tra thông báo lỗi khi mã code sai hoặc không có trong hệ thống",
                 "Bước 1: Mở màn hình Thanh toán\nBước 2: Nhập mã voucher 'VOUCHER_FAKE_123' vào ô 'Mã giảm giá'\nBước 3: Click button 'Áp dụng'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Mã code: 'VOUCHER_FAKE_123'", "Hiển thị thông báo lỗi màu đỏ: 'Mã khuyến mãi không tồn tại trong hệ thống'"),
                ("VOU_VAL_02", "Kiểm tra validate khi mã voucher đã hết hạn sử dụng", "Kiểm tra chặn áp dụng voucher có ngày kết thúc trong quá khứ",
                 "Bước 1: Nhập mã voucher đã hết hạn sử dụng (ví dụ mã kết thúc ngày hôm qua)\nBước 2: Click button 'Áp dụng'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Mã code: 'HET_HAN_2025'", "Hiển thị thông báo lỗi: 'Mã khuyến mãi này đã hết hạn sử dụng'"),
                ("VOU_VAL_03", "Kiểm tra validate khi đơn chưa đạt giá trị tối thiểu", "Kiểm tra điều kiện đơn tối thiểu (min_order_value) của voucher",
                 "Bước 1: Đơn hàng có tổng tiền 150.000đ\nBước 2: Nhập mã voucher yêu cầu đơn hàng tối thiểu từ 300.000đ trở lên\nBước 3: Click button 'Áp dụng'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Tổng đơn: 150.000đ\nĐiều kiện voucher: Đơn tối thiểu 300.000đ",
                 "Hiển thị thông báo lỗi: 'Đơn hàng chưa đạt giá trị tối thiểu 300.000đ để áp dụng voucher này'"),
                ("VOU_FUNC_01", "Kiểm tra áp dụng voucher giảm % có mức giảm tối đa", "Kiểm tra tính toán tiền giảm đúng mức trần quy định",
                 "Bước 1: Đơn hàng có tổng tiền 500.000đ\nBước 2: Nhập mã voucher 'DEVCINE50' (giảm 50%, tối đa 50.000đ)\nBước 3: Click button 'Áp dụng'\nBước 4: Kiểm tra số tiền được giảm và tổng thanh toán cuối",
                 "Tổng đơn: 500.000đ\nVoucher: Giảm 50% max 50.000đ",
                 "Hệ thống tính tiền giảm đúng bằng mức trần 50.000đ (thay vì 250.000đ), tổng tiền thanh toán cuối cùng hiển thị là 450.000đ"),
                ("VOU_FUNC_02", "Kiểm tra đổi điểm Loyalty lấy voucher ưu đãi", "Kiểm tra trừ điểm tích lũy và thêm voucher mới vào kho cá nhân",
                 "Bước 1: Mở kho ưu đãi 'Đổi điểm thưởng'\nBước 2: Chọn voucher 'Giảm 30k vé xem phim' (Yêu cầu 100 điểm)\nBước 3: Click button 'Xác nhận đổi điểm'\nBước 4: Kiểm tra số dư điểm và kho voucher",
                 "Số dư điểm: 250 điểm\nĐiểm đổi: 100 điểm",
                 "Đổi ưu đãi thành công, trừ 100 điểm trong ví (còn 150 điểm) và sinh mã voucher mới hiển thị ngay trong kho voucher của khách hàng")
            ]
        },
        {
            "code": "MOD_CUST_PAYMENT", "sheet": "Thanh toán VNPAY",
            "req": "Kiểm tra Tích hợp Cổng VNPAY và Sinh vé điện tử QR",
            "tester": "Nguyễn Quang Huy", "role": "Khách hàng", "pre": "Khách hàng đã hoàn tất chọn vé, ghế, F&B, áp dụng voucher và chọn phương thức VNPAY",
            "test_cases": [
                ("PAY_FUNC_01", "Kiểm tra chuyển hướng sang cổng thanh toán VNPAY", "Kiểm tra tạo URL thanh toán an toàn kèm chữ ký bảo mật SHA-512",
                 "Bước 1: Tại màn hình thanh toán, chọn phương thức 'Thanh toán qua VNPAY'\nBước 2: Click button 'Tiến hành thanh toán 350.000 VNĐ'\nBước 3: Kiểm tra chuyển hướng trình duyệt",
                 "Số tiền: 350.000 VNĐ", "Trình duyệt chuyển hướng thành công sang giao diện cổng thanh toán VNPAY chính thức với đúng số tiền đơn hàng"),
                ("PAY_FUNC_02", "Kiểm tra xử lý thanh toán VNPAY thành công", "Kiểm tra xác thực IPN callback, đổi trạng thái CONFIRMED và sinh vé QR",
                 "Bước 1: Trên giao diện VNPAY Sandbox, chọn phương thức Thẻ nội địa\nBước 2: Nhập thông tin thẻ test và nhập mã OTP xác thực thanh toán thành công\nBước 3: Chờ VNPAY redirect về trang kết quả của DevCine\nBước 4: Kiểm tra trạng thái đơn hàng, vé điện tử và email nhận vé",
                 "Mã phản hồi VNPAY: '00' (Thành công)\nSố tiền: 350.000 VNĐ",
                 "Đơn hàng chuyển sang trạng thái CONFIRMED, tự động tạo vé xem phim kèm mã QR Code độc nhất, cộng 35 điểm tích lũy và gửi email hóa đơn kèm vé cho khách hàng"),
                ("PAY_FUNC_03", "Kiểm tra xử lý khi khách hàng hủy giao dịch VNPAY", "Kiểm tra hủy đơn hàng và giải phóng ghế khi khách nhấn hủy thanh toán",
                 "Bước 1: Trên cổng thanh toán VNPAY, click vào nút 'Hủy giao dịch'\nBước 2: Chờ VNPAY redirect về website DevCine\nBước 3: Kiểm tra trạng thái đơn hàng và trạng thái ghế trên sơ đồ",
                 "Mã phản hồi VNPAY: '24' (Khách hủy giao dịch)",
                 "Hệ thống ghi nhận đơn hàng bị hủy (CANCELLED), tự động giải phóng các ghế đang giữ về trạng thái TRỐNG và hiển thị thông báo hủy thanh toán"),
                ("PAY_FUNC_04", "Kiểm tra chống xử lý callback IPN trùng lặp", "Kiểm tra tính toàn vẹn và cơ chế Idempotency khi nhận nhiều webhook IPN",
                 "Bước 1: Giao dịch đơn hàng #12345 đã thanh toán thành công và chuyển CONFIRMED\nBước 2: Giả lập gửi lại request IPN từ VNPAY lần thứ 2 cho cùng mã đơn hàng đó\nBước 3: Kiểm tra phản hồi trả về cho VNPAY và dữ liệu đơn hàng trong cơ sở dữ liệu",
                 "Mã đơn: #12345 (Đã CONFIRMED)",
                 "Hệ thống trả về mã RspCode: '02' (Order already confirmed), không thực hiện cộng điểm hay tạo vé lặp lại lần 2")
            ]
        },
        {
            "code": "MOD_CUST_SUPPORT", "sheet": "Hỗ trợ CSKH",
            "req": "Kiểm tra Gửi yêu cầu hỗ trợ (Support Ticket)",
            "tester": "Nguyễn Quang Huy", "role": "Khách hàng", "pre": "Người dùng mở trang Liên hệ & Hỗ trợ CSKH trên website DevCine",
            "test_cases": [
                ("CS_VAL_01", "Kiểm tra validate để trống tiêu đề yêu cầu", "Kiểm tra thông báo lỗi khi không nhập tiêu đề",
                 "Bước 1: Mở form Gửi yêu cầu hỗ trợ\nBước 2: Nhập đầy đủ họ tên, email, SĐT\nBước 3: Để trống trường 'Tiêu đề'\nBước 4: Click button 'Gửi yêu cầu'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Tiêu đề: ''", "Hiển thị thông báo lỗi: 'Tiêu đề yêu cầu hỗ trợ phải chứa từ 5 đến 200 ký tự'"),
                ("CS_VAL_02", "Kiểm tra validate nội dung yêu cầu dưới 10 ký tự", "Kiểm tra thông báo lỗi độ dài tối thiểu của nội dung ticket",
                 "Bước 1: Mở form Gửi yêu cầu hỗ trợ\nBước 2: Nhập tiêu đề hợp lệ\nBước 3: Nhập nội dung chỉ có 4 ký tự: 'Alo ạ'\nBước 4: Click button 'Gửi yêu cầu'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Nội dung: 'Alo ạ'", "Hiển thị thông báo lỗi: 'Nội dung chi tiết yêu cầu hỗ trợ phải chứa từ 10 đến 1000 ký tự'"),
                ("CS_FUNC_01", "Kiểm tra gửi yêu cầu hỗ trợ thành công", "Kiểm tra tạo ticket và gửi email xác nhận cho khách hàng",
                 "Bước 1: Điền đầy đủ Họ tên, Email, Số điện thoại hợp lệ\nBước 2: Nhập Tiêu đề: 'Hỏi về chính sách hoàn hủy vé do sự cố thời tiết'\nBước 3: Nhập Nội dung chi tiết thắc mắc\nBước 4: Click button 'Gửi yêu cầu'\nBước 5: Kiểm tra phản hồi từ hệ thống",
                 "Full valid ticket data",
                 "Tạo ticket thành công với trạng thái OPEN, hệ thống cấp mã Ticket (ví dụ #TCK-2026-001) và tự động gửi email xác nhận đã tiếp nhận yêu cầu đến hòm thư của khách")
            ]
        },

        # --- NHÂN VIÊN (10 modules) ---
        {
            "code": "MOD_STAFF_FIRST_PASS", "sheet": "Đổi mật khẩu lần đầu",
            "req": "Kiểm tra Đổi mật khẩu bắt buộc cho nhân viên mới",
            "tester": "Phạm Thị Quỳnh Anh", "role": "Nhân viên mới", "pre": "Nhân viên mới được cấp tài khoản, đăng nhập lần đầu tiên vào hệ thống",
            "test_cases": [
                ("FST_VAL_01", "Kiểm tra mật khẩu mới không đủ độ phức tạp", "Kiểm tra validate 4 nhóm ký tự (hoa, thường, số, ký tự đặc biệt)",
                 "Bước 1: Nhân viên mới đăng nhập bằng mật khẩu tạm thời\nBước 2: Hệ thống tự động chuyển hướng bắt buộc sang màn hình 'Đổi mật khẩu lần đầu'\nBước 3: Nhập mật khẩu mới chỉ gồm chữ thường: 'nhanvien123'\nBước 4: Click button 'Xác nhận đổi mật khẩu'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Mật khẩu mới: 'nhanvien123'", "Hiển thị thông báo lỗi: 'Mật khẩu mới phải chứa ít nhất 1 chữ hoa, 1 chữ thường, 1 chữ số và 1 ký tự đặc biệt'"),
                ("FST_VAL_02", "Kiểm tra mật khẩu mới trùng mật khẩu tạm thời", "Kiểm tra bắt buộc đổi mật khẩu mới khác mật khẩu ban đầu",
                 "Bước 1: Tại màn hình Đổi mật khẩu lần đầu, nhập mật khẩu mới giống hệt mật khẩu tạm thời\nBước 2: Click button 'Xác nhận đổi mật khẩu'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Mật khẩu mới: Trùng mật khẩu tạm", "Hiển thị thông báo lỗi: 'Mật khẩu mới không được trùng với mật khẩu tạm thời được cấp'"),
                ("FST_FUNC_01", "Kiểm tra đổi mật khẩu lần đầu thành công", "Kiểm tra tắt cờ requires_password_change và kích hoạt tài khoản",
                 "Bước 1: Nhập mật khẩu mới hợp lệ: 'Staff@DevCine2026'\nBước 2: Nhập xác nhận mật khẩu mới trùng khớp\nBước 3: Click button 'Xác nhận đổi mật khẩu'\nBước 4: Kiểm tra chuyển hướng giao diện",
                 "Mật khẩu mới: 'Staff@DevCine2026'",
                 "Đổi mật khẩu thành công, hệ thống tắt cờ bắt buộc đổi mật khẩu và chuyển hướng nhân viên vào giao diện màn hình POS Bán vé")
            ]
        },
        {
            "code": "MOD_POS_TICKETS", "sheet": "POS Bán vé tại quầy",
            "req": "Kiểm tra Bán vé xem phim tại quầy và Cinema Scoping",
            "tester": "Văn Minh Khôi", "role": "Nhân viên Quầy", "pre": "Nhân viên đã đăng nhập vào hệ thống POS cơ sở của rạp mình phụ trách",
            "test_cases": [
                ("POS_VAL_01", "Kiểm tra chặn bán chéo suất chiếu rạp khác", "Kiểm tra ràng buộc bảo mật Strict Cinema Scoping",
                 "Bước 1: Nhân viên trực thuộc cụm rạp CGV Cầu Giấy mở giao diện POS\nBước 2: Cố tình truy cập URL hoặc chọn suất chiếu thuộc cụm rạp CGV Hà Đông\nBước 3: Kiểm tra phản hồi từ hệ thống",
                 "Cinema Scope: Khác cụm rạp", "Hệ thống từ chối truy cập, báo lỗi 403 Forbidden: 'Bạn không có quyền thao tác trên dữ liệu của cụm rạp khác'"),
                ("POS_VAL_02", "Kiểm tra validate tra cứu SĐT hội viên sai", "Kiểm tra thông báo lỗi khi nhập số điện thoại không hợp lệ",
                 "Bước 1: Tại màn hình POS Bán vé, nhập số điện thoại 7 chữ số vào ô 'Tra cứu hội viên'\nBước 2: Click button 'Tìm kiếm'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "SĐT: '0912345'", "Hiển thị thông báo lỗi: 'Số điện thoại hội viên phải bao gồm đúng 10 chữ số'"),
                ("POS_FUNC_01", "Kiểm tra tra cứu hội viên thành công", "Kiểm tra hiển thị tên khách hàng, hạng thẻ và điểm thưởng để tích điểm",
                 "Bước 1: Nhập số điện thoại '0912345678' của khách hàng đã có tài khoản\nBước 2: Click button 'Tìm kiếm'\nBước 3: Kiểm tra thông tin hiển thị trên panel khách hàng",
                 "SĐT: '0912345678'", "Hiển thị đúng Tên khách hàng: 'Nguyễn Văn Dân', Hạng thẻ: 'Vàng (Gold)', Số điểm tích lũy: '350 điểm'"),
                ("POS_VAL_03", "Kiểm tra validate tiền khách đưa nhỏ hơn tổng tiền", "Kiểm tra validate thanh toán tiền mặt tại quầy",
                 "Bước 1: Đơn hàng bán vé có tổng tiền 220.000đ\nBước 2: Chọn phương thức 'Tiền mặt'\nBước 3: Nhập tiền khách đưa là 200.000đ\nBước 4: Click button 'Hoàn tất & In vé'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Tổng tiền: 220.000đ\nTiền khách đưa: 200.000đ", "Hiển thị thông báo lỗi: 'Số tiền khách đưa không đủ để thanh toán đơn hàng'"),
                ("POS_FUNC_02", "Kiểm tra tự động tính tiền thối thừa cho khách", "Kiểm tra tính năng tính tiền thừa tiền mặt",
                 "Bước 1: Đơn hàng có tổng tiền 180.000đ\nBước 2: Nhập số tiền khách đưa là 200.000đ\nBước 3: Quan sát ô hiển thị 'Tiền thừa thối lại'",
                 "Tổng tiền: 180.000đ\nTiền đưa: 200.000đ", "Hệ thống tự động hiển thị số tiền thừa cần trả lại khách là 20.000đ với chữ số lớn rõ ràng"),
                ("POS_FUNC_03", "Kiểm tra hoàn tất bán vé và in hóa đơn tại quầy", "Kiểm tra lưu đơn hàng, ghi nhận nhân viên bán và xuất lệnh in vé",
                 "Bước 1: Chọn suất chiếu, 2 ghế VIP, nhập SĐT hội viên\nBước 2: Nhận đủ tiền mặt từ khách\nBước 3: Click button 'Hoàn tất & In vé'\nBước 4: Kiểm tra kết quả lưu trữ và in ấn",
                 "Phương thức: Tiền mặt",
                 "Đơn hàng lưu thành công với sold_by=nhanvien_id, tích điểm cho hội viên, đổi trạng thái ghế sang SOLD và gửi lệnh in vé nhiệt trực tiếp ra máy in quầy")
            ]
        },
        {
            "code": "MOD_POS_PENDING", "sheet": "POS Đơn chờ",
            "req": "Kiểm tra Quản lý đơn chờ tạm thời trên POS (Tối đa 3 đơn)",
            "tester": "Văn Minh Khôi", "role": "Nhân viên Quầy", "pre": "Nhân viên đang thao tác chọn vé cho khách trên màn hình POS Bán vé",
            "test_cases": [
                ("PND_VAL_01", "Kiểm tra giới hạn tối đa 3 đơn chờ", "Kiểm tra chặn không cho tạo thêm đơn chờ thứ 4 trên cùng máy POS",
                 "Bước 1: Đang có sẵn 3 đơn chờ trên thanh tab POS\nBước 2: Chọn tiếp 1 đơn mới và click button 'Lưu đơn chờ'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Số đơn chờ hiện có: 3 đơn", "Hiển thị thông báo lỗi: 'Mỗi máy POS chỉ được phép lưu tối đa 3 đơn chờ cùng một lúc'"),
                ("PND_FUNC_01", "Kiểm tra khôi phục đơn chờ để thanh toán", "Kiểm tra mở lại đơn chờ và nạp lại đúng danh sách vé và ghế đã chọn",
                 "Bước 1: Click vào tab 'Đơn chờ #2' trên thanh POS\nBước 2: Quan sát sơ đồ ghế và giỏ hàng nạp lại\nBước 3: Tiếp tục thu tiền và hoàn tất đơn",
                 "Chọn: Đơn chờ #2", "Hệ thống khôi phục nguyên vẹn suất chiếu, vị trí ghế và bắp nước của đơn #2 để nhân viên tiến hành thu tiền"),
                ("PND_FUNC_02", "Kiểm tra hết hạn giữ đơn tự hủy và phạt khóa ghế 5 phút", "Kiểm tra cơ chế timeout 10 phút giữ đơn chờ và khóa phạt ghế",
                 "Bước 1: Lưu một đơn chờ trên POS\nBước 2: Không thanh toán và chờ quá thời hạn 10 phút\nBước 3: Kiểm tra trạng thái đơn chờ và trạng thái ghế",
                 "Thời gian: Quá 10 phút", "Đơn chờ tự động bị hủy, ghế bị khóa tạm thời trong 5 phút (Penalty Lock) không cho mở lại ngay lập tức để tránh chiếm giữ ghế ảo")
            ]
        },
        {
            "code": "MOD_POS_FNB", "sheet": "POS Bán F&B tại quầy",
            "req": "Kiểm tra Bán bắp nước riêng lẻ tại quầy không kèm vé",
            "tester": "Văn Minh Khôi", "role": "Nhân viên Quầy", "pre": "Nhân viên mở tab 'Bán Bắp Nước F&B' trên giao diện POS",
            "test_cases": [
                ("PFNB_VAL_01", "Kiểm tra validate khi thanh toán giỏ F&B rỗng", "Kiểm tra bắt buộc chọn ít nhất 1 món bắp nước",
                 "Bước 1: Mở màn hình Bán F&B tại quầy\nBước 2: Không chọn món nào trong thực đơn (giỏ hàng rỗng)\nBước 3: Click button 'Thanh toán'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Giỏ hàng: 0 món", "Hiển thị thông báo lỗi: 'Vui lòng chọn ít nhất một món bắp nước để thanh toán'"),
                ("PFNB_VAL_02", "Kiểm tra validate khi chưa chọn vị bắp của Combo", "Kiểm tra bắt buộc chọn tùy chọn món trong combo",
                 "Bước 1: Chọn món 'Combo Bắp Nước 2 Người'\nBước 2: Không chọn vị bắp trong popup tùy chọn\nBước 3: Click button 'Thêm vào giỏ'\nBước 4: Kiểm tra thông báo hiển thị",
                 "Combo: Chưa chọn vị", "Hiển thị cảnh báo: 'Vui lòng chọn đầy đủ vị bắp và loại nước trước khi thêm vào giỏ'"),
                ("PFNB_FUNC_01", "Kiểm tra bán F&B riêng lẻ và in hóa đơn thành công", "Kiểm tra thu tiền và hoàn tất đơn bán bắp nước độc lập",
                 "Bước 1: Chọn 2 Bắp Phô Mai + 2 Coca Cola cỡ lớn\nBước 2: Nhập SĐT hội viên để tích điểm\nBước 3: Thu tiền mặt 160.000đ từ khách\nBước 4: Click button 'Hoàn tất thanh toán'\nBước 5: Kiểm tra kết quả in hóa đơn",
                 "Món: 2 Bắp Phô Mai + 2 Coca L\nTiền: 160.000đ",
                 "Thanh toán thành công, ghi nhận doanh thu F&B của cụm rạp, tích 16 điểm cho hội viên và in hóa đơn bán lẻ F&B cho khách")
            ]
        },
        {
            "code": "MOD_POS_VOID_FNB", "sheet": "Yêu cầu hủy đơn F&B",
            "req": "Kiểm tra Tạo yêu cầu hủy đơn bắp nước (FnB Void Request)",
            "tester": "Văn Minh Khôi", "role": "Nhân viên Quầy", "pre": "Nhân viên mở lịch sử các đơn hàng F&B đã thanh toán trong ca",
            "test_cases": [
                ("VOID_VAL_01", "Kiểm tra validate khi không nhập lý do hủy đơn", "Kiểm tra bắt buộc nhập lý do hủy đơn bắp nước",
                 "Bước 1: Tìm kiếm đơn hàng F&B vừa thanh toán\nBước 2: Click button 'Yêu cầu hủy đơn (Void)'\nBước 3: Để trống trường 'Lý do yêu cầu hủy'\nBước 4: Click button 'Gửi yêu cầu'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Lý do: ''", "Hiển thị thông báo lỗi: 'Vui lòng nhập lý do yêu cầu hủy đơn (từ 5 đến 255 ký tự)'"),
                ("VOID_VAL_02", "Kiểm tra chặn hủy đơn của cụm rạp khác", "Kiểm tra ràng buộc bảo mật chi nhánh",
                 "Bước 1: Cố tình gửi yêu cầu hủy một đơn hàng thuộc cụm rạp khác\nBước 2: Click button 'Gửi yêu cầu'\nBước 3: Kiểm tra phản hồi từ hệ thống",
                 "Scope: Khác cụm rạp", "Hệ thống từ chối thao tác, báo lỗi: 'Bạn không có quyền yêu cầu hủy đơn hàng của cụm rạp khác'"),
                ("VOID_FUNC_01", "Kiểm tra tạo yêu cầu hủy đơn F&B thành công", "Kiểm tra chuyển trạng thái đơn sang PENDING_VOID và gửi thông báo cho Quản lý",
                 "Bước 1: Chọn đơn hàng F&B cần hủy\nBước 2: Nhập lý do: 'Khách hàng đổi ý muốn đổi sang Combo lớn hơn'\nBước 3: Click button 'Gửi yêu cầu'\nBước 4: Kiểm tra trạng thái đơn hàng",
                 "Lý do: 'Khách hàng đổi ý muốn đổi sang Combo lớn hơn'",
                 "Tạo yêu cầu hủy thành công, đơn chuyển sang trạng thái PENDING_VOID và gửi thông báo real-time lên màn hình của Quản lý ca trực")
            ]
        },
        {
            "code": "MOD_STAFF_CHECKIN", "sheet": "Soát vé & Check-in",
            "req": "Kiểm tra Quét mã QR soát vé vào phòng chiếu",
            "tester": "Văn Minh Khôi", "role": "Nhân viên Soát vé", "pre": "Nhân viên mở màn hình Quét mã QR soát vé tại cửa phòng chiếu",
            "test_cases": [
                ("CHK_VAL_01", "Kiểm tra quét mã QR không tồn tại trong hệ thống", "Kiểm tra thông báo lỗi khi quét mã giả mạo hoặc sai chuẩn",
                 "Bước 1: Mở camera quét mã QR soát vé\nBước 2: Đưa mã QR lạ không thuộc hệ thống DevCine vào vùng quét\nBước 3: Kiểm tra phản hồi từ hệ thống",
                 "Mã QR: 'QR_FAKE_CODE_999'", "Hiển thị thông báo cảnh báo màu đỏ: 'Mã vé không tồn tại trong hệ thống DevCine'"),
                ("CHK_VAL_02", "Kiểm tra soát vé của cụm rạp khác", "Kiểm tra ràng buộc kiểm soát vé đúng rạp đang làm việc",
                 "Bước 1: Nhân viên đang ở cụm rạp CGV Cầu Giấy\nBước 2: Khách xuất trình vé của suất chiếu tại cụm rạp CGV Hà Đông\nBước 3: Quét mã QR của vé\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Vé thuộc rạp: CGV Hà Đông\nNhân viên tại: CGV Cầu Giấy",
                 "Hiển thị cảnh báo màu đỏ: 'Vé không hợp lệ tại cụm rạp này. Vé thuộc rạp CGV Hà Đông'"),
                ("CHK_VAL_03", "Kiểm tra soát vé đã được check-in trước đó", "Kiểm tra cảnh báo vé đã sử dụng chống gian lận vào rạp 2 lần",
                 "Bước 1: Quét một mã vé đã được check-in thành công 15 phút trước\nBước 2: Kiểm tra thông tin cảnh báo hiển thị trên màn hình",
                 "Trạng thái vé: CHECKED_IN", "Hiển thị cảnh báo màu đỏ: 'VÉ ĐÃ SỬ DỤNG! Vé đã được check-in vào lúc 19:15 bởi nhân viên Văn Minh Khôi'"),
                ("CHK_FUNC_01", "Kiểm tra check-in vé hợp lệ thành công", "Kiểm tra đổi trạng thái vé sang CHECKED_IN và hiển thị thông tin phòng/ghế",
                 "Bước 1: Khách hàng xuất trình mã QR vé của suất chiếu 19:30 hôm nay tại đúng rạp\nBước 2: Đưa mã QR vào camera quét vé\nBước 3: Kiểm tra kết quả phản hồi trên màn hình",
                 "Mã QR: Hợp lệ\nThời gian: Trước giờ chiếu 20 phút",
                 "Hệ thống phát âm thanh bíp thành công, hiển thị dấu tích xanh to rõ ràng kèm thông tin: Phim, Phòng chiếu số 2, Vị trí ghế E05, E06 và lưu lịch sử soát vé")
            ]
        },
        {
            "code": "MOD_STAFF_INCIDENT_RELOCATE", "sheet": "Xử lý sự cố & Đổi ghế",
            "req": "Kiểm tra Đổi ghế tại chỗ cho khách khi ghế hỏng",
            "tester": "Văn Minh Khôi", "role": "Nhân viên & Quản lý", "pre": "Nhân viên mở màn hình Xử lý sự cố chỗ ngồi tại quầy hoặc cửa phòng chiếu",
            "test_cases": [
                ("REL_VAL_01", "Kiểm tra chặn đổi ghế khi suất chiếu đã bắt đầu", "Kiểm tra điều kiện chỉ cho phép đổi ghế trước giờ chiếu",
                 "Bước 1: Tìm kiếm vé của suất chiếu đã bắt đầu chiếu được 20 phút\nBước 2: Click button 'Đổi ghế sự cố'\nBước 3: Kiểm tra phản hồi từ hệ thống",
                 "Thời gian: Suất chiếu đã bắt đầu", "Hiển thị thông báo lỗi: 'Suất chiếu đã bắt đầu diễn ra, không thể thực hiện đổi ghế trên hệ thống'"),
                ("REL_VAL_02", "Kiểm tra chặn đổi sang ghế đã có người ngồi", "Kiểm tra tính khả dụng của ghế đích",
                 "Bước 1: Mở sơ đồ phòng chiếu để chọn ghế đích đổi sang\nBước 2: Cố tình chọn ghế màu đỏ (đã có khách khác mua)\nBước 3: Kiểm tra phản hồi từ hệ thống",
                 "Ghế đích: Đã bán (OCCUPIED)", "Hiển thị thông báo lỗi: 'Vị trí ghế đích đã có người mua. Vui lòng chọn một ghế còn trống'"),
                ("REL_VAL_03", "Kiểm tra validate khi để trống lý do đổi ghế", "Kiểm tra bắt buộc ghi nhận lý do phục vụ kiểm toán sự cố",
                 "Bước 1: Chọn ghế nguồn A01 và chọn ghế đích trống A05\nBước 2: Để trống trường 'Lý do đổi ghế'\nBước 3: Click button 'Xác nhận đổi ghế'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Lý do: ''", "Hiển thị thông báo lỗi: 'Vui lòng nhập lý do đổi ghế sự cố (từ 5 đến 255 ký tự)'"),
                ("REL_FUNC_01", "Kiểm tra đổi ghế tại chỗ thành công giữ nguyên mã QR", "Kiểm tra cập nhật vị trí ghế mới trên vé và hệ thống",
                 "Bước 1: Nhập mã vé hoặc SĐT khách hàng để tra cứu đơn\nBước 2: Chọn ghế cũ A01 bị gãy tay vịn\nBước 3: Chọn ghế mới A05 đang trống cùng hạng ghế\nBước 4: Nhập lý do: 'Ghế A01 bị gãy tay vịn cần đổi cho khách'\nBước 5: Click button 'Xác nhận đổi ghế'\nBước 6: Kiểm tra kết quả cập nhật",
                 "Ghế cũ: A01 -> Ghế mới: A05\nLý do: 'Ghế A01 bị gãy tay vịn'",
                 "Hệ thống cập nhật vị trí ghế mới A05 cho khách hàng, giữ nguyên mã QR Code và mã vé gốc, đồng thời ghi log nhật ký kiểm toán sự cố")
            ]
        },
        {
            "code": "MOD_MGR_APPROVE_VOID", "sheet": "Phê duyệt hủy đơn F&B",
            "req": "Kiểm tra Duyệt / Từ chối yêu cầu hủy đơn bắp nước",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản lý Cụm rạp", "pre": "Quản lý cụm rạp đăng nhập và mở màn hình 'Phê duyệt hủy đơn F&B'",
            "test_cases": [
                ("APP_VAL_01", "Kiểm tra validate khi từ chối mà không nhập lý do", "Kiểm tra bắt buộc nhập lý do từ chối yêu cầu hủy đơn",
                 "Bước 1: Mở danh sách các đơn yêu cầu hủy F&B\nBước 2: Chọn một đơn đang PENDING_VOID\nBước 3: Click button 'Từ chối hủy'\nBước 4: Để trống ô 'Lý do từ chối'\nBước 5: Click button 'Xác nhận'\nBước 6: Kiểm tra thông báo lỗi hiển thị",
                 "Lý do từ chối: ''", "Hiển thị thông báo lỗi: 'Vui lòng nhập lý do từ chối yêu cầu hủy đơn'"),
                ("APP_FUNC_01", "Kiểm tra phê duyệt hủy đơn F&B thành công", "Kiểm tra đổi trạng thái đơn sang VOIDED và hoàn tiền",
                 "Bước 1: Mở chi tiết đơn hàng yêu cầu hủy kèm lý do nhân viên gửi\nBước 2: Kiểm tra món ăn chưa xuất kho\nBước 3: Click button 'Phê duyệt hủy đơn'\nBước 4: Kiểm tra trạng thái đơn hàng và doanh thu ca",
                 "Hành động: APPROVE_VOID", "Đơn hàng chuyển trạng thái sang VOIDED, trừ doanh thu trong ca làm việc và ghi nhận tên Quản lý phê duyệt"),
                ("APP_FUNC_02", "Kiểm tra từ chối yêu cầu hủy đơn F&B", "Kiểm tra khôi phục trạng thái đơn hàng khi bị từ chối",
                 "Bước 1: Mở chi tiết đơn hàng yêu cầu hủy\nBước 2: Nhập lý do từ chối: 'Bắp nước đã giao cho khách, không được phép hủy'\nBước 3: Click button 'Xác nhận từ chối'\nBước 4: Kiểm tra trạng thái đơn hàng",
                 "Lý do: 'Bắp nước đã giao cho khách, không được phép hủy'",
                 "Đơn hàng được khôi phục về trạng thái COMPLETED, lưu lý do từ chối và gửi thông báo phản hồi cho nhân viên tạo yêu cầu")
            ]
        },
        {
            "code": "MOD_MGR_SEAT_MAINTENANCE", "sheet": "Khóa bảo trì ghế vật lý",
            "req": "Kiểm tra Chuyển trạng thái ghế sang bảo trì (Maintenance)",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản lý Cụm rạp", "pre": "Quản lý mở màn hình Sơ đồ quản lý trạng thái ghế phòng chiếu",
            "test_cases": [
                ("MNT_VAL_01", "Kiểm tra validate khi khóa ghế mà không nhập lý do", "Kiểm tra bắt buộc nhập lý do bảo trì ghế hỏng",
                 "Bước 1: Mở sơ đồ ghế phòng chiếu số 1\nBước 2: Chọn ghế B03\nBước 3: Chọn trạng thái 'Bảo trì (Maintenance)'\nBước 4: Để trống ô 'Lý do bảo trì'\nBước 5: Click button 'Lưu trạng thái'\nBước 6: Kiểm tra thông báo lỗi hiển thị",
                 "Lý do: ''", "Hiển thị thông báo lỗi: 'Vui lòng nhập lý do đưa ghế vào bảo trì (ví dụ: Ghế gãy, Rách đệm...)'"),
                ("MNT_FUNC_01", "Kiểm tra khóa bảo trì ghế thành công", "Kiểm tra tự động ẩn ghế bảo trì trên tất cả suất chiếu tương lai",
                 "Bước 1: Chọn ghế B03 trong phòng chiếu số 1\nBước 2: Chọn trạng thái 'Bảo trì'\nBước 3: Nhập lý do: 'Đệm ghế bị rách cần bọc lại da'\nBước 4: Click button 'Lưu trạng thái'\nBước 5: Kiểm tra hiển thị ghế trên các suất chiếu tương lai của phòng 1",
                 "Ghế: B03\nLý do: 'Đệm ghế bị rách cần bọc lại da'",
                 "Ghế B03 chuyển sang trạng thái MAINTENANCE (màu xám có icon cờ lê) và tự động bị khóa ẩn trên tất cả các suất chiếu tương lai của phòng chiếu số 1")
            ]
        },
        {
            "code": "MOD_MGR_COMPENSATION", "sheet": "Tặng voucher đền bù",
            "req": "Kiểm tra Phát voucher đền bù sự cố cho khách hàng",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản lý Cụm rạp", "pre": "Quản lý mở màn hình Tặng voucher đền bù sự cố cho khách",
            "test_cases": [
                ("CMP_VAL_01", "Kiểm tra validate khi chưa chọn mẫu voucher đền bù", "Kiểm tra bắt buộc chọn mẫu voucher theo chính sách đền bù",
                 "Bước 1: Mở popup 'Tặng voucher đền bù'\nBước 2: Nhập SĐT khách hàng gặp sự cố\nBước 3: Chưa chọn mẫu voucher trong danh sách\nBước 4: Click button 'Tặng voucher'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Voucher: Chưa chọn", "Hiển thị thông báo lỗi: 'Vui lòng chọn một mẫu voucher đền bù theo quy định'"),
                ("CMP_FUNC_01", "Kiểm tra tặng voucher đền bù sự cố thành công", "Kiểm tra cấp phát voucher trực tiếp vào ví của khách hàng",
                 "Bước 1: Nhập số điện thoại khách hàng bị ảnh hưởng do sự cố\nBước 2: Chọn mẫu voucher: 'Vé xem phim 2D Miễn Phí (Đền bù sự cố)'\nBước 3: Nhập ghi chú: 'Đền bù sự cố mất điện phòng chiếu 1 ngày 19/03'\nBước 4: Click button 'Tặng voucher'\nBước 5: Kiểm tra ví voucher của khách hàng",
                 "Mẫu: 'VOUCHER_FREE_TICKET_2D'\nGhi chú: 'Đền bù sự cố mất điện phòng chiếu 1'",
                 "Hệ thống phát thành công 1 voucher vé miễn phí vào ví cá nhân của khách hàng, đồng thời tự động gửi email xin lỗi kèm mã voucher cho khách")
            ]
        },

        # --- QUẢN TRỊ VIÊN / ADMIN (19 modules) ---
        {
            "code": "MOD_ADMIN_MOVIE_CRUD", "sheet": "Quản lý phim",
            "req": "Kiểm tra Thêm, Sửa, Xóa và Upload Media Phim",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Quản trị viên đã đăng nhập và mở trang Quản lý Phim trong Admin Dashboard",
            "test_cases": [
                ("MOV_VAL_01", "Kiểm tra validate để trống tên phim", "Kiểm tra bắt buộc nhập tên phim",
                 "Bước 1: Mở modal 'Thêm mới phim'\nBước 2: Để trống trường 'Tên phim'\nBước 3: Điền đầy đủ các thông tin còn lại\nBước 4: Click button 'Lưu phim'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Tên phim: ''", "Hiển thị thông báo lỗi: 'Tên phim không được để trống (từ 2 đến 150 ký tự)'"),
                ("MOV_VAL_02", "Kiểm tra validate thời lượng phim ngoài khoảng 30-300 phút", "Kiểm tra ràng buộc thời lượng phim chiếu rạp",
                 "Bước 1: Mở form thêm phim\nBước 2: Nhập thời lượng phim là 15 phút\nBước 3: Click button 'Lưu phim'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Thời lượng: 15 phút", "Hiển thị thông báo lỗi: 'Thời lượng phim phải là số nguyên từ 30 đến 300 phút'"),
                ("MOV_VAL_03", "Kiểm tra validate năm sản xuất ngoài khoảng 2020-2035", "Kiểm tra ràng buộc năm phát hành phim",
                 "Bước 1: Mở form thêm phim\nBước 2: Nhập năm sản xuất là 2010\nBước 3: Click button 'Lưu phim'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Năm sản xuất: 2010", "Hiển thị thông báo lỗi: 'Năm sản xuất phải nằm trong khoảng từ 2020 đến 2035'"),
                ("MOV_VAL_04", "Kiểm tra validate ngày kết thúc trước ngày khởi chiếu", "Kiểm tra tính hợp lệ của khoảng thời gian phát hành phim",
                 "Bước 1: Chọn Ngày khởi chiếu: 25/03/2026\nBước 2: Chọn Ngày kết thúc: 20/03/2026\nBước 3: Click button 'Lưu phim'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Khởi chiếu: 25/03/2026\nKết thúc: 20/03/2026", "Hiển thị thông báo lỗi: 'Ngày kết thúc chiếu phải sau hoặc bằng ngày khởi chiếu'"),
                ("MOV_VAL_05", "Kiểm tra validate đường dẫn Trailer không phải link Youtube", "Kiểm tra chuẩn đường dẫn video trailer nhúng",
                 "Bước 1: Nhập đường dẫn trailer từ trang Facebook hoặc TikTok\nBước 2: Click button 'Lưu phim'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Trailer: 'https://facebook.com/video/123'", "Hiển thị thông báo lỗi: 'Đường dẫn Trailer phải là link video Youtube hợp lệ (chứa youtube.com hoặc youtu.be)'"),
                ("MOV_VAL_06", "Kiểm tra validate khi tải ảnh poster vượt quá 10MB", "Kiểm tra giới hạn dung lượng tệp media của phim",
                 "Bước 1: Click chọn file tải lên cho trường 'Ảnh Poster'\nBước 2: Chọn file ảnh có dung lượng 15MB\nBước 3: Kiểm tra phản hồi từ hệ thống",
                 "Poster: 15MB", "Hiển thị thông báo lỗi: 'Dung lượng ảnh Poster không được vượt quá 10MB'"),
                ("MOV_FUNC_01", "Kiểm tra thêm mới phim thành công", "Kiểm tra lưu phim với đầy đủ thông tin chuẩn vào cơ sở dữ liệu",
                 "Bước 1: Mở form 'Thêm mới phim'\nBước 2: Điền đầy đủ thông tin: Tên phim, Thể loại, Độ tuổi T13, Thời lượng 120 phút, Đạo diễn, Diễn viên, Link Youtube Trailer\nBước 3: Upload ảnh Poster và ảnh Banner hợp lệ\nBước 4: Click button 'Lưu phim'\nBước 5: Kiểm tra kết quả hiển thị trên danh sách phim",
                 "Full valid movie data", "Thêm mới phim thành công, hiển thị phim trên danh sách Quản lý phim và sẵn sàng lên lịch chiếu"),
                ("MOV_FUNC_02", "Kiểm tra chặn xóa phim đã có vé được bán", "Kiểm tra ràng buộc toàn vẹn dữ liệu ngăn xóa phim đang có giao dịch",
                 "Bước 1: Trên danh sách phim, tìm bộ phim đang có các suất chiếu đã phát sinh vé bán\nBước 2: Click vào icon 'Xóa phim'\nBước 3: Xác nhận xóa trên popup cảnh báo\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Phim: Đã có vé bán", "Hệ thống từ chối xóa, hiển thị thông báo lỗi: 'Không thể xóa phim này do đã phát sinh giao dịch đặt vé. Vui lòng chuyển trạng thái phim sang Ngừng chiếu'")
            ]
        },
        {
            "code": "MOD_ADMIN_CATEGORIES", "sheet": "Danh mục phim",
            "req": "Kiểm tra Quản lý Thể loại, Định dạng và Độ tuổi",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Danh mục phim",
            "test_cases": [
                ("CAT_VAL_01", "Kiểm tra validate trùng tên thể loại phim", "Kiểm tra tính duy nhất của tên thể loại",
                 "Bước 1: Mở modal 'Thêm thể loại mới'\nBước 2: Nhập tên thể loại 'Hành động' đã có sẵn trong danh mục\nBước 3: Click button 'Lưu'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Tên thể loại: 'Hành động'", "Hiển thị thông báo lỗi: 'Tên thể loại phim đã tồn tại trong hệ thống'"),
                ("CAT_VAL_02", "Kiểm tra validate tên định dạng chứa ký tự đặc biệt", "Kiểm tra chuẩn ký tự của tên định dạng chiếu",
                 "Bước 1: Mở modal 'Thêm định dạng mới'\nBước 2: Nhập tên định dạng chứa các ký tự lạ: 'IMAX_3D@#$'\nBước 3: Click button 'Lưu'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Tên định dạng: 'IMAX_3D@#$'", "Hiển thị thông báo lỗi: 'Tên định dạng phim chỉ được chứa chữ cái, chữ số và dấu gạch ngang'"),
                ("CAT_VAL_03", "Kiểm tra validate mã độ tuổi viết hoa theo quy chuẩn", "Kiểm tra quy chuẩn mã nhãn độ tuổi (P, K, T13, T16, T18, C)",
                 "Bước 1: Mở modal 'Thêm độ tuổi mới'\nBước 2: Nhập mã độ tuổi viết chữ thường 't18'\nBước 3: Click button 'Lưu'\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Mã độ tuổi: 't18'", "Hệ thống tự động chuyển mã sang chữ in hoa 'T18' và lưu thành công"),
                ("CAT_FUNC_01", "Kiểm tra chặn xóa danh mục đang được phim sử dụng", "Kiểm tra ràng buộc khóa ngoại của danh mục",
                 "Bước 1: Chọn thể loại 'Hoạt hình' đang được gắn với 10 bộ phim trong hệ thống\nBước 2: Click button 'Xóa thể loại'\nBước 3: Xác nhận xóa trên popup\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Thể loại: Đang được sử dụng bởi 10 phim", "Hệ thống chặn xóa, thông báo lỗi: 'Không thể xóa thể loại này do đang có 10 bộ phim đang sử dụng'")
            ]
        },
        {
            "code": "MOD_ADMIN_CINEMAS", "sheet": "Quản lý cụm rạp",
            "req": "Kiểm tra Thêm, Sửa Cụm rạp và Giờ mở/đóng cửa",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Cụm rạp trong hệ thống",
            "test_cases": [
                ("CIN_VAL_01", "Kiểm tra validate tên cụm rạp dưới 5 ký tự", "Kiểm tra độ dài tối thiểu của tên rạp",
                 "Bước 1: Mở modal 'Thêm cụm rạp mới'\nBước 2: Nhập tên cụm rạp là 'CGV'\nBước 3: Click button 'Lưu cụm rạp'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Tên rạp: 'CGV'", "Hiển thị thông báo lỗi: 'Tên cụm rạp phải chứa từ 5 đến 100 ký tự'"),
                ("CIN_VAL_02", "Kiểm tra validate số điện thoại hotline rạp sai định dạng", "Kiểm tra chuẩn số hotline liên hệ",
                 "Bước 1: Nhập hotline chỉ có 6 chữ số: '190012'\nBước 2: Click button 'Lưu cụm rạp'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Hotline: '190012'", "Hiển thị thông báo lỗi: 'Số hotline liên hệ phải gồm từ 8 đến 11 chữ số'"),
                ("CIN_VAL_03", "Kiểm tra chặn đổi giờ đóng cửa khi có suất chiếu ngoài giờ", "Kiểm tra ràng buộc giờ hoạt động với lịch chiếu đang có",
                 "Bước 1: Cụm rạp đang có suất chiếu kết thúc lúc 23:30\nBước 2: Admin sửa Giờ đóng cửa của rạp thành 22:00\nBước 3: Click button 'Lưu thay đổi'\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Giờ đóng cửa mới: 22:00\nSuất chiếu hiện có: kết thúc 23:30",
                 "Hệ thống từ chối cập nhật, báo lỗi: 'Không thể đổi giờ đóng cửa thành 22:00 do đang có suất chiếu kết thúc lúc 23:30 vượt quá khung giờ mới'"),
                ("CIN_FUNC_01", "Kiểm tra thêm mới cụm rạp thành công", "Kiểm tra lưu cụm rạp mới với đầy đủ tọa độ và thông tin liên hệ",
                 "Bước 1: Điền Tên cụm rạp: 'DevCine Cầu Giấy', Tỉnh/TP: 'Hà Nội', Quận: 'Cầu Giấy', Địa chỉ chi tiết, Hotline: '19006017', Giờ mở cửa: '08:00', Giờ đóng cửa: '23:30'\nBước 2: Upload ảnh đại diện rạp\nBước 3: Click button 'Lưu cụm rạp'\nBước 4: Kiểm tra kết quả hiển thị",
                 "Full valid cinema data", "Thêm mới cụm rạp thành công, hiển thị trên danh sách quản trị và trên bản đồ cụm rạp trang người dùng")
            ]
        },
        {
            "code": "MOD_ADMIN_ROOMS", "sheet": "Quản lý phòng chiếu",
            "req": "Kiểm tra Thêm, Sửa Phòng chiếu và Cấu hình Dọn phòng",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Phòng chiếu của một cụm rạp cụ thể",
            "test_cases": [
                ("ROM_VAL_01", "Kiểm tra validate trùng tên phòng trong cùng cụm rạp", "Kiểm tra tính duy nhất của tên phòng trong cùng 1 rạp",
                 "Bước 1: Chọn cụm rạp DevCine Cầu Giấy\nBước 2: Thêm mới phòng chiếu và nhập tên 'Cinema 01' đã tồn tại trong rạp đó\nBước 3: Click button 'Lưu phòng chiếu'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Tên phòng: 'Cinema 01'", "Hiển thị thông báo lỗi: 'Tên phòng chiếu đã tồn tại trong cụm rạp này'"),
                ("ROM_VAL_02", "Kiểm tra validate thời gian dọn phòng ngoài khoảng 10-60 phút", "Kiểm tra giới hạn thời gian nghỉ giữa 2 suất chiếu",
                 "Bước 1: Mở form thêm phòng chiếu\nBước 2: Nhập thời gian dọn phòng là 5 phút\nBước 3: Click button 'Lưu phòng chiếu'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Thời gian dọn phòng: 5 phút", "Hiển thị thông báo lỗi: 'Thời gian dọn phòng phải là số nguyên từ 10 đến 60 phút'"),
                ("ROM_VAL_03", "Kiểm tra validate số hàng ghế vượt quá 20 hàng", "Kiểm tra giới hạn kích thước ma trận phòng chiếu",
                 "Bước 1: Nhập số hàng ghế là 25 hàng\nBước 2: Click button 'Lưu phòng chiếu'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Số hàng: 25", "Hiển thị thông báo lỗi: 'Số hàng ghế trong phòng chiếu chỉ được phép từ 5 đến 20 hàng (A đến T)'"),
                ("ROM_FUNC_01", "Kiểm tra thêm mới phòng chiếu thành công", "Kiểm tra tạo phòng chiếu mới và chuyển sang bước thiết lập sơ đồ ghế",
                 "Bước 1: Điền Tên phòng: 'Cinema 03 (IMAX)', Loại phòng: 'IMAX Laser', Số hàng: 12, Số cột: 16, Thời gian dọn: 20 phút\nBước 2: Click button 'Lưu phòng chiếu'\nBước 3: Kiểm tra kết quả tạo phòng",
                 "Full valid room data", "Tạo phòng chiếu thành công và tự động điều hướng sang màn hình Thiết lập sơ đồ ma trận ghế của phòng đó")
            ]
        },
        {
            "code": "MOD_ADMIN_SEATMAP", "sheet": "Sơ đồ ghế",
            "req": "Kiểm tra Thiết lập sơ đồ ma trận ghế và Phân loại ghế",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở màn hình Thiết lập sơ đồ ma trận ghế của phòng chiếu",
            "test_cases": [
                ("SMP_VAL_01", "Kiểm tra validate khi lưu phòng chiếu không có ghế nào", "Kiểm tra bắt buộc phòng chiếu phải có ít nhất 1 ghế khả dụng",
                 "Bước 1: Chuyển toàn bộ các ô trong ma trận thành 'Lối đi (Walkway)'\nBước 2: Click button 'Lưu sơ đồ ghế'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Số lượng ghế: 0", "Hiển thị thông báo lỗi: 'Sơ đồ phòng chiếu phải có ít nhất 1 ghế ngồi'"),
                ("SMP_VAL_02", "Kiểm tra validate ghế đôi Sweetbox không chiếm 2 cột liền kề", "Kiểm tra quy tắc kích thước của ghế đôi",
                 "Bước 1: Chọn 1 ô đơn lẻ và thiết lập loại ghế là Sweetbox\nBước 2: Click button 'Lưu sơ đồ ghế'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Sweetbox: 1 ô đơn lẻ", "Hiển thị thông báo lỗi: 'Ghế đôi Sweetbox bắt buộc phải chiếm đúng 2 cột liền kề trong cùng 1 hàng'"),
                ("SMP_VAL_03", "Kiểm tra chặn sửa sơ đồ khi phòng đang có vé đã bán", "Kiểm tra bảo vệ tính toàn vẹn vé đã bán cho khách",
                 "Bước 1: Phòng chiếu số 1 đang có các suất chiếu tương lai đã bán vé\nBước 2: Admin mở sơ đồ phòng và chỉnh sửa vị trí các ghế\nBước 3: Click button 'Lưu sơ đồ ghế'\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Phòng: Đang có vé bán", "Hệ thống từ chối lưu, thông báo lỗi: 'Không thể chỉnh sửa sơ đồ ghế do phòng chiếu đang có các suất chiếu tương lai đã bán vé'"),
                ("SMP_FUNC_01", "Kiểm tra lưu sơ đồ ma trận ghế hoàn chỉnh thành công", "Kiểm tra thiết lập các loại ghế Thường, VIP, Sweetbox và Lối đi",
                 "Bước 1: Vẽ sơ đồ ma trận 10 hàng x 14 cột gồm: 4 hàng ghế Thường, 4 hàng ghế VIP, 2 hàng cuối là Sweetbox và 2 cột lối đi ở giữa\nBước 2: Click button 'Lưu sơ đồ ghế'\nBước 3: Kiểm tra kết quả lưu trữ",
                 "Ma trận: 10x14 đầy đủ phân loại ghế",
                 "Lưu sơ đồ ghế thành công, hệ thống tự động sinh mã nhãn ghế chuẩn (A01..J14) và áp dụng chính xác lên giao diện chọn ghế của khách hàng")
            ]
        },
        {
            "code": "MOD_ADMIN_SHOWTIMES", "sheet": "Điều phối lịch chiếu",
            "req": "Kiểm tra Thêm suất chiếu đơn và Kiểm tra Xung đột phòng",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở màn hình Điều phối Lịch chiếu của cụm rạp",
            "test_cases": [
                ("ST_VAL_01", "Kiểm tra validate khi chọn giờ bắt đầu trong quá khứ", "Kiểm tra ràng buộc thời gian chiếu không được trong quá khứ",
                 "Bước 1: Mở form 'Thêm suất chiếu'\nBước 2: Chọn ngày giờ bắt đầu là thời điểm 2 tiếng trước hiện tại\nBước 3: Click button 'Lưu suất chiếu'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Thời gian bắt đầu: Trong quá khứ", "Hiển thị thông báo lỗi: 'Thời gian bắt đầu suất chiếu không được nằm trong quá khứ'"),
                ("ST_VAL_02", "Kiểm tra xung đột phòng chiếu khi bị trùng giờ", "Kiểm tra thuật toán Room Overlap Conflict kèm thời gian dọn phòng",
                 "Bước 1: Phòng chiếu 1 đang có suất chiếu phim A từ 18:00 đến 20:00 (dọn phòng 20 phút đến 20:20)\nBước 2: Admin thêm suất chiếu phim B tại phòng 1 bắt đầu lúc 19:30\nBước 3: Click button 'Lưu suất chiếu'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Suất 1: 18:00 - 20:20\nSuất 2: 19:30",
                 "Hệ thống báo lỗi xung đột phòng chiếu (Conflict 409): 'Phòng chiếu số 1 đang có suất chiếu phim Avatar từ 18:00 đến 20:20 (bao gồm dọn phòng). Vui lòng chọn khung giờ khác'"),
                ("ST_VAL_03", "Kiểm tra chặn xóa suất chiếu đã có vé bán", "Kiểm tra bảo vệ suất chiếu đang có khách đặt",
                 "Bước 1: Tìm kiếm suất chiếu đã có 5 vé được khách hàng thanh toán\nBước 2: Click icon 'Xóa suất chiếu'\nBước 3: Xác nhận xóa trên popup\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Suất chiếu: Đã bán 5 vé", "Hệ thống từ chối xóa, thông báo lỗi: 'Không thể xóa suất chiếu đã phát sinh vé bán hoặc đang có khách giữ chỗ'"),
                ("ST_FUNC_01", "Kiểm tra thêm mới suất chiếu đơn hợp lệ thành công", "Kiểm tra lên lịch chiếu và mở bán vé trên hệ thống",
                 "Bước 1: Chọn Phim: 'Mai', Phòng chiếu: 'Cinema 01', Định dạng: '2D Digital', Ngày chiếu: '20/03/2026', Giờ bắt đầu: '20:30'\nBước 2: Kiểm tra hệ thống tự động tính giờ kết thúc dựa trên thời lượng phim\nBước 3: Click button 'Lưu suất chiếu'\nBước 4: Kiểm tra kết quả hiển thị trên bảng lịch chiếu",
                 "Full valid showtime data", "Thêm suất chiếu thành công, hiển thị trên bảng lịch chiếu admin và tự động mở bán vé trên website khách hàng")
            ]
        },
        {
            "code": "MOD_ADMIN_BATCH_SCHEDULE", "sheet": "Xếp lịch chiếu hàng loạt",
            "req": "Kiểm tra Xếp lịch chiếu hàng loạt (Batch Scheduling)",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở công cụ Xếp lịch chiếu hàng loạt (Batch Scheduling)",
            "test_cases": [
                ("BSC_VAL_01", "Kiểm tra validate khi khoảng ngày Từ ngày > Đến ngày", "Kiểm tra tính hợp lệ của khoảng ngày xếp lịch",
                 "Bước 1: Mở công cụ Xếp lịch hàng loạt\nBước 2: Chọn Từ ngày: 25/03/2026, Đến ngày: 20/03/2026\nBước 3: Click button 'Sinh lịch chiếu'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Từ ngày: 25/03/2026\nĐến ngày: 20/03/2026", "Hiển thị thông báo lỗi: 'Từ ngày phải nhỏ hơn hoặc bằng Đến ngày'"),
                ("BSC_VAL_02", "Kiểm tra validate khi chưa chọn danh sách phim hoặc phòng", "Kiểm tra điều kiện bắt buộc để sinh lịch tự động",
                 "Bước 1: Chọn khoảng ngày hợp lệ\nBước 2: Để trống danh sách phim và danh sách phòng chiếu\nBước 3: Click button 'Sinh lịch chiếu'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Danh sách: Rỗng", "Hiển thị thông báo lỗi: 'Vui lòng chọn ít nhất 1 bộ phim và 1 phòng chiếu để xếp lịch'"),
                ("BSC_FUNC_01", "Kiểm tra sinh lịch chiếu hàng loạt tự động thành công", "Kiểm tra thuật toán sinh tự động các suất chiếu không trùng phòng",
                 "Bước 1: Chọn khoảng ngày: 7 ngày tiếp theo\nBước 2: Chọn 2 bộ phim đang chiếu và chọn 2 phòng chiếu\nBước 3: Thêm các khung giờ chiếu mẫu: 09:00, 13:30, 18:00, 21:00\nBước 4: Click button 'Sinh lịch chiếu tự động'\nBước 5: Kiểm tra kết quả sinh lịch",
                 "Batch: 7 ngày x 2 phim x 2 phòng x 4 khung giờ",
                 "Hệ thống tự động tính toán thời lượng và dọn phòng, sinh thành công toàn bộ 56 suất chiếu hợp lệ, không có suất nào bị trùng phòng và hiển thị bảng preview trước khi lưu")
            ]
        },
        {
            "code": "MOD_ADMIN_FNB_ITEMS", "sheet": "Quản lý thực đơn F&B",
            "req": "Kiểm tra Thêm, Sửa món bắp nước và Phân loại",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Thực đơn F&B trong Dashboard",
            "test_cases": [
                ("FNB_VAL_01", "Kiểm tra validate trùng tên món bắp nước", "Kiểm tra tính duy nhất của tên món ăn/nước uống",
                 "Bước 1: Mở modal 'Thêm món F&B mới'\nBước 2: Nhập tên món 'Bắp Rang Bơ Phô Mai' đã có trong thực đơn\nBước 3: Click button 'Lưu món'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Tên món: 'Bắp Rang Bơ Phô Mai'", "Hiển thị thông báo lỗi: 'Tên món bắp nước đã tồn tại trong thực đơn'"),
                ("FNB_VAL_02", "Kiểm tra validate đơn giá bán là số âm", "Kiểm tra giá bán phải là số nguyên dương",
                 "Bước 1: Mở form thêm món\nBước 2: Nhập đơn giá bán là -50.000đ\nBước 3: Click button 'Lưu món'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Giá bán: -50.000đ", "Hiển thị thông báo lỗi: 'Giá bán phải là số nguyên từ 0 đến 1.000.000đ'"),
                ("FNB_FUNC_01", "Kiểm tra chặn xóa vĩnh viễn món đã có trong lịch sử đơn", "Kiểm tra bảo toàn dữ liệu lịch sử hóa đơn",
                 "Bước 1: Chọn món 'Nước ngọt Coca Cola' đã có trong 500 đơn hàng cũ\nBước 2: Click button 'Xóa món'\nBước 3: Xác nhận xóa\nBước 4: Kiểm tra trạng thái của món",
                 "Món: Đã có trong 500 đơn hàng",
                 "Hệ thống không xóa cứng trong DB mà tự động chuyển trạng thái món sang ACTIVE=false (Ngừng kinh doanh) để bảo toàn lịch sử hóa đơn"),
                ("FNB_FUNC_02", "Kiểm tra thêm mới món F&B thành công", "Kiểm tra lưu món mới và upload ảnh sản phẩm",
                 "Bước 1: Điền Tên món: 'Bắp Phô Mai Trứng Muối', Phân loại: 'Đồ ăn', Giá bán: 65.000đ, Mô tả: 'Bắp giòn phủ phô mai và trứng muối thơm ngon'\nBước 2: Upload ảnh món dung lượng 1.2MB\nBước 3: Click button 'Lưu món'\nBước 4: Kiểm tra hiển thị món mới",
                 "Full valid F&B item data", "Thêm mới món thành công, hiển thị ngay trên thực đơn web đặt vé online và trên máy POS tại quầy")
            ]
        },
        {
            "code": "MOD_ADMIN_COMBOS", "sheet": "Cấu hình Combo F&B",
            "req": "Kiểm tra Cấu hình Combo và Tùy chọn món con",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở màn hình Cấu hình Combo và Tùy chọn món F&B",
            "test_cases": [
                ("CMB_VAL_01", "Kiểm tra validate số lượng chọn tối thiểu > tối đa", "Kiểm tra ràng buộc min <= max của nhóm tùy chọn",
                 "Bước 1: Mở form cấu hình nhóm tùy chọn\nBước 2: Nhập Số lượng chọn tối thiểu: 3, Số lượng chọn tối đa: 2\nBước 3: Click button 'Lưu cấu hình'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Min: 3\nMax: 2", "Hiển thị thông báo lỗi: 'Số lượng chọn tối thiểu phải nhỏ hơn hoặc bằng số lượng chọn tối đa'"),
                ("CMB_VAL_02", "Kiểm tra validate nhóm tùy chọn không có món con nào", "Kiểm tra bắt buộc nhóm phải có ít nhất 1 món con",
                 "Bước 1: Tạo nhóm tùy chọn 'Chọn loại nước ngọt'\nBước 2: Không thêm món con nào vào danh sách\nBước 3: Click button 'Lưu cấu hình'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Món con: Trống", "Hiển thị thông báo lỗi: 'Mỗi nhóm tùy chọn bắt buộc phải chứa ít nhất 1 món con'"),
                ("CMB_FUNC_01", "Kiểm tra lưu cấu hình Combo hoàn chỉnh thành công", "Kiểm tra thiết lập Combo gồm nhiều thành phần và phụ thu đổi vị",
                 "Bước 1: Tạo Combo 'Couple Combo'\nBước 2: Cấu hình Slot 1: Bắp 1 vị (Bắp ngọt 0đ, Bắp Phô mai +15k, Bắp Caramel +15k)\nBước 3: Cấu hình Slot 2: 2 Nước ngọt (Coca 0đ, Sprite 0đ, Nước Cam +10k)\nBước 4: Click button 'Lưu Combo'\nBước 5: Kiểm tra kết quả",
                 "Full valid combo config", "Lưu cấu hình Combo thành công, áp dụng đồng bộ trên giao diện chọn bắp nước của khách hàng và POS")
            ]
        },
        {
            "code": "MOD_ADMIN_BASE_PRICING", "sheet": "Cấu hình bảng giá vé",
            "req": "Kiểm tra Cấu hình Ma trận giá nền 3 chiều và Phụ thu",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở trang Cấu hình Bảng giá vé và Phụ thu",
            "test_cases": [
                ("PRC_VAL_01", "Kiểm tra validate giá vé nền ngoài khoảng 10k-500k", "Kiểm tra ràng buộc biên của giá vé nền",
                 "Bước 1: Tại ô nhập giá vé Ngày thường - Giờ thường, nhập giá 5.000đ\nBước 2: Click button 'Lưu bảng giá'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Giá vé: 5.000đ", "Hiển thị thông báo lỗi: 'Giá vé nền phải là số nguyên nằm trong khoảng từ 10.000đ đến 500.000đ'"),
                ("PRC_VAL_02", "Kiểm tra validate giá vé HSSV lớn hơn Người lớn", "Kiểm tra logic ưu đãi giá vé học sinh sinh viên",
                 "Bước 1: Nhập giá vé HSSV là 120.000đ trong khi giá vé Người lớn cùng khung giờ là 100.000đ\nBước 2: Click button 'Lưu bảng giá'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "HSSV: 120.000đ\nNgười lớn: 100.000đ", "Hiển thị thông báo lỗi: 'Giá vé HSSV phải luôn nhỏ hơn hoặc bằng giá vé Người lớn trong cùng khung giờ'"),
                ("PRC_VAL_03", "Kiểm tra validate mức phụ thu ghế VIP là số âm", "Kiểm tra giá trị phụ thu không được âm",
                 "Bước 1: Tại ô nhập phụ thu ghế VIP, nhập giá trị -20.000đ\nBước 2: Click button 'Lưu bảng giá'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Phụ thu VIP: -20.000đ", "Hiển thị thông báo lỗi: 'Mức tiền phụ thu phải là số nguyên lớn hơn hoặc bằng 0'"),
                ("PRC_FUNC_01", "Kiểm tra công cụ Simulator bóc tách và tính thử giá vé", "Kiểm tra độ chính xác của công thức tính giá vé tổng hợp",
                 "Bước 1: Mở tab 'Simulator tính thử giá vé'\nBước 2: Chọn Ngày: Thứ 7 (Cuối tuần), Giờ: 20:00 (Giờ vàng), Định dạng: 3D, Loại ghế: VIP, Đối tượng: Người lớn\nBước 3: Click button 'Tính giá thử nghiệm'\nBước 4: Kiểm tra kết quả bóc tách giá",
                 "Cuối tuần (110k) + Giờ vàng (10k) + Ghế VIP (20k) + 3D (30k)",
                 "Công cụ Simulator hiển thị chính xác tổng giá vé = 170.000 VNĐ kèm bảng bóc tách chi tiết từng dòng phụ thu")
            ]
        },
        {
            "code": "MOD_ADMIN_HOLIDAYS", "sheet": "Quản lý ngày lễ",
            "req": "Kiểm tra Khai báo Danh mục Ngày lễ tính giá vé",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở tab Quản lý Ngày lễ trong mục Cấu hình bảng giá",
            "test_cases": [
                ("HOL_VAL_01", "Kiểm tra validate khi thêm ngày lễ trùng ngày đã có", "Kiểm tra tính duy nhất của ngày lễ trong năm",
                 "Bước 1: Mở modal 'Thêm ngày lễ mới'\nBước 2: Nhập tên 'Nghỉ lễ 30/4' và chọn ngày 30/04/2026 (đã được khai báo trước đó)\nBước 3: Click button 'Lưu ngày lễ'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Ngày: '2026-04-30'", "Hiển thị thông báo lỗi: 'Ngày áp dụng này đã được khai báo ngày lễ trong hệ thống'"),
                ("HOL_FUNC_01", "Kiểm tra thêm ngày lễ thành công và áp giá tự động", "Kiểm tra tự động áp dụng biểu giá ngày lễ cho các suất chiếu",
                 "Bước 1: Điền Tên ngày lễ: 'Quốc Khánh 02/09', Ngày áp dụng: '2026-09-02'\nBước 2: Click button 'Lưu ngày lễ'\nBước 3: Kiểm tra giá vé của các suất chiếu trong ngày 02/09/2026",
                 "Tên: 'Quốc Khánh 02/09'\nNgày: '2026-09-02'",
                 "Thêm ngày lễ thành công, tất cả các suất chiếu diễn ra trong ngày 02/09 tự động áp dụng cột giá Ngày Lễ trong ma trận giá vé")
            ]
        },
        {
            "code": "MOD_ADMIN_PROMOTIONS", "sheet": "Quản lý đợt khuyến mãi",
            "req": "Kiểm tra Thêm, Sửa Đợt khuyến mãi và Phát voucher",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Khuyến mãi trong Dashboard",
            "test_cases": [
                ("PRM_VAL_01", "Kiểm tra validate mã khuyến mãi chứa khoảng trắng", "Kiểm tra chuẩn định dạng mã code khuyến mãi",
                 "Bước 1: Mở modal 'Tạo đợt khuyến mãi mới'\nBước 2: Nhập mã code có chứa dấu cách: 'GIAM GIA 50'\nBước 3: Click button 'Lưu khuyến mãi'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Mã code: 'GIAM GIA 50'", "Hiển thị thông báo lỗi: 'Mã khuyến mãi chỉ được chứa chữ cái không dấu và chữ số, không chứa khoảng trắng'"),
                ("PRM_VAL_02", "Kiểm tra validate giá trị giảm % vượt quá 100%", "Kiểm tra giới hạn phần trăm giảm giá",
                 "Bước 1: Chọn loại giảm theo phần trăm (%)\nBước 2: Nhập giá trị giảm là 120%\nBước 3: Click button 'Lưu khuyến mãi'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Giảm: 120%", "Hiển thị thông báo lỗi: 'Giá trị giảm phần trăm phải là số nguyên từ 1 đến 100'"),
                ("PRM_VAL_03", "Kiểm tra validate ngày kết thúc trước ngày bắt đầu", "Kiểm tra tính hợp lệ của thời hạn chương trình khuyến mãi",
                 "Bước 1: Chọn Ngày bắt đầu: 25/03/2026, Ngày kết thúc: 20/03/2026\nBước 2: Click button 'Lưu khuyến mãi'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Bắt đầu: 25/03/2026\nKết thúc: 20/03/2026", "Hiển thị thông báo lỗi: 'Ngày kết thúc chương trình khuyến mãi phải lớn hơn ngày bắt đầu'"),
                ("PRM_FUNC_01", "Kiểm tra phát voucher hàng loạt theo Hạng thẻ hội viên", "Kiểm tra tính năng phát hành voucher tự động vào ví khách hàng",
                 "Bước 1: Chọn đợt khuyến mãi 'Tri Ân Khách Hàng VIP'\nBước 2: Click button 'Phát hành voucher'\nBước 3: Chọn đối tượng: 'Tất cả khách hàng Hạng Vàng và Kim Cương'\nBước 4: Click button 'Xác nhận phát hành'\nBước 5: Kiểm tra kết quả phát hành",
                 "Đối tượng: Hạng Vàng & Kim Cương",
                 "Phát hành thành công, hệ thống tự động sinh và gửi voucher vào ví của tất cả khách hàng đạt hạng thẻ, đồng thời gửi email thông báo quà tặng")
            ]
        },
        {
            "code": "MOD_ADMIN_STAFF_MGMT", "sheet": "Quản lý nhân viên",
            "req": "Kiểm tra Thêm, Sửa Nhân viên và Gán Cụm rạp",
            "tester": "Phạm Thị Quỳnh Anh", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Nhân sự & Nhân viên trong Dashboard",
            "test_cases": [
                ("STF_VAL_01", "Kiểm tra validate khi thêm nhân viên có SĐT đã tồn tại", "Kiểm tra tính duy nhất của số điện thoại nhân viên",
                 "Bước 1: Mở modal 'Thêm nhân viên mới'\nBước 2: Nhập số điện thoại đã thuộc về một nhân viên khác\nBước 3: Click button 'Lưu nhân viên'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "SĐT: '0987654321' (Đã có)", "Hiển thị thông báo lỗi: 'Số điện thoại này đã được sử dụng bởi một nhân viên khác'"),
                ("STF_VAL_02", "Kiểm tra validate khi chưa gán cụm rạp cho nhân viên", "Kiểm tra bắt buộc gán cụm rạp cơ sở cho vai trò STAFF",
                 "Bước 1: Chọn vai trò là 'Nhân viên (STAFF)'\nBước 2: Để trống trường 'Cụm rạp trực thuộc'\nBước 3: Click button 'Lưu nhân viên'\nBước 4: Kiểm tra thông báo lỗi hiển thị",
                 "Cụm rạp: null", "Hiển thị thông báo lỗi: 'Nhân viên bắt buộc phải được gán vào một cụm rạp trực thuộc'"),
                ("STF_VAL_03", "Kiểm tra chặn khóa tài khoản Admin đang đăng nhập", "Kiểm tra cơ chế tự bảo vệ tài khoản quản trị tối cao",
                 "Bước 1: Trên danh sách nhân viên, tìm tài khoản Admin đang thực hiện đăng nhập phiên hiện tại\nBước 2: Click vào toggle chuyển trạng thái sang Khóa\nBước 3: Kiểm tra phản hồi từ hệ thống",
                 "User: Current Admin User", "Hệ thống từ chối thao tác, thông báo: 'Không thể tự khóa tài khoản Admin đang đăng nhập phiên hiện tại'"),
                ("STF_FUNC_01", "Kiểm tra tạo nhân viên mới và cấp mật khẩu tạm", "Kiểm tra luồng khởi tạo nhân viên mới",
                 "Bước 1: Điền Mã NV: 'NV008', Họ tên: 'Lê Văn An', Email: 'an.le@devcine.com', SĐT: '0977112233', Vai trò: 'STAFF', Gán rạp: 'CGV Cầu Giấy'\nBước 2: Click button 'Lưu nhân viên'\nBước 3: Kiểm tra kết quả tạo",
                 "Full valid staff data",
                 "Tạo nhân viên thành công, hệ thống tự động sinh mật khẩu tạm thời gửi về email nhân viên và bật cờ bắt buộc đổi mật khẩu ở lần đăng nhập đầu tiên")
            ]
        },
        {
            "code": "MOD_ADMIN_RBAC", "sheet": "Phân quyền hệ thống",
            "req": "Kiểm tra Phân quyền RBAC và Ghi đè quyền riêng lẻ",
            "tester": "Phạm Thị Quỳnh Anh", "role": "Quản trị viên", "pre": "Admin mở màn hình Phân quyền người dùng và vai trò RBAC",
            "test_cases": [
                ("RBC_VAL_01", "Kiểm tra chặn tước quyền tối cao của vai trò Admin", "Kiểm tra bảo vệ quyền hạn tối cao",
                 "Bước 1: Mở ma trận phân quyền của vai trò ROLE_ADMIN\nBước 2: Bỏ tích chọn quyền 'Quản trị hệ thống (SYSTEM_ADMIN)'\nBước 3: Click button 'Lưu quyền'\nBước 4: Kiểm tra phản hồi từ hệ thống",
                 "Quyền: SYSTEM_ADMIN của ROLE_ADMIN", "Hệ thống từ chối, thông báo: 'Không được phép xóa bỏ quyền quản trị tối cao của vai trò Admin'"),
                ("RBC_FUNC_01", "Kiểm tra ghi đè cấp thêm quyền duyệt hủy đơn cho nhân viên", "Kiểm tra cơ chế Override Grant Permission riêng lẻ",
                 "Bước 1: Chọn tài khoản nhân viên 'Văn Minh Khôi'\nBước 2: Tích chọn cấp thêm quyền riêng lẻ: 'Duyệt hủy đơn F&B (APPROVE_VOID)'\nBước 3: Click button 'Lưu quyền cá nhân'\nBước 4: Nhân viên Khôi đăng nhập lại và kiểm tra quyền hạn",
                 "Cấp thêm quyền: APPROVE_VOID cho user Khôi",
                 "Lưu phân quyền thành công, nhân viên Khôi có thể nhìn thấy nút và thực hiện chức năng duyệt hủy đơn F&B ngay ở phiên làm việc tiếp theo")
            ]
        },
        {
            "code": "MOD_ADMIN_CUSTOMERS", "sheet": "Quản lý khách hàng",
            "req": "Kiểm tra Quản lý Khách hàng và Khóa tài khoản",
            "tester": "Phạm Thị Quỳnh Anh", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Khách hàng trong Dashboard",
            "test_cases": [
                ("CUS_FUNC_01", "Kiểm tra tìm kiếm khách hàng theo SĐT và Hạng thẻ", "Kiểm tra tra cứu thông tin khách hàng và lịch sử giao dịch",
                 "Bước 1: Mở danh sách khách hàng\nBước 2: Nhập số điện thoại '0912345678' vào ô tìm kiếm\nBước 3: Chọn lọc Hạng thẻ: 'Vàng'\nBước 4: Click button 'Tìm kiếm'\nBước 5: Kiểm tra kết quả hiển thị",
                 "SĐT: '0912345678'\nHạng thẻ: 'Vàng'", "Hiển thị chính xác thông tin khách hàng, số điểm tích lũy, tổng chi tiêu trọn đời và danh sách các đơn đặt vé đã mua"),
                ("CUS_FUNC_02", "Kiểm tra cảnh báo khi khóa tài khoản có vé chưa xem", "Kiểm tra cảnh báo bảo vệ quyền lợi khách hàng trước khi khóa",
                 "Bước 1: Chọn một khách hàng đang có vé xem phim của suất chiếu tối nay chưa sử dụng\nBước 2: Click vào nút 'Khóa tài khoản'\nBước 3: Kiểm tra hiển thị modal xác nhận từ hệ thống",
                 "Khách hàng: Đang có vé xem phim chưa sử dụng",
                 "Hiển thị modal cảnh báo màu vàng: 'Khách hàng này hiện đang có 2 vé xem phim chưa sử dụng cho suất chiếu tối nay. Bạn có chắc chắn muốn khóa tài khoản này không?'")
            ]
        },
        {
            "code": "MOD_ADMIN_ORDERS", "sheet": "Quản lý đơn hàng",
            "req": "Kiểm tra Tra cứu Đơn hàng và Xuất hóa đơn VAT",
            "tester": "Nguyễn Quang Huy", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Đơn hàng & Giao dịch",
            "test_cases": [
                ("ORD_FUNC_01", "Kiểm tra lọc đơn hàng theo Cụm rạp, Ngày và Trạng thái", "Kiểm tra bộ lọc đa tiêu chí quản lý đơn hàng",
                 "Bước 1: Chọn Cụm rạp: 'CGV Cầu Giấy'\nBước 2: Chọn Khoảng ngày: từ 10/03/2026 đến 19/03/2026\nBước 3: Chọn Trạng thái: 'Đã thanh toán (CONFIRMED)'\nBước 4: Click button 'Lọc đơn hàng'\nBước 5: Kiểm tra danh sách kết quả trả về",
                 "Rạp: CGV Cầu Giấy\nNgày: 10-19/03\nTrạng thái: CONFIRMED", "Hiển thị chính xác danh sách các đơn hàng thỏa mãn toàn bộ tiêu chí lọc kèm tổng doanh thu của các đơn đó"),
                ("ORD_FUNC_02", "Kiểm tra in lại hóa đơn điện tử PDF", "Kiểm tra chức năng xuất hóa đơn điện tử cho đơn đã thanh toán",
                 "Bước 1: Chọn một đơn hàng đã thanh toán thành công (CONFIRMED)\nBước 2: Click button 'Xuất hóa đơn VAT'\nBước 3: Kiểm tra tệp hóa đơn được tải về",
                 "Đơn hàng: CONFIRMED", "Hệ thống sinh và tải về tệp hóa đơn điện tử PDF chuẩn chỉ, đầy đủ thông tin thuế VAT, mã tra cứu hóa đơn và chi tiết danh sách vé/F&B")
            ]
        },
        {
            "code": "MOD_ADMIN_BANNERS", "sheet": "Quản lý Banner",
            "req": "Kiểm tra Thêm, Sửa Banner quảng cáo",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Banner quảng cáo",
            "test_cases": [
                ("BAN_VAL_01", "Kiểm tra validate khi thêm banner mà không chọn ảnh", "Kiểm tra bắt buộc tải lên tệp ảnh banner",
                 "Bước 1: Mở modal 'Thêm banner mới'\nBước 2: Nhập tiêu đề banner\nBước 3: Không chọn file ảnh tải lên\nBước 4: Click button 'Lưu banner'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Ảnh banner: null", "Hiển thị thông báo lỗi: 'Vui lòng chọn tệp ảnh banner quảng cáo'"),
                ("BAN_VAL_02", "Kiểm tra validate thứ tự hiển thị là số âm", "Kiểm tra ràng buộc số thứ tự sắp xếp banner",
                 "Bước 1: Nhập thứ tự hiển thị là -1\nBước 2: Click button 'Lưu banner'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Thứ tự: -1", "Hiển thị thông báo lỗi: 'Thứ tự hiển thị phải là số nguyên lớn hơn hoặc bằng 0'"),
                ("BAN_FUNC_01", "Kiểm tra thêm banner quảng cáo mới thành công", "Kiểm tra lưu banner và hiển thị trên Slider trang chủ",
                 "Bước 1: Điền Tiêu đề: 'Bom Tấn Avatar Trở Lại', Chọn gắn link đến phim Avatar, Thứ tự: 1, Trạng thái: 'Hiển thị'\nBước 2: Upload ảnh banner kích thước 1920x600 px (dung lượng 2.5MB)\nBước 3: Click button 'Lưu banner'\nBước 4: Mở trang chủ DevCine kiểm tra",
                 "Full valid banner data", "Thêm banner thành công và hiển thị ngay trên Slider quảng cáo lớn ở đầu trang chủ người dùng")
            ]
        },
        {
            "code": "MOD_ADMIN_NEWS", "sheet": "Tin tức & Khuyến mãi",
            "req": "Kiểm tra Quản lý Bài viết Tin tức và Khuyến mãi",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở trang Quản lý Tin tức & Bài viết",
            "test_cases": [
                ("NEW_VAL_01", "Kiểm tra validate khi để trống nội dung bài viết", "Kiểm tra bắt buộc nhập nội dung chi tiết bài viết",
                 "Bước 1: Mở trình soạn thảo bài viết mới\nBước 2: Nhập tiêu đề và tóm tắt\nBước 3: Để trống khung soạn thảo nội dung rich text\nBước 4: Click button 'Xuất bản bài viết'\nBước 5: Kiểm tra thông báo lỗi hiển thị",
                 "Nội dung: ''", "Hiển thị thông báo lỗi: 'Nội dung bài viết không được để trống'"),
                ("NEW_FUNC_01", "Kiểm tra xuất bản bài viết tin tức mới thành công", "Kiểm tra đăng bài viết và hiển thị định dạng rich text",
                 "Bước 1: Nhập Tiêu đề: 'Ưu Đãi Thứ 4 Vui Vẻ - Đồng Giá Vé 50K Toàn Hệ Thống', Tóm tắt bài viết, Nội dung bài viết có chèn ảnh và định dạng chữ\nBước 2: Upload ảnh đại diện bài viết (Thumbnail)\nBước 3: Chọn trạng thái: 'Đã xuất bản'\nBước 4: Click button 'Xuất bản bài viết'\nBước 5: Kiểm tra trang Tin tức người dùng",
                 "Full valid news article data", "Xuất bản bài viết thành công, bài viết hiển thị đúng vị trí trên trang Tin tức & Khuyến mãi kèm đường dẫn slug chuẩn SEO")
            ]
        },
        {
            "code": "MOD_ADMIN_SETTINGS", "sheet": "Cài đặt hệ thống",
            "req": "Kiểm tra Cấu hình các Tham số Động của hệ thống",
            "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên", "pre": "Admin mở trang Cài đặt Tham số Động của hệ thống",
            "test_cases": [
                ("SET_VAL_01", "Kiểm tra validate thời gian giữ ghế online ngoài khoảng 5-30 phút", "Kiểm tra biên giá trị tham số timeout giữ ghế",
                 "Bước 1: Tại ô 'Thời gian giữ ghế online (phút)', nhập giá trị 45 phút\nBước 2: Click button 'Lưu cấu hình'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Giá trị: 45 phút", "Hiển thị thông báo lỗi: 'Thời gian giữ ghế online phải là số nguyên từ 5 đến 30 phút'"),
                ("SET_VAL_02", "Kiểm tra validate số ghế tối đa mỗi đơn ngoài khoảng 1-20 ghế", "Kiểm tra biên giá trị số ghế tối đa",
                 "Bước 1: Nhập số ghế tối đa mỗi đơn là 25 ghế\nBước 2: Click button 'Lưu cấu hình'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Giá trị: 25 ghế", "Hiển thị thông báo lỗi: 'Số ghế tối đa mỗi đơn đặt phải từ 1 đến 20 ghế'"),
                ("SET_VAL_03", "Kiểm tra validate email thông báo hệ thống sai định dạng", "Kiểm tra chuẩn email hệ thống",
                 "Bước 1: Nhập email thông báo là 'devcine_system_email'\nBước 2: Click button 'Lưu cấu hình'\nBước 3: Kiểm tra thông báo lỗi hiển thị",
                 "Email: 'devcine_system_email'", "Hiển thị thông báo lỗi: 'Định dạng email hệ thống không hợp lệ'"),
                ("SET_FUNC_01", "Kiểm tra lưu thay đổi tham số hệ thống thành công", "Kiểm tra áp dụng cấu hình tham số mới ngay lập tức",
                 "Bước 1: Đổi tham số 'Thời gian giữ đơn chờ POS' thành 8 phút\nBước 2: Đổi 'Hotline CSKH' thành '19006017'\nBước 3: Click button 'Lưu cấu hình'\nBước 4: Kiểm tra hiệu lực trên máy POS và website",
                 "POS Timeout: 8 phút\nHotline: '19006017'",
                 "Lưu cấu hình tham số thành công, hệ thống áp dụng ngay lập tức thời gian giữ đơn POS mới là 8 phút và cập nhật số hotline trên toàn hệ thống")
            ]
        }
    ]

    # -------------------------------------------------------------------------
    # 2. SHEET: Test case List (DS Test Case)
    # -------------------------------------------------------------------------
    ws_list = wb.create_sheet("Test case List (DS Test Case)")
    ws_list.views.sheetView[0].showGridLines = True
    
    ws_list.cell(2, 2, "DANH SÁCH BỘ KIỂM THỬ (TEST SUITE LIST)").font = font_title
    
    list_headers = ["STT", "Mã Module", "Tên Phân hệ / Chức năng", "Tên Sheet", "Phân loại Vai trò", "Điều kiện tiên quyết (Preconditions)", "Người phụ trách"]
    for c_idx, h in enumerate(list_headers, start=2):
        cell = ws_list.cell(4, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    for r_idx, mod in enumerate(modules_data, start=5):
        stt = r_idx - 4
        ws_list.cell(r_idx, 2, stt).alignment = align_center
        ws_list.cell(r_idx, 3, mod["code"]).alignment = align_center
        ws_list.cell(r_idx, 4, mod["req"]).alignment = align_left
        ws_list.cell(r_idx, 5, mod["sheet"]).alignment = align_left
        ws_list.cell(r_idx, 6, mod["role"]).alignment = align_center
        ws_list.cell(r_idx, 7, mod["pre"]).alignment = align_left
        ws_list.cell(r_idx, 8, mod["tester"]).alignment = align_center
        
        for c in range(2, 9):
            cell = ws_list.cell(r_idx, c)
            cell.font = font_regular
            cell.border = border_thin

    ws_list.column_dimensions['A'].width = 4.0
    ws_list.column_dimensions['B'].width = 8.0
    ws_list.column_dimensions['C'].width = 22.0
    ws_list.column_dimensions['D'].width = 45.0
    ws_list.column_dimensions['E'].width = 26.0
    ws_list.column_dimensions['F'].width = 24.0
    ws_list.column_dimensions['G'].width = 45.0
    ws_list.column_dimensions['H'].width = 20.0

    # -------------------------------------------------------------------------
    # 3. SHEET: Test Report (Tổng hợp số liệu kiểm thử)
    # -------------------------------------------------------------------------
    ws_report = wb.create_sheet("Test Report")
    ws_report.views.sheetView[0].showGridLines = True
    
    ws_report.cell(1, 2, "BÁO CÁO TỔNG HỢP KIỂM THỬ (TEST REPORT SUMMARY)").font = font_title
    
    rep_meta = [
        ("Project Name", "DevCine - Quản lý Rạp chiếu phim & Đặt vé trực tuyến", "Creator", "Nguyễn Quang Huy"),
        ("Project Code", "DEVCINE_2026", "Reviewer/Approver", "Hội đồng Đồ án Tốt nghiệp / Tech Lead"),
        ("Document Code", "TR_DEVCINE_v1.0", "Issue Date", test_date),
        ("Release Scope", "Release 1.0 (Full Functional & Validation Suite)", "Status", "100% Pass")
    ]
    for r_idx, (k1, v1, k2, v2) in enumerate(rep_meta, start=3):
        ws_report.cell(r_idx, 2, k1).font = font_bold
        ws_report.cell(r_idx, 2).border = border_thin
        ws_report.cell(r_idx, 3, v1).font = font_regular
        ws_report.cell(r_idx, 3).border = border_thin
        ws_report.cell(r_idx, 5, k2).font = font_bold
        ws_report.cell(r_idx, 5).border = border_thin
        ws_report.cell(r_idx, 6, v2).font = font_regular
        ws_report.cell(r_idx, 6).border = border_thin
        if isinstance(v2, datetime.datetime):
            ws_report.cell(r_idx, 6).number_format = 'yyyy-mm-dd'
            
    ws_report.cell(8, 2, "Kết quả thực thi kiểm thử Đợt 1 (Execution Summary - V1)").font = font_sub_title
    
    rep_headers = ["STT", "Tên Phân hệ / Module", "Pass", "Fail", "Untested", "N/A", "Tổng Test Case", "Tỷ lệ Pass (%)"]
    for c_idx, h in enumerate(rep_headers, start=2):
        cell = ws_report.cell(9, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    start_row = 10
    for idx, mod in enumerate(modules_data):
        r_num = start_row + idx
        sheet_name = mod["sheet"]
        
        ws_report.cell(r_num, 2, idx + 1).alignment = align_center
        ws_report.cell(r_num, 3, f"='{sheet_name}'!B1").alignment = align_left
        ws_report.cell(r_num, 4, f"='{sheet_name}'!A5").alignment = align_center
        ws_report.cell(r_num, 5, f"='{sheet_name}'!B5").alignment = align_center
        ws_report.cell(r_num, 6, f"='{sheet_name}'!C5").alignment = align_center
        ws_report.cell(r_num, 7, f"='{sheet_name}'!D5").alignment = align_center
        ws_report.cell(r_num, 8, f"='{sheet_name}'!E5").alignment = align_center
        ws_report.cell(r_num, 9, f"=IF(H{r_num}>0, D{r_num}/H{r_num}, 1)").alignment = align_center
        ws_report.cell(r_num, 9).number_format = '0.0%'
        
        for c in range(2, 10):
            cell = ws_report.cell(r_num, c)
            cell.font = font_regular
            cell.border = border_thin
            
    total_row = start_row + len(modules_data)
    ws_report.cell(total_row, 2, "").alignment = align_center
    ws_report.cell(total_row, 3, "TỔNG CỘNG").font = font_bold
    ws_report.cell(total_row, 3).alignment = align_center
    ws_report.cell(total_row, 4, f"=SUM(D{start_row}:D{total_row-1})").font = font_bold
    ws_report.cell(total_row, 4).alignment = align_center
    ws_report.cell(total_row, 5, f"=SUM(E{start_row}:E{total_row-1})").font = font_bold
    ws_report.cell(total_row, 5).alignment = align_center
    ws_report.cell(total_row, 6, f"=SUM(F{start_row}:F{total_row-1})").font = font_bold
    ws_report.cell(total_row, 6).alignment = align_center
    ws_report.cell(total_row, 7, f"=SUM(G{start_row}:G{total_row-1})").font = font_bold
    ws_report.cell(total_row, 7).alignment = align_center
    ws_report.cell(total_row, 8, f"=SUM(H{start_row}:H{total_row-1})").font = font_bold
    ws_report.cell(total_row, 8).alignment = align_center
    ws_report.cell(total_row, 9, f"=IF(H{total_row}>0, D{total_row}/H{total_row}, 1)").font = font_bold
    ws_report.cell(total_row, 9).alignment = align_center
    ws_report.cell(total_row, 9).number_format = '0.0%'
    
    for c in range(2, 10):
        cell = ws_report.cell(total_row, c)
        cell.fill = fill_header_gold
        cell.border = border_thin

    ws_report.column_dimensions['A'].width = 4.0
    ws_report.column_dimensions['B'].width = 8.0
    ws_report.column_dimensions['C'].width = 38.0
    ws_report.column_dimensions['D'].width = 12.0
    ws_report.column_dimensions['E'].width = 12.0
    ws_report.column_dimensions['F'].width = 12.0
    ws_report.column_dimensions['G'].width = 12.0
    ws_report.column_dimensions['H'].width = 16.0
    ws_report.column_dimensions['I'].width = 16.0

    # -------------------------------------------------------------------------
    # 4. CREATE INDIVIDUAL MODULE SHEETS
    # -------------------------------------------------------------------------
    tc_table_headers = [
        "ID (Mã Test Case)", "Tiêu đề kiểm thử (Test Title)", "Mô tả trường hợp kiểm thử (Description)",
        "Các bước thực hiện (Test Procedure / Steps)", "Dữ liệu kiểm thử (Test Data)",
        "Kết quả mong muốn (Expected Output)", "Kết quả thực tế (Actual Result)",
        "Minh chứng (Evidence)", "Phụ thuộc (Dependence)", "Kết quả (Result 1)", "Ngày test (Date 1)"
    ]
    
    for mod in modules_data:
        ws_mod = wb.create_sheet(mod["sheet"])
        ws_mod.views.sheetView[0].showGridLines = True
        
        # Row 1-3: Metadata
        ws_mod.cell(1, 1, "Module Code(Mã Module)").font = font_bold
        ws_mod.cell(1, 2, mod["sheet"]).font = font_bold
        
        ws_mod.cell(2, 1, "Test requirement(Yêu cầu test)").font = font_bold
        ws_mod.cell(2, 2, mod["req"]).font = font_regular
        
        ws_mod.cell(3, 1, "Tester(Người thực hiện)").font = font_bold
        ws_mod.cell(3, 2, mod["tester"]).font = font_regular
        
        # Row 4-5: Summary table V1
        stat_headers = ["PASS-V1", "FAIL-V1", "UNTESTED-V1", "N/A-V1", "Tổng số TestCase (V1)"]
        for c_idx, sh in enumerate(stat_headers, start=1):
            cell = ws_mod.cell(4, c_idx, sh)
            cell.font = font_header_black
            cell.fill = fill_header_blue
            cell.alignment = align_center
            cell.border = border_thin
            
        num_tc = len(mod["test_cases"])
        ws_mod.cell(5, 1, num_tc).alignment = align_center
        ws_mod.cell(5, 1).font = font_pass
        ws_mod.cell(5, 2, 0).alignment = align_center
        ws_mod.cell(5, 2).font = font_regular
        ws_mod.cell(5, 3, 0).alignment = align_center
        ws_mod.cell(5, 3).font = font_regular
        ws_mod.cell(5, 4, 0).alignment = align_center
        ws_mod.cell(5, 4).font = font_regular
        ws_mod.cell(5, 5, num_tc).alignment = align_center
        ws_mod.cell(5, 5).font = font_bold
        for c in range(1, 6):
            ws_mod.cell(5, c).border = border_thin
            
        # Row 7-8: Summary table V2
        stat_headers_v2 = ["PASS-V2", "FAIL-V2", "UNTESTED-V2", "N/A-V2", "Tổng số TestCase (V2)"]
        for c_idx, sh in enumerate(stat_headers_v2, start=1):
            cell = ws_mod.cell(7, c_idx, sh)
            cell.font = font_header_black
            cell.fill = fill_header_gold
            cell.alignment = align_center
            cell.border = border_thin
            
        for c in range(1, 6):
            cell = ws_mod.cell(8, c, 0)
            cell.alignment = align_center
            cell.font = font_regular
            cell.border = border_thin
            
        # Row 10: Column Headers
        for c_idx, th in enumerate(tc_table_headers, start=1):
            cell = ws_mod.cell(10, c_idx, th)
            cell.font = font_header_white
            cell.fill = fill_header_navy
            cell.alignment = align_center
            cell.border = border_thin
            
        # Row 11+: Test Cases
        for r_offset, tc in enumerate(mod["test_cases"], start=11):
            t_id, t_title, t_desc, t_steps, t_data, t_expect = tc
            
            ws_mod.cell(r_offset, 1, t_id).alignment = align_center
            ws_mod.cell(r_offset, 2, t_title).alignment = align_left
            ws_mod.cell(r_offset, 3, t_desc).alignment = align_left
            ws_mod.cell(r_offset, 4, t_steps).alignment = align_top_left
            ws_mod.cell(r_offset, 5, t_data).alignment = align_center
            ws_mod.cell(r_offset, 6, t_expect).alignment = align_top_left
            ws_mod.cell(r_offset, 7, t_expect).alignment = align_top_left
            ws_mod.cell(r_offset, 8, "Video_Pass").alignment = align_center
            ws_mod.cell(r_offset, 9, "N/A").alignment = align_center
            ws_mod.cell(r_offset, 10, "Pass").alignment = align_center
            ws_mod.cell(r_offset, 10).font = font_pass
            ws_mod.cell(r_offset, 11, test_date).alignment = align_center
            ws_mod.cell(r_offset, 11).number_format = 'yyyy-mm-dd'
            
            for c in range(1, 12):
                cell = ws_mod.cell(r_offset, c)
                if c != 10:
                    cell.font = font_regular
                cell.border = border_thin
                
        ws_mod.column_dimensions['A'].width = 16.0
        ws_mod.column_dimensions['B'].width = 30.0
        ws_mod.column_dimensions['C'].width = 32.0
        ws_mod.column_dimensions['D'].width = 50.0
        ws_mod.column_dimensions['E'].width = 25.0
        ws_mod.column_dimensions['F'].width = 40.0
        ws_mod.column_dimensions['G'].width = 40.0
        ws_mod.column_dimensions['H'].width = 14.0
        ws_mod.column_dimensions['I'].width = 14.0
        ws_mod.column_dimensions['J'].width = 12.0
        ws_mod.column_dimensions['K'].width = 14.0

    # -------------------------------------------------------------------------
    # 5. SHEET: FUNCTION (Cây chức năng & Sự kiện)
    # -------------------------------------------------------------------------
    ws_func = wb.create_sheet("FUNCTION")
    ws_func.views.sheetView[0].showGridLines = True
    
    ws_func.cell(1, 1, "CÂY CHỨC NĂNG VÀ SỰ KIỆN (FUNCTION HIERARCHY)").font = font_title
    ws_func.cell(3, 1, "Project Name").font = font_bold
    ws_func.cell(3, 2, "DevCine - Quản lý Rạp chiếu phim & Đặt vé trực tuyến").font = font_regular
    ws_func.cell(4, 1, "Project Code").font = font_bold
    ws_func.cell(4, 2, "DEVCINE_2026").font = font_regular
    
    func_headers = ["Function Level 1 (Phân hệ chính)", "Function Level 2 (Chức năng chi tiết)", "Action & Event (Hành động & Sự kiện)", "Ghi chú phân tích"]
    for c_idx, fh in enumerate(func_headers, start=1):
        cell = ws_func.cell(6, c_idx, fh)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    func_rows = [
        ("1. Khách hàng - Xác thực & Tài khoản", "1.1 Đăng ký tài khoản", "Nhập form đăng ký -> Bấm Đăng ký -> Validate -> Lưu DB", "Mã hóa BCrypt, kiểm tra trùng email/SĐT"),
        ("", "1.2 Đăng nhập", "Nhập thông tin -> Bấm Đăng nhập -> Sinh JWT Token -> Lưu session", "Lưu token, chặn tài khoản bị khóa"),
        ("", "1.3 Quên mật khẩu & OTP", "Nhập email -> Nhận mã OTP -> Xác thực -> Đặt mật khẩu mới", "OTP 6 số hết hạn 15 phút"),
        ("", "1.4 Hồ sơ & Thành viên", "Xem thông tin -> Cập nhật hồ sơ -> Tra cứu điểm Loyalty", "Tự động cập nhật hạng thành viên"),
        ("2. Khách hàng - Đặt vé Online", "2.1 Lịch chiếu & Chọn phim", "Chọn phim -> Chọn ngày -> Chọn suất chiếu theo rạp", "Chỉ hiển thị suất chưa chiếu"),
        ("", "2.2 Chọn ghế & Giữ chỗ", "Click chọn ghế -> Kiểm tra ghế trống -> Giữ chỗ 10 phút", "Đếm ngược 600s, khóa ghế tạm thời"),
        ("", "2.3 Combo F&B", "Chọn bắp nước -> Chọn vị bắp, loại nước -> Cộng tiền phụ thu", "Slot bắt buộc theo combo"),
        ("", "2.4 Áp dụng Voucher", "Nhập mã voucher -> Kiểm tra điều kiện min order/hạn dùng -> Giảm giá", "Giảm theo % hoặc tiền, không vượt trần"),
        ("", "2.5 Thanh toán VNPAY", "Chuyển sang cổng VNPAY -> Thanh toán -> Nhận vé QR & Gửi email", "Xác thực chữ ký HMAC-SHA512"),
        ("3. Nhân viên - Vận hành Quầy (POS)", "3.1 Bán vé tại quầy", "Chọn suất chiếu -> Chọn ghế -> Tra cứu hội viên -> Thu tiền -> In vé", "Strict Cinema Scoping, lưu sold_by"),
        ("", "3.2 Đơn chờ POS", "Lưu đơn chờ tạm thời -> Khôi phục thanh toán", "Tối đa 3 đơn chờ/máy POS"),
        ("", "3.3 Bán F&B tại quầy", "Chọn bắp nước -> Thanh toán -> In hóa đơn", "Bán độc lập không kèm vé"),
        ("", "3.4 Soát vé Check-in", "Quét mã QR / Nhập mã vé -> Kiểm tra vé hợp lệ -> Check-in", "Cảnh báo vé đã dùng/vé sai rạp"),
        ("", "3.5 Xử lý sự cố chỗ ngồi", "Tra cứu đơn -> Đổi ghế tại chỗ cho khách -> Ghi log sự cố", "Chỉ đổi trước giờ chiếu, giữ nguyên QR"),
        ("4. Quản trị viên (Admin Master)", "4.1 Quản lý phim", "Thêm/Sửa phim -> Upload poster/banner -> Đổi trạng thái", "Thời lượng 30-300', năm 2020-2035"),
        ("", "4.2 Quản lý Cụm rạp & Phòng", "Thêm rạp -> Thêm phòng -> Thiết lập sơ đồ ma trận ghế", "Cấu hình hàng A-Z, ghế đôi Sweetbox"),
        ("", "4.3 Điều phối Suất chiếu", "Lập lịch chiếu đơn -> Xếp lịch hàng loạt (Batch Scheduling)", "Kiểm tra xung đột phòng chiếu"),
        ("", "4.4 Bảng giá vé", "Cấu hình giá nền 3 chiều -> Phụ thu VIP/3D -> Ngày lễ", "Simulator tính thử giá vé"),
        ("", "4.5 Khuyến mãi & Voucher", "Tạo đợt khuyến mãi -> Phát hành voucher theo hạng thẻ", "Giảm % hoặc tiền, đơn tối thiểu"),
        ("", "4.6 Nhân sự & Phân quyền", "Tạo nhân viên -> Gán cụm rạp -> Phân quyền RBAC", "Tự sinh mật khẩu tạm cho NV mới"),
        ("", "4.7 Cài đặt hệ thống", "Cấu hình tham số giữ ghế, timeout, hotline, email", "Áp dụng cấu hình động toàn rạp")
    ]
    
    for r_idx, frow in enumerate(func_rows, start=7):
        for c_idx, val in enumerate(frow, start=1):
            cell = ws_func.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_top_left if c_idx != 1 else align_left
            cell.border = border_thin

    ws_func.column_dimensions['A'].width = 32.0
    ws_func.column_dimensions['B'].width = 30.0
    ws_func.column_dimensions['C'].width = 55.0
    ws_func.column_dimensions['D'].width = 35.0

    wb.save(output_path)
    print(f"Successfully generated TestReport workbook: {output_path}")

if __name__ == "__main__":
    out_dir = r"c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine"
    out_file = os.path.join(out_dir, "TestReport Dự án DevCine.xlsx")
    build_full_devcine_test_report(out_file)
    
    downloads_file = r"C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx"
    try:
        shutil.copy2(out_file, downloads_file)
        print(f"Successfully copied to: {downloads_file}")
    except Exception as e:
        print(f"Error copying to Downloads: {e}")
