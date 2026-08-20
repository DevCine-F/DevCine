# -*- coding: utf-8 -*-
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("TestReport Dự án DevCine.xlsx", data_only=False)
ws = wb["POS Bán F&B tại quầy"]

print("=== SHEET: POS Bán F&B tại quầy ===")
print("Max rows:", ws.max_row)
for r in range(1, ws.max_row + 1):
    c1 = ws.cell(r, 1).value
    c2 = ws.cell(r, 2).value
    c3 = ws.cell(r, 3).value
    c8 = ws.cell(r, 8).value
    c9 = ws.cell(r, 9).value
    print(f"Row {r:02d}: A={c1} | B={str(c2)[:35]} | I={c9} | H={c8}")

print("\nFormulas in Row 5:")
print(f"A5: {ws.cell(5, 1).value}")
print(f"B5: {ws.cell(5, 2).value}")
print(f"C5: {ws.cell(5, 3).value}")

ws_tr = wb["Test Report"]
print(f"\nTest Report Row 44 (POS Bán F&B): C44={ws_tr.cell(44, 3).value} | D44={ws_tr.cell(44, 4).value}")
print(f"Test Report Total Row 45: C45={ws_tr.cell(45, 3).value} | D45={ws_tr.cell(45, 4).value}")
