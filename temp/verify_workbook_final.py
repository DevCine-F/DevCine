# -*- coding: utf-8 -*-
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("TestReport Dự án DevCine.xlsx", data_only=False)
print("TOTAL SHEETS:", len(wb.sheetnames))
print("\nLIST OF ALL SHEETS:")
for i, s in enumerate(wb.sheetnames, start=1):
    ws = wb[s]
    max_r = ws.max_row
    print(f"{i:02d}. {s} (Max rows: {max_r})")

ws_tr = wb["Test Report"]
print("\nTEST REPORT ROW 4 to LAST:")
for r in range(3, ws_tr.max_row + 1):
    row_vals = [str(ws_tr.cell(r, c).value) for c in range(1, 9)]
    print(f"Row {r:02d}: " + " | ".join(row_vals))

print("\nVerification completed successfully!")
