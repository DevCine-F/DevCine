# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine\TestReport Dự án DevCine.xlsx', data_only=False)

print("Workbook sheet count:", len(wb.sheetnames))
print("Sheets list:", wb.sheetnames)

ws = wb['Quản lý phim']
print("\n--- QUẢN LÝ PHIM SHEET AUDIT ---")
print("Sheet title:", ws.title)
print("Cell A5 formula:", ws['A5'].value)
print("Cell B5 formula:", ws['B5'].value)
print("Cell H1-H4:", [ws[f'H{i}'].value for i in range(1, 5)])

sections = []
tcs = []
for r in range(11, ws.max_row + 1):
    val_a = str(ws.cell(r, 1).value or "")
    val_b = str(ws.cell(r, 2).value or "")
    if val_a.startswith('MOV_'):
        tcs.append((r, val_a, val_b))
    elif val_a:
        sections.append((r, val_a))

print(f"Total Test Cases in Quản lý phim: {len(tcs)}")
print("Sections found:")
for s in sections:
    print(f"  Row {s[0]}: {s[1]}")

print(f"\nFirst 5 test cases:")
for t in tcs[:5]:
    print(f"  {t[1]}: {t[2]}")

print(f"\nEdit section test cases:")
edit_tcs = [t for t in tcs if 'EDIT' in t[1]]
print(f"Total Edit Test cases: {len(edit_tcs)}")
for t in edit_tcs[:5]:
    print(f"  {t[1]}: {t[2]}")

ws_tr = wb['Test Report']
print("\n--- TEST REPORT SHEET AUDIT ---")
print("Total rows in Test Report:", ws_tr.max_row)
for r in range(3, ws_tr.max_row + 1):
    c1 = ws_tr.cell(r, 1).value
    c2 = ws_tr.cell(r, 2).value
    c3 = ws_tr.cell(r, 3).value
    c7 = ws_tr.cell(r, 7).value
    c8 = ws_tr.cell(r, 8).value
    if c2 == 'Quản lý phim' or c1 == 'TỔNG CỘNG':
        print(f"  Row {r}: {c1} | {c2} | {c3} | Rate: {c7} | Status: {c8}")
