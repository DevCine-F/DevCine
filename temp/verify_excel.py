# -*- coding: utf-8 -*-
"""
Verify the generated TestReport Excel file.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine\TestReport Dự án DevCine.xlsx")
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")

ws_rbac = wb["Phân quyền hệ thống"]
print("\n--- Sheet 'Phân quyền hệ thống' ---")
for r in range(1, 40):
    row_vals = [ws_rbac.cell(row=r, column=c).value for c in range(1, 12)]
    if any(row_vals):
        print(f"Row {r:02d}: A={row_vals[0]} | B={row_vals[1]} | C={row_vals[2]} | H={row_vals[7]}")
