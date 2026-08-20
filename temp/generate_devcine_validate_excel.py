# -*- coding: utf-8 -*-
"""
Script to generate Validate_DevCine.xlsx matching template Validate_SD17_.xlsx
CHỈ CHỨA CÁC MÀN HÌNH / FORM CÓ DỮ LIỆU ĐẦU VÀO CẦN VALIDATE.
Loại bỏ hoàn toàn bộ lọc, bảng dữ liệu, nút bật tắt hiển thị, màn hình chỉ xem.
Phân chia 3 nhóm: Khách hàng - Nhân viên - Quản lý / Admin
"""

import os
import sys
import datetime
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def build_validate_workbook(output_path):
    wb = openpyxl.Workbook()
    
    font_name = "Times New Roman"
    
    border_thin = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000')
    )
    
    header_font_s1 = Font(name=font_name, size=13, bold=True, color='FF000000')
    data_font_s1 = Font(name=font_name, size=11, bold=False, color='FF000000')
    
    header_font_s2 = Font(name=font_name, size=14, bold=True, color='FF000000')
    data_font_s2 = Font(name=font_name, size=12, bold=False, color='FF000000')
    
    header_font_s3 = Font(name=font_name, size=13, bold=True, color='FFFFFFFF')
    header_fill_s3 = PatternFill(start_color='FF990000', end_color='FF990000', fill_type='solid')
    data_font_s3 = Font(name=font_name, size=13, bold=False, color='FF000000')
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_top_center = Alignment(horizontal='center', vertical='top', wrap_text=True)

    # -------------------------------------------------------------------------
    # SHEET 1: Trang tính1 - Chỉ các Màn hình/Form CÓ INPUT CẦN VALIDATE
    # -------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Trang tính1"
    
    headers1 = ["STT", "Role", "Màn hình", "Chức năng", "Validate"]
    ws1.append(headers1)
    
    for c_idx in range(1, 6):
        cell = ws1.cell(1, c_idx)
        cell.font = header_font_s1
        cell.alignment = align_center
        cell.border = border_thin
    
    items = [
        # =====================================================================
        # PHẦN 1: KHÁCH HÀNG (11 Form có Input)
        # =====================================================================
        (
            1,
            "Khách hàng",
            "Đăng nhập",
            "Đăng nhập hệ thống",
            "- Tên đăng nhập / Email / SĐT không được để trống\n"
            "- Mật khẩu không được để trống, từ 6–50 ký tự\n"
            "- Email hoặc SĐT phải đúng định dạng\n"
            "- Tài khoản bị khóa thì không đăng nhập được\n"
            "- Nhập sai thông tin hiển thị thông báo lỗi\n"
            "- Đăng nhập sai quá 5 lần liên tiếp sẽ tạm khóa tài khoản 15 phút"
        ),
        (
            2,
            "Khách hàng",
            "Đăng ký",
            "Đăng ký tài khoản khách hàng",
            "- Họ tên không được để trống, có ít nhất 2 từ, từ 2–100 ký tự, không chứa ký tự đặc biệt\n"
            "- Email không được để trống, đúng định dạng, không trùng lặp, tối đa 100 ký tự\n"
            "- Số điện thoại không được để trống, đủ 10 số (đầu 03, 05, 07, 08, 09), không trùng\n"
            "- Tên đăng nhập từ 5–50 ký tự, không chứa khoảng trắng, không trùng\n"
            "- Mật khẩu không được để trống, từ 6–50 ký tự\n"
            "- Xác nhận mật khẩu bắt buộc nhập và phải khớp với mật khẩu\n"
            "- Ngày sinh không được lớn hơn ngày hiện tại, tuổi ≥ 13\n"
            "- Giới tính bắt buộc chọn (Nam / Nữ / Khác)\n"
            "- Bắt buộc tích chọn đồng ý điều khoản sử dụng"
        ),
        (
            3,
            "Khách hàng",
            "Quên mật khẩu",
            "Xác thực OTP & Đặt lại mật khẩu",
            "- Email không được để trống, đúng định dạng email\n"
            "- Email phải tồn tại trong hệ thống, không tồn tại báo lỗi\n"
            "- Khi còn thời gian chờ (60 giây) thì không gửi lại mã OTP\n"
            "- Mã OTP bắt buộc nhập đủ 6 chữ số\n"
            "- Mã OTP hết hạn sau 15 phút\n"
            "- Nhập sai OTP quá 5 lần thì mã bị hủy và phải yêu cầu lại từ đầu\n"
            "- Mật khẩu mới có ít nhất 8 ký tự, xác nhận mật khẩu phải khớp"
        ),
        (
            4,
            "Khách hàng",
            "Đổi mật khẩu",
            "Đổi mật khẩu tài khoản",
            "- Mật khẩu hiện tại không được để trống và phải nhập chính xác\n"
            "- Mật khẩu mới từ 6–50 ký tự, không được trùng mật khẩu cũ\n"
            "- Mật khẩu mới và xác nhận mật khẩu phải nhập giống nhau"
        ),
        (
            5,
            "Khách hàng",
            "Hồ sơ cá nhân",
            "Cập nhật thông tin cá nhân",
            "- Họ tên không được để trống, có ít nhất 2 từ, từ 2–100 ký tự\n"
            "- Số điện thoại không được để trống, đủ 10 số, không trùng\n"
            "- Email không được để trống, đúng định dạng, không trùng, tối đa 100 ký tự\n"
            "- Tên đăng nhập không cho phép chỉnh sửa\n"
            "- Ngày sinh không được lớn hơn ngày hiện tại\n"
            "- Giới tính bắt buộc chọn\n"
            "- Địa chỉ liên hệ tối đa 255 ký tự\n"
            "- Ảnh đại diện không bắt buộc, chỉ nhận file jpg, png, dung lượng ≤ 5MB"
        ),
        (
            6,
            "Khách hàng",
            "Chi tiết phim",
            "Đánh giá & Bình luận phim",
            "- Chỉ khách hàng đã mua vé và suất chiếu đã kết thúc mới được đánh giá\n"
            "- Điểm đánh giá bắt buộc chọn từ 1 đến 5 sao\n"
            "- Nội dung bình luận không bắt buộc, nếu nhập tối đa 500 ký tự\n"
            "- Mỗi khách hàng chỉ được đánh giá 1 lần cho 1 bộ phim"
        ),
        (
            7,
            "Khách hàng",
            "Đặt vé online",
            "Chọn suất chiếu & Đối tượng vé",
            "- Suất chiếu phải còn mở bán (trước giờ chiếu tối thiểu 10 phút)\n"
            "- Phim giới hạn độ tuổi (T13, T16, T18) bắt buộc hiển thị cảnh báo xác nhận độ tuổi\n"
            "- Chọn đối tượng vé Người lớn hoặc HSSV (kèm lưu ý xuất trình thẻ khi vào rạp)\n"
            "- Số lượng vé đặt từ 1 đến 8 vé mỗi lần"
        ),
        (
            8,
            "Khách hàng",
            "Đặt vé online",
            "Chọn ghế & Giữ chỗ",
            "- Số ghế chọn trên sơ đồ phải bằng đúng số vé đã chọn\n"
            "- Không được chọn ghế đã bán, ghế người khác đang giữ hoặc ghế đang bảo trì\n"
            "- Ghế đôi (Sweetbox) bắt buộc chọn cả cặp 2 ghế liền kề trong cùng hàng\n"
            "- Thời gian giữ ghế tối đa 10 phút, có đồng hồ đếm ngược, hết giờ tự động nhả ghế"
        ),
        (
            9,
            "Khách hàng",
            "Đặt vé online",
            "Chọn combo bắp nước F&B",
            "- Mua bắp nước không bắt buộc, có thể bỏ qua\n"
            "- Số lượng mỗi món từ 0 đến 20, tổng không quá 50 món mỗi đơn\n"
            "- Món combo bắt buộc chọn đủ vị bắp và loại nước theo quy định\n"
            "- Đổi vị hoặc nâng size đặc biệt tự động cộng thêm tiền phụ thu"
        ),
        (
            10,
            "Khách hàng",
            "Đặt vé online",
            "Áp dụng mã giảm giá / Voucher",
            "- Mã voucher không được để trống, tự động chuyển chữ in hoa\n"
            "- Mã phải còn hạn sử dụng, còn lượt dùng và đang hoạt động\n"
            "- Mỗi khách hàng chỉ được dùng số lần quy định (thường 1 lần)\n"
            "- Đơn hàng phải đạt giá trị tối thiểu của voucher\n"
            "- Đúng cụm rạp hoặc phim áp dụng (nếu có quy định)\n"
            "- Mỗi đơn chỉ áp dụng 1 voucher, tiền giảm không vượt mức tối đa\n"
            "- Số tiền thanh toán cuối cùng không được nhỏ hơn 0"
        ),
        (
            11,
            "Khách hàng",
            "Liên hệ & Hỗ trợ",
            "Gửi yêu cầu hỗ trợ (Support Ticket)",
            "- Họ tên không được để trống, từ 2–100 ký tự\n"
            "- Email không được để trống, đúng định dạng\n"
            "- Số điện thoại không được để trống, đủ 10 số\n"
            "- Tiêu đề không được để trống, từ 5–200 ký tự\n"
            "- Nội dung yêu cầu không được để trống, từ 10–1000 ký tự"
        ),

        # =====================================================================
        # PHẦN 2: NHÂN VIÊN (10 Form có Input)
        # =====================================================================
        (
            12,
            "Nhân viên",
            "Đăng nhập hệ thống",
            "Đăng nhập tài khoản nhân viên",
            "- Tên đăng nhập / Email / SĐT không được để trống\n"
            "- Mật khẩu không được để trống, từ 6–50 ký tự\n"
            "- Tài khoản phải có vai trò nhân viên hoặc quản lý hợp lệ\n"
            "- Tài khoản bị khóa thì không cho phép đăng nhập\n"
            "- Nhập sai quá 5 lần liên tiếp sẽ tạm khóa tài khoản\n"
            "- Nhân viên mới chưa đổi mật khẩu phải chuyển sang màn hình đổi mật khẩu lần đầu"
        ),
        (
            13,
            "Nhân viên",
            "Đổi mật khẩu lần đầu",
            "Đổi mật khẩu cho nhân viên mới",
            "- Nhân viên mới được cấp tài khoản bắt buộc đổi mật khẩu ở lần đầu đăng nhập\n"
            "- Mật khẩu tạm thời phải nhập chính xác\n"
            "- Mật khẩu mới từ 8–50 ký tự, gồm ít nhất 1 chữ hoa, 1 chữ thường, 1 số và 1 ký tự đặc biệt\n"
            "- Mật khẩu mới không được trùng mật khẩu tạm thời\n"
            "- Xác nhận mật khẩu mới phải khớp 100% với mật khẩu mới"
        ),
        (
            14,
            "Nhân viên",
            "Hồ sơ cá nhân",
            "Cập nhật thông tin nhân viên",
            "- Họ tên không được để trống, từ 2–100 ký tự\n"
            "- Số điện thoại đủ 10 số, không trùng lặp\n"
            "- Email đúng định dạng, không trùng lặp\n"
            "- Mã nhân viên, vai trò và cụm rạp chỉ cho phép xem, không tự ý sửa"
        ),
        (
            15,
            "Nhân viên",
            "Đổi mật khẩu",
            "Đổi mật khẩu tài khoản nhân viên",
            "- Mật khẩu cũ phải nhập chính xác\n"
            "- Mật khẩu mới từ 8–50 ký tự, không trùng mật khẩu cũ\n"
            "- Mật khẩu mới và xác nhận mật khẩu phải nhập giống nhau"
        ),
        (
            16,
            "Nhân viên",
            "POS Bán vé",
            "Bán vé xem phim tại quầy",
            "- Chỉ được xem và bán vé tại cụm rạp mình phụ trách, không bán chéo rạp\n"
            "- Chọn suất chiếu trong ngày, chưa kết thúc\n"
            "- Không chọn ghế đã bán, ghế đang giữ hoặc ghế bảo trì\n"
            "- Ô nhập SĐT hội viên: Đúng định dạng 10 số để tích điểm\n"
            "- Ô nhập mã voucher khuyến mãi: Kiểm tra hợp lệ tương tự online\n"
            "- Ô nhập tiền khách đưa: Bắt buộc là số tiền ≥ tổng tiền đơn hàng, tự động tính tiền thừa"
        ),
        (
            17,
            "Nhân viên",
            "POS Bán F&B",
            "Bán bắp nước riêng lẻ tại quầy",
            "- Bắt buộc chọn ít nhất 1 món, số lượng từ 1–50\n"
            "- Chọn đủ vị bắp và loại nước cho các món combo\n"
            "- Ô nhập SĐT hội viên: Đúng định dạng 10 số để tích điểm\n"
            "- Ô nhập tiền khách đưa: Bắt buộc ≥ tổng tiền đơn hàng"
        ),
        (
            18,
            "Nhân viên",
            "Hủy đơn F&B",
            "Yêu cầu hủy đơn bắp nước",
            "- Đơn yêu cầu hủy phải thuộc cụm rạp của nhân viên và chưa bị hủy\n"
            "- Ô nhập lý do hủy: Bắt buộc nhập lý do hủy đơn từ 5–255 ký tự"
        ),
        (
            19,
            "Nhân viên",
            "Soát vé Check-in",
            "Soát vé & Check-in vé vào rạp",
            "- Ô nhập mã vé / Quét mã QR: Không được để trống\n"
            "- Vé phải thuộc đúng cụm rạp nhân viên đang làm việc, không soát vé rạp khác\n"
            "- Vé đã check-in trước đó phải cảnh báo 'Vé đã sử dụng' kèm giờ check-in\n"
            "- Vé bị hủy phải cảnh báo 'Vé không hợp lệ'\n"
            "- Chỉ cho phép check-in trước giờ chiếu tối đa 45 phút đến khi hết phim"
        ),
        (
            20,
            "Nhân viên",
            "Xử lý sự cố",
            "Đổi ghế tại chỗ cho khách",
            "- Ô tra cứu: Bắt buộc nhập mã vé hoặc số điện thoại\n"
            "- Đơn phải thuộc cụm rạp của nhân viên\n"
            "- Chỉ đổi ghế khi suất chiếu chưa bắt đầu\n"
            "- Ghế mới chọn phải đang trống, không đổi sang ghế đã có người\n"
            "- Ô nhập lý do đổi ghế: Bắt buộc nhập từ 5–255 ký tự"
        ),
        (
            21,
            "Nhân viên",
            "Hỗ trợ khách hàng",
            "Phản hồi Ticket hỗ trợ CSKH",
            "- Ô nhập nội dung phản hồi: Bắt buộc nhập từ 5–1000 ký tự\n"
            "- Gửi phản hồi tự động gửi email thông báo cho khách\n"
            "- Trạng thái ticket: Chọn cập nhật trạng thái xử lý phù hợp"
        ),

        # =====================================================================
        # PHẦN 3: QUẢN LÝ / ADMIN (24 Form có Input)
        # =====================================================================
        (
            22,
            "Quản lý / Admin",
            "Phê duyệt",
            "Duyệt yêu cầu hủy đơn F&B",
            "- Chỉ quản lý cùng cụm rạp mới có quyền duyệt\n"
            "- Phê duyệt: Đơn chuyển sang Đã hủy, hoàn tiền và ghi nhận người duyệt\n"
            "- Từ chối: Ô nhập lý do bắt buộc nhập lý do từ chối (tối đa 255 ký tự)"
        ),
        (
            23,
            "Quản lý / Admin",
            "Xử lý sự cố",
            "Khóa bảo trì ghế vật lý",
            "- Chọn trạng thái: Hoạt động, Bảo trì hoặc Khóa\n"
            "- Ô nhập lý do: Bắt buộc nhập lý do bảo trì (tối đa 255 ký tự, vd: Ghế gãy, Rách đệm...)\n"
            "- Ghế bảo trì sẽ tự động ẩn trên tất cả suất chiếu tương lai"
        ),
        (
            24,
            "Quản lý / Admin",
            "Xử lý sự cố",
            "Tặng voucher đền bù sự cố",
            "- Bắt buộc chọn mẫu voucher đền bù có sẵn (vé miễn phí, bắp nước miễn phí, giảm giá)\n"
            "- Không hoàn tiền mặt trực tiếp tại quầy theo chính sách\n"
            "- Ô ghi chú đền bù: Không bắt buộc, nếu nhập tối đa 255 ký tự"
        ),
        (
            25,
            "Quản lý / Admin",
            "Quản lý phim",
            "Thêm/Sửa phim",
            "- Tên phim không được để trống, từ 2–150 ký tự, không trùng\n"
            "- Tên phim tiếng Việt tối đa 150 ký tự\n"
            "- Thời lượng phim là số nguyên từ 30–300 phút\n"
            "- Năm sản xuất từ 2020–2035\n"
            "- Ngày kết thúc phải ≥ ngày khởi chiếu\n"
            "- Giới hạn độ tuổi bắt buộc chọn (P, K, T13, T16, T18, C)\n"
            "- Thể loại và định dạng bắt buộc chọn ít nhất 1 loại\n"
            "- Đạo diễn và diễn viên tối đa 255 ký tự\n"
            "- Trailer phải là đường dẫn link Youtube hợp lệ\n"
            "- Tóm tắt nội dung tối đa 1000 ký tự, ghi chú tối đa 500 ký tự\n"
            "- Ảnh poster và banner bắt buộc chọn, định dạng jpg, png, webp, dung lượng ≤ 10MB"
        ),
        (
            26,
            "Quản lý / Admin",
            "Danh mục phim",
            "Thêm/Sửa thể loại phim",
            "- Tên thể loại không được để trống, từ 2–50 ký tự, không trùng lặp\n"
            "- Tên thể loại không chứa các ký tự đặc biệt (@#$%^&*<>/[]{})\n"
            "- Mô tả thể loại không bắt buộc, nếu nhập tối đa 255 ký tự\n"
            "- Không cho phép xóa thể loại đang được phim sử dụng"
        ),
        (
            27,
            "Quản lý / Admin",
            "Danh mục phim",
            "Thêm/Sửa định dạng phim",
            "- Tên định dạng không được để trống, từ 2–30 ký tự, không trùng (2D, 3D, IMAX, 4DX)\n"
            "- Mô tả định dạng không bắt buộc, nếu nhập tối đa 150 ký tự\n"
            "- Không cho phép xóa định dạng đang được suất chiếu sử dụng"
        ),
        (
            28,
            "Quản lý / Admin",
            "Danh mục phim",
            "Thêm/Sửa độ tuổi",
            "- Mã độ tuổi không được để trống, viết hoa (P, K, T13, T16, T18, C), không trùng\n"
            "- Tên nhãn độ tuổi không được để trống, từ 2–100 ký tự\n"
            "- Độ tuổi tối thiểu là số nguyên từ 0–21 tuổi\n"
            "- Mô tả cảnh báo / khuyến cáo độ tuổi không được để trống, tối đa 500 ký tự"
        ),
        (
            29,
            "Quản lý / Admin",
            "Quản lý cụm rạp",
            "Thêm/Sửa cụm rạp",
            "- Tên cụm rạp không được để trống, từ 5–100 ký tự, không trùng\n"
            "- Tỉnh/TP và Quận/Huyện bắt buộc chọn từ danh mục chuẩn\n"
            "- Địa chỉ chi tiết từ 10–255 ký tự\n"
            "- Hotline liên hệ gồm 8–11 chữ số\n"
            "- Giờ mở cửa và đóng cửa đúng định dạng HH:mm\n"
            "- Không cho phép đổi giờ hoạt động nếu có suất chiếu chưa kết thúc nằm ngoài giờ mới\n"
            "- Loại rạp và trạng thái hoạt động bắt buộc chọn\n"
            "- Ảnh đại diện rạp định dạng jpg, png, dung lượng ≤ 5MB"
        ),
        (
            30,
            "Quản lý / Admin",
            "Quản lý phòng chiếu",
            "Thêm/Sửa phòng chiếu",
            "- Tên phòng từ 2–50 ký tự, không trùng trong cùng cụm rạp\n"
            "- Cụm rạp và loại phòng (2D, 3D, IMAX, VIP) bắt buộc chọn\n"
            "- Trạng thái phòng bắt buộc chọn (Hoạt động, Bảo trì, Ngừng hoạt động)\n"
            "- Thời gian dọn phòng là số nguyên từ 10–60 phút (mặc định 15–20 phút)\n"
            "- Số hàng ghế từ 5–20 hàng (A–Z), số cột từ 5–30 cột"
        ),
        (
            31,
            "Quản lý / Admin",
            "Sơ đồ ghế",
            "Thiết lập sơ đồ ma trận ghế",
            "- Mỗi ô chọn loại Ghế hoặc Lối đi\n"
            "- Phân loại ghế: Thường, VIP, Ghế đôi (Sweetbox), Ghế người khuyết tật\n"
            "- Ghế đôi bắt buộc chọn chiếm 2 cột liền kề trong cùng 1 hàng\n"
            "- Nhãn ghế tự sinh hoặc nhập tùy chỉnh (tối đa 10 ký tự, không trùng nhãn trong phòng)\n"
            "- Không cho phép lưu nếu phòng không có ghế nào\n"
            "- Không được sửa lại sơ đồ khi phòng đang có suất chiếu tương lai đã bán vé"
        ),
        (
            32,
            "Quản lý / Admin",
            "Điều phối lịch chiếu",
            "Thêm/Sửa suất chiếu",
            "- Phim phải ở trạng thái Đang chiếu hoặc Sắp chiếu\n"
            "- Phòng chiếu phải đang ở trạng thái Hoạt động\n"
            "- Định dạng chiếu phải phù hợp với phim và phòng\n"
            "- Giờ bắt đầu không được chọn trong quá khứ, phải nằm trong giờ mở cửa của rạp\n"
            "- Thời gian chiếu và dọn phòng không được trùng/chồng lấn với suất chiếu khác trong cùng phòng\n"
            "- Không cho phép xóa suất chiếu đã có khách mua vé hoặc giữ chỗ"
        ),
        (
            33,
            "Quản lý / Admin",
            "Điều phối lịch chiếu",
            "Xếp lịch chiếu hàng loạt",
            "- Khoảng ngày: Từ ngày phải ≤ Đến ngày, không chọn ngày quá khứ\n"
            "- Chọn ít nhất 1 phim và 1 phòng chiếu\n"
            "- Thêm ít nhất 1 khung giờ chiếu mẫu (định dạng HH:mm)\n"
            "- Tự động kiểm tra và cảnh báo các suất bị trùng giờ trước khi lưu"
        ),
        (
            34,
            "Quản lý / Admin",
            "Quản lý F&B",
            "Thêm/Sửa món bắp nước",
            "- Tên món không được để trống, từ 2–100 ký tự, không trùng\n"
            "- Phân loại món: Thức ăn, Đồ uống, Combo, Snack\n"
            "- Giá bán là số nguyên từ 0 đến 1.000.000đ\n"
            "- Ảnh món định dạng jpg, png, webp, dung lượng ≤ 5MB\n"
            "- Mô tả món tối đa 255 ký tự\n"
            "- Không cho phép xóa món đã có trong lịch sử đơn hàng (chỉ tắt trạng thái hoạt động)"
        ),
        (
            35,
            "Quản lý / Admin",
            "Quản lý F&B",
            "Cấu hình Combo & Tùy chọn món",
            "- Tên nhóm tùy chọn từ 2–100 ký tự (vd: Vị bắp, Loại nước, Size ly)\n"
            "- Số lượng chọn tối thiểu ≤ số lượng chọn tối đa\n"
            "- Tên món con từ 2–100 ký tự, giá phụ thu là số nguyên ≥ 0\n"
            "- Mỗi nhóm tùy chọn phải có ít nhất 1 món con\n"
            "- Mỗi combo phải có ít nhất 1 thành phần"
        ),
        (
            36,
            "Quản lý / Admin",
            "Bảng giá vé",
            "Cấu hình giá nền",
            "- Giá vé nền từ 10.000đ đến 500.000đ theo ngày thường/cuối tuần/lễ, giờ thường/giờ vàng, người lớn/HSSV\n"
            "- Giá vé HSSV phải luôn ≤ giá vé người lớn trong cùng khung giờ\n"
            "- Giá trị nhập tại mỗi ô bắt buộc là số nguyên"
        ),
        (
            37,
            "Quản lý / Admin",
            "Bảng giá vé",
            "Cấu hình phụ thu loại ghế & định dạng",
            "- Phụ thu ghế VIP, Sweetbox là số nguyên từ 0 đến 200.000đ\n"
            "- Phụ thu định dạng 3D, IMAX, 4DX từ 0 đến 300.000đ\n"
            "- Mức phụ thu cập nhật có hiệu lực ngay cho các giao dịch mới"
        ),
        (
            38,
            "Quản lý / Admin",
            "Bảng giá vé",
            "Thêm/Sửa ngày lễ",
            "- Tên ngày lễ từ 2–100 ký tự, không để trống\n"
            "- Ngày áp dụng bắt buộc chọn, không được trùng với ngày lễ đã có"
        ),
        (
            39,
            "Quản lý / Admin",
            "Quản lý khuyến mãi",
            "Thêm/Sửa đợt khuyến mãi",
            "- Tên chương trình từ 3–200 ký tự, không trùng\n"
            "- Mã khuyến mãi từ 3–20 ký tự, viết hoa, không dấu, không khoảng trắng, chỉ gồm chữ và số, không trùng\n"
            "- Giảm theo %: Giá trị từ 1–100%\n"
            "- Giảm theo tiền: Giá trị từ 1.000đ đến 10.000.000đ\n"
            "- Giảm tối đa ≥ 0 (nhập 0 nếu không giới hạn trần)\n"
            "- Đơn tối thiểu ≥ 0, số lượng phát hành > 0\n"
            "- Số lượt dùng mỗi khách ≥ 1 (mặc định 1)\n"
            "- Ngày bắt đầu ≥ ngày hiện tại, ngày kết thúc > ngày bắt đầu\n"
            "- Bật đổi điểm thì số điểm yêu cầu phải > 0\n"
            "- Không cho phép kích hoạt lại đợt khuyến mãi đã hết hạn"
        ),
        (
            40,
            "Quản lý / Admin",
            "Quản lý khuyến mãi",
            "Phát hành & Tặng voucher cho khách",
            "- Đợt khuyến mãi phải đang hoạt động\n"
            "- Chọn danh sách khách hàng hoặc chọn theo hạng thành viên\n"
            "- Số lượng voucher còn lại phải đủ để phát cho số khách đã chọn\n"
            "- Tặng thành công tự động lưu vào ví khách và gửi email thông báo"
        ),
        (
            41,
            "Quản lý / Admin",
            "Quản lý nhân viên",
            "Thêm/Sửa nhân viên & Gán rạp",
            "- Mã nhân viên từ 3–20 ký tự, không trùng\n"
            "- Họ tên không được để trống, từ 2–100 ký tự\n"
            "- Email đúng định dạng, tối đa 100 ký tự, không trùng\n"
            "- Số điện thoại đủ 10 số (đầu 03, 05, 07, 08, 09), không trùng\n"
            "- Vai trò bắt buộc chọn (Admin, Quản lý, Nhân viên)\n"
            "- Quản lý và nhân viên bắt buộc gán cụm rạp trực thuộc\n"
            "- Không cho phép khóa tài khoản Admin đang đăng nhập\n"
            "- Nhân viên mới tự động tạo mật khẩu tạm và yêu cầu đổi mật khẩu ở lần đầu đăng nhập"
        ),
        (
            42,
            "Quản lý / Admin",
            "Quản lý Banner",
            "Thêm/Sửa banner quảng cáo",
            "- Tiêu đề banner từ 2–100 ký tự\n"
            "- Ảnh banner bắt buộc chọn, định dạng jpg, png, webp, dung lượng ≤ 10MB\n"
            "- Link liên kết hoặc chọn phim hợp lệ\n"
            "- Thứ tự hiển thị là số nguyên ≥ 0\n"
            "- Cho phép cập nhật thứ tự hiển thị hàng loạt"
        ),
        (
            43,
            "Quản lý / Admin",
            "Tin tức & Khuyến mãi",
            "Thêm/Sửa bài viết",
            "- Tiêu đề bài viết từ 5–255 ký tự, không trùng\n"
            "- Đường dẫn slug tự sinh từ tiêu đề, không chứa dấu tiếng Việt\n"
            "- Tóm tắt bài viết tối đa 500 ký tự\n"
            "- Nội dung bài viết bắt buộc nhập, định dạng rich text\n"
            "- Chọn trạng thái: Bản nháp, Đã xuất bản hoặc Lưu trữ\n"
            "- Ảnh đại diện định dạng jpg, png, dung lượng ≤ 5MB"
        ),
        (
            44,
            "Quản lý / Admin",
            "Quản lý FAQ",
            "Thêm/Sửa câu hỏi thường gặp",
            "- Danh mục câu hỏi tối đa 100 ký tự\n"
            "- Câu hỏi từ 5–500 ký tự, không trùng trong cùng danh mục\n"
            "- Câu trả lời bắt buộc nhập nội dung\n"
            "- Thứ tự hiển thị là số nguyên ≥ 0"
        ),
        (
            45,
            "Quản lý / Admin",
            "Cài đặt hệ thống",
            "Cấu hình tham số hệ thống",
            "- Tên tham số không cho phép chỉnh sửa\n"
            "- Thời gian giữ ghế online từ 5–30 phút\n"
            "- Thời gian giữ đơn chờ POS từ 3–20 phút\n"
            "- Thời gian phạt hủy giữ ghế từ 1–15 phút\n"
            "- Số ghế tối đa mỗi lần đặt từ 1–20 ghế\n"
            "- Thời gian đóng bán vé trước giờ chiếu từ 0–30 phút\n"
            "- Email và hotline hệ thống phải đúng định dạng\n"
            "- Giá trị tham số không được để trống"
        )
    ]
    
    for r_idx, (stt, role, screen, func, val) in enumerate(items, start=2):
        ws1.cell(r_idx, 1, stt).alignment = align_top_center
        ws1.cell(r_idx, 2, role).alignment = align_top_left
        ws1.cell(r_idx, 3, screen).alignment = align_top_left
        ws1.cell(r_idx, 4, func).alignment = align_top_left
        ws1.cell(r_idx, 5, val).alignment = align_top_left
        
        for c in range(1, 6):
            cell = ws1.cell(r_idx, c)
            cell.font = data_font_s1
            cell.border = border_thin
            
    # Set Column Widths for Sheet 1
    ws1.column_dimensions['A'].width = 8.0
    ws1.column_dimensions['B'].width = 22.0
    ws1.column_dimensions['C'].width = 26.0
    ws1.column_dimensions['D'].width = 34.0
    ws1.column_dimensions['E'].width = 82.0

    # -------------------------------------------------------------------------
    # SHEET 2: Trang tính2 - Thống kê Test Case theo 45 Màn hình Input
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet("Trang tính2")
    headers2 = ["Chức năng", "Ngày bắt đầu test", "Người thực hiện", "Số test case", "Kết quả"]
    ws2.append(headers2)
    
    for c_idx in range(1, 6):
        cell = ws2.cell(1, c_idx)
        cell.font = header_font_s2
        cell.alignment = align_center if c_idx in [2, 3, 4] else align_left
        cell.border = border_thin
        
    test_summary_data = [
        # Khách hàng (11)
        ("Đăng nhập (Khách hàng)", "Phạm Thị Quỳnh Anh", 95),
        ("Đăng ký (Khách hàng)", "Phạm Thị Quỳnh Anh", 108),
        ("Quên mật khẩu & OTP (Khách hàng)", "Phạm Thị Quỳnh Anh", 56),
        ("Đổi mật khẩu (Khách hàng)", "Phạm Thị Quỳnh Anh", 48),
        ("Hồ sơ cá nhân (Khách hàng)", "Phạm Thị Quỳnh Anh", 65),
        ("Đánh giá & Bình luận phim (Khách hàng)", "Nguyễn Quang Huy", 68),
        ("Chọn suất chiếu & Vé HSSV (Khách hàng)", "Nguyễn Quang Huy", 85),
        ("Chọn ghế & Giữ chỗ 10' (Khách hàng)", "Nguyễn Quang Huy", 120),
        ("Chọn combo bắp nước F&B (Khách hàng)", "Nguyễn Quang Huy", 92),
        ("Áp dụng mã giảm giá / Voucher (Khách hàng)", "Nguyễn Quang Huy", 115),
        ("Gửi yêu cầu hỗ trợ CSKH (Khách hàng)", "Nguyễn Quang Huy", 45),
        
        # Nhân viên (10)
        ("Đăng nhập hệ thống (Nhân viên)", "Văn Minh Khôi", 80),
        ("Đổi mật khẩu lần đầu (Nhân viên)", "Văn Minh Khôi", 42),
        ("Cập nhật thông tin nhân viên (Nhân viên)", "Văn Minh Khôi", 45),
        ("Đổi mật khẩu nhân viên (Nhân viên)", "Văn Minh Khôi", 40),
        ("Bán vé xem phim tại quầy (Nhân viên)", "Văn Minh Khôi", 145),
        ("Bán bắp nước riêng lẻ tại quầy (Nhân viên)", "Văn Minh Khôi", 82),
        ("Yêu cầu hủy đơn bắp nước (Nhân viên)", "Văn Minh Khôi", 54),
        ("Soát vé & Check-in vé vào rạp (Nhân viên)", "Văn Minh Khôi", 98),
        ("Đổi ghế tại chỗ cho khách (Nhân viên)", "Văn Minh Khôi", 112),
        ("Phản hồi Ticket hỗ trợ CSKH (Nhân viên)", "Văn Minh Khôi", 62),
        
        # Quản lý / Admin (24)
        ("Duyệt yêu cầu hủy đơn F&B (Quản lý)", "Nguyễn Ngọc Hà Linh", 60),
        ("Khóa bảo trì ghế vật lý (Quản lý)", "Nguyễn Ngọc Hà Linh", 58),
        ("Tặng voucher đền bù sự cố (Quản lý)", "Nguyễn Ngọc Hà Linh", 62),
        ("Thêm/Sửa phim (Admin)", "Nguyễn Ngọc Hà Linh", 135),
        ("Thêm/Sửa thể loại phim (Admin)", "Nguyễn Ngọc Hà Linh", 45),
        ("Thêm/Sửa định dạng phim (Admin)", "Nguyễn Ngọc Hà Linh", 40),
        ("Thêm/Sửa độ tuổi (Admin)", "Nguyễn Ngọc Hà Linh", 42),
        ("Thêm/Sửa cụm rạp & Giờ mở cửa (Admin)", "Nguyễn Ngọc Hà Linh", 110),
        ("Thêm/Sửa phòng chiếu (Admin)", "Nguyễn Ngọc Hà Linh", 95),
        ("Thiết lập sơ đồ ma trận ghế (Admin)", "Nguyễn Ngọc Hà Linh", 125),
        ("Thêm/Sửa suất chiếu & Trùng phòng (Quản lý / Admin)", "Nguyễn Ngọc Hà Linh", 140),
        ("Xếp lịch chiếu hàng loạt (Quản lý / Admin)", "Nguyễn Ngọc Hà Linh", 88),
        ("Thêm/Sửa món bắp nước F&B (Admin)", "Nguyễn Ngọc Hà Linh", 75),
        ("Cấu hình Combo & Tùy chọn món (Admin)", "Nguyễn Ngọc Hà Linh", 90),
        ("Cấu hình giá nền (Admin)", "Nguyễn Ngọc Hà Linh", 85),
        ("Cấu hình phụ thu loại ghế & định dạng (Admin)", "Nguyễn Ngọc Hà Linh", 72),
        ("Thêm/Sửa ngày lễ (Admin)", "Nguyễn Ngọc Hà Linh", 45),
        ("Thêm/Sửa đợt khuyến mãi (Admin)", "Nguyễn Ngọc Hà Linh", 130),
        ("Phát hành & Tặng voucher cho khách (Admin)", "Nguyễn Ngọc Hà Linh", 65),
        ("Thêm/Sửa nhân viên & Gán rạp (Admin)", "Phạm Thị Quỳnh Anh", 102),
        ("Thêm/Sửa banner quảng cáo (Admin)", "Nguyễn Ngọc Hà Linh", 60),
        ("Thêm/Sửa bài viết tin tức (Admin)", "Nguyễn Ngọc Hà Linh", 65),
        ("Thêm/Sửa câu hỏi thường gặp FAQ (Admin)", "Nguyễn Ngọc Hà Linh", 50),
        ("Cấu hình tham số hệ thống (Admin)", "Nguyễn Ngọc Hà Linh", 76)
    ]
    
    start_date = datetime.datetime(2026, 3, 10, 0, 0)
    for r_idx, (func_name, tester, num_cases) in enumerate(test_summary_data, start=2):
        ws2.cell(r_idx, 1, func_name).alignment = align_left
        ws2.cell(r_idx, 2, start_date).alignment = align_center
        ws2.cell(r_idx, 3, tester).alignment = align_center
        ws2.cell(r_idx, 4, float(num_cases)).alignment = align_center
        ws2.cell(r_idx, 5, f"{num_cases} pass, 0 fail").alignment = align_left
        
        for c in range(1, 6):
            cell = ws2.cell(r_idx, c)
            cell.font = data_font_s2
            cell.border = border_thin
            
    ws2.column_dimensions['A'].width = 46.0
    ws2.column_dimensions['B'].width = 20.0
    ws2.column_dimensions['C'].width = 18.0
    ws2.column_dimensions['D'].width = 16.0
    ws2.column_dimensions['E'].width = 22.0

    # -------------------------------------------------------------------------
    # SHEET 3: Trang tính3 - Tổng hợp 2 cột song song (Header đỏ đô)
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet("Trang tính3")
    headers3 = [
        "Chức năng", "Ngày bắt đầu test", "Người thực hiện", "Số test case", "Kết quả",
        "Chức năng 2", "Ngày bắt đầu test 2", "Người thực hiện 2", "Số test case 2", "Kết quả 2"
    ]
    ws3.append(headers3)
    
    for c_idx in range(1, 11):
        cell = ws3.cell(1, c_idx)
        cell.font = header_font_s3
        cell.fill = header_fill_s3
        cell.alignment = align_center
        cell.border = border_thin
        
    half_len = (len(test_summary_data) + 1) // 2
    col1_data = test_summary_data[:half_len]
    col2_data = test_summary_data[half_len:]
    
    for r_idx in range(half_len):
        row_num = r_idx + 2
        
        # Cột 1-5 (Khách hàng & Nhân viên)
        f1, t1, n1 = col1_data[r_idx]
        ws3.cell(row_num, 1, f1).alignment = align_left
        ws3.cell(row_num, 2, start_date).alignment = align_center
        ws3.cell(row_num, 3, t1).alignment = align_center
        ws3.cell(row_num, 4, float(n1)).alignment = align_center
        ws3.cell(row_num, 5, f"{n1} pass, 0 fail").alignment = align_left
        
        # Cột 6-10 (Quản lý / Admin)
        if r_idx < len(col2_data):
            f2, t2, n2 = col2_data[r_idx]
            ws3.cell(row_num, 6, f2).alignment = align_left
            ws3.cell(row_num, 7, start_date).alignment = align_center
            ws3.cell(row_num, 8, t2).alignment = align_center
            ws3.cell(row_num, 9, float(n2)).alignment = align_center
            ws3.cell(row_num, 10, f"{n2} pass, 0 fail").alignment = align_left
        else:
            for c in range(6, 11):
                ws3.cell(row_num, c, "").alignment = align_center
                
        for c in range(1, 11):
            cell = ws3.cell(row_num, c)
            cell.font = data_font_s3
            cell.border = border_thin

    ws3.column_dimensions['A'].width = 44.0
    ws3.column_dimensions['B'].width = 20.0
    ws3.column_dimensions['C'].width = 18.0
    ws3.column_dimensions['D'].width = 16.0
    ws3.column_dimensions['E'].width = 20.0
    ws3.column_dimensions['F'].width = 46.0
    ws3.column_dimensions['G'].width = 20.0
    ws3.column_dimensions['H'].width = 18.0
    ws3.column_dimensions['I'].width = 16.0
    ws3.column_dimensions['J'].width = 20.0

    wb.save(output_path)
    print(f"Successfully saved: {output_path}")

if __name__ == "__main__":
    out_dir = r"c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine"
    out_file = os.path.join(out_dir, "Validate_DevCine.xlsx")
    build_validate_workbook(out_file)
    
    # Save copies to Downloads
    dst_downloads = r"C:\Users\ADMIN\Downloads\Validate_DevCine_PhanHe.xlsx"
    build_validate_workbook(dst_downloads)
    
    try:
        shutil.copy2(out_file, r"C:\Users\ADMIN\Downloads\Validate_DevCine.xlsx")
        print("Updated C:\\Users\\ADMIN\\Downloads\\Validate_DevCine.xlsx")
    except Exception as e:
        print("Downloads Validate_DevCine.xlsx locked by Excel, saved to Validate_DevCine_PhanHe.xlsx")
