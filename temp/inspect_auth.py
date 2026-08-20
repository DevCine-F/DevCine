# -*- coding: utf-8 -*-
import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook("TestReport Dự án DevCine.xlsx", data_only=False)

print("=== SHEET: Đăng nhập ===")
ws_dn = wb["Đăng nhập"]
print("Max row:", ws_dn.max_row)
print("Row 10:", [ws_dn.cell(10, c).value for c in range(1, 10)])
for r in range(11, ws_dn.max_row + 1):
    c1 = ws_dn.cell(r, 1).value
    c2 = ws_dn.cell(r, 2).value
    if c1 and ("KIỂM TRA" in str(c1) or "CHỨC NĂNG" in str(c1) or "KỸ THUẬT" in str(c1)):
        print(f"--- [HEADER] Row {r}: {c1}")
    elif c1:
        print(f"  Row {r}: {c1} | {str(c2)[:40]}")

print("\n=== SHEET: Đăng ký ===")
ws_reg = wb["Đăng ký"]
print("Max row:", ws_reg.max_row)
for r in range(11, ws_reg.max_row + 1):
    c1 = ws_reg.cell(r, 1).value
    c2 = ws_reg.cell(r, 2).value
    if c1 and ("KIỂM TRA" in str(c1) or "CHỨC NĂNG" in str(c1) or "KỸ THUẬT" in str(c1)):
        print(f"--- [HEADER] Row {r}: {c1}")
    elif c1:
        print(f"  Row {r}: {c1} | {str(c2)[:40]}")
