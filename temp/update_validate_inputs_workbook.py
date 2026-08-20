# -*- coding: utf-8 -*-
"""
Update Validate_DevCine_InputsOnly.xlsx to synchronize 100% with TestReport Dự án DevCine.xlsx:
- 14 Consolidated Functional Modules
- Exact DOM & Backend Validation Rules for all Fields & Modals
- Accurate Test Case Counts & QA Tester Assignments
- Professional Excel Styling (Times New Roman, Navy & Green headers, thin borders, auto wrap)
"""

import os
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

def build_validate_inputs_workbook():
    target_path = r'C:\Users\ADMIN\Downloads\Validate_DevCine_InputsOnly.xlsx'
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # remove default sheet

    font_name = "Times New Roman"
    font_title = Font(name=font_name, size=15, bold=True, color='FF002060')
    font_header_white = Font(name=font_name, size=11, bold=True, color='FFFFFFFF')
    font_header_black = Font(name=font_name, size=11, bold=True, color='FF000000')
    font_module_group = Font(name=font_name, size=11, bold=True, color='FF002060')
    font_bold = Font(name=font_name, size=11, bold=True, color='FF000000')
    font_regular = Font(name=font_name, size=11, bold=False, color='FF000000')
    font_pass = Font(name=font_name, size=11, bold=True, color='FF008000')

    fill_header_navy = PatternFill(start_color='FF002060', end_color='FF002060', fill_type='solid')
    fill_header_green = PatternFill(start_color='FFC5E0B3', end_color='FFC5E0B3', fill_type='solid')
    fill_group_blue = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
    fill_zebra = PatternFill(start_color='FFF9FBFD', end_color='FFF9FBFD', fill_type='solid')

    border_thin = Border(
        left=Side(style='thin', color='FFD9D9D9'), right=Side(style='thin', color='FFD9D9D9'),
        top=Side(style='thin', color='FFD9D9D9'), bottom=Side(style='thin', color='FFD9D9D9')
    )
    border_header = Border(
        left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000')
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_top_center = Alignment(horizontal='center', vertical='top', wrap_text=True)

    # =========================================================================
    # SHEET 1: QUY TẮC VALIDATE ĐẦU VÀO (INPUT VALIDATION SPECS)
    # =========================================================================
    ws1 = wb.create_sheet("Quy tắc Validate Đầu vào")
    ws1.views.sheetView[0].showGridLines = True

    ws1.cell(1, 1, "BẢNG TỔNG HỢP RÀNG BUỘC DỮ LIỆU ĐẦU VÀO & QUY TẮC VALIDATE HỆ THỐNG DEVCINE").font = font_title
    ws1.merge_cells("A1:F1")
    ws1.row_dimensions[1].height = 30

    headers_ws1 = [
        "STT", "Phân hệ nghiệp vụ (Consolidated Module)", "Màn hình / Chức năng",
        "Vai trò áp dụng", "Các trường dữ liệu đầu vào (Input Fields)",
        "Ràng buộc Validate & Quy tắc Nghiệp vụ (Validation Rules & Constraints)"
    ]

    for c_idx, h in enumerate(headers_ws1, start=1):
        cell = ws1.cell(3, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_header
    ws1.row_dimensions[3].height = 28

    validation_data = [
        # 1. Xác thực & Tài khoản
        (1, "Xác thực & Tài khoản", "Đăng nhập hệ thống", "Khách hàng & Nội bộ",
         "• Số điện thoại hoặc Email (identifier)\n• Mật khẩu (password)",
         "- Số điện thoại hoặc Email không được để trống (báo lỗi inline 'Vui lòng nhập số điện thoại hoặc email.')\n"
         "- Tự động nhận diện Email (nếu có ký tự '@') vs Số điện thoại\n"
         "- Email: Phải đúng định dạng regex ^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\\.[a-zA-Z]{2,}$\n"
         "- Số điện thoại: Chuẩn hóa bỏ khoảng trắng, tự đổi đầu '+84' thành '0', đủ 10 chữ số (đầu 03, 05, 07, 08, 09)\n"
         "- Mật khẩu: Bắt buộc nhập; có nút icon con mắt bật/tắt hiển thị mật khẩu\n"
         "- Xác thực sai: Báo lỗi chung 'Số điện thoại/email hoặc mật khẩu không chính xác.' (chống dò quét tài khoản)\n"
         "- Tài khoản bị khóa: Báo lỗi 'Tài khoản của bạn đã bị khóa. Vui lòng liên hệ CSKH.'"),

        (2, "Xác thực & Tài khoản", "Quên mật khẩu & OTP 3 bước", "Khách hàng",
         "• Bước 1: Email tài khoản\n• Bước 2: Mã OTP 6 chữ số\n• Bước 3: Mật khẩu mới, Xác nhận mật khẩu",
         "- Bước 1: Email đúng định dạng; hệ thống gửi OTP 6 số và khóa nút đếm ngược Cooldown 30s chống spam\n"
         "- Bước 2: Mã OTP bắt buộc đúng 6 chữ số (tự động lọc bỏ ký tự chữ regex \\D), hiệu lực trong 10 phút\n"
         "- Bước 3: Mật khẩu mới từ 8–32 ký tự, gồm ít nhất 1 chữ hoa, 1 chữ thường, 1 số và 1 ký tự đặc biệt\n"
         "- Xác nhận mật khẩu mới phải khớp 100% với mật khẩu mới\n"
         "- Thành công: Thông báo Toast và tự động điền sẵn Email tại Form đăng nhập"),

        (3, "Xác thực & Tài khoản", "Đăng ký tài khoản hội viên", "Khách vãng lai",
         "• Họ và tên (fullName)\n• Email của bạn (email)\n• Số điện thoại (phone)\n• Mật khẩu (password)",
         "- Họ và tên: Bắt buộc, từ 2–50 ký tự, chỉ gồm chữ cái có dấu và khoảng trắng\n"
         "- Email: Bắt buộc, đúng định dạng chuẩn, không được trùng với tài khoản đã tồn tại trong CSDL\n"
         "- Số điện thoại: Bắt buộc, đủ 10 số (đầu 03, 05, 07, 08, 09), tự động chặn nhập ký tự chữ, không trùng\n"
         "- Mật khẩu: 8–32 ký tự, gồm chữ hoa, chữ thường, số và ký tự đặc biệt\n"
         "- Nút 'ĐĂNG KÝ >' tự động bị vô hiệu hóa (disabled) khi form chưa hợp lệ và sáng màu vàng khi điền đủ 4 trường"),

        (4, "Xác thực & Tài khoản", "Hồ sơ cá nhân & Đổi mật khẩu", "Khách hàng",
         "• Họ tên, Ngày sinh, Giới tính, Tỉnh/TP\n• Mật khẩu hiện tại, Mật khẩu mới, Xác nhận MK",
         "- Họ tên: 2–50 ký tự; Ngày sinh: Không được chọn ngày trong tương lai (≤ ngày hiện tại, tuổi ≥ 13)\n"
         "- Mật khẩu hiện tại: Bắt buộc nhập chính xác\n"
         "- Mật khẩu mới: 8–32 ký tự, chuẩn mạnh, không được trùng mật khẩu cũ\n"
         "- Xác nhận mật khẩu mới phải trùng khớp 100%"),

        # 2. Đặt vé trực tuyến
        (5, "Đặt vé trực tuyến", "Chọn suất chiếu, Sơ đồ ghế & Giữ chỗ", "Khách hàng",
         "• Số lượng vé Người lớn & U22/HSSV\n• Vị trí ghế trên ma trận phòng chiếu",
         "- Tổng số lượng vé từ 1 đến 8 vé mỗi giao dịch (nút [+] bị khóa khi đạt 8 vé)\n"
         "- Số ghế click chọn trên sơ đồ phải bằng đúng tổng số vé đã khai báo\n"
         "- Chặn tuyệt đối chọn ghế đã bán (SOLD), ghế đang bảo trì (MAINTENANCE) hoặc người khác đang giữ\n"
         "- Ghế Sweetbox đôi bắt buộc chọn đồng thời cả cặp 2 chỗ liền kề trong cùng hàng\n"
         "- Quy tắc Orphan Seat Rule: Chặn không cho phép để lại 1 ghế trống cô lập ở đầu dãy hoặc giữa các ghế\n"
         "- Đồng hồ đếm ngược giữ chỗ 10 phút (10:00 -> 00:00), hết giờ tự động nhả khóa ghế Redis real-time"),

        (6, "Đặt vé trực tuyến", "Chọn Combo bắp nước & Tùy chọn vị F&B", "Khách hàng",
         "• Số lượng phần Combo\n• Tùy chọn vị bắp (Ngọt, Phô mai, Trứng muối)\n• Tùy chọn nước (Coca, Sprite, Fanta)",
         "- Bước chọn F&B không bắt buộc, cho phép bấm 'TIẾP TỤC ➔' để bỏ qua nếu không mua\n"
         "- Số lượng mỗi gói combo từ 0 đến 10 phần\n"
         "- Modal FnbOptionModal: Bắt buộc chọn đủ vị bắp và loại nước theo cấu hình combo\n"
         "- Đổi vị đặc biệt (Phô mai +15k, Trứng muối +15k) tự động cộng phụ thu vào tổng tiền"),

        (7, "Đặt vé trực tuyến", "Thanh toán VNPAY, VietQR & Sinh vé QR", "Khách hàng",
         "• Phương thức thanh toán (VNPAY / VietQR)\n• Thông tin thẻ/OTP ngân hàng",
         "- VNPAY: Chuyển hướng sang cổng VNPAY, xử lý phản hồi mã giao dịch 00 (Thành công), 24 (Hủy), 51 (Không đủ số dư)\n"
         "- VietQR: Tự sinh mã QR động chuẩn VietQR chứa đúng số tiền, STK rạp và mã đơn hàng\n"
         "- Backend xác thực chữ ký bảo mật SHA512 trên Webhook IPN trước khi xuất vé\n"
         "- Thành công: Chuyển sang BookingSuccessView, sinh mã vé QR Code độc nhất và tích điểm hội viên"),

        # 3. Lịch sử vé & Voucher
        (8, "Lịch sử vé & Voucher", "Lịch sử đặt vé & Chi tiết mã vé QR", "Khách hàng",
         "• Bộ lọc trạng thái vé (Tất cả / Đã thanh toán / Đã sử dụng / Đã hủy)",
         "- Hiển thị danh sách đơn vé phân trang, sắp xếp theo thời gian mới nhất\n"
         "- Modal Chi tiết vé: Hiển thị đầy đủ Poster, Tên phim, Rạp, Phòng, Suất chiếu, Ghế, Combo F&B và Mã QR vé nét cao"),

        (9, "Lịch sử vé & Voucher", "Ví voucher cá nhân & Kích hoạt mã", "Khách hàng",
         "• Ô nhập mã Voucher khuyến mãi\n• Tab phân loại (Còn hạn / Đã dùng / Hết hạn)",
         "- Mã voucher: Bắt buộc nhập, tự động chuyển in hoa, không khoảng trắng\n"
         "- Kiểm tra điều kiện: Mã hợp lệ, còn hạn sử dụng, chưa từng sử dụng, đạt giá trị đơn hàng tối thiểu"),

        # 4. Tương tác & Đánh giá
        (10, "Tương tác & Đánh giá", "Đánh giá sao, Bình luận & Góp ý CSKH", "Khách hàng",
         "• Số sao đánh giá (1–5 sao)\n• Nội dung nhận xét (reviewComment)\n• Form liên hệ CSKH (Họ tên, Email, Tiêu đề, Nội dung)",
         "- Đánh giá sao: Bắt buộc chọn từ 1 đến 5 sao; Nội dung nhận xét từ 10–500 ký tự\n"
         "- Khách vãng lai chưa đăng nhập: Modal nhắc đăng nhập, tự động lưu bản nháp đánh giá vào sessionStorage\n"
         "- Bình luận thảo luận: Nội dung từ 2–300 ký tự, tự động chặn spam và từ ngữ thô tục\n"
         "- Form CSKH: Họ tên 2–50 ký tự, Email đúng chuẩn, Tiêu đề 5–100 ký tự, Nội dung 10–1000 ký tự"),

        # 5. Tổng quan (Dashboard)
        (11, "Tổng quan (Dashboard)", "Báo cáo doanh thu & Thống kê KPI", "Quản trị viên",
         "• Bộ lọc thời gian (Hôm nay, Tuần này, Tháng này, Năm nay)\n• Custom Month Picker (Tháng MM/YYYY)\n• Dropdown Cụm rạp",
         "- Month Picker: Chặn không cho phép chọn các tháng trong tương lai (> tháng hiện tại)\n"
         "- Bộ lọc Cụm rạp: Cho phép lọc riêng từng chi nhánh hoặc xem toàn hệ thống\n"
         "- Trạng thái rỗng: Xử lý hiển thị 0đ, biểu đồ phẳng mượt mà khi không có dữ liệu phát sinh\n"
         "- Nút '📥 XUẤT BÁO CÁO EXCEL': Tải về file .xlsx đầy đủ các bảng dữ liệu thống kê"),

        # 6. Bán hàng tại quầy (POS)
        (12, "Bán hàng tại quầy (POS)", "Bán vé, Đơn chờ & Bán bắp nước tại quầy", "Nhân viên",
         "• Chọn suất, chọn ghế trên POS\n• 3 Tab Đơn chờ (Tab 1, Tab 2, Tab 3)\n• Tra cứu hội viên (SĐT/Mã thẻ)\n• Áp voucher/điểm\n• Phương thức thanh toán (Tiền mặt / Chuyển khoản QR / Thẻ)",
         "- Sơ đồ POS: Khóa ghế tức thì khi nhân viên chọn, chặn trùng ghế với khách online\n"
         "- Quản lý 3 Tab Đơn chờ: Cho phép lưu tạm tối đa 3 đơn cùng lúc khi khách đổi ý hoặc chờ chuyển khoản\n"
         "- Tra cứu hội viên qua SĐT: Tự động hiển thị tên khách, hạng thẻ, số điểm tích lũy và gợi ý trừ điểm\n"
         "- In vé nhiệt / Hóa đơn: Tự động gửi lệnh in hóa đơn thanh toán và phiếu nhận món bắp nước"),

        # 7. Kiểm soát vé (Check-in)
        (13, "Kiểm soát vé (Check-in)", "Soát vé qua mã QR & Camera", "Nhân viên",
         "• Quét mã QR qua Camera / Máy quét 2D\n• Ô nhập mã vé thủ công (Ticket Code)",
         "- Tự động giải mã chuỗi mã vé QR độc nhất từ hệ thống DevCine\n"
         "- Vé hợp lệ: Hiển thị Banner XANH LÁ 'CHECK-IN THÀNH CÔNG', phát âm thanh BEEP ngắn, cập nhật trạng thái USED\n"
         "- Vé đã sử dụng trước đó: Hiển thị Banner ĐỎ 'VÉ ĐÃ SỬ DỤNG' kèm thời gian check-in lần trước, phát tiếng CÒI CẢNH BÁO\n"
         "- Vé sai suất chiếu / Sai phòng / Vé giả mạo: Từ chối ngay lập tức và hiển thị lý do vi phạm rõ ràng"),

        # 8. Sự cố & Hóa đơn
        (14, "Sự cố & Hóa đơn", "Quản lý hóa đơn, Hủy đơn & Đổi ghế sự cố", "Nhân viên & Quản lý",
         "• Mã đơn hàng / Hóa đơn\n• Lý do yêu cầu hủy đơn bắp nước F&B\n• Ghế gặp sự cố vật lý -> Ghế đổi ngang VIP\n• Khóa bảo trì ghế",
         "- Hủy đơn F&B: Nhân viên gửi lý do hủy -> Quản lý rạp duyệt Void và hoàn tiền cho khách\n"
         "- Sự cố chỗ ngồi: Đổi sang ghế trống tương đương hoặc nâng cấp miễn phí lên ghế VIP, in lại phiếu đổi chỗ\n"
         "- Khóa bảo trì ghế: Gắn cờ MAINTENANCE cho ghế hỏng vật lý (gãy tay vịn, rách đệm), khóa trên toàn bộ Web & POS"),

        # 9. Quản lý Phim & Danh mục
        (15, "Quản lý Phim & Danh mục", "Modal Thêm phim mới (Add Movie)", "Quản trị viên",
         "• Tên phim (*), Thời lượng (*), Trailer URL (*), Đạo diễn, Diễn viên\n• Quốc gia, Năm SX (*), Ngôn ngữ, Loại hình (*), Thể loại (*), Định dạng (*)\n• Phân loại độ tuổi (*), Trạng thái (*), Ngày khởi chiếu (*), Ngày kết thúc (*)\n• Poster tỷ lệ 2/3 (*), Banner ngang 16/9, Tóm tắt nội dung (*) (50–1000 ký tự)",
         "- Tên phim: Bắt buộc, từ 2–150 ký tự; Thời lượng: Số nguyên từ 1–300 phút\n"
         "- Trailer URL: Bắt buộc, đúng định dạng link YouTube (youtube.com hoặc youtu.be)\n"
         "- Năm sản xuất: Số nguyên từ 1900 đến năm hiện tại + 5\n"
         "- Lưới chọn Thể loại: Bắt buộc chọn ít nhất 1 thể loại; Định dạng: Bắt buộc chọn ít nhất 1 định dạng (2D, 3D, IMAX)\n"
         "- Ngày khởi chiếu & kết thúc: Bắt buộc; Ngày kết thúc phải ≥ Ngày khởi chiếu\n"
         "- Tóm tắt nội dung: Bắt buộc, từ 50 đến 1000 ký tự (có bộ đếm ký tự real-time)\n"
         "- Poster: Bắt buộc upload, định dạng ảnh jpg/png/webp, dung lượng ≤ 8MB, tự động hiển thị khung xem trước 2/3"),

        (16, "Quản lý Phim & Danh mục", "Modal Chỉnh sửa phim & Danh mục", "Quản trị viên",
         "• Dữ liệu cũ của phim\n• Thể loại, Đạo diễn, Diễn viên, Định dạng, Banner",
         "- Load đầy đủ 100% dữ liệu cũ vào Form\n"
         "- Khóa ô 'Ngày khởi chiếu' và hạn chế đổi trạng thái khi phim đang có suất chiếu kích hoạt trong hệ thống\n"
         "- Quản lý Banner: Tiêu đề 2–100 ký tự, Ảnh banner ≤ 10MB, Thứ tự hiển thị là số nguyên ≥ 0"),

        # 10. Cụm rạp & Lịch chiếu
        (17, "Cụm rạp & Lịch chiếu", "Cụm rạp, Phòng chiếu & Thiết kế sơ đồ ghế", "Quản trị viên",
         "• Tên cụm rạp (*), Địa chỉ (*), Hotline (*), Giờ mở/đóng cửa (*)\n• Tên phòng (*), Loại phòng (*), Thời gian dọn phòng (10–60p)\n• Trình dựng sơ đồ ghế SeatMapBuilder (Ghế Thường, VIP, Sweetbox, Lối đi)",
         "- Tên cụm rạp: 2–100 ký tự, không trùng; Hotline: 8–11 chữ số\n"
         "- Giờ mở/đóng cửa: Đúng định dạng HH:mm, Giờ mở cửa < Giờ đóng cửa\n"
         "- Thời gian dọn phòng: Số nguyên từ 10 đến 60 phút (mặc định 15–20 phút)\n"
         "- SeatMapBuilder: Hàng A–Z, Cột 1–20; Ghế Sweetbox tự gộp 2 ô; Lối đi tự động bỏ qua khi đánh số nhãn ghế\n"
         "- Chặn lưu và cảnh báo khi sửa sơ đồ ghế của phòng chiếu đang có suất chiếu đã bán vé"),

        (18, "Cụm rạp & Lịch chiếu", "Quản lý Lịch chiếu & Chặn trùng giờ", "Quản trị viên & Quản lý",
         "• Suất chiếu: Chọn Phim, Phòng chiếu, Ngày chiếu, Giờ bắt đầu (HH:mm), Định dạng, Giá vé áp dụng",
         "- Giờ bắt đầu không được chọn trong quá khứ, phải nằm trong khung giờ mở cửa của rạp\n"
         "- Thuật toán Overlap Guard: Tự động tính Giờ kết thúc = Giờ bắt đầu + Thời lượng phim + Thời gian dọn phòng\n"
         "- Chặn tuyệt đối trùng/chồng lấn giờ chiếu trong cùng 1 phòng chiếu và cảnh báo tức thì\n"
         "- Không cho phép xóa suất chiếu đã có khách hàng mua vé hoặc giữ chỗ"),

        # 11. Thực đơn F&B
        (19, "Thực đơn F&B", "Món bắp nước, Combo & Bảng tùy chọn Topping", "Quản lý",
         "• Tên món (*), Phân loại (Thức ăn/Nước/Combo/Snack), Đơn giá (*), Ảnh món\n• Nhóm tùy chọn (Vị bắp, Loại nước), Tên topping, Giá phụ thu",
         "- Tên món: 2–100 ký tự, không trùng tên; Đơn giá: Số nguyên từ 0 đến 1.000.000đ\n"
         "- Ảnh món: jpg/png/webp, dung lượng ≤ 5MB; Mô tả tối đa 255 ký tự\n"
         "- Không cho phép xóa món đã phát sinh trong lịch sử đơn hàng (chỉ tắt trạng thái hoạt động)\n"
         "- Giá phụ thu topping: Số nguyên ≥ 0đ; Mỗi combo bắt buộc có ít nhất 1 món thành phần"),

        # 12. Giá vé & Khuyến mãi
        (20, "Giá vé & Khuyến mãi", "Bảng giá vé, Khuyến mãi & Voucher, Loyalty", "Quản trị viên",
         "• Giá nền theo ngày thường/cuối tuần/lễ, giờ thường/giờ vàng, Người lớn/HSSV\n• Phụ thu ghế VIP, Sweetbox, 3D, IMAX\n• Tên khuyến mãi, Mã voucher, % hoặc tiền giảm, Đơn tối thiểu, Hạn dùng\n• Cấu hình tích điểm hạng thẻ Loyalty (Standard, VIP, Diamond)",
         "- Giá vé nền: Số nguyên từ 10.000đ đến 500.000đ; Giá HSSV luôn ≤ Giá Người lớn trong cùng khung giờ\n"
         "- Phụ thu ghế: VIP (0–200k), Sweetbox (0–200k); Phụ thu định dạng 3D/IMAX (0–300k)\n"
         "- Mã voucher: 3–20 ký tự, viết hoa, không dấu, không khoảng trắng, chỉ gồm chữ và số, không trùng\n"
         "- Giảm %: 1–100%; Giảm tiền: 1.000đ – 10.000.000đ; Đơn tối thiểu ≥ 0đ; Giảm tối đa ≥ 0đ\n"
         "- Ngày kết thúc khuyến mãi phải > Ngày bắt đầu\n"
         "- Hạng thẻ Loyalty: Cấu hình mốc chi tiêu nâng hạng và % tích điểm thưởng tự động khi mua vé"),

        # 13. Khách hàng & CSKH
        (21, "Khách hàng & CSKH", "Quản lý hội viên, Khóa tài khoản & FAQ", "Quản trị viên",
         "• Tra cứu hội viên (SĐT, Email, Tên)\n• Trạng thái tài khoản (ACTIVE / LOCKED)\n• Tiếp nhận góp ý, Danh mục FAQ, Câu hỏi & Câu trả lời",
         "- Modal xác nhận trước khi Khóa tài khoản: Hiển thị cảnh báo xác nhận ngăn thao tác bấm nhầm\n"
         "- Khóa tài khoản: Chuyển trạng thái LOCKED (Badge đỏ), ngay lập tức chặn đăng nhập trên Web\n"
         "- Mở khóa tài khoản: Chuyển ACTIVE (Badge xanh), khôi phục quyền đăng nhập bình thường\n"
         "- Câu hỏi FAQ: Tiêu đề 5–500 ký tự, Câu trả lời bắt buộc nhập nội dung rich text"),

        # 14. Quản trị Hệ thống (RBAC)
        (22, "Quản trị Hệ thống (RBAC)", "Quản lý nhân sự, Phân quyền RBAC & Audit Logs", "Quản trị viên",
         "• Họ tên, Email nội bộ, SĐT, Gán rạp, Vai trò (ADMIN / MANAGER / STAFF)\n• Ma trận 4 Tab quyền (TỔNG QUAN, NGHIỆP VỤ, NỘI DUNG, HỆ THỐNG)\n• Override Allow/Deny theo từng nhân viên\n• Nhật ký hệ thống Audit Logs",
         "- Họ tên: 2–50 ký tự; Email: Đúng chuẩn regex nội bộ, không trùng; SĐT: Đủ 10 chữ số\n"
         "- Quản lý và Nhân viên bắt buộc gán cụm rạp trực thuộc\n"
         "- Superuser Guard: Khóa toàn bộ quyền của vai trò ADMIN (Badge 'ADMIN TOÀN QUYỀN 🔒', read-only) chống tự tước quyền\n"
         "- Cảnh báo thay đổi chưa lưu khi chuyển đổi giữa các vai trò\n"
         "- Backend Enforce RBAC: Trả về mã lỗi HTTP 403 Forbidden nếu gọi API ngoài quyền hạn\n"
         "- Audit Logs Immutability: Bảng nhật ký chỉ đọc (Read-only), chặn toàn bộ thao tác Sửa/Xóa log để đảm bảo minh bạch")
    ]

    cur_r = 4
    for item in validation_data:
        stt, mod_name, screen, role, inputs, rules = item
        ws1.cell(cur_r, 1, stt).alignment = align_top_center
        ws1.cell(cur_r, 2, mod_name).alignment = align_top_left
        ws1.cell(cur_r, 2).font = font_bold
        ws1.cell(cur_r, 3, screen).alignment = align_top_left
        ws1.cell(cur_r, 3).font = font_bold
        ws1.cell(cur_r, 4, role).alignment = align_top_left
        ws1.cell(cur_r, 5, inputs).alignment = align_top_left
        ws1.cell(cur_r, 6, rules).alignment = align_top_left

        for c_idx in range(1, 7):
            cell = ws1.cell(cur_r, c_idx)
            cell.border = border_thin
            if c_idx not in [2, 3]:
                cell.font = font_regular
        
        # dynamic row height based on content
        lines = max(inputs.count('\n') + 1, rules.count('\n') + 1)
        ws1.row_dimensions[cur_r].height = max(45, lines * 16)
        cur_r += 1

    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 26
    ws1.column_dimensions["C"].width = 32
    ws1.column_dimensions["D"].width = 22
    ws1.column_dimensions["E"].width = 40
    ws1.column_dimensions["F"].width = 75

    # =========================================================================
    # SHEET 2: KẾ HOẠCH & TIẾN ĐỘ KIỂM THỬ (14 CONSOLIDATED MODULES)
    # =========================================================================
    ws2 = wb.create_sheet("Kế hoạch & Kết quả Kiểm thử")
    ws2.views.sheetView[0].showGridLines = True

    ws2.cell(1, 1, "DANH SÁCH 14 PHÂN HỆ KIỂM THỬ & SỐ LƯỢNG TEST CASE DỰ ÁN DEVCINE").font = font_title
    ws2.merge_cells("A1:G1")
    ws2.row_dimensions[1].height = 30

    headers_ws2 = [
        "STT", "Mã Phân hệ (Module Code)", "Tên Phân hệ (Consolidated Module)",
        "Ngày bắt đầu test", "Người thực hiện (Tester)", "Số lượng Test Case", "Kết quả thực tế (Execution Result)"
    ]

    for c_idx, h in enumerate(headers_ws2, start=1):
        cell = ws2.cell(3, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_header
    ws2.row_dimensions[3].height = 28

    modules_summary_data = [
        (1, "MOD_AUTH_ACCOUNT", "Xác thực & Tài khoản", "2026-03-10", "Phạm Thị Quỳnh Anh", 131, "131 Passed (100%)"),
        (2, "MOD_ONLINE_BOOKING", "Đặt vé trực tuyến", "2026-03-10", "Nguyễn Quang Huy", 32, "32 Passed (100%)"),
        (3, "MOD_HISTORY_VOUCHERS", "Lịch sử vé & Voucher", "2026-03-10", "Nguyễn Quang Huy", 38, "38 Passed (100%)"),
        (4, "MOD_INTERACTION_REVIEWS", "Tương tác & Đánh giá", "2026-03-10", "Nguyễn Quang Huy", 86, "86 Passed (100%)"),
        (5, "MOD_ADMIN_DASHBOARD", "Tổng quan (Dashboard)", "2026-03-10", "Nguyễn Ngọc Hà Linh", 15, "15 Passed (100%)"),
        (6, "MOD_ADMIN_POS", "Bán hàng tại quầy (POS)", "2026-03-10", "Văn Minh Khôi", 73, "73 Passed (100%)"),
        (7, "MOD_ADMIN_CHECKIN", "Kiểm soát vé (Check-in)", "2026-03-10", "Văn Minh Khôi", 21, "21 Passed (100%)"),
        (8, "MOD_ADMIN_INCIDENTS_BOOKINGS", "Sự cố & Hóa đơn", "2026-03-10", "Văn Minh Khôi", 56, "56 Passed (100%)"),
        (9, "MOD_ADMIN_MOVIES_CAT", "Quản lý Phim & Danh mục", "2026-03-10", "Nguyễn Ngọc Hà Linh", 258, "258 Passed (100%)"),
        (10, "MOD_ADMIN_CINEMAS_SCHED", "Cụm rạp & Lịch chiếu", "2026-03-10", "Nguyễn Ngọc Hà Linh", 108, "108 Passed (100%)"),
        (11, "MOD_ADMIN_FNB", "Thực đơn F&B", "2026-03-10", "Nguyễn Quang Huy", 73, "73 Passed (100%)"),
        (12, "MOD_ADMIN_PRICING_PROMO", "Giá vé & Khuyến mãi", "2026-03-10", "Nguyễn Ngọc Hà Linh", 143, "143 Passed (100%)"),
        (13, "MOD_ADMIN_CUSTOMERS_SUPPORT", "Khách hàng & CSKH", "2026-03-10", "Phạm Thị Quỳnh Anh", 63, "63 Passed (100%)"),
        (14, "MOD_ADMIN_SYSTEM_RBAC", "Quản trị Hệ thống (RBAC)", "2026-03-10", "Phạm Thị Quỳnh Anh", 115, "115 Passed (100%)")
    ]

    cur_r = 4
    for row in modules_summary_data:
        stt, code, name, start_d, tester, tc_count, res = row
        ws2.cell(cur_r, 1, stt).alignment = align_center
        ws2.cell(cur_r, 2, code).alignment = align_left
        ws2.cell(cur_r, 2).font = font_bold
        ws2.cell(cur_r, 3, name).alignment = align_left
        ws2.cell(cur_r, 3).font = font_bold
        ws2.cell(cur_r, 4, start_d).alignment = align_center
        ws2.cell(cur_r, 5, tester).alignment = align_left
        ws2.cell(cur_r, 6, tc_count).alignment = align_center
        ws2.cell(cur_r, 6).font = font_bold
        ws2.cell(cur_r, 7, res).alignment = align_center
        ws2.cell(cur_r, 7).font = font_pass

        for c_idx in range(1, 8):
            cell = ws2.cell(cur_r, c_idx)
            cell.border = border_thin
            if c_idx not in [2, 3, 6, 7]:
                cell.font = font_regular
        ws2.row_dimensions[cur_r].height = 24
        cur_r += 1

    # Total Row
    ws2.cell(cur_r, 1, "TỔNG CỘNG").font = font_bold
    ws2.cell(cur_r, 1).alignment = align_center
    ws2.cell(cur_r, 1).border = border_header
    ws2.merge_cells(start_row=cur_r, start_column=1, end_row=cur_r, end_column=5)
    for c in range(1, 6):
        ws2.cell(cur_r, c).border = border_header
        ws2.cell(cur_r, c).fill = fill_group_blue

    ws2.cell(cur_r, 6, "=SUM(F4:F17)").font = font_bold
    ws2.cell(cur_r, 6).alignment = align_center
    ws2.cell(cur_r, 6).border = border_header
    ws2.cell(cur_r, 6).fill = fill_group_blue

    ws2.cell(cur_r, 7, "1.212 / 1.212 (100%)").font = font_pass
    ws2.cell(cur_r, 7).alignment = align_center
    ws2.cell(cur_r, 7).border = border_header
    ws2.cell(cur_r, 7).fill = fill_group_blue
    ws2.row_dimensions[cur_r].height = 26

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 36
    ws2.column_dimensions["C"].width = 32
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 24
    ws2.column_dimensions["F"].width = 20
    ws2.column_dimensions["G"].width = 26

    # =========================================================================
    # SHEET 3: PHÂN CÔNG THÀNH VIÊN & TIẾN ĐỘ QA
    # =========================================================================
    ws3 = wb.create_sheet("Phân công & Tiến độ QA")
    ws3.views.sheetView[0].showGridLines = True

    ws3.cell(1, 1, "BẢNG PHÂN CÔNG NHIỆM VỤ & TIẾN ĐỘ THỰC HIỆN CỦA ĐỘI NGŨ TESTER").font = font_title
    ws3.merge_cells("A1:F1")
    ws3.row_dimensions[1].height = 30

    headers_ws3 = [
        "STT", "Họ và tên Tester", "Mã SV / Vai trò", "Các phân hệ phụ trách", "Tổng số Test Case", "Trạng thái hoàn thành"
    ]

    for c_idx, h in enumerate(headers_ws3, start=1):
        cell = ws3.cell(3, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_header
    ws3.row_dimensions[3].height = 28

    tester_data = [
        (1, "Nguyễn Quang Huy", "PH12345 (Test Lead)",
         "• Đặt vé trực tuyến (32 TCs)\n• Lịch sử vé & Voucher (38 TCs)\n• Tương tác & Đánh giá (86 TCs)\n• Thực đơn F&B (73 TCs)",
         229, "Hoàn thành 100% (Pass: 229, Fail: 0)"),

        (2, "Văn Minh Khôi", "PH12346 (Tester / QA)",
         "• Bán hàng tại quầy POS (73 TCs)\n• Kiểm soát vé Check-in (21 TCs)\n• Sự cố & Hóa đơn (56 TCs)",
         150, "Hoàn thành 100% (Pass: 150, Fail: 0)"),

        (3, "Phạm Thị Quỳnh Anh", "PH12347 (Tester / QA)",
         "• Xác thực & Tài khoản (131 TCs)\n• Khách hàng & CSKH (63 TCs)\n• Quản trị Hệ thống RBAC (115 TCs)",
         309, "Hoàn thành 100% (Pass: 309, Fail: 0)"),

        (4, "Nguyễn Ngọc Hà Linh", "PH12348 (Tester / QA)",
         "• Tổng quan Dashboard (15 TCs)\n• Quản lý Phim & Danh mục (258 TCs)\n• Cụm rạp & Lịch chiếu (108 TCs)\n• Giá vé & Khuyến mãi (143 TCs)",
         524, "Hoàn thành 100% (Pass: 524, Fail: 0)")
    ]

    cur_r = 4
    for row in tester_data:
        stt, name, role, scope, total_tc, status = row
        ws3.cell(cur_r, 1, stt).alignment = align_top_center
        ws3.cell(cur_r, 2, name).alignment = align_top_left
        ws3.cell(cur_r, 2).font = font_bold
        ws3.cell(cur_r, 3, role).alignment = align_top_left
        ws3.cell(cur_r, 3).font = font_bold
        ws3.cell(cur_r, 4, scope).alignment = align_top_left
        ws3.cell(cur_r, 5, total_tc).alignment = align_top_center
        ws3.cell(cur_r, 5).font = font_bold
        ws3.cell(cur_r, 6, status).alignment = align_top_center
        ws3.cell(cur_r, 6).font = font_pass

        for c_idx in range(1, 7):
            cell = ws3.cell(cur_r, c_idx)
            cell.border = border_thin
            if c_idx not in [2, 3, 5, 6]:
                cell.font = font_regular

        lines = scope.count('\n') + 1
        ws3.row_dimensions[cur_r].height = max(35, lines * 18)
        cur_r += 1

    # Total Row
    ws3.cell(cur_r, 1, "TỔNG CỘNG TOÀN BỘ DỰ ÁN").font = font_bold
    ws3.cell(cur_r, 1).alignment = align_center
    ws3.merge_cells(start_row=cur_r, start_column=1, end_row=cur_r, end_column=4)
    for c in range(1, 5):
        ws3.cell(cur_r, c).border = border_header
        ws3.cell(cur_r, c).fill = fill_group_blue

    ws3.cell(cur_r, 5, "=SUM(E4:E7)").font = font_bold
    ws3.cell(cur_r, 5).alignment = align_center
    ws3.cell(cur_r, 5).border = border_header
    ws3.cell(cur_r, 5).fill = fill_group_blue

    ws3.cell(cur_r, 6, "1.212 / 1.212 Test Cases (100%)").font = font_pass
    ws3.cell(cur_r, 6).alignment = align_center
    ws3.cell(cur_r, 6).border = border_header
    ws3.cell(cur_r, 6).fill = fill_group_blue
    ws3.row_dimensions[cur_r].height = 26

    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 24
    ws3.column_dimensions["D"].width = 45
    ws3.column_dimensions["E"].width = 20
    ws3.column_dimensions["F"].width = 35

    wb.save(target_path)
    print(f"File updated successfully: {target_path}")

    # Also save a copy to project root for backup
    project_copy = r'c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine\Validate_DevCine_InputsOnly.xlsx'
    wb.save(project_copy)
    print(f"File synced to project folder: {project_copy}")

if __name__ == '__main__':
    build_validate_inputs_workbook()
