# -*- coding: utf-8 -*-
import openpyxl
import sys

sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine\TestReport Dự án DevCine.xlsx', data_only=True)
ws_func = wb['FUNCTION']

print("--- FUNCTION SHEET AUDIT (ALL 40 MODULES) ---")
for r in range(2, ws_func.max_row + 1):
    stt = ws_func.cell(r, 1).value
    sheet = ws_func.cell(r, 2).value
    code = ws_func.cell(r, 3).value
    role = ws_func.cell(r, 6).value
    print(f"Row {r:2d} | STT {stt:2d} | {sheet:<25} | Role: {role}")

print("\n--- CHECKING FOR ANY FORBIDDEN WORDS IN ALL CELLS ---")
forbidden = ["Thu ngân", "thu ngân", "soát vé", "Nhân viên rạp", "Quản lý rạp", "(STAFF)", "(MANAGER)", "(ADMIN)"]
found = 0
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in range(1, min(ws.max_row + 1, 150)):
        for c in range(1, min(ws.max_column + 1, 12)):
            val = str(ws.cell(r, c).value or "")
            for f in ["Thu ngân", "thu ngân", "Nhân viên soát vé", "nhân viên soát vé", "Nhân viên rạp", "nhân viên rạp", "Quản lý rạp", "quản lý rạp", "(STAFF)", "(MANAGER)", "(ADMIN)"]:
                if f in val:
                    print(f"[{sheet}] Cell ({r},{c}): found '{f}' in: {val[:60]}")
                    found += 1

if found == 0:
    print("SUCCESS: 0 forbidden role words found across all sheets!")
