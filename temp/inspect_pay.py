# -*- coding: utf-8 -*-
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine\TestReport Dự án DevCine.xlsx")
ws = wb["Thanh toán VNPAY"]

print("=== SHEET 'Thanh toán VNPAY' ===")
for r in range(1, ws.max_row + 1):
    c1 = ws.cell(row=r, column=1).value
    c2 = ws.cell(row=r, column=2).value
    c3 = ws.cell(row=r, column=3).value
    c8 = ws.cell(row=r, column=8).value
    if c1 or c2:
        print(f"Row {r:02d}: A={c1} | B={c2} | C={c3} | H={c8}")
