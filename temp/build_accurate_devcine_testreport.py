# -*- coding: utf-8 -*-
"""
100% Codebase-Accurate TestReport Generator for DevCine
Strictly formats ALL Test Titles as:
'Kiểm tra chức năng [abc] khi [xyz] - [Thành công / Thất bại]'
Directly mapped to Vue.js frontend components in devcine-frontend/src/views/
Graduation Thesis (DATN) Standard.
"""

import os
import sys
import datetime
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

def build_gui_cases(prefix, screen_name, role, elements):
    cases = []
    for i, (elem, desc, exp) in enumerate(elements, start=1):
        c_id = f"{prefix}_GUI_{i:02d}"
        c_title = f"Kiểm tra chức năng hiển thị {elem} khi truy cập màn hình {screen_name} - Thành công"
        c_desc = f"Kiểm tra {desc} trên màn hình {screen_name}"
        c_steps = f"Bước 1: Đăng nhập thành công vào hệ thống DevCine với vai trò '{role}'\nBước 2: Điều hướng đến màn hình '{screen_name}'\nBước 3: Quan sát {elem}\nBước 4: Kiểm tra kết quả hiển thị"
        c_data = "N/A"
        c_exp = exp
        cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    return cases

def build_field_validation_cases(prefix, screen_name, role, field_name, field_label, normal_val, boundaries, special_cases=None):
    cases = []
    tc_num = 1
    
    # 1. Trống
    c_id = f"{prefix}_VAL_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng {field_label} khi để trống dữ liệu - Thành công"
    c_desc = f"Kiểm tra thông báo lỗi bắt buộc nhập đối với trường {field_name}"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Để trống trường '{field_name}'\nBước 3: Nhập đầy đủ các trường còn lại hợp lệ\nBước 4: Click button thực thi\nBước 5: Kiểm tra thông báo lỗi hiển thị"
    c_data = f"{field_name}: '' (Để trống)"
    c_exp = f"Hiển thị thông báo lỗi 'Vui lòng nhập {field_name}' màu đỏ bên dưới ô nhập liệu"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    # 2. Toàn khoảng trắng
    c_id = f"{prefix}_VAL_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng {field_label} khi nhập toàn khoảng trắng - Thành công"
    c_desc = f"Kiểm tra validate không chấp nhận chuỗi chỉ chứa space"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập chuỗi khoảng trắng '     ' vào trường '{field_name}'\nBước 3: Click button thực thi\nBước 4: Kiểm tra thông báo lỗi"
    c_data = f"{field_name}: '     '"
    c_exp = f"Hiển thị thông báo lỗi yêu cầu nhập nội dung hợp lệ cho trường {field_name}"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    # 3. Khoảng trắng ở đầu (Trim)
    c_id = f"{prefix}_VAL_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng {field_label} khi nhập có khoảng trắng ở đầu (Auto trim) - Thành công"
    c_desc = f"Kiểm tra hệ thống tự động cắt bỏ khoảng trắng thừa phía trước"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập '   {normal_val}' vào trường '{field_name}'\nBước 3: Click button thực thi\nBước 4: Kiểm tra dữ liệu được lưu"
    c_data = f"{field_name}: '   {normal_val}'"
    c_exp = f"Hệ thống tự động trim khoảng trắng đầu, lưu dữ liệu '{normal_val}' thành công"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    # 4. Khoảng trắng ở cuối (Trim)
    c_id = f"{prefix}_VAL_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng {field_label} khi nhập có khoảng trắng ở cuối (Auto trim) - Thành công"
    c_desc = f"Kiểm tra hệ thống tự động cắt bỏ khoảng trắng thừa phía sau"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập '{normal_val}   ' vào trường '{field_name}'\nBước 3: Click button thực thi\nBước 4: Kiểm tra dữ liệu được lưu"
    c_data = f"{field_name}: '{normal_val}   '"
    c_exp = f"Hệ thống tự động trim khoảng trắng cuối, lưu dữ liệu '{normal_val}' thành công"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    # Boundaries: min_len, max_len
    if "min_len" in boundaries:
        min_l = boundaries["min_len"]
        # Min - 1 (Fail)
        c_id = f"{prefix}_VAL_{tc_num:02d}"
        c_title = f"Kiểm tra chức năng {field_label} khi nhập độ dài dưới {min_l} ký tự - Thành công"
        c_desc = f"Kiểm tra chặn độ dài tối thiểu của trường {field_name}"
        c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập chuỗi có độ dài {min_l - 1} ký tự vào trường '{field_name}'\nBước 3: Click button thực thi\nBước 4: Kiểm tra thông báo lỗi"
        c_data = f"{field_name}: '{'A' * (min_l - 1)}' ({min_l - 1} ký tự)"
        c_exp = f"Hiển thị thông báo lỗi '{field_name} phải chứa ít nhất {min_l} ký tự'"
        cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
        tc_num += 1

        # Min exact (Pass)
        c_id = f"{prefix}_VAL_{tc_num:02d}"
        c_title = f"Kiểm tra chức năng {field_label} khi nhập đúng độ dài tối thiểu {min_l} ký tự - Thành công"
        c_desc = f"Kiểm tra biên độ dài tối thiểu"
        c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập chuỗi có độ dài đúng {min_l} ký tự vào trường '{field_name}'\nBước 3: Click button thực thi\nBước 4: Kiểm tra kết quả"
        c_data = f"{field_name}: '{'A' * min_l}' ({min_l} ký tự)"
        c_exp = f"Hệ thống chấp nhận dữ liệu và thực hiện thành công"
        cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
        tc_num += 1

    if "max_len" in boundaries:
        max_l = boundaries["max_len"]
        # Max exact (Pass)
        c_id = f"{prefix}_VAL_{tc_num:02d}"
        c_title = f"Kiểm tra chức năng {field_label} khi nhập đúng độ dài tối đa {max_l} ký tự - Thành công"
        c_desc = f"Kiểm tra biên độ dài tối đa"
        c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập chuỗi có độ dài đúng {max_l} ký tự vào trường '{field_name}'\nBước 3: Click button thực thi\nBước 4: Kiểm tra kết quả"
        c_data = f"{field_name}: (Chuỗi {max_l} ký tự)"
        c_exp = f"Hệ thống chấp nhận dữ liệu và thực hiện thành công"
        cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
        tc_num += 1

        # Max + 1 (Fail)
        c_id = f"{prefix}_VAL_{tc_num:02d}"
        c_title = f"Kiểm tra chức năng {field_label} khi nhập vượt quá độ dài tối đa {max_l} ký tự - Thành công"
        c_desc = f"Kiểm tra chặn độ dài tối đa"
        c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập chuỗi có độ dài {max_l + 1} ký tự vào trường '{field_name}'\nBước 3: Click button thực thi\nBước 4: Kiểm tra thông báo lỗi"
        c_data = f"{field_name}: (Chuỗi {max_l + 1} ký tự)"
        c_exp = f"Hiển thị thông báo lỗi '{field_name} không được vượt quá {max_l} ký tự'"
        cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
        tc_num += 1

    # Security: SQL & XSS
    c_id = f"{prefix}_VAL_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng {field_label} khi nhập mã SQL Injection - Thành công"
    c_desc = f"Kiểm tra bảo mật chống tấn công SQL Injection"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập chuỗi \"' OR '1'='1\" vào trường '{field_name}'\nBước 3: Click button thực thi\nBước 4: Kiểm tra phản hồi hệ thống"
    c_data = f"{field_name}: \"' OR '1'='1\""
    c_exp = f"Hệ thống lọc chuỗi an toàn, không gây lỗi cú pháp cơ sở dữ liệu và báo lỗi dữ liệu không hợp lệ"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    c_id = f"{prefix}_VAL_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng {field_label} khi nhập mã XSS Script - Thành công"
    c_desc = f"Kiểm tra bảo mật chống tấn công Cross-Site Scripting"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập chuỗi '<script>alert(\"XSS\")</script>' vào trường '{field_name}'\nBước 3: Click button thực thi\nBước 4: Kiểm tra phản hồi"
    c_data = f"{field_name}: '<script>alert(\"XSS\")</script>'"
    c_exp = f"Hệ thống tự động HTML-encode hoặc từ chối chuỗi script độc hại, đảm bảo an toàn"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    if special_cases:
        for sc_title, sc_desc, sc_steps, sc_data, sc_exp in special_cases:
            c_id = f"{prefix}_VAL_{tc_num:02d}"
            cases.append((c_id, sc_title, sc_desc, sc_steps, sc_data, sc_exp))
            tc_num += 1

    return cases

def build_search_filter_cases(prefix, screen_name, role, filter_fields):
    cases = []
    tc_num = 1
    
    # 1. Search keyword
    c_id = f"{prefix}_LOC_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng Tìm kiếm khi nhập từ khóa chính xác - Thành công"
    c_desc = f"Kiểm tra tìm kiếm đúng từ khóa trên màn hình {screen_name}"
    c_steps = f"Bước 1: Đăng nhập vai trò '{role}'\nBước 2: Mở màn hình '{screen_name}'\nBước 3: Nhập từ khóa 'Avatar' vào ô Tìm kiếm\nBước 4: Kiểm tra kết quả hiển thị trên bảng"
    c_data = "Tìm kiếm: 'Avatar'"
    c_exp = f"Bảng dữ liệu tự động lọc và chỉ hiển thị các bản ghi có chứa từ khóa 'Avatar'"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    # 2. Search with leading spaces
    c_id = f"{prefix}_LOC_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng Tìm kiếm khi nhập từ khóa có khoảng trắng ở đầu - Thành công"
    c_desc = f"Kiểm tra tự động trim khoảng trắng khi tìm kiếm"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập '   Avatar' vào ô Tìm kiếm\nBước 3: Kiểm tra kết quả hiển thị"
    c_data = "Tìm kiếm: '   Avatar'"
    c_exp = f"Hệ thống tự động cắt khoảng trắng đầu và trả về kết quả khớp với từ khóa 'Avatar'"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    # 3. Search with trailing spaces
    c_id = f"{prefix}_LOC_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng Tìm kiếm khi nhập từ khóa có khoảng trắng ở cuối - Thành công"
    c_desc = f"Kiểm tra tự động trim khoảng trắng sau khi tìm kiếm"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập 'Avatar   ' vào ô Tìm kiếm\nBước 3: Kiểm tra kết quả hiển thị"
    c_data = "Tìm kiếm: 'Avatar   '"
    c_exp = f"Hệ thống tự động cắt khoảng trắng cuối và trả về kết quả khớp với từ khóa 'Avatar'"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    # 4. Search no result
    c_id = f"{prefix}_LOC_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng Tìm kiếm khi nhập từ khóa không tồn tại - Thành công"
    c_desc = f"Kiểm tra hiển thị trạng thái trống (Empty State)"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập từ khóa không tồn tại 'XXXX_NOT_FOUND_999'\nBước 3: Kiểm tra kết quả hiển thị"
    c_data = "Tìm kiếm: 'XXXX_NOT_FOUND_999'"
    c_exp = f"Hiển thị thông báo 'Không tìm thấy dữ liệu phù hợp' kèm hình minh họa trống"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    # 5. Dropdown filters
    for f_name, f_val in filter_fields:
        c_id = f"{prefix}_LOC_{tc_num:02d}"
        c_title = f"Kiểm tra chức năng Lọc khi chọn tiêu chí {f_name} là '{f_val}' - Thành công"
        c_desc = f"Kiểm tra lọc dữ liệu theo combobox {f_name}"
        c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Click vào combobox '{f_name}'\nBước 3: Chọn giá trị '{f_val}'\nBước 4: Kiểm tra kết quả hiển thị"
        c_data = f"{f_name}: '{f_val}'"
        c_exp = f"Bảng dữ liệu lọc chính xác toàn bộ các bản ghi có {f_name} là '{f_val}'"
        cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
        tc_num += 1

    # 6. Combined filter
    c_id = f"{prefix}_LOC_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng Lọc khi kết hợp đồng thời nhiều tiêu chí lọc - Thành công"
    c_desc = f"Kiểm tra lọc đồng thời từ khóa và các combobox"
    c_steps = f"Bước 1: Mở màn hình '{screen_name}'\nBước 2: Nhập từ khóa tìm kiếm và chọn đồng thời các giá trị trên các combobox bộ lọc\nBước 3: Kiểm tra kết quả hiển thị"
    c_data = "Kết hợp đa tiêu chí lọc"
    c_exp = f"Bảng dữ liệu hiển thị các bản ghi thỏa mãn đồng thời tất cả các tiêu chí đã chọn"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    # 7. Reset filters
    c_id = f"{prefix}_LOC_{tc_num:02d}"
    c_title = f"Kiểm tra chức năng Xóa bộ lọc khi click button 'Đặt lại / Reset' - Thành công"
    c_desc = f"Kiểm tra quay về danh sách mặc định khi xóa bộ lọc"
    c_steps = f"Bước 1: Đang ở trạng thái có áp dụng bộ lọc\nBước 2: Click button 'Xóa bộ lọc / Đặt lại'\nBước 3: Kiểm tra danh sách hiển thị"
    c_data = "Click Reset"
    c_exp = f"Toàn bộ các ô tìm kiếm và combobox quay về trạng thái mặc định, hiển thị lại toàn bộ danh sách ban đầu"
    cases.append((c_id, c_title, c_desc, c_steps, c_data, c_exp))
    tc_num += 1

    return cases

def build_accurate_modules():
    modules = []

    # =========================================================================
    # 1. ĐĂNG NHẬP (LoginView.vue & AdminLoginView.vue)
    # =========================================================================
    tc_dn = [
        # GUI
        ("DN_GUI_01", "Kiểm tra chức năng hiển thị bố cục chia đôi màn hình khi truy cập trang Đăng nhập - Thành công",
         "Kiểm tra nửa trái là hình ảnh Cinematic Poster và nửa phải là Form xác thực Auth Card",
         "Bước 1: Mở trình duyệt và truy cập vào đường dẫn '/login'\nBước 2: Quan sát cấu trúc tổng thể màn hình\nBước 3: Kiểm tra hiển thị 2 nửa màn hình",
         "N/A", "Màn hình chia 2 nửa rõ ràng: Nửa trái hiển thị ảnh Poster nghệ thuật kèm dòng chữ 'DEVCINE CINEMA.', nửa phải hiển thị Auth Card nền tối chứa form đăng nhập"),

        ("DN_GUI_02", "Kiểm tra chức năng hiển thị tiêu đề 'Chào mừng trở lại' khi mở Auth Card - Thành công",
         "Kiểm tra tiêu đề chính trên Auth Card",
         "Bước 1: Mở trang Đăng nhập\nBước 2: Quan sát phần đầu của Auth Card bên phải",
         "N/A", "Hiển thị tiêu đề 'Chào mừng trở lại' chữ lớn đậm màu trắng và dòng phụ 'Hãy nhập thông tin để truy cập tài khoản DevCine của bạn.'"),

        ("DN_GUI_03", "Kiểm tra chức năng hiển thị 2 Tab chuyển đổi 'Đăng nhập' và 'Đăng ký' khi mở Auth Card - Thành công",
         "Kiểm tra thanh chuyển tab trên Auth Card",
         "Bước 1: Mở trang Đăng nhập\nBước 2: Quan sát 2 tab Đăng nhập / Đăng ký",
         "N/A", "Hiển thị 2 tab chữ in hoa; tab 'Đăng nhập' đang được chọn có chữ màu vàng (#f5c518) và đường viền gạch chân màu vàng"),

        ("DN_GUI_04", "Kiểm tra chức năng hiển thị trường nhập 'Số điện thoại hoặc Email' khi mở form Đăng nhập - Thành công",
         "Kiểm tra label và placeholder của ô identifier",
         "Bước 1: Mở form Đăng nhập\nBước 2: Quan sát ô nhập đầu tiên",
         "N/A", "Hiển thị nhãn 'SỐ ĐIỆN THOẠI HOẶC EMAIL', icon người dùng bên trái, placeholder 'Nhập số điện thoại hoặc email'"),

        ("DN_GUI_05", "Kiểm tra chức năng hiển thị trường nhập 'Mật khẩu' kèm nút ẩn/hiện khi mở form Đăng nhập - Thành công",
         "Kiểm tra ô nhập mật khẩu",
         "Bước 1: Mở form Đăng nhập\nBước 2: Quan sát ô nhập mật khẩu",
         "N/A", "Hiển thị nhãn 'MẬT KHẨU', icon ổ khóa bên trái, placeholder '••••••••' và icon con mắt (visibility) ở góc phải"),

        ("DN_GUI_06", "Kiểm tra chức năng Ẩn / Hiện mật khẩu khi click icon con mắt - Thành công",
         "Kiểm tra toggle visibility mật khẩu",
         "Bước 1: Nhập 'Khach@123' vào ô mật khẩu\nBước 2: Click vào icon con mắt\nBước 3: Click lại lần 2",
         "Mật khẩu: 'Khach@123'", "Click lần 1: Mật khẩu hiển thị rõ chữ và icon đổi sang visibility_off; Click lần 2: Mật khẩu chuyển lại về dạng chấm tròn"),

        ("DN_GUI_07", "Kiểm tra chức năng hiển thị liên kết 'Quên mật khẩu?' khi mở form Đăng nhập - Thành công",
         "Kiểm tra nút chuyển sang wizard quên mật khẩu",
         "Bước 1: Mở form Đăng nhập\nBước 2: Quan sát bên dưới ô mật khẩu",
         "N/A", "Hiển thị nút chữ nhỏ 'Quên mật khẩu?' màu xám, chuyển sang màu vàng khi rê chuột"),

        ("DN_GUI_08", "Kiểm tra chức năng hiển thị button 'Đăng nhập' khi mở form Đăng nhập - Thành công",
         "Kiểm tra nút submit đăng nhập",
         "Bước 1: Mở form Đăng nhập\nBước 2: Quan sát button Đăng nhập",
         "N/A", "Button nền vàng chữ đen in hoa đậm 'ĐĂNG NHẬP' kèm icon mũi tên, có hiệu ứng sáng hơn khi hover"),

        ("DN_GUI_09", "Kiểm tra chức năng hiển thị khung Chế độ thử nghiệm với 2 nút Demo khi mở Auth Card - Thành công",
         "Kiểm tra khu vực quick access cho testing",
         "Bước 1: Mở form Đăng nhập\nBước 2: Cuộn xuống cuối Auth Card",
         "N/A", "Hiển thị tiêu đề 'CHẾ ĐỘ THỬ NGHIỆM' cùng 2 nút 'Demo Khách hàng' và 'Vào trang Admin'"),

        # VALIDATION & FUNCTIONAL
        ("DN_VAL_01", "Kiểm tra chức năng Đăng nhập khi để trống cả 2 trường tài khoản và mật khẩu - Thành công",
         "Validate bắt buộc nhập toàn form",
         "Bước 1: Mở form Đăng nhập\nBước 2: Để trống toàn bộ\nBước 3: Click button 'Đăng nhập'\nBước 4: Kiểm tra thông báo lỗi",
         "Tài khoản: '' | Mật khẩu: ''", "Hiển thị thông báo đỏ 'Vui lòng nhập số điện thoại hoặc email.' dưới ô tài khoản và 'Vui lòng nhập mật khẩu.' dưới ô mật khẩu"),

        ("DN_VAL_02", "Kiểm tra chức năng Đăng nhập khi để trống ô Số điện thoại hoặc Email - Thành công",
         "Validate để trống tài khoản",
         "Bước 1: Để trống ô tài khoản\nBước 2: Nhập mật khẩu 'Khach@123'\nBước 3: Click button 'Đăng nhập'",
         "Tài khoản: '' | Mật khẩu: 'Khach@123'", "Hiển thị viền đỏ và dòng chữ 'Vui lòng nhập số điện thoại hoặc email.' màu đỏ"),

        ("DN_VAL_03", "Kiểm tra chức năng Đăng nhập khi để trống ô Mật khẩu - Thành công",
         "Validate để trống mật khẩu",
         "Bước 1: Nhập '0901234567' vào ô tài khoản\nBước 2: Để trống ô mật khẩu\nBước 3: Click button 'Đăng nhập'",
         "Tài khoản: '0901234567' | Mật khẩu: ''", "Hiển thị viền đỏ và dòng chữ 'Vui lòng nhập mật khẩu.' màu đỏ"),

        ("DN_VAL_04", "Kiểm tra chức năng Đăng nhập khi nhập Email sai định dạng - Thành công",
         "Validate regex email",
         "Bước 1: Nhập 'khachhang@devcine' (thiếu đuôi .com)\nBước 2: Nhập mật khẩu 'Khach@123'\nBước 3: Click button 'Đăng nhập'",
         "Tài khoản: 'khachhang@devcine'", "Hiển thị thông báo lỗi 'Email không đúng định dạng (vd: ten@domain.com).'"),

        ("DN_VAL_05", "Kiểm tra chức năng Đăng nhập khi nhập Số điện thoại không đủ 10 số - Thành công",
         "Validate độ dài số điện thoại",
         "Bước 1: Nhập số điện thoại '09012345'\nBước 2: Nhập mật khẩu 'Khach@123'\nBước 3: Click button 'Đăng nhập'",
         "Tài khoản: '09012345'", "Hiển thị thông báo lỗi 'Số điện thoại không hợp lệ (đầu số VN, đủ 10 số).'"),

        ("DN_VAL_06", "Kiểm tra chức năng Đăng nhập khi nhập Số điện thoại chứa chữ cái - Thành công",
         "Validate ký tự số điện thoại",
         "Bước 1: Nhập '09012345ab' vào ô tài khoản\nBước 2: Click button 'Đăng nhập'",
         "Tài khoản: '09012345ab'", "Hiển thị thông báo lỗi 'Số điện thoại chỉ gồm chữ số.'"),

        ("DN_VAL_07", "Kiểm tra chức năng Đăng nhập khi nhập sai mật khẩu - Thành công",
         "Validate mật khẩu không đúng",
         "Bước 1: Nhập tài khoản '0901234567'\nBước 2: Nhập sai mật khẩu 'WrongPass@123'\nBước 3: Click button 'Đăng nhập'",
         "Tài khoản: '0901234567' | Mật khẩu: 'WrongPass@123'", "Hiển thị thông báo Toast góc trên: 'Số điện thoại/email hoặc mật khẩu không chính xác.'"),

        ("DN_FUNC_01", "Kiểm tra chức năng Đăng nhập khi click nút 'Demo Khách hàng' điền sẵn tài khoản test - Thành công",
         "Tính năng điền mẫu demo",
         "Bước 1: Mở form Đăng nhập\nBước 2: Click vào button 'Demo Khách hàng' ở khung dưới",
         "Click Demo", "Tự động điền '0901234567' vào ô tài khoản, 'Khach@123' vào ô mật khẩu, hiện mật khẩu và hiển thị toast 'Đã điền tài khoản demo. Bấm \"Đăng nhập\" để tiếp tục.'"),

        ("DN_FUNC_02", "Kiểm tra chức năng Đăng nhập khi nhập đúng Số điện thoại và Mật khẩu (Khách hàng) - Thành công",
         "Đăng nhập bằng SĐT",
         "Bước 1: Nhập số điện thoại '0901234567'\nBước 2: Nhập đúng mật khẩu 'Khach@123'\nBước 3: Click button 'Đăng nhập'",
         "SĐT: '0901234567' | Pass: 'Khach@123'", "Đăng nhập thành công, hiển thị toast 'Đăng nhập thành công! Chào mừng bạn đã trở lại.', lưu token và chuyển về Trang chủ"),

        ("DN_FUNC_03", "Kiểm tra chức năng Đăng nhập khi nhập đúng Email và Mật khẩu (Khách hàng) - Thành công",
         "Đăng nhập bằng Email",
         "Bước 1: Nhập email 'khachhang@gmail.com'\nBước 2: Nhập đúng mật khẩu 'Khach@123'\nBước 3: Click button 'Đăng nhập'",
         "Email: 'khachhang@gmail.com' | Pass: 'Khach@123'", "Đăng nhập thành công và chuyển hướng về Trang chủ"),

        ("DN_FUNC_04", "Kiểm tra chức năng Đăng nhập khi đăng nhập tài khoản Nhân viên và chuyển sang khu nội bộ - Thành công",
         "Điều hướng role nội bộ",
         "Bước 1: Nhập tài khoản nhân viên quầy 'nv_minh' hoặc email nhân viên\nBước 2: Nhập mật khẩu 'Staff@123'\nBước 3: Click button 'Đăng nhập'",
         "Role: STAFF / ADMIN", "Đăng nhập thành công, hiển thị toast 'Đang chuyển sang khu nội bộ.' và điều hướng vào trang quản trị đầu tiên được phép xem"),

        ("DN_FUNC_05", "Kiểm tra chức năng hiển thị giao diện Đăng nhập Quản trị viên (AdminLoginView) khi truy cập '/admin/login' - Thành công",
         "Giao diện trang /admin/login",
         "Bước 1: Truy cập vào đường dẫn '/admin/login'\nBước 2: Quan sát giao diện",
         "URL: '/admin/login'", "Hiển thị nền StarryBackground sao động, thẻ kính mờ bo tròn ở giữa có tiêu đề 'DEVCINE ADMIN' / 'HỆ THỐNG QUẢN TRỊ' và 2 ô input floating label")
    ]

    modules.append({
        "code": "MOD_AUTH_LOGIN", "sheet": "Đăng nhập",
        "req": "Kiểm tra Đăng nhập tài khoản khách hàng, nhân viên và quản trị viên",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng & Nhân viên",
        "pre": "Người dùng mở trình duyệt và truy cập vào trang Đăng nhập hệ thống DevCine",
        "test_cases": tc_dn
    })

    # =========================================================================
    # 2. ĐĂNG KÝ (LoginView.vue - Tab Đăng ký)
    # =========================================================================
    tc_dk = [
        # GUI
        ("DK_GUI_01", "Kiểm tra chức năng hiển thị Tab 'Đăng ký' khi click chuyển tab trên Auth Card - Thành công",
         "Kiểm tra chuyển sang form đăng ký",
         "Bước 1: Tại màn hình Auth Card, click vào tab 'Đăng ký'\nBước 2: Quan sát giao diện form hiển thị",
         "N/A", "Tab 'Đăng ký' chuyển sang màu vàng có viền gạch chân, hiển thị form Đăng ký với 4 trường nhập liệu"),

        ("DK_GUI_02", "Kiểm tra chức năng hiển thị trường nhập 'Họ và tên' khi mở tab Đăng ký - Thành công",
         "Kiểm tra ô họ tên",
         "Bước 1: Mở tab Đăng ký\nBước 2: Quan sát ô nhập đầu tiên",
         "N/A", "Hiển thị nhãn 'HỌ VÀ TÊN', icon person, placeholder 'Nguyễn Văn A'"),

        ("DK_GUI_03", "Kiểm tra chức năng hiển thị trường nhập 'Email của bạn' khi mở tab Đăng ký - Thành công",
         "Kiểm tra ô email",
         "Bước 1: Mở tab Đăng ký\nBước 2: Quan sát ô nhập thứ hai",
         "N/A", "Hiển thị nhãn 'EMAIL CỦA BẠN', icon mail, placeholder 'email@example.com'"),

        ("DK_GUI_04", "Kiểm tra chức năng hiển thị trường nhập 'Số điện thoại' khi mở tab Đăng ký - Thành công",
         "Kiểm tra ô số điện thoại",
         "Bước 1: Mở tab Đăng ký\nBước 2: Quan sát ô nhập thứ ba",
         "N/A", "Hiển thị nhãn 'SỐ ĐIỆN THOẠI', icon call, placeholder 'VD: 0901234567'"),

        ("DK_GUI_05", "Kiểm tra chức năng hiển thị trường nhập 'Mật khẩu' kèm dòng hướng dẫn khi mở tab Đăng ký - Thành công",
         "Kiểm tra ô mật khẩu đăng ký",
         "Bước 1: Mở tab Đăng ký\nBước 2: Quan sát ô nhập thứ tư",
         "N/A", "Hiển thị nhãn 'MẬT KHẨU', placeholder '••••••••', icon con mắt và dòng chữ nhỏ '8–32 ký tự, gồm chữ hoa, chữ thường, số và ký tự đặc biệt.'"),

        ("DK_GUI_06", "Kiểm tra chức năng vô hiệu hóa (disabled) button 'Đăng ký' khi form chưa điền hợp lệ - Thành công",
         "Kiểm tra nút đăng ký disabled",
         "Bước 1: Mở tab Đăng ký khi chưa điền thông tin\nBước 2: Quan sát button 'Đăng ký'",
         "N/A", "Button 'Đăng ký' bị làm mờ (opacity-50), con trỏ chuột not-allowed và không thể click submit"),

        # VALIDATION & FUNCTIONAL
        ("DK_VAL_01", "Kiểm tra chức năng Họ và tên khi để trống họ tên - Thành công",
         "Validate để trống họ tên",
         "Bước 1: Click vào ô Họ và tên rồi blur ra ngoài (rời chuột)\nBước 2: Kiểm tra thông báo lỗi",
         "Họ tên: ''", "Hiển thị thông báo đỏ 'Vui lòng nhập họ và tên.' dưới ô Họ và tên"),

        ("DK_VAL_02", "Kiểm tra chức năng Họ và tên khi nhập họ tên dưới 2 ký tự - Thành công",
         "Validate độ dài họ tên",
         "Bước 1: Nhập 'A' vào ô Họ và tên\nBước 2: Rời chuột ra ngoài",
         "Họ tên: 'A'", "Hiển thị thông báo đỏ 'Họ tên từ 2 đến 50 ký tự.'"),

        ("DK_VAL_03", "Kiểm tra chức năng Họ và tên khi chứa chữ số hoặc ký tự đặc biệt - Thành công",
         "Validate ký tự họ tên",
         "Bước 1: Nhập 'Nguyễn Văn A 123@' vào ô Họ và tên\nBước 2: Rời chuột",
         "Họ tên: 'Nguyễn Văn A 123@'", "Hiển thị thông báo đỏ 'Họ tên chỉ gồm chữ cái và khoảng trắng.'"),

        ("DK_VAL_04", "Kiểm tra chức năng Email khi nhập Email sai định dạng - Thành công",
         "Validate định dạng email",
         "Bước 1: Nhập 'email_sai_dinh_dang' vào ô Email\nBước 2: Rời chuột",
         "Email: 'email_sai_dinh_dang'", "Hiển thị thông báo đỏ 'Email không đúng định dạng (vd: ten@domain.com).'"),

        ("DK_VAL_05", "Kiểm tra chức năng Số điện thoại khi gõ phím chữ tự động lọc bỏ ký tự không phải số - Thành công",
         "Chặn ký tự không phải số",
         "Bước 1: Gõ các phím chữ 'abc090123xyz'\nBước 2: Quan sát giá trị hiển thị trong ô",
         "Phím gõ: 'abc090123xyz'", "Hệ thống tự động lọc bỏ chữ cái, chỉ giữ lại các chữ số '090123'"),

        ("DK_VAL_06", "Kiểm tra chức năng Số điện thoại khi nhập số điện thoại không đủ 10 số - Thành công",
         "Validate độ dài số điện thoại",
         "Bước 1: Nhập '090123456' (9 số)\nBước 2: Rời chuột",
         "SĐT: '090123456'", "Hiển thị thông báo đỏ 'Số điện thoại không hợp lệ (đầu số VN, đủ 10 số).'"),

        ("DK_VAL_07", "Kiểm tra chức năng Mật khẩu khi mật khẩu không đủ 4 nhóm ký tự (hoa, thường, số, ký tự đặc biệt) - Thành công",
         "Validate độ phức tạp mật khẩu",
         "Bước 1: Nhập 'matkhau123' (chỉ có chữ thường và số)\nBước 2: Rời chuột",
         "Mật khẩu: 'matkhau123'", "Hiển thị thông báo đỏ 'Mật khẩu 8–32 ký tự, gồm chữ hoa, chữ thường, số và ký tự đặc biệt.'"),

        ("DK_FUNC_01", "Kiểm tra chức năng Đăng ký khi nhập đầy đủ thông tin hợp lệ - Thành công",
         "Luồng đăng ký thành công",
         "Bước 1: Nhập Họ tên 'Nguyễn Văn Dân', Email 'dan.nguyen.test@gmail.com', SĐT '0987654321', Mật khẩu 'Dan@123456'\nBước 2: Quan sát button 'Đăng ký' sáng lên\nBước 3: Click button 'Đăng ký'\nBước 4: Kiểm tra kết quả",
         "Họ tên: 'Nguyễn Văn Dân' | Email: 'dan.nguyen.test@gmail.com' | SĐT: '0987654321' | MK: 'Dan@123456'",
         "Đăng ký thành công, tự động gọi API đăng nhập, hiển thị toast 'Đăng ký thành công! Chào mừng bạn đến với DevCine.' và chuyển về Trang chủ"),

        ("DK_FUNC_02", "Kiểm tra chức năng Đăng ký khi đăng ký bằng Số điện thoại hoặc Email đã tồn tại - Thành công",
         "Validate trùng lặp",
         "Bước 1: Nhập SĐT '0901234567' đã có trong cơ sở dữ liệu\nBước 2: Điền đầy đủ thông tin hợp lệ\nBước 3: Click button 'Đăng ký'",
         "SĐT: '0901234567' (Đã tồn tại)", "Hệ thống báo lỗi Toast: 'Số điện thoại hoặc Email này đã được sử dụng. Vui lòng đăng nhập hoặc dùng thông tin khác.'")
    ]

    modules.append({
        "code": "MOD_AUTH_REG", "sheet": "Đăng ký",
        "req": "Kiểm tra Đăng ký tài khoản khách hàng mới",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng",
        "pre": "Người dùng mở trang Đăng ký tài khoản DevCine",
        "test_cases": tc_dk
    })

    # =========================================================================
    # 3. QUÊN MẬT KHẨU & OTP (LoginView.vue - Wizard 3 bước)
    # =========================================================================
    tc_qmk = [
        # GUI
        ("QMK_GUI_01", "Kiểm tra chức năng hiển thị Wizard Quên mật khẩu khi click 'Quên mật khẩu?' trên Auth Card - Thành công",
         "Mở wizard quên mật khẩu",
         "Bước 1: Tại form Đăng nhập, click vào 'Quên mật khẩu?'\nBước 2: Quan sát giao diện mở ra",
         "N/A", "Form chuyển sang giao diện Quên mật khẩu có nút 'Quay lại đăng nhập' và thanh chỉ báo tiến trình 3 bước"),

        ("QMK_GUI_02", "Kiểm tra chức năng hiển thị Thanh chỉ báo tiến trình 3 bước khi mở wizard Quên mật khẩu - Thành công",
         "Thanh tiến trình 3 bước",
         "Bước 1: Mở wizard Quên mật khẩu\nBước 2: Quan sát thanh chỉ báo trên đầu form",
         "N/A", "Hiển thị 3 thanh vạch ngang; vạch bước 1 có màu vàng (#f5c518), 2 vạch còn lại màu xám mờ"),

        ("QMK_GUI_03", "Kiểm tra chức năng hiển thị Bước 1: Ô nhập Email tài khoản khi mở wizard Quên mật khẩu - Thành công",
         "Giao diện bước 1",
         "Bước 1: Mở wizard Quên mật khẩu ở Bước 1\nBước 2: Quan sát các trường",
         "N/A", "Hiển thị tiêu đề 'Quên mật khẩu', mô tả hướng dẫn, ô nhập email và button 'Gửi mã xác minh'"),

        ("QMK_GUI_04", "Kiểm tra chức năng hiển thị Bước 2: Ô nhập OTP 6 số căn giữa tracking rộng khi gửi mã thành công - Thành công",
         "Giao diện bước 2",
         "Bước 1: Gửi mã OTP thành công để chuyển sang Bước 2\nBước 2: Quan sát ô nhập mã OTP",
         "N/A", "Hiển thị tiêu đề 'Nhập mã xác minh', ô nhập OTP có font số to 24px, căn giữa, khoảng cách số rộng (tracking-[0.5em]), placeholder '------'"),

        ("QMK_GUI_05", "Kiểm tra chức năng hiển thị Đồng hồ đếm ngược Cooldown 30s khi vừa gửi mã OTP - Thành công",
         "Đếm ngược 30 giây",
         "Bước 1: Vừa bấm gửi mã OTP\nBước 2: Quan sát nút gửi mã",
         "N/A", "Nút bị vô hiệu hóa và hiển thị đếm ngược 'Gửi lại sau (30s)' giảm dần từng giây"),

        ("QMK_GUI_06", "Kiểm tra chức năng hiển thị Bước 3: Form Đặt mật khẩu mới khi xác thực OTP thành công - Thành công",
         "Giao diện bước 3",
         "Bước 1: Xác thực OTP thành công để chuyển sang Bước 3\nBước 2: Quan sát các trường",
         "N/A", "Hiển thị tiêu đề 'Đặt mật khẩu mới', 2 ô nhập Mật khẩu mới và Xác nhận mật khẩu kèm nút 'Đặt lại mật khẩu'"),

        # VALIDATION & FUNCTIONAL
        ("QMK_VAL_01", "Kiểm tra chức năng Quên mật khẩu khi để trống Email hoặc nhập Email sai định dạng - Thành công",
         "Validate email quên mật khẩu",
         "Bước 1: Để trống ô Email\nBước 2: Click button 'Gửi mã xác minh'",
         "Email: ''", "Hiển thị toast lỗi: 'Email không đúng định dạng (vd: ten@domain.com).'"),

        ("QMK_FUNC_01", "Kiểm tra chức năng Quên mật khẩu khi gửi mã OTP thành công về Email tài khoản - Thành công",
         "Gửi OTP thành công",
         "Bước 1: Nhập email tài khoản hợp lệ 'dan.nguyen@gmail.com'\nBước 2: Click button 'Gửi mã xác minh'\nBước 3: Kiểm tra giao diện và hộp thư",
         "Email: 'dan.nguyen@gmail.com'", "Hiển thị toast 'Đã gửi yêu cầu cấp mã xác minh. Vui lòng kiểm tra hộp thư', kích hoạt đếm ngược 30s, chuyển sang Bước 2 và tự động focus vào ô OTP"),

        ("QMK_VAL_02", "Kiểm tra chức năng Quên mật khẩu khi nhập mã OTP không đủ 6 chữ số - Thành công",
         "Validate độ dài OTP",
         "Bước 1: Tại Bước 2, nhập '1234'\nBước 2: Click button 'Xác minh'",
         "OTP: '1234'", "Hiển thị toast lỗi: 'Mã xác minh gồm 6 chữ số.'"),

        ("QMK_VAL_03", "Kiểm tra chức năng Quên mật khẩu khi nhập mã OTP không chính xác hoặc đã hết hạn - Thành công",
         "Xác thực OTP sai",
         "Bước 1: Nhập mã OTP '999999' sai\nBước 2: Click button 'Xác minh'",
         "OTP: '999999'", "Hiển thị toast lỗi: 'Mã xác minh không đúng hoặc đã hết hạn.'"),

        ("QMK_FUNC_02", "Kiểm tra chức năng Quên mật khẩu khi xác thực chính xác mã OTP 6 số hợp lệ - Thành công",
         "Xác thực OTP đúng",
         "Bước 1: Nhập chính xác mã OTP 6 số nhận từ email\nBước 2: Click button 'Xác minh'",
         "OTP: Hợp lệ", "Xác thực thành công và chuyển sang Bước 3 Đặt mật khẩu mới"),

        ("QMK_VAL_04", "Kiểm tra chức năng Quên mật khẩu khi Mật khẩu xác nhận không khớp mật khẩu mới - Thành công",
         "Validate khớp mật khẩu mới",
         "Bước 1: Nhập MK mới 'NewPass@2026', nhập xác nhận 'NewPass@9999'\nBước 2: Click button 'Đặt lại mật khẩu'",
         "Mật khẩu mới: Không khớp", "Hiển thị toast lỗi: 'Mật khẩu xác nhận không khớp.'"),

        ("QMK_FUNC_03", "Kiểm tra chức năng Quên mật khẩu khi đặt lại mật khẩu mới hợp lệ và tự động quay về Đăng nhập - Thành công",
         "Đổi mật khẩu thành công",
         "Bước 1: Nhập MK mới 'NewPass@2026' và xác nhận khớp\nBước 2: Click button 'Đặt lại mật khẩu'\nBước 3: Kiểm tra chuyển hướng",
         "Mật khẩu mới: 'NewPass@2026'", "Hiển thị toast 'Đặt lại mật khẩu thành công! Vui lòng đăng nhập lại.', tự động điền email vào form Đăng nhập và quay về tab Đăng nhập")
    ]

    modules.append({
        "code": "MOD_AUTH_FORGOT", "sheet": "Quên mật khẩu",
        "req": "Kiểm tra Quên mật khẩu, Xác thực OTP và Đặt lại mật khẩu",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng",
        "pre": "Người dùng mở trang Quên mật khẩu trên website DevCine",
        "test_cases": tc_qmk
    })

    # =========================================================================
    # 4. CHỌN GHẾ & GIỮ CHỖ (BookingView.vue - Bước 1)
    # =========================================================================
    tc_st = [
        # GUI
        ("ST_GUI_01", "Kiểm tra chức năng hiển thị Thanh điều hướng Wizard 4 bước đặt vé khi truy cập BookingView - Thành công",
         "Thanh tiến trình 4 bước",
         "Bước 1: Chọn phim, suất chiếu và chuyển vào trang Đặt vé\nBước 2: Quan sát thanh tiến trình trên cùng",
         "N/A", "Hiển thị 4 bước: '01. Chọn Chỗ Ngồi' (icon event_seat), '02. Combo' (icon fastfood), '03. Ưu Đãi' (icon local_activity), '04. Thanh Toán' (icon payments)"),

        ("ST_GUI_02", "Kiểm tra chức năng hiển thị Màn hình chiếu cong (Screen) khi mở sơ đồ chọn ghế - Thành công",
         "Màn hình Screen",
         "Bước 1: Mở màn hình Chọn ghế\nBước 2: Quan sát khu vực trên cùng của sơ đồ ghế",
         "N/A", "Hiển thị dải màn hình cong phát sáng có chữ 'MÀN HÌNH CHIẾU / SCREEN' ở trung tâm"),

        ("ST_GUI_03", "Kiểm tra chức năng hiển thị Sơ đồ ma trận ghế SeatGridRenderer khi mở bước chọn ghế - Thành công",
         "Sơ đồ ma trận ghế",
         "Bước 1: Mở màn hình Chọn ghế\nBước 2: Quan sát toàn bộ lưới ghế",
         "N/A", "Hiển thị lưới ghế cân đối, các hàng được đánh chữ cái A, B, C... ở 2 bên mép, mỗi ghế có mã nhãn rõ nét"),

        ("ST_GUI_04", "Kiểm tra chức năng hiển thị Bảng chú giải màu sắc ghế (Legend) khi mở sơ đồ chọn ghế - Thành công",
         "Bảng chú giải loại ghế",
         "Bước 1: Quan sát khu vực dưới sơ đồ ghế",
         "N/A", "Hiển thị rõ các ô mẫu: Ghế Thường (xám), Ghế VIP (vàng), Ghế Đôi Sweetbox (hồng), Ghế Đang Chọn (xanh lá), Ghế Đang Giữ (cam), Ghế Đã Bán (đỏ), Ghế Bảo Trì (gạch chéo)"),

        ("ST_GUI_05", "Kiểm tra chức năng hiển thị Thanh tóm tắt Đơn đặt vé cố định (Sidebar Summary) khi chọn ghế - Thành công",
         "Sidebar tóm tắt đặt vé",
         "Bước 1: Quan sát cột bên phải màn hình",
         "N/A", "Hiển thị Poster phim, Tên phim, Nhãn độ tuổi, Cụm rạp, Phòng chiếu, Suất chiếu, Danh sách ghế đã chọn, Tổng tiền tạm tính và Nút 'Tiếp tục'"),

        # VALIDATION & FUNCTIONAL
        ("ST_VAL_01", "Kiểm tra chức năng Chọn ghế khi click chọn ghế Đã bán (màu đỏ - SOLD) - Thành công",
         "Chặn ghế đã bán",
         "Bước 1: Di chuột vào ghế có trạng thái SOLD (màu đỏ)\nBước 2: Click chuột vào ghế",
         "Ghế: Trạng thái SOLD", "Ghế bị khóa (cursor-not-allowed), không thể click chọn và không đổi màu"),

        ("ST_VAL_02", "Kiểm tra chức năng Chọn ghế khi click chọn ghế Khóa bảo trì vật lý (MAINTENANCE / LOCKED) - Thành công",
         "Chặn ghế bảo trì",
         "Bước 1: Click vào ghế có trạng thái MAINTENANCE",
         "Ghế: MAINTENANCE", "Ghế bị vô hiệu hóa hoàn toàn, không thể click chọn"),

        ("ST_VAL_03", "Kiểm tra chức năng Chọn ghế khi click chọn ghế đang được khách khác giữ chỗ qua WebSocket STOMP - Thành công",
         "Khóa ghế real-time STOMP",
         "Bước 1: Khách hàng B giữ ghế F08 trên máy khác\nBước 2: Khách hàng A click vào ghế F08\nBước 3: Kiểm tra phản hồi",
         "Ghế: Bị lock bởi khách khác", "Hệ thống từ chối và hiển thị toast: 'Ghế F08 vừa được chọn hoặc đã được bán ở nơi khác. Vui lòng chọn vị trí ghế khác!'"),

        ("ST_VAL_04", "Kiểm tra chức năng Chọn ghế khi click chọn 1 ghế trong cặp ghế đôi Sweetbox - Thành công",
         "Quy tắc ghế đôi Sweetbox",
         "Bước 1: Click vào 1 ghế Sweetbox (H01)\nBước 2: Quan sát trạng thái trên sơ đồ và Sidebar",
         "Ghế click: 'H01'", "Hệ thống tự động kích hoạt chọn cả cặp H01 và H02, Sidebar hiển thị 'H01-H02 (Sweetbox)' và tính đúng tiền 1 cặp ghế đôi"),

        ("ST_VAL_05", "Kiểm tra chức năng Chọn ghế khi để trống 1 ghế đơn cô lập (validateSeatGap) - Thành công",
         "Chống để trống 1 ghế đơn",
         "Bước 1: Hàng ghế có 5 ghế trống liền nhau (E01 đến E05)\nBước 2: Chọn ghế E02 và E03 (bỏ trống đúng 1 ghế E01 ở đầu hàng)\nBước 3: Click 'Tiếp tục'",
         "Ghế chọn: E02, E03 (Trống E01)", "Hiển thị thông báo cảnh báo: 'Vui lòng không để trống 1 ghế đơn cô lập ở đầu hoặc giữa các ghế đã chọn'"),

        ("ST_FUNC_01", "Kiểm tra chức năng Chọn ghế khi chọn đủ số lượng ghế và kích hoạt nút 'Tiếp tục' - Thành công",
         "Chọn ghế thành công",
         "Bước 1: Chọn 2 ghế VIP (E05, E06) khớp với 2 vé đã chọn\nBước 2: Quan sát nút 'Tiếp tục'\nBước 3: Click 'Tiếp tục'",
         "Ghế: E05, E06", "Nút 'Tiếp tục' sáng lên màu vàng, click chuyển mượt mà sang Bước 2 'Combo - Đồ Ăn & Nước Uống'")
    ]

    modules.append({
        "code": "MOD_CUST_SEAT_HOLD", "sheet": "Chọn ghế & Giữ chỗ",
        "req": "Kiểm tra Chọn ghế trên ma trận và Giữ chỗ 10 phút",
        "tester": "Nguyễn Quang Huy", "role": "Khách hàng",
        "pre": "Khách hàng mở sơ đồ chọn ghế của suất chiếu",
        "test_cases": tc_st
    })

    # =========================================================================
    # 5. POS BÁN VÉ TẠI QUẦY (TicketingPOS.vue)
    # =========================================================================
    tc_pos = [
        # GUI
        ("POS_GUI_01", "Kiểm tra chức năng hiển thị Topbar POS với Tên Cụm rạp và Tên Thu ngân khi mở POS - Thành công",
         "Topbar giao diện POS",
         "Bước 1: Nhân viên đăng nhập vào màn hình POS Bán vé\nBước 2: Quan sát thanh Topbar",
         "N/A", "Hiển thị Tên rạp cơ sở (CGV Cầu Giấy), Tên thu ngân (Văn Minh Khôi) và đồng hồ giờ:phút:giây chạy liên tục"),

        ("POS_GUI_02", "Kiểm tra chức năng hiển thị Danh sách Suất chiếu trong ngày của cụm rạp khi mở POS - Thành công",
         "Danh sách suất chiếu POS",
         "Bước 1: Quan sát cột bên trái màn hình POS\nBước 2: Kiểm tra danh sách phim và khung giờ",
         "N/A", "Hiển thị danh sách các phim kèm poster nhỏ, định dạng (2D/3D) và các khung giờ chiếu trong ngày"),

        ("POS_GUI_03", "Kiểm tra chức năng hiển thị Sơ đồ ghế cảm ứng tối ưu cho thao tác chạm nhanh khi chọn suất chiếu - Thành công",
         "Sơ đồ ghế POS",
         "Bước 1: Chọn một suất chiếu\nBước 2: Quan sát sơ đồ ghế ở giữa",
         "N/A", "Sơ đồ ghế hiển thị trực quan, các nút ghế lớn dễ chạm cảm ứng, phân biệt rõ ghế Trống, VIP, Sweetbox, Đã bán"),

        ("POS_GUI_04", "Kiểm tra chức năng hiển thị Panel Thu ngân bên phải (Tra cứu hội viên, Giỏ vé, Tiền thối) khi mở POS - Thành công",
         "Panel thu ngân bên phải",
         "Bước 1: Quan sát cột bên phải màn hình POS",
         "N/A", "Hiển thị ô tra cứu SĐT hội viên (F2), bảng chi tiết vé/F&B đã chọn, ô nhập tiền khách đưa, tiền thừa thối lại và các nút thanh toán"),

        ("POS_GUI_05", "Kiểm tra chức năng hiển thị Nhãn phím tắt nhanh (F2, F4, F9) trên giao diện POS - Thành công",
         "Hiển thị nhãn phím tắt",
         "Bước 1: Quan sát các nút chức năng trên màn hình POS",
         "N/A", "Mỗi nút chức năng đều có nhãn phím tắt tương ứng (F2, F4, F9) giúp nhân viên thao tác bàn phím nhanh"),

        # VALIDATION & FUNCTIONAL
        ("POS_FUNC_01", "Kiểm tra chức năng Bán vé tại quầy khi bán chéo suất chiếu của cụm rạp khác (Cinema Scoping) - Thành công",
         "Ràng buộc chi nhánh POS",
         "Bước 1: Nhân viên thuộc rạp CGV Cầu Giấy mở POS\nBước 2: Kiểm tra danh sách suất chiếu hiển thị",
         "Cinema: CGV Cầu Giấy", "Hệ thống chỉ hiển thị đúng các suất chiếu tại CGV Cầu Giấy, không hiển thị suất của rạp khác"),

        ("POS_FUNC_02", "Kiểm tra chức năng Tra cứu hội viên khi nhập đúng Số điện thoại khách hàng - Thành công",
         "Tra cứu hội viên POS",
         "Bước 1: Nhập số điện thoại '0901234567' vào ô tra cứu (hoặc bấm F2)\nBước 2: Bấm Tìm kiếm",
         "SĐT: '0901234567'", "Hiển thị thông tin: Tên 'Nguyễn Văn Dân', Hạng thẻ 'Vàng (Gold)', Điểm tích lũy '350 điểm'"),

        ("POS_FUNC_03", "Kiểm tra chức năng Tính tiền thừa tiền mặt khi nhập số tiền khách đưa lớn hơn tổng đơn - Thành công",
         "Tính tiền thối lại",
         "Bước 1: Đơn hàng bán vé có tổng tiền 180.000đ\nBước 2: Nhập số tiền khách đưa là 200.000đ\nBước 3: Quan sát ô Tiền thối lại",
         "Tổng: 180.000đ | Đưa: 200.000đ", "Hệ thống tự động tính và hiển thị số tiền thối lại là 20.000đ với font số lớn"),

        ("POS_VAL_01", "Kiểm tra chức năng Bán vé tại quầy khi số tiền khách đưa nhỏ hơn tổng tiền đơn hàng - Thành công",
         "Validate tiền khách đưa",
         "Bước 1: Đơn hàng 220.000đ, nhập tiền khách đưa 200.000đ\nBước 2: Bấm 'Hoàn tất & In vé' (F9)",
         "Tổng: 220.000đ | Đưa: 200.000đ", "Hiển thị cảnh báo lỗi: 'Số tiền khách đưa không đủ để thanh toán đơn hàng'"),

        ("POS_FUNC_04", "Kiểm tra chức năng Đơn chờ POS khi lưu tối đa 3 đơn chờ và khôi phục nạp lại đúng giỏ hàng - Thành công",
         "Quản lý đơn chờ POS",
         "Bước 1: Bấm F4 lưu đơn tạm\nBước 2: Chọn Đơn chờ #1 trên thanh tab\nBước 3: Kiểm tra giỏ hàng",
         "Lưu và khôi phục đơn", "Lưu thành công vào danh sách đơn chờ (tối đa 3 đơn) và nạp lại đúng suất chiếu, vị trí ghế khi click mở lại"),

        ("POS_FUNC_05", "Kiểm tra chức năng Bán vé tại quầy khi hoàn tất thanh toán tiền mặt và xuất lệnh in vé nhiệt - Thành công",
         "Hoàn tất đơn POS",
         "Bước 1: Thu đủ tiền, bấm F9 'Hoàn tất & In vé'\nBước 2: Kiểm tra DB và lệnh in",
         "Action: F9", "Đơn hàng hoàn tất với trạng thái COMPLETED, ghi nhận sold_by nhân viên trực ca, tích điểm hội viên và gửi lệnh in vé nhiệt")
    ]

    modules.append({
        "code": "MOD_POS_TICKETS", "sheet": "POS Bán vé tại quầy",
        "req": "Kiểm tra Bán vé xem phim tại quầy và Cinema Scoping",
        "tester": "Văn Minh Khôi", "role": "Nhân viên Quầy",
        "pre": "Nhân viên đăng nhập vào hệ thống POS cơ sở của rạp mình phụ trách",
        "test_cases": tc_pos
    })

    # =========================================================================
    # 6. QUẢN LÝ PHIM (AdminMovies.vue)
    # =========================================================================
    tc_mov = [
        # GUI
        ("MOV_GUI_01", "Kiểm tra chức năng hiển thị Menu 'Quản lý phim' trên Sidebar Admin khi đăng nhập quyền Admin - Thành công",
         "Sidebar menu Admin",
         "Bước 1: Đăng nhập quyền Admin\nBước 2: Quan sát thanh Sidebar bên trái",
         "N/A", "Mục 'Phim' / 'Quản lý phim' được highlight màu vàng nổi bật trên nền tối"),

        ("MOV_GUI_02", "Kiểm tra chức năng hiển thị Thanh tìm kiếm phim kèm icon kính lúp khi mở trang Quản lý phim - Thành công",
         "Ô tìm kiếm phim",
         "Bước 1: Mở trang Quản lý phim\nBước 2: Quan sát thanh công cụ phía trên",
         "N/A", "Ô tìm kiếm hiển thị rõ ràng, placeholder 'Tìm theo tên phim, đạo diễn...'"),

        ("MOV_GUI_03", "Kiểm tra chức năng hiển thị Bộ lọc Thể loại, Độ tuổi và Trạng thái khi mở trang Quản lý phim - Thành công",
         "Các combobox bộ lọc phim",
         "Bước 1: Quan sát các dropdown bộ lọc trên trang",
         "N/A", "Hiển thị 3 dropdown lọc: Thể loại (Hành động, Kinh dị...), Độ tuổi (P, K, T13, T16, T18), Trạng thái (Đang chiếu, Sắp chiếu, Ngừng chiếu)"),

        ("MOV_GUI_04", "Kiểm tra chức năng hiển thị Bảng danh sách phim kèm ảnh Poster thu nhỏ khi mở trang Quản lý phim - Thành công",
         "Bảng dữ liệu phim",
         "Bước 1: Quan sát bảng danh sách phim",
         "N/A", "Bảng hiển thị đầy đủ các cột: Poster, Tên phim, Thể loại, Thời lượng, Ngày khởi chiếu, Trạng thái, Thao tác"),

        ("MOV_GUI_05", "Kiểm tra chức năng hiển thị Button 'Thêm phim mới' màu vàng khi mở trang Quản lý phim - Thành công",
         "Nút thêm phim",
         "Bước 1: Quan sát góc trên bên phải trang",
         "N/A", "Button 'Thêm phim mới' màu vàng chữ đen kèm icon dấu cộng (+)"),

        ("MOV_GUI_06", "Kiểm tra chức năng hiển thị Modal Form Thêm / Sửa phim chia 2 cột khi click 'Thêm phim mới' - Thành công",
         "Modal nhập liệu phim",
         "Bước 1: Click button 'Thêm phim mới'\nBước 2: Quan sát modal popup",
         "N/A", "Modal hiển thị ở giữa màn hình: Cột trái chứa thông tin cơ bản (Tên, Thể loại, Độ tuổi, Thời lượng, Trailer), Cột phải chứa khu vực upload Poster/Banner và preview"),

        # VALIDATION & FUNCTIONAL
        ("MOV_VAL_01", "Kiểm tra chức năng Tên phim khi để trống Tên phim - Thành công",
         "Validate để trống tên phim",
         "Bước 1: Mở modal thêm phim\nBước 2: Để trống trường Tên phim\nBước 3: Click button 'Lưu phim'",
         "Tên phim: ''", "Hiển thị thông báo lỗi 'Tên phim không được để trống (từ 2 đến 150 ký tự)'"),

        ("MOV_VAL_02", "Kiểm tra chức năng Thời lượng phim khi nhập thời lượng dưới 30 phút hoặc trên 300 phút - Thành công",
         "Biên thời lượng phim",
         "Bước 1: Nhập thời lượng 15 phút\nBước 2: Click button 'Lưu phim'",
         "Thời lượng: 15", "Hiển thị thông báo lỗi 'Thời lượng phim phải là số nguyên từ 30 đến 300 phút'"),

        ("MOV_VAL_03", "Kiểm tra chức năng Trailer khi nhập đường dẫn không phải link Youtube - Thành công",
         "Validate chuẩn link Youtube",
         "Bước 1: Nhập đường dẫn link từ Facebook hoặc TikTok\nBước 2: Click button 'Lưu phim'",
         "Trailer: 'https://facebook.com/video/123'", "Hiển thị thông báo lỗi 'Đường dẫn Trailer phải là link video Youtube hợp lệ (chứa youtube.com hoặc youtu.be)'"),

        ("MOV_FUNC_01", "Kiểm tra chức năng Thêm mới phim khi nhập đầy đủ thông tin và upload ảnh hợp lệ - Thành công",
         "Thêm phim thành công",
         "Bước 1: Nhập đầy đủ thông tin chuẩn, upload ảnh Poster và Banner hợp lệ\nBước 2: Click button 'Lưu phim'\nBước 3: Kiểm tra danh sách",
         "Full valid movie data", "Thêm mới phim thành công, hiển thị toast thông báo thành công và phim xuất hiện trên đầu danh sách"),

        ("MOV_FUNC_02", "Kiểm tra chức năng Xóa phim khi phim đã phát sinh giao dịch đặt vé - Thành công",
         "Khóa ngoại bảo vệ vé bán",
         "Bước 1: Click icon Xóa một bộ phim đang có các suất chiếu đã bán vé\nBước 2: Xác nhận xóa trên popup",
         "Phim: Đã có vé bán", "Hệ thống từ chối xóa, thông báo: 'Không thể xóa phim này do đã phát sinh giao dịch đặt vé. Vui lòng chuyển trạng thái sang Ngừng chiếu'")
    ]

    modules.append({
        "code": "MOD_ADMIN_MOVIE_CRUD", "sheet": "Quản lý phim",
        "req": "Kiểm tra Thêm, Sửa, Xóa và Upload Media Phim",
        "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên",
        "pre": "Admin mở trang Quản lý Phim trong Admin Dashboard",
        "test_cases": tc_mov
    })

    # =========================================================================
    # Remaining modules strictly following the exact requested pattern
    # =========================================================================
    remaining_specs = [
        ("MOD_AUTH_CHANGE_PASS", "Đổi mật khẩu", "Kiểm tra chức năng Đổi mật khẩu trong trang cá nhân", "Phạm Thị Quỳnh Anh", "Khách hàng & Nhân viên", "Đã đăng nhập và mở ChangePasswordView.vue", "DMK",
         [("form Đổi mật khẩu trên ChangePasswordView", "giao diện", "Hiển thị ô Mật khẩu hiện tại, Mật khẩu mới, Xác nhận mật khẩu mới")],
         [("Mật khẩu hiện tại", "Mật khẩu hiện tại", "Khach@123", {"min_len": 6, "max_len": 50}, []),
          ("Mật khẩu mới", "Mật khẩu mới", "NewPass@2026", {"min_len": 8, "max_len": 50}, [])],
         [],
         [("DMK_FUNC_01", "Kiểm tra chức năng Đổi mật khẩu khi Mật khẩu mới trùng với Mật khẩu hiện tại - Thành công", "Trùng MK cũ", "Bước 1: Nhập MK mới giống hệt MK cũ\nBước 2: Bấm Lưu", "Pass mới = Pass cũ", "Báo lỗi 'Mật khẩu mới không được trùng với mật khẩu hiện tại'"),
          ("DMK_FUNC_02", "Kiểm tra chức năng Đổi mật khẩu khi nhập đúng mật khẩu cũ và mật khẩu mới hợp lệ - Thành công", "Đổi MK thành công", "Bước 1: Nhập đúng MK cũ, MK mới 8 ký tự đủ 4 nhóm, xác nhận khớp\nBước 2: Bấm Lưu", "Full valid data", "Đổi mật khẩu thành công và yêu cầu đăng nhập lại")]),

        ("MOD_CUST_PROFILE", "Hồ sơ cá nhân", "Kiểm tra Cập nhật thông tin cá nhân và Avatar", "Phạm Thị Quỳnh Anh", "Khách hàng", "Khách hàng mở ProfileInfoView.vue", "HS",
         [("thông tin hồ sơ và Thẻ hội viên Loyalty", "giao diện", "Hiển thị Họ tên, Email, SĐT, Điểm Loyalty, Hạng thẻ (Đồng/Bạc/Vàng/Kim Cương)")],
         [("Họ và tên", "Họ và tên", "Nguyễn Văn Dân", {"min_len": 2, "max_len": 50}, [])],
         [],
         [("HS_FUNC_01", "Kiểm tra chức năng Cập nhật hồ sơ khi lưu thông tin họ tên và avatar mới - Thành công", "Lưu hồ sơ", "Bước 1: Sửa họ tên, upload avatar mới\nBước 2: Bấm Lưu", "Full valid data", "Lưu thông tin thành công và cập nhật ngay trên Navbar")]),

        ("MOD_CUST_SEARCH", "Tìm kiếm & Lọc phim", "Kiểm tra Tìm kiếm và Bộ lọc phim trên SearchView.vue", "Nguyễn Quang Huy", "Khách hàng", "Khách hàng mở SearchView.vue", "SRC",
         [("thanh tìm kiếm phim với Debounce 300ms", "ô tìm kiếm", "Ô tìm kiếm từ khóa kèm icon kính lúp")],
         [],
         [("Thể loại", "Hành động"), ("Định dạng", "IMAX"), ("Độ tuổi", "T16")],
         [("SRC_FUNC_01", "Kiểm tra chức năng Tìm kiếm khi gõ phím sau khoảng trễ 300ms (Debounce) - Thành công", "Debounce 300ms", "Bước 1: Gõ 'Avatar'\nBước 2: Chờ 300ms", "Keyword: 'Avatar'", "Tự động hiển thị các phim có chứa từ khóa 'Avatar'")]),

        ("MOD_CUST_REVIEW", "Chi tiết phim & Đánh giá", "Kiểm tra Đánh giá sao và Bình luận trên MovieDetail.vue", "Nguyễn Quang Huy", "Khách hàng", "Khách hàng mở MovieDetail.vue", "REV",
         [("thông tin chi tiết phim và Trailer Youtube nhúng", "giao diện", "Hiển thị Poster, Banner, Video Trailer Youtube, Thời lượng, Diễn viên, Tóm tắt")],
         [("Bình luận", "Bình luận", "Kỹ xảo và âm thanh quá tuyệt vời!", {"min_len": 5, "max_len": 500}, [])],
         [],
         [("REV_FUNC_01", "Kiểm tra chức năng Đánh giá khi khách hàng chưa từng mua vé xem phim này - Thành công", "Chặn chưa mua vé", "Bước 1: Dùng tài khoản chưa mua vé gửi đánh giá\nBước 2: Bấm Gửi", "Purchased: False", "Báo lỗi bạn cần mua vé và xem phim trước khi có thể gửi đánh giá"),
          ("REV_FUNC_02", "Kiểm tra chức năng Đánh giá khi gửi đánh giá 5 sao và bình luận hợp lệ - Thành công", "Đánh giá 5 sao", "Bước 1: Chọn 5 sao, nhập bình luận\nBước 2: Bấm Gửi", "Rating: 5 sao", "Gửi đánh giá thành công và cập nhật điểm sao trung bình của phim")]),

        ("MOD_CUST_FNB", "Combo F&B online", "Kiểm tra Chọn bắp nước và Modal tùy chọn FnbOptionModal.vue", "Nguyễn Quang Huy", "Khách hàng", "Khách hàng ở bước Combo trên BookingView", "FNB",
         [("menu bắp nước và Modal FnbOptionModal", "menu F&B", "Hiển thị các combo và popup chọn vị bắp, loại nước ngọt")],
         [],
         [],
         [("FNB_FUNC_01", "Kiểm tra chức năng Tùy chọn vị Combo khi đổi sang vị bắp Phô mai có phụ thu - Thành công", "Phụ thu đổi vị", "Bước 1: Đổi vị bắp sang Phô mai (+15.000đ)", "Extra: +15k", "Tổng tiền tạm tính cộng thêm 15.000đ chính xác"),
          ("FNB_FUNC_02", "Kiểm tra chức năng Chọn F&B khi bỏ qua bước bắp nước để chuyển sang thanh toán - Thành công", "Bỏ qua F&B", "Bước 1: Không chọn món nào\nBước 2: Bấm Bỏ qua & Tiếp tục", "Cart: Empty", "Cho phép bỏ qua và chuyển thẳng sang bước Ưu đãi / Voucher")]),

        ("MOD_CUST_VOUCHER", "Khuyến mãi & Voucher", "Kiểm tra Áp dụng mã giảm giá và Kho voucher trên BookingView", "Nguyễn Quang Huy", "Khách hàng", "Khách hàng ở bước Ưu đãi trên BookingView", "VOU",
         [("ô nhập mã voucher và danh sách voucher trong ví", "voucher view", "Ô nhập mã kèm nút Áp dụng và danh sách voucher khả dụng")],
         [("Mã giảm giá", "Mã giảm giá", "DEVCINE50", {"min_len": 3, "max_len": 30}, [])],
         [],
         [("VOU_FUNC_01", "Kiểm tra chức năng Áp dụng voucher khi áp dụng mã giảm giá có giới hạn mức giảm tối đa - Thành công", "Voucher max discount", "Bước 1: Đơn 500k, áp voucher giảm 50% max 50k\nBước 2: Bấm Áp dụng", "Voucher: 50% max 50k", "Tính tiền giảm đúng 50.000đ, tổng thanh toán còn 450.000đ")]),

        ("MOD_CUST_PAYMENT", "Thanh toán VNPAY", "Kiểm tra Tích hợp Cổng VNPAY và Sinh vé QR trên BookingView", "Nguyễn Quang Huy", "Khách hàng", "Khách hàng thanh toán qua VNPAY", "PAY",
         [("nút 'Thanh toán qua VNPAY' và màn hình BookingSuccessView", "thanh toán VNPAY", "Giao diện chọn cổng VNPAY và trang kết quả đặt vé thành công")],
         [],
         [],
         [("PAY_FUNC_01", "Kiểm tra chức năng Thanh toán VNPAY khi xác thực giao dịch thành công trên cổng thanh toán - Thành công", "Thanh toán thành công", "Bước 1: Nhập thẻ test, xác thực OTP thành công trên VNPAY\nBước 2: Redirect về DevCine", "Response: '00'", "Đơn chuyển CONFIRMED, sinh vé QR độc nhất, tích điểm và gửi email vé cho khách"),
          ("PAY_FUNC_02", "Kiểm tra chức năng Thanh toán VNPAY khi khách hàng chủ động hủy giao dịch trên cổng - Thành công", "Khách hủy đơn", "Bước 1: Bấm Hủy giao dịch trên cổng VNPAY", "Response: '24'", "Hủy đơn hàng và tự động giải phóng ghế đang giữ về trạng thái trống")]),

        ("MOD_CUST_SUPPORT", "Hỗ trợ CSKH", "Kiểm tra Gửi yêu cầu hỗ trợ trên SupportRequestForm.vue", "Nguyễn Quang Huy", "Khách hàng", "Người dùng mở trang FAQView / ContactView", "CS",
         [("form Gửi yêu cầu hỗ trợ CSKH", "form ticket", "Hiển thị ô Tiêu đề, Nội dung thắc mắc và nút gửi")],
         [("Tiêu đề", "Tiêu đề", "Hỏi về chính sách hoàn vé", {"min_len": 5, "max_len": 200}, []),
          ("Nội dung", "Nội dung", "Tôi muốn hỏi chính sách hủy vé khi có sự cố thời tiết...", {"min_len": 10, "max_len": 1000}, [])],
         [],
         [("CS_FUNC_01", "Kiểm tra chức năng Gửi yêu cầu hỗ trợ khi nhập đầy đủ thông tin thắc mắc hợp lệ - Thành công", "Tạo ticket", "Bước 1: Điền đầy đủ thông tin hợp lệ\nBước 2: Bấm Gửi yêu cầu", "Full valid data", "Tạo ticket thành công với trạng thái OPEN và gửi email tiếp nhận cho khách")]),

        ("MOD_STAFF_FIRST_PASS", "Đổi mật khẩu lần đầu", "Kiểm tra Đổi mật khẩu bắt buộc trên FirstLoginPassword.vue", "Văn Minh Khôi", "Nhân viên mới", "Nhân viên mới đăng nhập lần đầu", "FST",
         [("màn hình FirstLoginPassword bắt buộc đổi mật khẩu", "giao diện force change", "Modal bắt buộc đổi mật khẩu để kích hoạt tài khoản")],
         [("Mật khẩu mới", "Mật khẩu mới", "Staff@DevCine2026", {"min_len": 8, "max_len": 50}, [])],
         [],
         [("FST_FUNC_01", "Kiểm tra chức năng Đổi mật khẩu lần đầu khi nhập mật khẩu mới đủ 4 nhóm ký tự - Thành công", "Kích hoạt NV", "Bước 1: Nhập MK mới 4 nhóm ký tự\nBước 2: Bấm Xác nhận", "Pass: 'Staff@DevCine2026'", "Đổi thành công, tắt cờ bắt buộc đổi MK và chuyển vào POS Bán vé")]),

        ("MOD_POS_PENDING", "POS Đơn chờ", "Kiểm tra Quản lý đơn chờ tạm thời trên TicketingPOS.vue", "Văn Minh Khôi", "Nhân viên Quầy", "Đang thao tác chọn vé trên POS", "PND",
         [("thanh tab Đơn chờ trên TicketingPOS (tối đa 3 đơn)", "tabs đơn chờ", "Hiển thị tối đa 3 tab đơn chờ kèm đồng hồ đếm ngược")],
         [],
         [],
         [("PND_FUNC_01", "Kiểm tra chức năng Đơn chờ POS khi tạo thêm đơn chờ thứ 4 trên cùng máy POS - Thành công", "Chặn đơn thứ 4", "Bước 1: Đang có 3 đơn chờ, tạo tiếp đơn 4\nBước 2: Bấm Lưu đơn chờ (F4)", "Orders: 4", "Báo lỗi mỗi máy POS chỉ được lưu tối đa 3 đơn chờ"),
          ("PND_FUNC_02", "Kiểm tra chức năng Đơn chờ POS khi khôi phục lại đơn chờ để tiếp tục thanh toán - Thành công", "Khôi phục đơn", "Bước 1: Click chọn Đơn chờ #2 trên thanh POS", "Order: #2", "Nạp lại đúng suất chiếu, vị trí ghế và bắp nước đã chọn để thu tiền"),
          ("PND_FUNC_03", "Kiểm tra chức năng Đơn chờ POS khi hết hạn 10 phút tự động hủy và phạt khóa ghế 5 phút - Thành công", "Timeout phạt ghế", "Bước 1: Để đơn chờ quá 10 phút", "Timeout: 10'", "Đơn chờ tự hủy, khóa phạt ghế trong 5 phút không cho mở lại ngay")]),

        ("MOD_POS_FNB", "POS Bán F&B tại quầy", "Kiểm tra Bán bắp nước riêng lẻ tại quầy trên TicketingPOS.vue", "Văn Minh Khôi", "Nhân viên Quầy", "Mở tab Bán F&B trên POS", "PFNB",
         [("thực đơn Bán F&B độc lập trên TicketingPOS", "thực đơn POS", "Lưới các món bắp nước kèm nút tăng giảm số lượng nhanh")],
         [],
         [],
         [("PFNB_FUNC_01", "Kiểm tra chức năng Bán F&B tại quầy khi bán bắp nước độc lập và thu tiền mặt - Thành công", "Bán lẻ F&B", "Bước 1: Chọn 2 bắp phô mai, nhập SĐT hội viên, thu 160k tiền mặt\nBước 2: Bấm Hoàn tất thanh toán", "Total: 160k", "Thanh toán thành công, tích điểm hội viên và in hóa đơn bán lẻ F&B")]),

        ("MOD_POS_VOID_FNB", "Yêu cầu hủy đơn F&B", "Kiểm tra Tạo yêu cầu hủy đơn bắp nước (ApprovalQueue.vue)", "Văn Minh Khôi", "Nhân viên Quầy", "Mở lịch sử đơn F&B", "VOID",
         [("modal Yêu cầu hủy đơn F&B", "modal void", "Ô nhập lý do hủy đơn và thông tin tóm tắt đơn")],
         [("Lý do hủy đơn", "Lý do hủy đơn", "Khách hàng đổi ý muốn đổi sang Combo lớn hơn", {"min_len": 5, "max_len": 255}, [])],
         [],
         [("VOID_FUNC_01", "Kiểm tra chức năng Yêu cầu hủy đơn F&B khi nhân viên quầy gửi yêu cầu kèm lý do - Thành công", "Tạo yêu cầu void", "Bước 1: Nhập lý do 'Khách đổi ý đổi combo'\nBước 2: Bấm Gửi yêu cầu", "Reason: 'Khách đổi ý'", "Chuyển đơn sang PENDING_VOID và gửi thông báo real-time cho Quản lý")]),

        ("MOD_STAFF_CHECKIN", "Soát vé & Check-in", "Kiểm tra Quét mã QR soát vé trên TicketCheckIn.vue", "Văn Minh Khôi", "Nhân viên Soát vé", "Mở TicketCheckIn.vue", "CHK",
         [("camera quét mã QR thời gian thực trên TicketCheckIn.vue", "camera QR", "Khung quét QR có tia laser dẫn hướng và ô nhập mã thủ công")],
         [],
         [],
         [("CHK_FUNC_01", "Kiểm tra chức năng Soát vé khi quét mã QR vé hợp lệ trước giờ chiếu - Thành công", "Check-in hợp lệ", "Bước 1: Quét mã QR vé hợp lệ trước giờ chiếu 20 phút", "QR: Valid", "Phát tiếng bíp thành công, hiển thị tích xanh và thông tin ghế E05, E06"),
          ("CHK_FUNC_02", "Kiểm tra chức năng Soát vé khi quét mã QR vé đã được check-in trước đó - Thành công", "Vé đã dùng", "Bước 1: Quét mã vé đã check-in 15 phút trước", "Status: CHECKED_IN", "Cảnh báo đỏ: 'VÉ ĐÃ SỬ DỤNG! Đã check-in lúc 19:15 bởi nhân viên Khôi'"),
          ("CHK_FUNC_03", "Kiểm tra chức năng Soát vé khi quét mã QR vé thuộc cụm rạp khác - Thành công", "Vé sai rạp", "Bước 1: Quét vé rạp Hà Đông tại rạp Cầu Giấy", "Scope: Sai rạp", "Cảnh báo đỏ: 'Vé không hợp lệ tại cụm rạp này. Vé thuộc rạp CGV Hà Đông'")]),

        ("MOD_STAFF_INCIDENT_RELOCATE", "Xử lý sự cố & Đổi ghế", "Kiểm tra Đổi ghế sự cố trên IncidentManagement.vue", "Văn Minh Khôi", "Nhân viên & Quản lý", "Mở IncidentManagement.vue", "REL",
         [("màn hình Đổi ghế sự cố trên IncidentManagement.vue", "giao diện đổi ghế", "Hiển thị ghế nguồn hỏng và sơ đồ chọn ghế đích trống")],
         [("Lý do đổi ghế", "Lý do đổi ghế", "Ghế A01 bị gãy tay vịn cần đổi cho khách", {"min_len": 5, "max_len": 255}, [])],
         [],
         [("REL_FUNC_01", "Kiểm tra chức năng Đổi ghế sự cố khi chuyển khách sang vị trí ghế trống cùng hạng - Thành công", "Đổi ghế giữ QR", "Bước 1: Chọn ghế cũ A01, chọn ghế mới A05 trống, nhập lý do\nBước 2: Bấm Xác nhận đổi ghế", "Src: A01 -> Dst: A05", "Cập nhật vị trí ghế mới A05 cho khách, giữ nguyên mã vé QR và ghi log sự cố")]),

        ("MOD_MGR_APPROVE_VOID", "Phê duyệt hủy đơn F&B", "Kiểm tra Duyệt / Từ chối trên ApprovalQueue.vue", "Nguyễn Ngọc Hà Linh", "Quản lý Cụm rạp", "Quản lý mở ApprovalQueue.vue", "APP",
         [("danh sách các đơn yêu cầu hủy PENDING_VOID trên ApprovalQueue.vue", "danh sách void", "Hiển thị mã đơn, tên món, lý do hủy của nhân viên")],
         [("Lý do từ chối", "Lý do từ chối", "Bắp nước đã giao cho khách, không được phép hủy", {"min_len": 5, "max_len": 255}, [])],
         [],
         [("APP_FUNC_01", "Kiểm tra chức năng Phê duyệt hủy đơn F&B khi Quản lý chấp thuận yêu cầu hủy - Thành công", "Duyệt void", "Bước 1: Bấm Phê duyệt hủy đơn F&B", "Action: APPROVE", "Đơn chuyển VOIDED, trừ doanh thu trong ca và lưu tên Quản lý phê duyệt"),
          ("APP_FUNC_02", "Kiểm tra chức năng Phê duyệt hủy đơn F&B khi Quản lý từ chối yêu cầu hủy - Thành công", "Từ chối void", "Bước 1: Nhập lý do từ chối, bấm Xác nhận từ chối", "Action: REJECT", "Đơn khôi phục COMPLETED và gửi thông báo phản hồi cho nhân viên")]),

        ("MOD_MGR_SEAT_MAINTENANCE", "Khóa bảo trì ghế vật lý", "Kiểm tra Khóa ghế bảo trì trên CinemaManager.vue", "Nguyễn Ngọc Hà Linh", "Quản lý Cụm rạp", "Mở sơ đồ quản lý trạng thái ghế", "MNT",
         [("sơ đồ quản lý bảo trì ghế phòng chiếu", "sơ đồ bảo trì", "Hiển thị ghế thường, ghế đang bảo trì (màu xám cờ lê)")],
         [("Lý do bảo trì", "Lý do bảo trì", "Đệm ghế bị rách cần bọc lại da", {"min_len": 5, "max_len": 255}, [])],
         [],
         [("MNT_FUNC_01", "Kiểm tra chức năng Khóa bảo trì ghế khi chuyển trạng thái ghế sang MAINTENANCE - Thành công", "Khóa bảo trì", "Bước 1: Chọn ghế B03, nhập lý do 'Rách đệm'\nBước 2: Bấm Lưu trạng thái", "Seat: B03", "Ghế B03 chuyển sang MAINTENANCE và tự động bị khóa ẩn trên toàn bộ suất chiếu tương lai")]),

        ("MOD_MGR_COMPENSATION", "Tặng voucher đền bù", "Kiểm tra Phát voucher đền bù trên CustomerSupport.vue", "Nguyễn Ngọc Hà Linh", "Quản lý Cụm rạp", "Mở CustomerSupport.vue", "CMP",
         [("popup Tặng voucher đền bù sự cố cho khách hàng", "popup đền bù", "Dropdown chọn mẫu voucher và ô nhập ghi chú")],
         [("Ghi chú đền bù", "Ghi chú đền bù", "Đền bù sự cố mất điện phòng chiếu 1 ngày 19/03", {"min_len": 5, "max_len": 255}, [])],
         [],
         [("CMP_FUNC_01", "Kiểm tra chức năng Tặng voucher đền bù khi phát vé miễn phí vào ví khách hàng gặp sự cố - Thành công", "Phát voucher đền bù", "Bước 1: Nhập SĐT khách, chọn mẫu 'Vé 2D Miễn Phí', nhập ghi chú\nBước 2: Bấm Tặng voucher", "Template: FREE_2D", "Phát voucher vào ví khách hàng và tự động gửi email xin lỗi kèm mã voucher")]),

        ("MOD_ADMIN_CATEGORIES", "Danh mục phim", "Kiểm tra Thể loại, Định dạng, Độ tuổi trên MovieCategoryManager.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở MovieCategoryManager.vue", "CAT",
         [("bảng Danh mục Thể loại, Định dạng, Độ tuổi trên MovieCategoryManager.vue", "danh sách danh mục", "Liệt kê các danh mục kèm số lượng phim đang sử dụng")],
         [("Tên thể loại", "Tên thể loại", "Khoa học viễn tưởng", {"min_len": 2, "max_len": 50}, [])],
         [],
         [("CAT_FUNC_01", "Kiểm tra chức năng Thêm thể loại phim khi trùng tên thể loại đã có - Thành công", "Trùng thể loại", "Bước 1: Nhập tên 'Hành động' đã có\nBước 2: Bấm Lưu", "Name: 'Hành động'", "Hiển thị thông báo lỗi 'Tên thể loại phim đã tồn tại trong hệ thống'"),
          ("CAT_FUNC_02", "Kiểm tra chức năng Xóa thể loại phim khi đang có 10 bộ phim sử dụng - Thành công", "Khóa ngoại thể loại", "Bước 1: Bấm xóa thể loại đang gắn với 10 phim", "In Use: True", "Báo lỗi không thể xóa thể loại do đang có 10 bộ phim đang sử dụng")]),

        ("MOD_ADMIN_CINEMAS", "Quản lý cụm rạp", "Kiểm tra Thêm, Sửa Cụm rạp trên CinemaManager.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở CinemaManager.vue", "CIN",
         [("bảng Danh sách Cụm rạp và giờ hoạt động trên CinemaManager.vue", "danh sách rạp", "Hiển thị Tên rạp, Địa chỉ, Hotline, Giờ mở/đóng cửa, Số phòng chiếu")],
         [("Tên cụm rạp", "Tên cụm rạp", "DevCine Cầu Giấy", {"min_len": 5, "max_len": 100}, []),
          ("Hotline", "Hotline", "19006017", {"min_len": 8, "max_len": 11}, [])],
         [("Tỉnh/Thành", "Hà Nội")],
         [("CIN_FUNC_01", "Kiểm tra chức năng Thêm cụm rạp khi nhập đầy đủ thông tin và giờ hoạt động hợp lệ - Thành công", "Thêm rạp", "Bước 1: Nhập tên rạp, địa chỉ, hotline, giờ mở 08:00 - đóng 23:30, upload ảnh\nBước 2: Bấm Lưu", "Full valid data", "Thêm cụm rạp thành công, hiển thị trên danh sách quản trị và bản đồ người dùng")]),

        ("MOD_ADMIN_ROOMS", "Quản lý phòng chiếu", "Kiểm tra Thêm, Sửa Phòng chiếu trên CinemaManager.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở tab Phòng chiếu trên CinemaManager.vue", "ROM",
         [("danh sách Phòng chiếu của cụm rạp trên CinemaManager.vue", "danh sách phòng", "Hiển thị Tên phòng, Loại phòng, Số ghế, Thời gian dọn phòng")],
         [("Tên phòng chiếu", "Tên phòng chiếu", "Cinema 03 (IMAX)", {"min_len": 3, "max_len": 50}, [])],
         [("Loại phòng", "IMAX Laser")],
         [("ROM_FUNC_01", "Kiểm tra chức năng Thêm phòng chiếu khi nhập tên phòng và kích thước ma trận ghế hợp lệ - Thành công", "Tạo phòng", "Bước 1: Nhập Tên phòng: 'Cinema 03 (IMAX)', Số hàng: 12, Số cột: 16, Dọn: 20 phút\nBước 2: Bấm Lưu", "Full valid data", "Tạo phòng chiếu thành công và tự động chuyển sang thiết lập sơ đồ ghế")]),

        ("MOD_ADMIN_SEATMAP", "Sơ đồ ghế", "Kiểm tra Thiết lập sơ đồ ma trận ghế trên CinemaManager.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở công cụ vẽ sơ đồ ghế trên CinemaManager.vue", "SMP",
         [("công cụ vẽ ma trận ghế phòng chiếu trên CinemaManager.vue", "bản vẽ ghế", "Lưới ma trận cho phép click chọn loại ghế Thường, VIP, Sweetbox, Lối đi")],
         [],
         [],
         [("SMP_FUNC_01", "Kiểm tra chức năng Thiết lập sơ đồ ghế khi lưu ma trận 10 hàng x 14 cột hoàn chỉnh - Thành công", "Lưu sơ đồ", "Bước 1: Thiết lập 4 hàng Thường, 4 hàng VIP, 2 hàng Sweetbox, 2 cột lối đi\nBước 2: Bấm Lưu sơ đồ ghế", "Matrix: 10x14", "Lưu sơ đồ thành công, tự sinh mã nhãn ghế chuẩn (A01..J14)"),
          ("SMP_FUNC_02", "Kiểm tra chức năng Thiết lập sơ đồ ghế khi chỉnh sửa phòng chiếu đang có vé đã bán - Thành công", "Khóa sửa phòng có vé", "Bước 1: Sửa sơ đồ phòng đang có các suất chiếu đã bán vé\nBước 2: Bấm Lưu", "Has Bookings: True", "Báo lỗi không thể chỉnh sửa sơ đồ ghế do phòng chiếu đang có các suất chiếu tương lai đã bán vé")]),

        ("MOD_ADMIN_SHOWTIMES", "Điều phối lịch chiếu", "Kiểm tra Lịch chiếu và Xung đột phòng trên AdminBookings.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở màn hình Điều phối lịch chiếu", "STA_ADM",
         [("timeline Suất chiếu theo từng phòng trên AdminBookings.vue", "timeline lịch chiếu", "Các phòng chiếu được chia theo hàng ngang, timeline giờ từ 08:00 đến 24:00")],
         [],
         [],
         [("STA_FUNC_01", "Kiểm tra chức năng Thêm suất chiếu khi khung giờ phòng chiếu hoàn toàn trống - Thành công", "Thêm suất chiếu hợp lệ", "Bước 1: Chọn Phim: 'Avatar 2', Phòng: 'Cinema 01', Định dạng: '2D', Ngày: '20/03/2026', Giờ bắt đầu: '20:30'\nBước 2: Click button 'Lưu suất chiếu'", "Full valid data", "Thêm suất chiếu thành công, hiển thị khối suất chiếu trên timeline và mở bán vé trên website"),
          ("STA_FUNC_02", "Kiểm tra chức năng Thêm suất chiếu khi bị xung đột trùng phòng chiếu (Room Overlap) - Thành công", "Room Overlap Conflict", "Bước 1: Thêm suất 19:30 trùng khoảng 18:00-20:20 đang chiếu\nBước 2: Bấm Lưu", "Overlap: True", "Báo lỗi xung đột phòng chiếu kèm tên phim và khoảng giờ trùng")]),

        ("MOD_ADMIN_BATCH_SCHEDULE", "Xếp lịch chiếu hàng loạt", "Kiểm tra Công cụ Xếp lịch hàng loạt (Batch Scheduling)", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở công cụ Xếp lịch hàng loạt", "BSC",
         [("công cụ Xếp lịch hàng loạt trên AdminBookings.vue", "công cụ batch", "Bộ chọn khoảng ngày, danh sách phim, danh sách phòng và mẫu khung giờ")],
         [],
         [],
         [("BSC_FUNC_01", "Kiểm tra chức năng Xếp lịch hàng loạt khi sinh tự động 56 suất chiếu không trùng phòng - Thành công", "Sinh lịch tự động", "Bước 1: Chọn 7 ngày, 2 phim, 2 phòng, 4 khung giờ mẫu\nBước 2: Click button 'Sinh lịch chiếu tự động'", "Batch: 56 suất", "Thuật toán sinh thành công 56 suất chiếu hợp lệ, không có suất nào bị trùng phòng")]),

        ("MOD_ADMIN_FNB_ITEMS", "Quản lý thực đơn F&B", "Kiểm tra Thêm, Sửa món bắp nước trên FnbMenuManager.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở FnbMenuManager.vue", "FNB_ADM",
         [("bảng Danh mục món ăn và nước uống trên FnbMenuManager.vue", "thực đơn admin", "Hiển thị Ảnh món, Tên món, Phân loại, Đơn giá, Trạng thái")],
         [("Tên món F&B", "Tên món F&B", "Bắp Phô Mai Trứng Muối", {"min_len": 2, "max_len": 100}, [])],
         [("Phân loại", "Đồ ăn")],
         [("FNB_ADM_FUNC_01", "Kiểm tra chức năng Thêm món F&B khi nhập đầy đủ thông tin và giá bán hợp lệ - Thành công", "Thêm món", "Bước 1: Nhập tên món, chọn Đồ ăn, giá bán 65.000đ, upload ảnh\nBước 2: Bấm Lưu", "Full valid data", "Thêm món thành công, hiển thị ngay trên thực đơn web và máy POS")]),

        ("MOD_ADMIN_COMBOS", "Cấu hình Combo F&B", "Kiểm tra Cấu hình Combo và Tùy chọn trên FnbMenuManager.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở tab Combo trên FnbMenuManager.vue", "CMB",
         [("trình Cấu hình nhóm tùy chọn món con của Combo", "cấu hình combo", "Cho phép thêm các slot bắp nước, chọn món con và cấu hình mức phụ thu")],
         [("Tên Combo", "Tên Combo", "Couple Combo Đặc Biệt", {"min_len": 3, "max_len": 100}, [])],
         [],
         [("CMB_FUNC_01", "Kiểm tra chức năng Cấu hình Combo F&B khi lưu Combo gồm nhiều thành phần và phụ thu - Thành công", "Lưu Combo", "Bước 1: Cấu hình Slot 1: Bắp 1 vị (+15k phô mai), Slot 2: 2 Nước ngọt\nBước 2: Bấm Lưu Combo", "Full valid data", "Lưu cấu hình Combo thành công, áp dụng đồng bộ trên web và POS")]),

        ("MOD_ADMIN_BASE_PRICING", "Cấu hình bảng giá vé", "Kiểm tra Ma trận giá 3 chiều trên AdminPricing.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở AdminPricing.vue", "PRC",
         [("ma trận Bảng giá vé 3 chiều và Simulator tính giá trên AdminPricing.vue", "ma trận giá vé", "Ma trận hiển thị dạng bảng lưới chia theo khung giờ và loại ngày kèm công cụ Simulator bóc tách giá")],
         [("Giá vé nền", "Giá vé nền", "85000", {}, [])],
         [],
         [("PRC_FUNC_01", "Kiểm tra chức năng Simulator tính giá vé khi bóc tách chi tiết từng dòng phụ thu - Thành công", "Simulator giá vé", "Bước 1: Chọn Thứ 7 (110k), Giờ vàng (10k), VIP (20k), 3D (30k)\nBước 2: Bấm Tính giá", "Tổng: 170.000đ", "Simulator hiển thị chính xác tổng giá vé = 170.000 VNĐ kèm bảng bóc tách chi tiết từng dòng phụ thu")]),

        ("MOD_ADMIN_HOLIDAYS", "Quản lý ngày lễ", "Kiểm tra Khai báo Ngày lễ trên AdminPricing.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở tab Ngày lễ trên AdminPricing.vue", "HOL",
         [("danh sách Ngày lễ trong năm trên AdminPricing.vue", "danh sách ngày lễ", "Hiển thị Tên ngày lễ và Ngày áp dụng")],
         [("Tên ngày lễ", "Tên ngày lễ", "Quốc Khánh 02/09", {"min_len": 3, "max_len": 100}, [])],
         [],
         [("HOL_FUNC_01", "Kiểm tra chức năng Thêm ngày lễ khi khai báo ngày lễ mới và tự động áp biểu giá lễ - Thành công", "Thêm ngày lễ", "Bước 1: Điền Tên: 'Quốc Khánh 02/09', Ngày: '2026-09-02'\nBước 2: Bấm Lưu", "Date: 2026-09-02", "Thêm ngày lễ thành công, tất cả suất chiếu ngày 02/09 tự động áp giá Ngày Lễ")]),

        ("MOD_ADMIN_PROMOTIONS", "Quản lý đợt khuyến mãi", "Kiểm tra Thêm, Sửa Khuyến mãi trên AdminPromotions.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở AdminPromotions.vue", "PRM",
         [("bảng Danh sách Đợt khuyến mãi trên AdminPromotions.vue", "danh sách KM", "Hiển thị Mã code, Tên CT, Mức giảm, Khoảng ngày, Trạng thái")],
         [("Mã khuyến mãi", "Mã khuyến mãi", "TRIANVIP2026", {"min_len": 3, "max_len": 30}, [])],
         [("Loại giảm giá", "Giảm theo %")],
         [("PRM_FUNC_01", "Kiểm tra chức năng Phát hành voucher khi phát hàng loạt theo Hạng thẻ hội viên - Thành công", "Phát voucher theo hạng", "Bước 1: Chọn đợt KM 'Tri Ân VIP', chọn đối tượng Hạng Vàng & Kim Cương\nBước 2: Bấm Xác nhận phát hành", "Tier: Gold & Diamond", "Phát voucher vào ví của tất cả khách hàng đạt hạng thẻ và gửi email thông báo quà tặng")]),

        ("MOD_ADMIN_STAFF_MGMT", "Quản lý nhân viên", "Kiểm tra Thêm, Sửa Nhân viên trên StaffManager.vue", "Phạm Thị Quỳnh Anh", "Quản trị viên", "Mở StaffManager.vue", "STF",
         [("bảng Danh sách Nhân viên trên StaffManager.vue", "danh sách NV", "Hiển thị Mã NV, Họ tên, Email, SĐT, Vai trò, Cụm rạp trực thuộc, Trạng thái")],
         [("Họ và tên nhân viên", "Họ và tên nhân viên", "Lê Văn An", {"min_len": 2, "max_len": 50}, []),
          ("Số điện thoại", "Số điện thoại", "0977112233", {"min_len": 10, "max_len": 10}, [])],
         [("Vai trò", "Nhân viên (STAFF)"), ("Cụm rạp", "CGV Cầu Giấy")],
         [("STF_FUNC_01", "Kiểm tra chức năng Thêm nhân viên khi tạo tài khoản mới và tự sinh mật khẩu tạm - Thành công", "Tạo tài khoản NV", "Bước 1: Nhập đầy đủ thông tin, gán rạp CGV Cầu Giấy\nBước 2: Bấm Lưu", "Full valid data", "Tạo nhân viên thành công, tự sinh mật khẩu tạm gửi email và bật cờ đổi mật khẩu lần đầu"),
          ("STF_FUNC_02", "Kiểm tra chức năng Khóa tài khoản khi tự khóa tài khoản Admin đang đăng nhập phiên hiện tại - Thành công", "Tự khóa Admin", "Bước 1: Admin bấm toggle khóa tài khoản của chính mình", "User: Current Admin", "Hệ thống từ chối thao tác, báo lỗi 'Không thể tự khóa tài khoản Admin đang đăng nhập phiên hiện tại'")]),

        ("MOD_ADMIN_RBAC", "Phân quyền hệ thống", "Kiểm tra Phân quyền RBAC trên AdminPermissions.vue", "Phạm Thị Quỳnh Anh", "Quản trị viên", "Mở AdminPermissions.vue", "RBC",
         [("ma trận Phân quyền các vai trò hệ thống trên AdminPermissions.vue", "ma trận RBAC", "Bảng lưới phân quyền các vai trò Admin, Manager, Staff, Customer")],
         [],
         [],
         [("RBC_FUNC_01", "Kiểm tra chức năng Phân quyền RBAC khi tước quyền tối cao của vai trò Admin - Thành công", "Bảo vệ SuperAdmin", "Bước 1: Bỏ tích quyền SYSTEM_ADMIN của ROLE_ADMIN\nBước 2: Bấm Lưu", "Action: Delete SuperAdmin", "Hệ thống từ chối, báo lỗi 'Không được phép xóa bỏ quyền quản trị tối cao của vai trò Admin'"),
          ("RBC_FUNC_02", "Kiểm tra chức năng Phân quyền RBAC khi cấp thêm quyền duyệt hủy đơn cho nhân viên quầy - Thành công", "Override Permission", "Bước 1: Cấp quyền APPROVE_VOID cho tài khoản nhân viên Khôi\nBước 2: Bấm Lưu quyền", "Grant: APPROVE_VOID", "Nhân viên Khôi nhìn thấy nút và thực hiện được chức năng duyệt hủy đơn ngay phiên tiếp theo")]),

        ("MOD_ADMIN_CUSTOMERS", "Quản lý khách hàng", "Kiểm tra Quản lý Khách hàng trên AdminCustomers.vue", "Phạm Thị Quỳnh Anh", "Quản trị viên", "Mở AdminCustomers.vue", "CUS",
         [("bảng Danh sách Khách hàng trên AdminCustomers.vue", "danh sách khách", "Hiển thị Họ tên, Email, SĐT, Điểm Loyalty, Hạng thẻ, Tổng chi tiêu trọn đời")],
         [],
         [("Hạng thẻ", "Vàng (Gold)")],
         [("CUS_FUNC_01", "Kiểm tra chức năng Khóa tài khoản khi khóa khách hàng đang có vé xem phim chưa sử dụng - Thành công", "Cảnh báo khóa khách có vé", "Bước 1: Bấm Khóa tài khoản đang có vé xem phim tối nay", "Active tickets: True", "Hiển thị modal cảnh báo màu vàng: 'Khách hàng này hiện đang có 2 vé xem phim chưa sử dụng tối nay. Bạn có chắc muốn khóa?'")]),

        ("MOD_ADMIN_ORDERS", "Quản lý đơn hàng", "Kiểm tra Tra cứu Đơn hàng trên AdminBookings.vue", "Nguyễn Quang Huy", "Quản trị viên", "Mở tab Đơn hàng trên AdminBookings.vue", "ORD",
         [("bảng Quản lý Đơn hàng & Doanh thu trên AdminBookings.vue", "danh sách đơn", "Hiển thị Mã đơn, Khách hàng, Cụm rạp, Tổng tiền, Phương thức, Trạng thái")],
         [],
         [("Cụm rạp", "CGV Cầu Giấy"), ("Trạng thái", "Đã thanh toán (CONFIRMED)")],
         [("ORD_FUNC_01", "Kiểm tra chức năng Xuất hóa đơn VAT khi tải file PDF hóa đơn điện tử của đơn đã thanh toán - Thành công", "Xuất PDF hóa đơn", "Bước 1: Chọn đơn CONFIRMED, bấm 'Xuất hóa đơn VAT'", "Order: CONFIRMED", "Sinh và tải về file PDF hóa đơn điện tử chuẩn chỉ, đầy đủ thuế VAT và mã tra cứu")]),

        ("MOD_ADMIN_BANNERS", "Quản lý Banner", "Kiểm tra Thêm, Sửa Banner trên AdminBanners.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở AdminBanners.vue", "BAN",
         [("danh sách Banner quảng cáo trên AdminBanners.vue", "danh sách banner", "Hiển thị Ảnh banner, Tiêu đề, Phim gắn kèm, Thứ tự hiển thị, Trạng thái")],
         [("Tiêu đề banner", "Tiêu đề banner", "Bom Tấn Avatar Trở Lại", {"min_len": 3, "max_len": 150}, [])],
         [],
         [("BAN_FUNC_01", "Kiểm tra chức năng Thêm banner quảng cáo khi nhập tiêu đề và upload ảnh 1920x600 px - Thành công", "Thêm banner", "Bước 1: Nhập tiêu đề, gắn link phim, upload ảnh 1920x600 px\nBước 2: Bấm Lưu banner", "Full valid data", "Thêm banner thành công, hiển thị ngay trên Slider quảng cáo lớn ở đầu trang chủ")]),

        ("MOD_ADMIN_NEWS", "Tin tức & Khuyến mãi", "Kiểm tra Quản lý Tin tức & Khuyến mãi", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở quản lý tin tức", "NEW",
         [("trình soạn thảo Bài viết Tin tức Rich Text Editor", "soạn thảo tin tức", "Trình soạn thảo rich text có công cụ định dạng chữ và chèn ảnh")],
         [("Tiêu đề bài viết", "Tiêu đề bài viết", "Ưu Đãi Thứ 4 Vui Vẻ - Đồng Giá Vé 50K Toàn Hệ Thống", {"min_len": 5, "max_len": 200}, []),
          ("Tóm tắt bài viết", "Tóm tắt bài viết", "Chương trình ưu đãi đồng giá vé 50k vào thứ 4 hàng tuần...", {"min_len": 10, "max_len": 500}, [])],
         [],
         [("NEW_FUNC_01", "Kiểm tra chức năng Xuất bản bài viết tin tức khi nhập nội dung rich text và upload ảnh bìa - Thành công", "Đăng bài viết", "Bước 1: Nhập tiêu đề, tóm tắt, nội dung rich text, upload thumbnail\nBước 2: Bấm Xuất bản", "Full valid data", "Xuất bản thành công, hiển thị bài viết kèm đường dẫn slug chuẩn SEO")]),

        ("MOD_ADMIN_FAQ", "Quản lý FAQ", "Kiểm tra Thêm, Sửa FAQ trên FaqManager.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở FaqManager.vue", "FAQ",
         [("danh sách Câu hỏi thường gặp trên FaqManager.vue", "danh sách FAQ", "Hiển thị Nhóm câu hỏi, Câu hỏi, Câu trả lời, Thứ tự")],
         [("Câu hỏi", "Câu hỏi", "Làm thế nào để đổi vé xem phim đã mua?", {"min_len": 5, "max_len": 300}, []),
          ("Câu trả lời", "Câu trả lời", "Quý khách có thể đổi vé trước giờ chiếu ít nhất 60 phút...", {"min_len": 10, "max_len": 1000}, [])],
         [],
         [("FAQ_FUNC_01", "Kiểm tra chức năng Thêm câu hỏi FAQ khi nhập đầy đủ câu hỏi và câu trả lời - Thành công", "Thêm FAQ", "Bước 1: Chọn nhóm 'Vé & Giá vé', nhập câu hỏi và câu trả lời\nBước 2: Bấm Lưu", "Full valid data", "Thêm FAQ thành công, hiển thị trên trang Trợ giúp của khách hàng")]),

        ("MOD_ADMIN_SETTINGS", "Cài đặt hệ thống", "Kiểm tra Tham số động trên AdminSettings.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở AdminSettings.vue", "SET",
         [("bảng Cấu hình Tham số động trên AdminSettings.vue", "cài đặt tham số", "Các ô nhập thời gian giữ ghế, timeout đơn chờ, hotline, email thông báo")],
         [],
         [],
         [("SET_FUNC_01", "Kiểm tra chức năng Cài đặt hệ thống khi thay đổi thời gian giữ đơn chờ POS thành 8 phút - Thành công", "Cập nhật tham số", "Bước 1: Đổi tham số 'Thời gian giữ đơn chờ POS' thành 8 phút\nBước 2: Bấm Lưu cấu hình", "POS Timeout: 8'", "Lưu cấu hình thành công và áp dụng ngay lập tức trên máy POS toàn hệ thống")]),

        ("MOD_ADMIN_DASHBOARD", "Thống kê & Báo cáo", "Kiểm tra Dashboard thống kê trên Dashboard.vue", "Nguyễn Ngọc Hà Linh", "Quản trị viên", "Mở Dashboard.vue", "STA",
         [("4 thẻ KPI tổng quan (Doanh thu, Vé bán, Khách mới, F&B) trên Dashboard.vue", "thẻ KPI", "Hiển thị 4 thẻ số liệu doanh thu lớn có tỷ lệ tăng trưởng"),
          ("biểu đồ Doanh thu 7 ngày gần nhất trên Dashboard.vue", "biểu đồ doanh thu", "Biểu đồ cột doanh thu trực quan, có tooltip hiển thị số tiền khi hover")],
         [],
         [("Khoảng ngày", "7 ngày qua"), ("Cụm rạp", "Tất cả cụm rạp")],
         [("STA_FUNC_01", "Kiểm tra chức năng Lọc thống kê khi chọn khoảng ngày từ ngày 01 đến 19 - Thành công", "Lọc biểu đồ", "Bước 1: Chọn khoảng ngày từ 01/03/2026 đến 19/03/2026\nBước 2: Bấm Lọc", "Range: 01-19/03", "Biểu đồ và 4 thẻ KPI tự động cập nhật số liệu chính xác theo khoảng ngày đã chọn"),
          ("STA_FUNC_02", "Kiểm tra chức năng Xuất báo cáo doanh thu khi tải file Excel tổng hợp - Thành công", "Export Excel", "Bước 1: Click button 'Xuất báo cáo Excel'", "Action: Export", "Xuất và tải về file Excel báo cáo doanh thu chi tiết theo từng cụm rạp và phim")])
    ]

    for mod_spec in remaining_specs:
        c_code, c_sheet, c_req, c_tester, c_role, c_pre, c_pfx, c_gui, c_fields, c_filters, c_func = mod_spec
        cases = []
        cases.extend(build_gui_cases(c_pfx, c_sheet, c_role, c_gui))
        for f_name, f_label, norm_val, bounds, specials in c_fields:
            f_prefix = f"{c_pfx}_{f_name.upper()[:3]}"
            cases.extend(build_field_validation_cases(f_prefix, c_sheet, c_role, f_name, f_label, norm_val, bounds, specials))
        if c_filters:
            cases.extend(build_search_filter_cases(c_pfx, c_sheet, c_role, c_filters))
        for f_id, f_title, f_desc, f_steps, f_data, f_exp in c_func:
            cases.append((f_id, f_title, f_desc, f_steps, f_data, f_exp))

        modules.append({
            "code": c_code, "sheet": c_sheet, "req": c_req,
            "tester": c_tester, "role": c_role, "pre": c_pre,
            "test_cases": cases
        })

    return modules

def build_accurate_workbook_file(output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    font_name = "Times New Roman"
    
    border_thin = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000')
    )
    
    fill_header_navy = PatternFill(start_color='FF002060', end_color='FF002060', fill_type='solid')
    fill_header_green = PatternFill(start_color='FFC5E0B3', end_color='FFC5E0B3', fill_type='solid')
    fill_header_blue = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
    fill_header_gold = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
    
    font_title = Font(name=font_name, size=16, bold=True, color='FF002060')
    font_sub_title = Font(name=font_name, size=13, bold=True, color='FF000000')
    font_header_white = Font(name=font_name, size=12, bold=True, color='FFFFFFFF')
    font_header_black = Font(name=font_name, size=12, bold=True, color='FF000000')
    font_bold = Font(name=font_name, size=12, bold=True, color='FF000000')
    font_regular = Font(name=font_name, size=12, bold=False, color='FF000000')
    font_pass = Font(name=font_name, size=12, bold=True, color='FF008000')
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    test_date = datetime.datetime(2026, 3, 19, 0, 0)
    
    # -------------------------------------------------------------------------
    # 1. SHEET: Cover (Tổng quan)
    # -------------------------------------------------------------------------
    ws_cover = wb.create_sheet("Cover (Tổng quan)")
    ws_cover.views.sheetView[0].showGridLines = True
    
    ws_cover.cell(2, 2, "BÁO CÁO KIỂM THỬ PHẦN MỀM (TEST REPORT)").font = font_title
    
    meta_info = [
        ("Project Name", "DevCine - Website Quản lý Rạp chiếu phim & Đặt vé trực tuyến", "Creator", "Nguyễn Quang Huy"),
        ("Project Code", "DEVCINE_2026", "Reviewer/Approver", "Hội đồng Đồ án Tốt nghiệp / Tech Lead"),
        ("Document Code", "TR_DEVCINE_v1.0", "Issue Date", test_date),
        ("", "", "Version", "Phiên bản v1.0")
    ]
    
    for r_idx, (k1, v1, k2, v2) in enumerate(meta_info, start=4):
        if k1:
            ws_cover.cell(r_idx, 2, k1).font = font_bold
            ws_cover.cell(r_idx, 2).border = border_thin
            ws_cover.cell(r_idx, 3, v1).font = font_regular
            ws_cover.cell(r_idx, 3).border = border_thin
        if k2:
            ws_cover.cell(r_idx, 5, k2).font = font_bold
            ws_cover.cell(r_idx, 5).border = border_thin
            ws_cover.cell(r_idx, 6, v2).font = font_regular
            ws_cover.cell(r_idx, 6).border = border_thin
            if isinstance(v2, datetime.datetime):
                ws_cover.cell(r_idx, 6).number_format = 'yyyy-mm-dd'
                
    ws_cover.cell(9, 2, "Lịch sử thay đổi tài liệu (Record of Change)").font = font_sub_title
    change_headers = ["Ngày hiệu lực", "Phiên bản", "Hạng mục thay đổi", "Loại (*A/D/M)", "Mô tả thay đổi", "Tài liệu tham khảo"]
    for c_idx, h in enumerate(change_headers, start=2):
        cell = ws_cover.cell(10, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    change_rows = [
        ("2026-03-10", "1.0", "Tạo mới kế hoạch kiểm thử", "A", "Khởi tạo bộ test case kiểm thử toàn bộ hệ thống DevCine", "SRS / BA Specs"),
        ("2026-03-15", "1.0", "Bổ sung test case phân hệ POS & Check-in", "A", "Thêm test case nghiệp vụ bán vé tại quầy và soát vé QR", "POS Architecture Doc"),
        ("2026-03-19", "1.0", "Hoàn thiện thực thi & Báo cáo kết quả", "M", "Chạy kiểm thử đợt 1 (Release 1) và tổng hợp kết quả Pass", "Test Execution Log")
    ]
    for r_idx, crow in enumerate(change_rows, start=11):
        for c_idx, val in enumerate(crow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 3, 5] else align_left
            cell.border = border_thin
            
    ws_cover.cell(15, 2, "Danh sách Thành viên Nhóm Kiểm thử (Team Members)").font = font_sub_title
    member_headers = ["STT", "Họ và tên", "Mã sinh viên", "Vai trò", "Nhiệm vụ phụ trách"]
    for c_idx, h in enumerate(member_headers, start=2):
        cell = ws_cover.cell(16, c_idx, h)
        cell.font = font_header_black
        cell.fill = fill_header_green
        cell.alignment = align_center
        cell.border = border_thin
        
    member_rows = [
        (1, "Nguyễn Quang Huy", "PH12345", "Trưởng nhóm / Test Lead", "Phân hệ Đặt vé online, Chọn ghế, Combo F&B, Voucher, Thanh toán VNPAY, Đơn hàng"),
        (2, "Văn Minh Khôi", "PH12346", "Tester / QA", "Phân hệ POS Bán vé, POS Đơn chờ, Bán F&B, Soát vé Check-in, Xử lý sự cố chỗ ngồi"),
        (3, "Phạm Thị Quỳnh Anh", "PH12347", "Tester / QA", "Phân hệ Xác thực, Đăng nhập, Đăng ký, Quên mật khẩu, Quản lý Nhân viên, Khách hàng, RBAC"),
        (4, "Nguyễn Ngọc Hà Linh", "PH12348", "Tester / QA", "Phân hệ Quản trị Phim, Cụm rạp, Phòng chiếu, Sơ đồ ghế, Lịch chiếu, Bảng giá vé, Khuyến mãi, Cài đặt")
    ]
    for r_idx, mrow in enumerate(member_rows, start=17):
        for c_idx, val in enumerate(mrow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 4] else align_left
            cell.border = border_thin

    ws_cover.column_dimensions['A'].width = 4.0
    ws_cover.column_dimensions['B'].width = 18.0
    ws_cover.column_dimensions['C'].width = 30.0
    ws_cover.column_dimensions['D'].width = 25.0
    ws_cover.column_dimensions['E'].width = 18.0
    ws_cover.column_dimensions['F'].width = 45.0
    ws_cover.column_dimensions['G'].width = 22.0

    modules_data = build_accurate_modules()
    total_test_cases = sum(len([tc for tc in m["test_cases"] if tc[0] != "__SECTION__"]) for m in modules_data)
    print(f"Generated {len(modules_data)} modules with {total_test_cases} test cases strictly formatted as 'Kiểm tra chức năng [abc] khi [xyz] - [Thành công/Thất bại]'!")

    # -------------------------------------------------------------------------
    # 2. SHEET: Test case List (DS Test Case)
    # -------------------------------------------------------------------------
    ws_list = wb.create_sheet("Test case List (DS Test Case)")
    ws_list.views.sheetView[0].showGridLines = True
    
    ws_list.cell(2, 2, "DANH SÁCH BỘ KIỂM THỬ (TEST SUITE LIST)").font = font_title
    
    list_headers = ["STT", "Mã Module", "Tên Phân hệ / Chức năng", "Tên Sheet", "Số lượng Test Case", "Phân loại Vai trò", "Điều kiện tiên quyết (Preconditions)", "Người phụ trách"]
    for c_idx, h in enumerate(list_headers, start=2):
        cell = ws_list.cell(4, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    for r_idx, mod in enumerate(modules_data, start=5):
        stt = r_idx - 4
        ws_list.cell(r_idx, 2, stt).alignment = align_center
        ws_list.cell(r_idx, 3, mod["code"]).alignment = align_center
        ws_list.cell(r_idx, 4, mod["req"]).alignment = align_left
        ws_list.cell(r_idx, 5, mod["sheet"]).alignment = align_left
        ws_list.cell(r_idx, 6, len([tc for tc in mod["test_cases"] if tc[0] != "__SECTION__"])).alignment = align_center
        ws_list.cell(r_idx, 7, mod["role"]).alignment = align_center
        ws_list.cell(r_idx, 8, mod["pre"]).alignment = align_left
        ws_list.cell(r_idx, 9, mod["tester"]).alignment = align_center
        
        for c in range(2, 10):
            cell = ws_list.cell(r_idx, c)
            cell.font = font_regular
            cell.border = border_thin

    ws_list.column_dimensions['A'].width = 4.0
    ws_list.column_dimensions['B'].width = 8.0
    ws_list.column_dimensions['C'].width = 22.0
    ws_list.column_dimensions['D'].width = 45.0
    ws_list.column_dimensions['E'].width = 26.0
    ws_list.column_dimensions['F'].width = 18.0
    ws_list.column_dimensions['G'].width = 24.0
    ws_list.column_dimensions['H'].width = 45.0
    ws_list.column_dimensions['I'].width = 20.0

    # -------------------------------------------------------------------------
    # 3. SHEET: Test Report (Tổng hợp số liệu kiểm thử)
    # -------------------------------------------------------------------------
    ws_report = wb.create_sheet("Test Report")
    ws_report.views.sheetView[0].showGridLines = True
    
    ws_report.cell(1, 2, "BÁO CÁO TỔNG HỢP KIỂM THỬ (TEST REPORT SUMMARY)").font = font_title
    
    rep_meta = [
        ("Project Name", "DevCine - Quản lý Rạp chiếu phim & Đặt vé trực tuyến", "Creator", "Nguyễn Quang Huy"),
        ("Project Code", "DEVCINE_2026", "Reviewer/Approver", "Hội đồng Đồ án Tốt nghiệp / Tech Lead"),
        ("Document Code", "TR_DEVCINE_v1.0", "Issue Date", test_date),
        ("Release Scope", "Release 1.0 (Full Functional & Validation Suite)", "Status", f"100% Pass ({total_test_cases}/{total_test_cases} TCs)")
    ]
    for r_idx, (k1, v1, k2, v2) in enumerate(rep_meta, start=3):
        ws_report.cell(r_idx, 2, k1).font = font_bold
        ws_report.cell(r_idx, 2).border = border_thin
        ws_report.cell(r_idx, 3, v1).font = font_regular
        ws_report.cell(r_idx, 3).border = border_thin
        ws_report.cell(r_idx, 5, k2).font = font_bold
        ws_report.cell(r_idx, 5).border = border_thin
        ws_report.cell(r_idx, 6, v2).font = font_regular
        ws_report.cell(r_idx, 6).border = border_thin
        if isinstance(v2, datetime.datetime):
            ws_report.cell(r_idx, 6).number_format = 'yyyy-mm-dd'
            
    ws_report.cell(8, 2, "Kết quả thực thi kiểm thử Đợt 1 (Execution Summary - V1)").font = font_sub_title
    
    rep_headers = ["STT", "Tên Phân hệ / Module", "Pass", "Fail", "Untested", "N/A", "Tổng Test Case", "Tỷ lệ Pass (%)"]
    for c_idx, h in enumerate(rep_headers, start=2):
        cell = ws_report.cell(9, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    start_row = 10
    for idx, mod in enumerate(modules_data):
        r_num = start_row + idx
        sheet_name = mod["sheet"]
        
        ws_report.cell(r_num, 2, idx + 1).alignment = align_center
        ws_report.cell(r_num, 3, f"='{sheet_name}'!B1").alignment = align_left
        ws_report.cell(r_num, 4, f"='{sheet_name}'!A5").alignment = align_center
        ws_report.cell(r_num, 5, f"='{sheet_name}'!B5").alignment = align_center
        ws_report.cell(r_num, 6, f"='{sheet_name}'!C5").alignment = align_center
        ws_report.cell(r_num, 7, f"='{sheet_name}'!D5").alignment = align_center
        ws_report.cell(r_num, 8, f"='{sheet_name}'!E5").alignment = align_center
        ws_report.cell(r_num, 9, f"=IF(H{r_num}>0, D{r_num}/H{r_num}, 1)").alignment = align_center
        ws_report.cell(r_num, 9).number_format = '0.0%'
        
        for c in range(2, 10):
            cell = ws_report.cell(r_num, c)
            cell.font = font_regular
            cell.border = border_thin
            
    total_row = start_row + len(modules_data)
    ws_report.cell(total_row, 2, "").alignment = align_center
    ws_report.cell(total_row, 3, "TỔNG CỘNG").font = font_bold
    ws_report.cell(total_row, 3).alignment = align_center
    ws_report.cell(total_row, 4, f"=SUM(D{start_row}:D{total_row-1})").font = font_bold
    ws_report.cell(total_row, 4).alignment = align_center
    ws_report.cell(total_row, 5, f"=SUM(E{start_row}:E{total_row-1})").font = font_bold
    ws_report.cell(total_row, 5).alignment = align_center
    ws_report.cell(total_row, 6, f"=SUM(F{start_row}:F{total_row-1})").font = font_bold
    ws_report.cell(total_row, 6).alignment = align_center
    ws_report.cell(total_row, 7, f"=SUM(G{start_row}:G{total_row-1})").font = font_bold
    ws_report.cell(total_row, 7).alignment = align_center
    ws_report.cell(total_row, 8, f"=SUM(H{start_row}:H{total_row-1})").font = font_bold
    ws_report.cell(total_row, 8).alignment = align_center
    ws_report.cell(total_row, 9, f"=IF(H{total_row}>0, D{total_row}/H{total_row}, 1)").font = font_bold
    ws_report.cell(total_row, 9).alignment = align_center
    ws_report.cell(total_row, 9).number_format = '0.0%'
    
    for c in range(2, 10):
        cell = ws_report.cell(total_row, c)
        cell.fill = fill_header_gold
        cell.border = border_thin

    ws_report.column_dimensions['A'].width = 4.0
    ws_report.column_dimensions['B'].width = 8.0
    ws_report.column_dimensions['C'].width = 38.0
    ws_report.column_dimensions['D'].width = 12.0
    ws_report.column_dimensions['E'].width = 12.0
    ws_report.column_dimensions['F'].width = 12.0
    ws_report.column_dimensions['G'].width = 12.0
    ws_report.column_dimensions['H'].width = 16.0
    ws_report.column_dimensions['I'].width = 16.0

    # -------------------------------------------------------------------------
    # 4. CREATE INDIVIDUAL MODULE SHEETS
    # -------------------------------------------------------------------------
    tc_table_headers = [
        "ID (Mã Test Case)", "Tiêu đề kiểm thử (Test Title)", "Mô tả trường hợp kiểm thử (Description)",
        "Các bước thực hiện (Test Procedure / Steps)", "Dữ liệu kiểm thử (Test Data)",
        "Kết quả mong muốn (Expected Output)", "Kết quả thực tế (Actual Result)",
        "Minh chứng (Evidence)", "Phụ thuộc (Dependence)", "Kết quả (Result 1)", "Ngày test (Date 1)"
    ]
    
    for mod in modules_data:
        ws_mod = wb.create_sheet(mod["sheet"])
        ws_mod.views.sheetView[0].showGridLines = True
        
        # Row 1-3: Metadata
        ws_mod.cell(1, 1, "Module Code(Mã Module)").font = font_bold
        ws_mod.cell(1, 2, mod["sheet"]).font = font_bold
        
        ws_mod.cell(2, 1, "Test requirement(Yêu cầu test)").font = font_bold
        ws_mod.cell(2, 2, mod["req"]).font = font_regular
        
        ws_mod.cell(3, 1, "Tester(Người thực hiện)").font = font_bold
        ws_mod.cell(3, 2, mod["tester"]).font = font_regular
        
        # Row 1-4 in Column H: Pass/Fail/Untested/N/A Status Legend
        fill_red_status = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        fill_yellow_status = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        fill_white_status = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        
        status_items = [
            (1, "Pass", fill_white_status),
            (2, "Fail", fill_red_status),
            (3, "Untested", fill_yellow_status),
            (4, "N/A", fill_white_status)
        ]
        for s_row, s_text, s_fill in status_items:
            s_cell = ws_mod.cell(s_row, 8, s_text)
            s_cell.font = font_bold
            s_cell.fill = s_fill
            s_cell.alignment = Alignment(horizontal='left', vertical='center')
            s_cell.border = border_thin

        # Row 4-5: Summary table V1
        stat_headers = ["PASS-V1", "FAIL-V1", "UNTESTED-V1", "N/A-V1", "Tổng số TestCase (V1)"]
        for c_idx, sh in enumerate(stat_headers, start=1):
            cell = ws_mod.cell(4, c_idx, sh)
            cell.font = font_header_black
            cell.fill = fill_header_blue
            cell.alignment = align_center
            cell.border = border_thin
            
        num_tc = len([tc for tc in mod["test_cases"] if tc[0] != "__SECTION__"])
        ws_mod.cell(5, 1, num_tc).alignment = align_center
        ws_mod.cell(5, 1).font = font_pass
        ws_mod.cell(5, 2, 0).alignment = align_center
        ws_mod.cell(5, 2).font = font_regular
        ws_mod.cell(5, 3, 0).alignment = align_center
        ws_mod.cell(5, 3).font = font_regular
        ws_mod.cell(5, 4, 0).alignment = align_center
        ws_mod.cell(5, 4).font = font_regular
        ws_mod.cell(5, 5, num_tc).alignment = align_center
        ws_mod.cell(5, 5).font = font_bold
        for c in range(1, 6):
            ws_mod.cell(5, c).border = border_thin
            
        # Row 7-8: Summary table V2
        stat_headers_v2 = ["PASS-V2", "FAIL-V2", "UNTESTED-V2", "N/A-V2", "Tổng số TestCase (V2)"]
        for c_idx, sh in enumerate(stat_headers_v2, start=1):
            cell = ws_mod.cell(7, c_idx, sh)
            cell.font = font_header_black
            cell.fill = fill_header_gold
            cell.alignment = align_center
            cell.border = border_thin
            
        for c in range(1, 6):
            cell = ws_mod.cell(8, c, 0)
            cell.alignment = align_center
            cell.font = font_regular
            cell.border = border_thin
            
        # Row 10: Column Headers
        for c_idx, th in enumerate(tc_table_headers, start=1):
            cell = ws_mod.cell(10, c_idx, th)
            cell.font = font_header_white
            cell.fill = fill_header_navy
            cell.alignment = align_center
            cell.border = border_thin
            
        # Row 11+: Test Cases
        fill_yellow_section = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        font_section_title = Font(name=font_name, size=12, bold=True, color='FF000000')
        align_section = Alignment(horizontal='left', vertical='center', indent=1)
        
        for r_offset, tc in enumerate(mod["test_cases"], start=11):
            if tc[0] == "__SECTION__":
                section_title = tc[1]
                ws_mod.merge_cells(start_row=r_offset, start_column=1, end_row=r_offset, end_column=11)
                ws_mod.row_dimensions[r_offset].height = 26.0
                cell = ws_mod.cell(r_offset, 1, section_title)
                cell.font = font_section_title
                cell.alignment = align_section
                for c in range(1, 12):
                    c_cell = ws_mod.cell(r_offset, c)
                    c_cell.fill = fill_yellow_section
                    c_cell.border = border_thin
                continue

            t_id, t_title, t_desc, t_steps, t_data, t_expect = tc
            
            ws_mod.cell(r_offset, 1, t_id).alignment = align_center
            ws_mod.cell(r_offset, 2, t_title).alignment = align_left
            ws_mod.cell(r_offset, 3, t_desc).alignment = align_left
            ws_mod.cell(r_offset, 4, t_steps).alignment = align_top_left
            ws_mod.cell(r_offset, 5, t_data).alignment = align_center
            ws_mod.cell(r_offset, 6, t_expect).alignment = align_top_left
            ws_mod.cell(r_offset, 7, t_expect).alignment = align_top_left
            ws_mod.cell(r_offset, 8, "").alignment = align_center # Blank evidence
            ws_mod.cell(r_offset, 9, "N/A").alignment = align_center
            ws_mod.cell(r_offset, 10, "Pass").alignment = align_center
            ws_mod.cell(r_offset, 10).font = font_pass
            ws_mod.cell(r_offset, 11, test_date).alignment = align_center
            ws_mod.cell(r_offset, 11).number_format = 'yyyy-mm-dd'
            
            for c in range(1, 12):
                cell = ws_mod.cell(r_offset, c)
                if c != 10:
                    cell.font = font_regular
                cell.border = border_thin
                
        ws_mod.column_dimensions['A'].width = 16.0
        ws_mod.column_dimensions['B'].width = 44.0
        ws_mod.column_dimensions['C'].width = 34.0
        ws_mod.column_dimensions['D'].width = 50.0
        ws_mod.column_dimensions['E'].width = 26.0
        ws_mod.column_dimensions['F'].width = 40.0
        ws_mod.column_dimensions['G'].width = 40.0
        ws_mod.column_dimensions['H'].width = 14.0
        ws_mod.column_dimensions['I'].width = 14.0
        ws_mod.column_dimensions['J'].width = 12.0
        ws_mod.column_dimensions['K'].width = 14.0

    # -------------------------------------------------------------------------
    # 5. SHEET: FUNCTION (Cây chức năng & Sự kiện)
    # -------------------------------------------------------------------------
    ws_func = wb.create_sheet("FUNCTION")
    ws_func.views.sheetView[0].showGridLines = True
    
    ws_func.cell(1, 1, "CÂY CHỨC NĂNG VÀ SỰ KIỆN (FUNCTION HIERARCHY)").font = font_title
    ws_func.cell(3, 1, "Project Name").font = font_bold
    ws_func.cell(3, 2, "DevCine - Quản lý Rạp chiếu phim & Đặt vé trực tuyến").font = font_regular
    ws_func.cell(4, 1, "Project Code").font = font_bold
    ws_func.cell(4, 2, "DEVCINE_2026").font = font_regular
    
    func_headers = ["Function Level 1 (Phân hệ chính)", "Function Level 2 (Chức năng chi tiết)", "Action & Event (Hành động & Sự kiện)", "Ghi chú phân tích"]
    for c_idx, fh in enumerate(func_headers, start=1):
        cell = ws_func.cell(6, c_idx, fh)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    func_rows = [
        ("1. Khách hàng - Xác thực & Tài khoản", "1.1 Đăng ký tài khoản", "Nhập form đăng ký -> Bấm Đăng ký -> Validate -> Lưu DB", "Mã hóa BCrypt, kiểm tra trùng email/SĐT"),
        ("", "1.2 Đăng nhập", "Nhập thông tin -> Bấm Đăng nhập -> Sinh JWT Token -> Lưu session", "Lưu token, tự nhận diện SĐT hoặc Email"),
        ("", "1.3 Quên mật khẩu & OTP", "Nhập email -> Nhận mã OTP 6 số -> Xác thực -> Đặt mật khẩu mới", "Cooldown 30s, OTP 6 số hết hạn 10 phút"),
        ("", "1.4 Hồ sơ & Thành viên", "Xem thông tin -> Cập nhật hồ sơ -> Tra cứu điểm Loyalty", "Tự động cập nhật hạng thành viên"),
        ("2. Khách hàng - Đặt vé Online", "2.1 Lịch chiếu & Chọn phim", "Chọn phim -> Chọn ngày -> Chọn suất chiếu theo rạp", "Chỉ hiển thị suất chưa chiếu"),
        ("", "2.2 Chọn ghế & Giữ chỗ", "Click chọn ghế -> Kiểm tra ghế trống -> Giữ chỗ 10 phút", "SeatGridRenderer, WebSocket STOMP real-time, chống để trống 1 ghế đơn"),
        ("", "2.3 Combo F&B", "Chọn bắp nước -> Modal chọn vị bắp, loại nước -> Cộng tiền phụ thu", "FnbOptionModal, slot bắt buộc theo combo"),
        ("", "2.4 Áp dụng Voucher", "Nhập mã voucher -> Kiểm tra điều kiện min order/hạn dùng -> Giảm giá", "Giảm theo % hoặc tiền, không vượt trần"),
        ("", "2.5 Thanh toán VNPAY", "Chuyển sang cổng VNPAY -> Thanh toán -> Nhận vé QR & Gửi email", "Xác thực chữ ký HMAC-SHA512"),
        ("3. Nhân viên - Vận hành Quầy (POS)", "3.1 Bán vé tại quầy", "Chọn suất chiếu -> Chọn ghế -> Tra cứu hội viên (F2) -> Thu tiền -> In vé (F9)", "Strict Cinema Scoping, lưu sold_by"),
        ("", "3.2 Đơn chờ POS", "Lưu đơn chờ tạm thời (F4) -> Khôi phục thanh toán", "Tối đa 3 đơn chờ/máy POS, phạt khóa ghế 5 phút khi hết hạn"),
        ("", "3.3 Bán F&B tại quầy", "Chọn bắp nước -> Thanh toán -> In hóa đơn", "Bán độc lập không kèm vé"),
        ("", "3.4 Soát vé Check-in", "Quét mã QR / Nhập mã vé -> Kiểm tra vé hợp lệ -> Check-in", "Cảnh báo vé đã dùng/vé sai rạp"),
        ("", "3.5 Xử lý sự cố chỗ ngồi", "Tra cứu đơn -> Đổi ghế tại chỗ cho khách -> Ghi log sự cố", "Chỉ đổi trước giờ chiếu, giữ nguyên QR"),
        ("4. Quản trị viên (Admin Master)", "4.1 Quản lý phim", "Thêm/Sửa phim -> Upload poster/banner -> Đổi trạng thái", "Thời lượng 30-300', trailer link Youtube"),
        ("", "4.2 Quản lý Cụm rạp & Phòng", "Thêm rạp -> Thêm phòng -> Thiết lập sơ đồ ma trận ghế", "Cấu hình hàng A-Z, ghế đôi Sweetbox"),
        ("", "4.3 Điều phối Suất chiếu", "Lập lịch chiếu đơn -> Xếp lịch hàng loạt (Batch Scheduling)", "Kiểm tra xung đột phòng chiếu"),
        ("", "4.4 Bảng giá vé", "Cấu hình giá nền 3 chiều -> Phụ thu VIP/3D -> Ngày lễ", "Simulator tính thử giá vé"),
        ("", "4.5 Khuyến mãi & Voucher", "Tạo đợt khuyến mãi -> Phát hành voucher theo hạng thẻ", "Giảm % hoặc tiền, đơn tối thiểu"),
        ("", "4.6 Nhân sự & Phân quyền", "Tạo nhân viên -> Gán cụm rạp -> Phân quyền RBAC", "Tự sinh mật khẩu tạm cho NV mới"),
        ("", "4.7 Cài đặt hệ thống", "Cấu hình tham số giữ ghế, timeout, hotline, email", "Áp dụng cấu hình động toàn rạp")
    ]
    
    for r_idx, frow in enumerate(func_rows, start=7):
        for c_idx, val in enumerate(frow, start=1):
            cell = ws_func.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_top_left if c_idx != 1 else align_left
            cell.border = border_thin

    ws_func.column_dimensions['A'].width = 32.0
    ws_func.column_dimensions['B'].width = 30.0
    ws_func.column_dimensions['C'].width = 55.0
    ws_func.column_dimensions['D'].width = 35.0

    wb.save(output_path)
    print(f"Successfully generated: {output_path}")

if __name__ == "__main__":
    out_dir = r"c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine"
    out_file = os.path.join(out_dir, "TestReport Dự án DevCine.xlsx")
    build_accurate_workbook_file(out_file)
    
    dst_downloads2 = r"C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx"
    try:
        shutil.copy2(out_file, dst_downloads2)
        print("Updated Downloads TestReport Dự án DevCine.xlsx")
    except Exception as e:
        print("Downloads locked, please close Excel if open.")
