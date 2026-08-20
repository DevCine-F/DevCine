# -*- coding: utf-8 -*-
"""
Audit all 43 sheets to verify test cases and ensure full compliance.
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r"c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine\TestReport Dự án DevCine.xlsx")

total_cases = 0
for name in wb.sheetnames:
    if name in ["Cover (Tổng quan)", "Test case List (DS Test Case)", "Test Report", "FUNCTION"]:
        continue
    ws = wb[name]
    case_ids = []
    for r in range(11, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and not str(v).startswith("KIỂM TRA") and not str(v).startswith("KỸ THUẬT"):
            case_ids.append(v)
    total_cases += len(case_ids)
    print(f"Sheet '{name}': {len(case_ids)} test cases (VD: {case_ids[0] if case_ids else 'N/A'})")

print(f"\n===> TOTAL FUNCTIONAL TEST CASES: {total_cases}")
