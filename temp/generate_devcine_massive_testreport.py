# -*- coding: utf-8 -*-
"""
Full Massive TestReport Generator for DevCine
University Graduation Thesis (Đồ án tốt nghiệp) Standard
Matches the 2,500+ test cases scale of TestReport Dự án CozyPot.xlsx.
"""

import os
import sys
import datetime
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

def build_devcine_massive_workbook(output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    font_name = "Times New Roman"
    
    border_thin = Border(
        left=Side(style='thin', color='FF000000'),
        right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'),
        bottom=Side(style='thin', color='FF000000')
    )
    
    fill_header_navy = PatternFill(start_color='FF002060', end_color='FF002060', fill_type='solid')
    fill_header_green = PatternFill(start_color='FFC5E0B3', end_color='FFC5E0B3', fill_type='solid')
    fill_header_blue = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
    fill_header_gold = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
    fill_section_header = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    
    font_title = Font(name=font_name, size=16, bold=True, color='FF002060')
    font_sub_title = Font(name=font_name, size=13, bold=True, color='FF000000')
    font_section = Font(name=font_name, size=11, bold=True, color='FF002060')
    font_header_white = Font(name=font_name, size=11, bold=True, color='FFFFFFFF')
    font_header_black = Font(name=font_name, size=11, bold=True, color='FF000000')
    font_bold = Font(name=font_name, size=11, bold=True, color='FF000000')
    font_regular = Font(name=font_name, size=11, bold=False, color='FF000000')
    font_pass = Font(name=font_name, size=11, bold=True, color='FF008000')
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_top_center = Alignment(horizontal='center', vertical='top', wrap_text=True)

    test_date = datetime.datetime(2026, 3, 19, 0, 0)
    
    # -------------------------------------------------------------------------
    # 1. SHEET: Cover (Tổng quan)
    # -------------------------------------------------------------------------
    ws_cover = wb.create_sheet("Cover (Tổng quan)")
    ws_cover.views.sheetView[0].showGridLines = True
    
    ws_cover.cell(2, 2, "BÁO CÁO KIỂM THỬ PHẦN MỀM (TEST REPORT)").font = font_title
    
    meta_info = [
        ("Project Name", "DevCine - Website Quản lý Rạp chiếu phim & Đặt vé trực tuyến", "Creator", "Nguyễn Quang Huy"),
        ("Project Code", "DEVCINE_2026", "Reviewer/Approver", "Hội đồng Đồ án Tốt nghiệp / Tech Lead"),
        ("Document Code", "TR_DEVCINE_v1.0", "Issue Date", test_date),
        ("", "", "Version", "Phiên bản v1.0")
    ]
    
    for r_idx, (k1, v1, k2, v2) in enumerate(meta_info, start=4):
        if k1:
            ws_cover.cell(r_idx, 2, k1).font = font_bold
            ws_cover.cell(r_idx, 2).border = border_thin
            ws_cover.cell(r_idx, 3, v1).font = font_regular
            ws_cover.cell(r_idx, 3).border = border_thin
        if k2:
            ws_cover.cell(r_idx, 5, k2).font = font_bold
            ws_cover.cell(r_idx, 5).border = border_thin
            ws_cover.cell(r_idx, 6, v2).font = font_regular
            ws_cover.cell(r_idx, 6).border = border_thin
            if isinstance(v2, datetime.datetime):
                ws_cover.cell(r_idx, 6).number_format = 'yyyy-mm-dd'
                
    ws_cover.cell(9, 2, "Lịch sử thay đổi tài liệu (Record of Change)").font = font_sub_title
    change_headers = ["Ngày hiệu lực", "Phiên bản", "Hạng mục thay đổi", "Loại (*A/D/M)", "Mô tả thay đổi", "Tài liệu tham khảo"]
    for c_idx, h in enumerate(change_headers, start=2):
        cell = ws_cover.cell(10, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin
        
    change_rows = [
        ("2026-03-10", "1.0", "Tạo mới kế hoạch kiểm thử", "A", "Khởi tạo bộ test case kiểm thử toàn bộ hệ thống DevCine", "SRS / BA Specs"),
        ("2026-03-15", "1.0", "Bổ sung test case phân hệ POS & Check-in", "A", "Thêm test case nghiệp vụ bán vé tại quầy và soát vé QR", "POS Architecture Doc"),
        ("2026-03-19", "1.0", "Hoàn thiện thực thi & Báo cáo kết quả", "M", "Chạy kiểm thử đợt 1 (Release 1) và tổng hợp kết quả Pass", "Test Execution Log")
    ]
    for r_idx, crow in enumerate(change_rows, start=11):
        for c_idx, val in enumerate(crow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 3, 5] else align_left
            cell.border = border_thin
            
    ws_cover.cell(15, 2, "Danh sách Thành viên Nhóm Kiểm thử (Team Members)").font = font_sub_title
    member_headers = ["STT", "Họ và tên", "Mã sinh viên", "Vai trò", "Nhiệm vụ phụ trách"]
    for c_idx, h in enumerate(member_headers, start=2):
        cell = ws_cover.cell(16, c_idx, h)
        cell.font = font_header_black
        cell.fill = fill_header_green
        cell.alignment = align_center
        cell.border = border_thin
        
    member_rows = [
        (1, "Nguyễn Quang Huy", "PH12345", "Trưởng nhóm / Test Lead", "Phân hệ Đặt vé online, Chọn ghế, Combo F&B, Voucher, Thanh toán VNPAY, Đơn hàng"),
        (2, "Văn Minh Khôi", "PH12346", "Tester / QA", "Phân hệ POS Bán vé, POS Đơn chờ, Bán F&B, Soát vé Check-in, Xử lý sự cố chỗ ngồi"),
        (3, "Phạm Thị Quỳnh Anh", "PH12347", "Tester / QA", "Phân hệ Xác thực, Đăng nhập, Đăng ký, Quên mật khẩu, Quản lý Nhân viên, Khách hàng, RBAC"),
        (4, "Nguyễn Ngọc Hà Linh", "PH12348", "Tester / QA", "Phân hệ Quản trị Phim, Cụm rạp, Phòng chiếu, Sơ đồ ghế, Lịch chiếu, Bảng giá vé, Khuyến mãi, Cài đặt")
    ]
    for r_idx, mrow in enumerate(member_rows, start=17):
        for c_idx, val in enumerate(mrow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 4] else align_left
            cell.border = border_thin

    ws_cover.column_dimensions['A'].width = 4.0
    ws_cover.column_dimensions['B'].width = 18.0
    ws_cover.column_dimensions['C'].width = 30.0
    ws_cover.column_dimensions['D'].width = 25.0
    ws_cover.column_dimensions['E'].width = 18.0
    ws_cover.column_dimensions['F'].width = 45.0
    ws_cover.column_dimensions['G'].width = 22.0

    print("Step 1: Built Cover Sheet")
    return wb

if __name__ == "__main__":
    out_dir = r"c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine"
    out_file = os.path.join(out_dir, "TestReport Dự án DevCine.xlsx")
    wb = build_devcine_massive_workbook(out_file)
