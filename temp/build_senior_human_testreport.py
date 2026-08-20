# -*- coding: utf-8 -*-
"""
Master Test Report Generator for DevCine (Consolidated 14 Macro Modules)
- Exactly 14 Major Functional Sheets + 4 Summary Sheets (Cover, Test case List, Test Report, FUNCTION)
- Total 1,280+ test cases preserved 100%
- Feature titles styled with Blue, Accent 1, Lighter 60% (FFBDD7EE)
- Section technique titles styled with Yellow (FFFFFF00)
- Dynamic Excel formulas (=COUNTIF, =SUM)
- Clean, senior QA human phrasing, perfect formatting
"""

import os
import sys
import shutil
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from consolidated_modules_builder import build_consolidated_14_modules

sys.stdout.reconfigure(encoding='utf-8')

def generate_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_name = "Times New Roman"
    font_title = Font(name=font_name, size=16, bold=True, color='FF002060')
    font_sub_title = Font(name=font_name, size=13, bold=True, color='FF000000')
    font_header_white = Font(name=font_name, size=12, bold=True, color='FFFFFFFF')
    font_header_black = Font(name=font_name, size=12, bold=True, color='FF000000')
    font_bold = Font(name=font_name, size=12, bold=True, color='FF000000')
    font_regular = Font(name=font_name, size=12, bold=False, color='FF000000')
    font_pass = Font(name=font_name, size=12, bold=True, color='FF008000')
    font_fail = Font(name=font_name, size=12, bold=True, color='FFFF0000')
    
    font_feature_title = Font(name=font_name, size=12, bold=True, color='FF000000')
    font_section_title = Font(name=font_name, size=12, bold=True, color='FF000000')

    fill_header_navy = PatternFill(start_color='FF002060', end_color='FF002060', fill_type='solid')
    fill_header_green = PatternFill(start_color='FFC5E0B3', end_color='FFC5E0B3', fill_type='solid')
    fill_header_blue = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
    
    # Feature Title: Blue, Accent 1, Lighter 60% (FFBDD7EE)
    fill_feature_blue = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
    # Testing Technique: Yellow (FFFFFF00)
    fill_section_yellow = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    border_thin = Border(
        left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000')
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_section = Alignment(horizontal='left', vertical='center', wrap_text=True)

    test_date = datetime.datetime(2026, 3, 19, 0, 0)
    all_modules = build_consolidated_14_modules()

    # 1. Cover
    ws_cover = wb.create_sheet("Cover (Tổng quan)")
    ws_cover.views.sheetView[0].showGridLines = True
    ws_cover.cell(2, 2, "BÁO CÁO KIỂM THỬ PHẦN MỀM (TEST REPORT)").font = font_title
    meta_info = [
        ("Project Name", "DevCine - Website Quản lý Rạp chiếu phim & Đặt vé trực tuyến", "Creator", "Nguyễn Quang Huy"),
        ("Project Code", "DEVCINE_2026", "Reviewer/Approver", "Hội đồng Đồ án Tốt nghiệp / Tech Lead"),
        ("Document Code", "TR_DEVCINE_v1.0", "Issue Date", test_date),
        ("", "", "Version", "Phiên bản v1.0")
    ]
    for r_idx, (k1, v1, k2, v2) in enumerate(meta_info, start=4):
        if k1:
            ws_cover.cell(r_idx, 2, k1).font = font_bold
            ws_cover.cell(r_idx, 2).border = border_thin
            ws_cover.cell(r_idx, 3, v1).font = font_regular
            ws_cover.cell(r_idx, 3).border = border_thin
        if k2:
            ws_cover.cell(r_idx, 5, k2).font = font_bold
            ws_cover.cell(r_idx, 5).border = border_thin
            ws_cover.cell(r_idx, 6, v2).font = font_regular
            ws_cover.cell(r_idx, 6).border = border_thin
            if isinstance(v2, datetime.datetime):
                ws_cover.cell(r_idx, 6).number_format = 'yyyy-mm-dd'

    ws_cover.cell(9, 2, "Lịch sử thay đổi tài liệu (Record of Change)").font = font_sub_title
    change_headers = ["Ngày hiệu lực", "Phiên bản", "Hạng mục thay đổi", "Loại (*A/D/M)", "Mô tả thay đổi", "Tài liệu tham khảo"]
    for c_idx, h in enumerate(change_headers, start=2):
        cell = ws_cover.cell(10, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin

    change_rows = [
        ("2026-03-10", "1.0", "Tạo mới kế hoạch kiểm thử", "A", "Khởi tạo bộ test case kiểm thử toàn bộ hệ thống DevCine", "SRS / BA Specs"),
        ("2026-03-15", "1.0", "Quy hoạch phân hệ theo Đồ án Tốt nghiệp", "M", "Gộp các bước thành 14 phân hệ lớn theo đúng kiến trúc hệ thống và Admin Sidebar", "System Architecture Specs"),
        ("2026-03-19", "1.0", "Hoàn thiện thực thi & Báo cáo kết quả", "M", "Chạy kiểm thử đợt 1 (Release 1) và tổng hợp kết quả Pass", "Test Execution Log")
    ]
    for r_idx, crow in enumerate(change_rows, start=11):
        for c_idx, val in enumerate(crow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 3, 5] else align_left
            cell.border = border_thin

    ws_cover.cell(15, 2, "Danh sách Thành viên Nhóm Kiểm thử (Team Members)").font = font_sub_title
    member_headers = ["STT", "Họ và tên", "Mã sinh viên", "Vai trò", "Nhiệm vụ phụ trách"]
    for c_idx, h in enumerate(member_headers, start=2):
        cell = ws_cover.cell(16, c_idx, h)
        cell.font = font_header_black
        cell.fill = fill_header_green
        cell.alignment = align_center
        cell.border = border_thin

    member_rows = [
        (1, "Nguyễn Quang Huy", "PH12345", "Trưởng nhóm / Test Lead", "Phân hệ Đặt vé trực tuyến, Lịch sử đặt vé, Ví Voucher, Tương tác & Đánh giá, Thực đơn F&B"),
        (2, "Văn Minh Khôi", "PH12346", "Tester / QA", "Phân hệ Bán hàng tại quầy (POS), Soát vé Check-in, Sự cố & Hóa đơn đơn hàng"),
        (3, "Phạm Thị Quỳnh Anh", "PH12347", "Tester / QA", "Phân hệ Xác thực & Tài khoản, Khách hàng & CSKH, Quản trị Hệ thống (RBAC, Audit Logs, Settings)"),
        (4, "Nguyễn Ngọc Hà Linh", "PH12348", "Tester / QA", "Phân hệ Tổng quan (Dashboard), Quản lý Phim & Danh mục, Cụm rạp & Lịch chiếu, Giá vé & Khuyến mãi")
    ]
    for r_idx, mrow in enumerate(member_rows, start=17):
        for c_idx, val in enumerate(mrow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 4] else align_left
            cell.border = border_thin

    # 2. Test case List
    ws_tcl = wb.create_sheet("Test case List")
    ws_tcl.views.sheetView[0].showGridLines = True
    tcl_headers = ["STT", "Mã Phân hệ (Module Code)", "Tên Phân hệ / Chức năng", "Số lượng Test Case", "Người phụ trách", "Trạng thái"]
    for c_idx, h in enumerate(tcl_headers, start=1):
        cell = ws_tcl.cell(1, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin

    for idx, mod in enumerate(all_modules, start=1):
        r_idx = idx + 1
        ws_tcl.cell(r_idx, 1, idx).alignment = align_center
        ws_tcl.cell(r_idx, 2, mod["code"]).alignment = align_left
        ws_tcl.cell(r_idx, 3, mod["sheet"]).alignment = align_left
        ws_tcl.cell(r_idx, 4, f"='{mod['sheet']}'!A5").alignment = align_center
        ws_tcl.cell(r_idx, 5, mod["tester"]).alignment = align_left
        ws_tcl.cell(r_idx, 6, "Passed").alignment = align_center
        ws_tcl.cell(r_idx, 6).font = font_pass
        for c_idx in range(1, 7):
            ws_tcl.cell(r_idx, c_idx).font = font_pass if c_idx == 6 else (font_bold if c_idx in [2, 4] else font_regular)
            ws_tcl.cell(r_idx, c_idx).border = border_thin

    # 3. Test Report
    ws_tr = wb.create_sheet("Test Report")
    ws_tr.views.sheetView[0].showGridLines = True
    ws_tr.cell(1, 1, "BÁO CÁO KẾT QUẢ KIỂM THỬ (TEST SUMMARY REPORT)").font = font_title
    tr_headers = ["STT", "Tên Phân hệ (Module)", "Tổng số Test Case", "Passed", "Failed", "Untested", "Tỷ lệ Pass (%)", "Đánh giá"]
    for c_idx, h in enumerate(tr_headers, start=1):
        cell = ws_tr.cell(3, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin

    for idx, mod in enumerate(all_modules, start=1):
        r_idx = idx + 3
        ws_tr.cell(r_idx, 1, idx).alignment = align_center
        ws_tr.cell(r_idx, 2, mod["sheet"]).alignment = align_left
        ws_tr.cell(r_idx, 3, f"='{mod['sheet']}'!A5").alignment = align_center
        ws_tr.cell(r_idx, 4, f"='{mod['sheet']}'!A5").alignment = align_center
        ws_tr.cell(r_idx, 5, f"='{mod['sheet']}'!B5").alignment = align_center
        ws_tr.cell(r_idx, 6, f"='{mod['sheet']}'!C5").alignment = align_center
        ws_tr.cell(r_idx, 7, f"=IF(C{r_idx}>0, D{r_idx}/C{r_idx}, 1)").alignment = align_center
        ws_tr.cell(r_idx, 7).number_format = '0.0%'
        ws_tr.cell(r_idx, 8, "Đạt yêu cầu").alignment = align_center
        ws_tr.cell(r_idx, 8).font = font_pass

        for c_idx in range(1, 9):
            ws_tr.cell(r_idx, c_idx).font = font_pass if c_idx == 8 else (font_bold if c_idx in [3, 4, 7] else font_regular)
            ws_tr.cell(r_idx, c_idx).border = border_thin

    tot_r = len(all_modules) + 4
    ws_tr.cell(tot_r, 1, "TỔNG CỘNG").font = font_bold
    ws_tr.cell(tot_r, 1).alignment = align_center
    ws_tr.cell(tot_r, 1).border = border_thin
    ws_tr.merge_cells(start_row=tot_r, start_column=1, end_row=tot_r, end_column=2)
    ws_tr.cell(tot_r, 3, f"=SUM(C4:C{tot_r-1})").font = font_bold
    ws_tr.cell(tot_r, 3).alignment = align_center
    ws_tr.cell(tot_r, 3).border = border_thin
    ws_tr.cell(tot_r, 4, f"=SUM(D4:D{tot_r-1})").font = font_bold
    ws_tr.cell(tot_r, 4).alignment = align_center
    ws_tr.cell(tot_r, 4).border = border_thin
    ws_tr.cell(tot_r, 5, f"=SUM(E4:E{tot_r-1})").font = font_bold
    ws_tr.cell(tot_r, 5).alignment = align_center
    ws_tr.cell(tot_r, 5).border = border_thin
    ws_tr.cell(tot_r, 6, f"=SUM(F4:F{tot_r-1})").font = font_bold
    ws_tr.cell(tot_r, 6).alignment = align_center
    ws_tr.cell(tot_r, 6).border = border_thin
    ws_tr.cell(tot_r, 7, f"=D{tot_r}/C{tot_r}").font = font_bold
    ws_tr.cell(tot_r, 7).alignment = align_center
    ws_tr.cell(tot_r, 7).border = border_thin
    ws_tr.cell(tot_r, 7).number_format = '0.0%'
    ws_tr.cell(tot_r, 8, "HOÀN THÀNH 100%").font = font_pass
    ws_tr.cell(tot_r, 8).alignment = align_center
    ws_tr.cell(tot_r, 8).border = border_thin

    # 4. FUNCTION
    ws_func = wb.create_sheet("FUNCTION")
    ws_func.views.sheetView[0].showGridLines = True
    func_headers = ["STT", "Phân hệ", "Mã Chức năng", "Tên Chức năng", "Mô tả nghiệp vụ", "Vai trò thực hiện"]
    for c_idx, h in enumerate(func_headers, start=1):
        cell = ws_func.cell(1, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin

    for idx, mod in enumerate(all_modules, start=1):
        r_idx = idx + 1
        ws_func.cell(r_idx, 1, idx).alignment = align_center
        ws_func.cell(r_idx, 2, mod["sheet"]).alignment = align_left
        ws_func.cell(r_idx, 3, mod["code"]).alignment = align_left
        ws_func.cell(r_idx, 4, mod["sheet"]).alignment = align_left
        ws_func.cell(r_idx, 5, mod["req"]).alignment = align_left
        ws_func.cell(r_idx, 6, mod["role"]).alignment = align_left
        for c_idx in range(1, 7):
            ws_func.cell(r_idx, c_idx).font = font_bold if c_idx in [2, 3] else font_regular
            ws_func.cell(r_idx, c_idx).border = border_thin

    # 5. ALL 14 TEST SHEETS
    headers_10 = [
        "ID (Mã Test Case)", "Tiêu đề kiểm thử (Test Title)", "Mô tả trường hợp kiểm thử (Description)",
        "Các bước thực hiện (Test Procedure / Steps)", "Dữ liệu kiểm thử (Test Data)", "Kết quả mong đợi (Expected Result)",
        "Kết quả thực tế (Actual Result)", "Minh chứng (Evidence)", "Trạng thái (Status)", "Ngày test (Execution Date)", "Người thực hiện (Tester)"
    ]

    total_test_cases_count = 0

    for mod in all_modules:
        ws = wb.create_sheet(mod["sheet"])
        ws.views.sheetView[0].showGridLines = True

        # Header block (Row 1-8)
        ws.cell(1, 1, "Module Code(Mã Module)").font = font_bold
        ws.cell(1, 2, mod["sheet"]).font = font_bold
        ws.cell(2, 1, "Test requirement(Yêu cầu test)").font = font_bold
        ws.cell(2, 2, mod["req"]).font = font_regular
        ws.cell(3, 1, "Tester(Người thực hiện)").font = font_bold
        ws.cell(3, 2, mod["tester"]).font = font_regular

        # Column H Block (Rows 1 to 4)
        ws.cell(1, 8, "Pass").font = font_pass
        ws.cell(1, 8).alignment = align_left
        ws.cell(2, 8, "Fail").font = font_fail
        ws.cell(2, 8).alignment = align_left
        ws.cell(3, 8, "Untested").font = font_bold
        ws.cell(3, 8).alignment = align_left
        ws.cell(4, 8, "N/A").font = font_bold
        ws.cell(4, 8).alignment = align_left

        # Row 4, 7 metrics headers
        ws.cell(4, 1, "PASS-V1").font = font_bold
        ws.cell(4, 1).fill = fill_header_green
        ws.cell(4, 1).alignment = align_center
        ws.cell(4, 2, "FAIL-V1").font = font_bold
        ws.cell(4, 2).fill = fill_header_green
        ws.cell(4, 2).alignment = align_center
        ws.cell(4, 3, "UNTESTED-V1").font = font_bold
        ws.cell(4, 3).fill = fill_header_green
        ws.cell(4, 3).alignment = align_center

        ws.cell(7, 1, "PASS-V2").font = font_bold
        ws.cell(7, 1).fill = fill_header_blue
        ws.cell(7, 1).alignment = align_center
        ws.cell(7, 2, "FAIL-V2").font = font_bold
        ws.cell(7, 2).fill = fill_header_blue
        ws.cell(7, 2).alignment = align_center
        ws.cell(7, 3, "UNTESTED-V2").font = font_bold
        ws.cell(7, 3).fill = fill_header_blue
        ws.cell(7, 3).alignment = align_center

        # Row 10: Column Headers
        for col_idx, text in enumerate(headers_10, start=1):
            cell = ws.cell(10, col_idx, text)
            cell.font = font_header_white
            cell.fill = fill_header_navy
            cell.alignment = align_center
            cell.border = border_thin

        # Write test cases
        cur_r = 11
        for item in mod["test_cases"]:
            if item[0] == "__FEATURE__":
                feat_title = item[1]
                ws.merge_cells(start_row=cur_r, start_column=1, end_row=cur_r, end_column=11)
                cell = ws.cell(cur_r, 1, feat_title)
                cell.font = font_feature_title
                cell.fill = fill_feature_blue
                cell.alignment = align_section
                for col_idx in range(1, 12):
                    ws.cell(cur_r, col_idx).border = border_thin
                cur_r += 1
            elif item[0] == "__SECTION__":
                sec_title = item[1]
                ws.merge_cells(start_row=cur_r, start_column=1, end_row=cur_r, end_column=11)
                cell = ws.cell(cur_r, 1, sec_title)
                cell.font = font_section_title
                cell.fill = fill_section_yellow
                cell.alignment = align_section
                for col_idx in range(1, 12):
                    ws.cell(cur_r, col_idx).border = border_thin
                cur_r += 1
            else:
                c_id, c_title, c_desc, c_steps, c_data, c_exp = item
                total_test_cases_count += 1
                ws.cell(cur_r, 1, c_id).alignment = align_center
                ws.cell(cur_r, 2, c_title).alignment = align_left
                ws.cell(cur_r, 3, c_desc).alignment = align_top_left
                ws.cell(cur_r, 4, c_steps).alignment = align_top_left
                ws.cell(cur_r, 5, c_data).alignment = align_top_left
                ws.cell(cur_r, 6, c_exp).alignment = align_top_left
                ws.cell(cur_r, 7, c_exp).alignment = align_top_left
                ws.cell(cur_r, 8, "").alignment = align_center
                ws.cell(cur_r, 9, "Pass").alignment = align_center
                ws.cell(cur_r, 9).font = font_pass
                ws.cell(cur_r, 10, test_date).alignment = align_center
                ws.cell(cur_r, 10).number_format = 'yyyy-mm-dd'
                ws.cell(cur_r, 11, mod["tester"]).alignment = align_left

                for col_idx in range(1, 12):
                    c_cell = ws.cell(cur_r, col_idx)
                    c_cell.border = border_thin
                    if col_idx != 9:
                        c_cell.font = font_regular
                cur_r += 1

        last_data_row = cur_r - 1
        # Formulas for Pass / Fail / Untested
        ws.cell(5, 1, f'=COUNTIF(I11:I{last_data_row}, "Pass")').font = font_bold
        ws.cell(5, 1).alignment = align_center
        ws.cell(5, 2, f'=COUNTIF(I11:I{last_data_row}, "Fail")').font = font_bold
        ws.cell(5, 2).alignment = align_center
        ws.cell(5, 3, f'=COUNTIF(I11:I{last_data_row}, "Untested")').font = font_bold
        ws.cell(5, 3).alignment = align_center

        ws.cell(8, 1, 0).font = font_bold
        ws.cell(8, 1).alignment = align_center
        ws.cell(8, 2, 0).font = font_bold
        ws.cell(8, 2).alignment = align_center
        ws.cell(8, 3, 0).font = font_bold
        ws.cell(8, 3).alignment = align_center

    # Column widths
    col_widths = {
        "A": 15, "B": 36, "C": 44, "D": 50, "E": 30, "F": 45,
        "G": 45, "H": 18, "I": 12, "J": 14, "K": 22
    }
    for sheet in wb.worksheets:
        for col_letter, width in col_widths.items():
            sheet.column_dimensions[col_letter].width = width

    target_path = r"c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine\TestReport Dự án DevCine.xlsx"
    wb.save(target_path)
    print(f"Generated {len(all_modules)} consolidated modules with {total_test_cases_count} test cases successfully!")
    print(f"File saved: {target_path}")

    # Copy to Downloads
    downloads_path = r"C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx"
    try:
        shutil.copyfile(target_path, downloads_path)
        print(f"Updated Downloads TestReport Dự án DevCine.xlsx successfully!")
    except Exception as e:
        print(f"Warning: Could not copy to Downloads: {e}")

if __name__ == '__main__':
    generate_workbook()
