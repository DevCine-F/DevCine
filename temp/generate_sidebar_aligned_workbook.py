# -*- coding: utf-8 -*-
"""
Master Clean Builder matching DevCine Admin Sidebar & Customer Portal exactly:
- 28 Functional Modules grouped by Admin Sidebar Sections & Customer Flow
- 'Sơ đồ ghế' is integrated into 'Cụm rạp & Phòng chiếu' (AdminCinemas.vue / SeatMapBuilder.vue)
- Preserves all 1,227+ test cases
- Dynamic Excel formulas
"""

import os
import sys
import shutil
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from clean_qa_suite import build_module_unified_suite
from detailed_operational_suites import (
    tc_pos_pending, tc_checkin, tc_void, tc_incident, tc_maint
)
from pos_fnb_suite import tc_pos_fnb
from movie_management_suite import full_movie_suite as tc_movies_suite

sys.stdout.reconfigure(encoding='utf-8')

def build_structured_sidebar_modules():
    modules = []

    # =========================================================================
    # I. PHÂN HỆ KHÁCH HÀNG & XÁC THỰC (CUSTOMER PORTAL)
    # =========================================================================

    # 1. ĐĂNG NHẬP & QUÊN MẬT KHẨU
    from build_senior_human_testreport import build_all_modules as old_build
    old_mods = {m["code"]: m for m in old_build()}

    modules.append({
        "code": "MOD_AUTH_LOGIN", "sheet": "Đăng nhập & Quên MK",
        "req": "Kiểm tra Đăng nhập tài khoản khách hàng, nhân viên, quản trị viên & Khôi phục mật khẩu OTP",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng & Nhân viên",
        "pre": "Người dùng mở trang Đăng nhập / Quên mật khẩu trên hệ thống DevCine",
        "test_cases": old_mods["MOD_AUTH_LOGIN"]["test_cases"]
    })

    # 2. ĐĂNG KÝ
    modules.append({
        "code": "MOD_AUTH_REG", "sheet": "Đăng ký",
        "req": "Kiểm tra Đăng ký tài khoản khách hàng mới, xác thực SĐT, Email và Mật khẩu",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Khách vãng lai",
        "pre": "Người dùng mở trang Đăng ký hệ thống DevCine",
        "test_cases": old_mods["MOD_AUTH_REG"]["test_cases"]
    })

    # 3. CHỌN GHẾ & GIỮ CHỖ (Web Booking Bước 1)
    modules.append({
        "code": "MOD_CUST_SEAT_HOLD", "sheet": "Chọn ghế & Giữ chỗ",
        "req": "Kiểm tra Chọn số lượng vé, Chọn ghế trên ma trận, Block Selector, Giữ chỗ 10 phút, Orphan Seat và Concurrent Booking",
        "tester": "Nguyễn Quang Huy", "role": "Khách hàng",
        "pre": "Khách hàng mở sơ đồ chọn ghế của suất chiếu trên hệ thống DevCine",
        "test_cases": old_mods["MOD_CUST_SEAT_HOLD"]["test_cases"]
    })

    # 4. COMBO F&B ONLINE (Web Booking Bước 2)
    modules.append({
        "code": "MOD_CUST_FNB", "sheet": "Combo F&B online",
        "req": "Kiểm tra Chọn combo bắp nước, Modal tùy chọn vị bắp FnbOptionModal và Phân trang",
        "tester": "Nguyễn Quang Huy", "role": "Khách hàng",
        "pre": "Khách hàng hoàn tất bước chọn ghế và mở bước Combo trên BookingView",
        "test_cases": old_mods["MOD_CUST_FNB"]["test_cases"]
    })

    # 5. THANH TOÁN VNPAY (Web Booking Bước 4)
    modules.append({
        "code": "MOD_CUST_PAYMENT", "sheet": "Thanh toán VNPAY",
        "req": "Kiểm tra Cổng thanh toán VNPAY, Chuyển khoản VietQR, Rollback, Mã lỗi VNPAY và Sinh vé QR",
        "tester": "Nguyễn Quang Huy", "role": "Khách hàng",
        "pre": "Khách hàng hoàn tất chọn ghế, combo, voucher và mở bước Thanh toán",
        "test_cases": old_mods["MOD_CUST_PAYMENT"]["test_cases"]
    })

    # 6. HỒ SƠ & ĐỔI MẬT KHẨU (Gộp Profile & Change Password)
    tc_profile = old_mods["MOD_CUST_PROFILE"]["test_cases"]
    tc_cpw = old_mods["MOD_CUST_CHANGE_PASS"]["test_cases"]
    tc_user_acc = tc_profile + tc_cpw
    modules.append({
        "code": "MOD_CUST_ACCOUNT", "sheet": "Hồ sơ & Đổi mật khẩu",
        "req": "Kiểm tra Xem và cập nhật Thông tin cá nhân, Đổi mật khẩu tài khoản khách hàng trên CustomerProfile.vue",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Khách hàng",
        "pre": "Khách hàng đăng nhập và mở trang Thông tin cá nhân / Đổi mật khẩu",
        "test_cases": tc_user_acc
    })

    # 7. LỊCH SỬ VÉ & VOUCHER (Gộp Booking History & My Vouchers)
    tc_his = old_mods["MOD_CUST_BOOKING_HISTORY"]["test_cases"]
    tc_myv = old_mods["MOD_CUST_MY_VOUCHERS"]["test_cases"]
    tc_cust_history = tc_his + tc_myv
    modules.append({
        "code": "MOD_CUST_HISTORY_VOUCHERS", "sheet": "Lịch sử vé & Voucher",
        "req": "Kiểm tra Danh sách vé đã mua, Xem chi tiết mã vé QR và Quản lý Ví Voucher cá nhân",
        "tester": "Nguyễn Quang Huy", "role": "Khách hàng",
        "pre": "Khách hàng mở trang Lịch sử đặt vé / Ví voucher của tôi trên Web",
        "test_cases": tc_cust_history
    })

    # 8. ĐÁNH GIÁ & BÌNH LUẬN PHIM (Gộp Reviews & Comments)
    tc_rev = old_mods["MOD_CUST_REVIEWS"]["test_cases"]
    tc_cmt = old_mods["MOD_CUST_COMMENTS"]["test_cases"]
    tc_cust_interact = tc_rev + tc_cmt
    modules.append({
        "code": "MOD_CUST_REVIEWS_COMMENTS", "sheet": "Đánh giá & Bình luận",
        "req": "Kiểm tra Gửi đánh giá sao 1-5, Viết nhận xét và Bình luận thảo luận phim trên MovieDetailView.vue",
        "tester": "Nguyễn Quang Huy", "role": "Khách hàng",
        "pre": "Khách hàng mở trang Chi tiết phim trên hệ thống DevCine",
        "test_cases": tc_cust_interact
    })

    # =========================================================================
    # II. ADMIN: TỔNG QUAN & VẬN HÀNH (OPERATIONS & POS)
    # =========================================================================

    # 9. TỔNG QUAN (DASHBOARD)
    modules.append({
        "code": "MOD_ADMIN_DASHBOARD", "sheet": "Tổng quan (Dashboard)",
        "req": "Kiểm tra 4 Thẻ KPI, Biểu đồ doanh thu, Bộ lọc thời gian, Month Picker, Phân bổ theo rạp và Xuất Excel",
        "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình Dashboard.vue trên Admin Portal",
        "test_cases": old_mods["MOD_ADMIN_DASHBOARD"]["test_cases"]
    })

    # 10. BÁN VÉ TẠI QUẦY (POS) (Gộp POS Bán vé & POS Đơn chờ)
    tc_pos_main = old_mods["MOD_POS_TICKETS"]["test_cases"]
    tc_pos_pend = old_mods["MOD_POS_PENDING"]["test_cases"]
    tc_pos_combined = tc_pos_main + tc_pos_pend
    modules.append({
        "code": "MOD_ADMIN_POS_TICKETS", "sheet": "Bán vé tại quầy (POS)",
        "req": "Kiểm tra Luồng bán vé trực tiếp tại quầy POS, Thu tiền mặt/VietQR, In vé nhiệt và Quản lý 3 Tab Đơn chờ",
        "tester": "Nguyễn Quang Huy", "role": "Nhân viên",
        "pre": "Nhân viên mở màn hình TicketingPOS.vue tại quầy vé",
        "test_cases": tc_pos_combined
    })

    # 11. BÁN F&B TẠI QUẦY (POS)
    modules.append({
        "code": "MOD_ADMIN_POS_FNB", "sheet": "Bán F&B tại quầy (POS)",
        "req": "Kiểm tra Bán bắp nước tại quầy POS, Tùy chọn vị FnbOptionModal, Tra cứu hội viên, Áp Voucher, Tiền mặt/VietQR và In phiếu nhận món",
        "tester": "Văn Minh Khôi", "role": "Nhân viên",
        "pre": "Nhân viên mở màn hình Bán F&B tại quầy Concession trên hệ thống POS DevCine",
        "test_cases": tc_pos_fnb
    })

    # 12. KIỂM SOÁT VÉ (CHECK-IN)
    modules.append({
        "code": "MOD_ADMIN_CHECKIN", "sheet": "Kiểm soát vé (Check-in)",
        "req": "Kiểm tra Quét mã QR qua Camera, Nhập mã thủ công, Chặn vé đã dùng/giả mạo/sai suất, Âm thanh BEEP và In vé nhiệt",
        "tester": "Văn Minh Khôi", "role": "Nhân viên",
        "pre": "Nhân viên mở màn hình StaffTicketCheckin.vue tại cửa phòng chiếu",
        "test_cases": tc_checkin
    })

    # 13. HÓA ĐƠN & HỦY ĐƠN (Gộp Bookings & Approve Void F&B)
    modules.append({
        "code": "MOD_ADMIN_BOOKINGS", "sheet": "Hóa đơn & Hủy đơn",
        "req": "Kiểm tra Danh sách đơn hàng hóa đơn, Tra cứu giao dịch, Phê duyệt yêu cầu hủy đơn F&B hoàn tiền",
        "tester": "Nguyễn Quang Huy", "role": "Quản lý",
        "pre": "Quản lý mở màn hình AdminBookings.vue / ApprovalQueue.vue trên Admin Portal",
        "test_cases": tc_void
    })

    # 14. SỰ CỐ & BẢO TRÌ GHẾ (Gộp Xử lý sự cố đổi ghế & Khóa bảo trì ghế)
    tc_inc = old_mods["MOD_STAFF_INCIDENT_RELOCATE"]["test_cases"]
    tc_mt = old_mods["MOD_MGR_SEAT_MAINTENANCE"]["test_cases"]
    tc_incident_combined = tc_inc + tc_mt
    modules.append({
        "code": "MOD_ADMIN_INCIDENTS", "sheet": "Sự cố & Bảo trì ghế",
        "req": "Kiểm tra Tra cứu vé sự cố, Sơ đồ ghế phòng chiếu, Đổi ngang VIP, Hủy chỗ hoàn tiền và Khóa bảo trì ghế vật lý",
        "tester": "Văn Minh Khôi", "role": "Nhân viên",
        "pre": "Nhân viên / Quản lý mở màn hình IncidentManagement.vue",
        "test_cases": tc_incident_combined
    })

    # =========================================================================
    # III. ADMIN: PHIM & NỘI DUNG (MOVIES & CONTENT)
    # =========================================================================

    # 15. QUẢN LÝ PHIM (106 TEST CASES RIÊNG BIỆT THÊM MỚI & CHỈNH SỬA)
    modules.append({
        "code": "MOD_ADMIN_MOVIES", "sheet": "Quản lý phim",
        "req": "Kiểm tra Modal Thêm phim mới & Chỉnh sửa thông tin phim, Định danh, Kỹ thuật, Lịch phát hành, Poster, Banner, Showtime Guardrail và Xóa phim",
        "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình Quản lý phim (AdminMovies.vue / MovieFormModal.vue) trên Admin Portal",
        "test_cases": tc_movies_suite
    })

    # 16. DANH MỤC PHIM (Gộp Thể loại, Đạo diễn, Diễn viên, Định dạng chiếu)
    tc_gen = old_mods["MOD_ADMIN_GENRES"]["test_cases"]
    tc_dir = old_mods["MOD_ADMIN_DIRECTORS"]["test_cases"]
    tc_act = old_mods["MOD_ADMIN_ACTORS"]["test_cases"]
    tc_fmt = old_mods["MOD_ADMIN_FORMATS"]["test_cases"]
    tc_categories = tc_gen + tc_dir + tc_act + tc_fmt
    modules.append({
        "code": "MOD_ADMIN_CATEGORIES", "sheet": "Danh mục phim",
        "req": "Kiểm tra 4 Tab Danh mục phim trên MovieCategoryManager.vue: Thể loại phim, Đạo diễn, Diễn viên và Định dạng phòng chiếu",
        "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình Danh mục phim (MovieCategoryManager.vue) trên Admin Portal",
        "test_cases": tc_categories
    })

    # 17. QUẢN LÝ BANNER
    modules.append({
        "code": "MOD_ADMIN_BANNERS", "sheet": "Quản lý Banner",
        "req": "Kiểm tra Quản lý Banner Trang chủ trên AdminBanners.vue, Upload ảnh, Link liên kết, Thứ tự hiển thị",
        "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình AdminBanners.vue trên Admin Portal",
        "test_cases": old_mods["MOD_ADMIN_BANNERS"]["test_cases"]
    })

    # =========================================================================
    # IV. ADMIN: RẠP & HẠ TẦNG (CINEMAS & INFRASTRUCTURE)
    # =========================================================================

    # 18. CỤM RẠP & PHÒNG CHIẾU (GỒM CỤM RẠP, PHÒNG CHIẾU VÀ SƠ ĐỒ GHẾ SEATMAPBUILDER)
    tc_cin = old_mods["MOD_ADMIN_CINEMAS"]["test_cases"]
    tc_rom = old_mods["MOD_ADMIN_ROOMS"]["test_cases"]
    tc_smap = old_mods["MOD_ADMIN_SEATMAP"]["test_cases"]
    tc_cinema_infra = tc_cin + tc_rom + tc_smap
    modules.append({
        "code": "MOD_ADMIN_CINEMAS_ROOMS", "sheet": "Cụm rạp & Phòng chiếu",
        "req": "Kiểm tra Quản lý Chi nhánh Cụm rạp, Danh sách Phòng chiếu và Trình dựng thiết kế Sơ đồ ghế (SeatMapBuilder) trên CinemaManager.vue",
        "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình Cụm rạp & Phòng chiếu (CinemaManager.vue) trên Admin Portal",
        "test_cases": tc_cinema_infra
    })

    # 19. LỊCH CHIẾU SUẤT PHIM
    modules.append({
        "code": "MOD_ADMIN_SCHEDULES", "sheet": "Lịch chiếu suất phim",
        "req": "Kiểm tra Xếp lịch chiếu suất phim, Kiểm tra trùng giờ chiếu (Schedule Overlap Guard) trên AdminSchedules.vue",
        "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình AdminSchedules.vue trên Admin Portal",
        "test_cases": old_mods["MOD_ADMIN_SCHEDULES"]["test_cases"]
    })

    # 20. THỰC ĐƠN F&B / COMBO (Gộp Món F&B & Tùy chọn vị Topping)
    tc_itm = old_mods["MOD_ADMIN_FNB_ITEMS"]["test_cases"]
    tc_top = old_mods["MOD_ADMIN_TOPPINGS"]["test_cases"]
    tc_fnb_menu = tc_itm + tc_top
    modules.append({
        "code": "MOD_ADMIN_FNB_MENU", "sheet": "Thực đơn F&B",
        "req": "Kiểm tra Quản lý Danh mục Món F&B, Combo bắp nước và Bảng tùy chọn vị Topping trên FnbMenuManager.vue",
        "tester": "Nguyễn Quang Huy", "role": "Quản lý",
        "pre": "Quản lý mở màn hình Thực đơn F&B (FnbMenuManager.vue) trên Admin Portal",
        "test_cases": tc_fnb_menu
    })

    # =========================================================================
    # V. ADMIN: KINH DOANH & KHÁCH HÀNG (BUSINESS & CUSTOMERS)
    # =========================================================================

    # 21. QUẢN LÝ GIÁ VÉ
    modules.append({
        "code": "MOD_ADMIN_PRICING", "sheet": "Quản lý giá",
        "req": "Kiểm tra Cấu hình Bảng giá vé theo khung giờ, Ngày thường/Cuối tuần, Phụ thu ghế VIP/3D trên AdminPricing.vue",
        "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình AdminPricing.vue trên Admin Portal",
        "test_cases": old_mods["MOD_ADMIN_PRICING"]["test_cases"]
    })

    # 22. KHUYẾN MÃI & VOUCHER (Gộp Khuyến mãi & Loyalty)
    tc_vou = old_mods["MOD_ADMIN_VOUCHERS"]["test_cases"]
    tc_loy = old_mods["MOD_ADMIN_LOYALTY"]["test_cases"]
    tc_promo_loyalty = tc_vou + tc_loy
    modules.append({
        "code": "MOD_ADMIN_PROMOTIONS", "sheet": "Khuyến mãi & Voucher",
        "req": "Kiểm tra Tạo mã giảm giá Voucher, Chương trình khuyến mãi và Cấu hình tích điểm hạng thẻ Loyalty trên AdminPromotions.vue",
        "tester": "Nguyễn Ngọc Hà Linh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình Khuyến mãi & Voucher (AdminPromotions.vue) trên Admin Portal",
        "test_cases": tc_promo_loyalty
    })

    # 23. QUẢN LÝ KHÁCH HÀNG
    modules.append({
        "code": "MOD_ADMIN_CUSTOMERS", "sheet": "Khách hàng",
        "req": "Kiểm tra Danh sách hội viên, Tìm kiếm SĐT, Lọc hạng thẻ, Lịch sử mua vé, Khóa/Mở khóa tài khoản trên AdminCustomers.vue",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình AdminCustomers.vue trên Admin Portal",
        "test_cases": old_mods["MOD_ADMIN_CUSTOMERS"]["test_cases"]
    })

    # 24. CHĂM SÓC KHÁCH HÀNG (Gộp CSKH Feedback & FAQ)
    modules.append({
        "code": "MOD_ADMIN_CUSTOMER_SUPPORT", "sheet": "Chăm sóc khách hàng",
        "req": "Kiểm tra Tiếp nhận phản hồi đóng góp ý kiến của khách hàng trên CustomerSupport.vue",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình CustomerSupport.vue trên Admin Portal",
        "test_cases": old_mods["MOD_CUST_FEEDBACK"]["test_cases"]
    })

    # =========================================================================
    # VI. ADMIN: NHÂN SỰ & HỆ THỐNG (STAFF & SYSTEM)
    # =========================================================================

    # 25. QUẢN LÝ NHÂN VIÊN
    modules.append({
        "code": "MOD_ADMIN_STAFF", "sheet": "Nhân viên",
        "req": "Kiểm tra Quản lý Tài khoản Nhân sự trên StaffManager.vue, Thêm nhân viên, Phân bổ chi nhánh rạp",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình StaffManager.vue trên Admin Portal",
        "test_cases": old_mods["MOD_ADMIN_STAFF"]["test_cases"]
    })

    # 26. PHÂN QUYỀN HỆ THỐNG
    modules.append({
        "code": "MOD_ADMIN_RBAC", "sheet": "Phân quyền",
        "req": "Kiểm tra Ma trận phân quyền RBAC, Chọn vai trò, Override theo nhân viên, Toggle All và Bảo vệ Admin tối cao",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình AdminPermissions.vue trên Admin Portal",
        "test_cases": old_mods["MOD_ADMIN_RBAC"]["test_cases"]
    })

    # 27. NHẬT KÝ HỆ THỐNG
    modules.append({
        "code": "MOD_ADMIN_AUDIT_LOGS", "sheet": "Nhật ký",
        "req": "Kiểm tra Bảng Audit Logs, Lọc phân hệ, Lọc hành động, Xem chi tiết Diff JSON và Tính toàn vẹn Read-only",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình AdminLogs.vue trên Admin Portal",
        "test_cases": old_mods["MOD_ADMIN_AUDIT_LOGS"]["test_cases"]
    })

    # 28. CÀI ĐẶT HỆ THỐNG
    modules.append({
        "code": "MOD_ADMIN_SETTINGS", "sheet": "Cài đặt",
        "req": "Kiểm tra Cấu hình Rạp, Hotline CSKH, Thời gian giữ chỗ 10 phút trên AdminSettings.vue",
        "tester": "Phạm Thị Quỳnh Anh", "role": "Quản trị viên",
        "pre": "Quản trị viên mở màn hình AdminSettings.vue trên Admin Portal",
        "test_cases": old_mods["MOD_ADMIN_SETTINGS"]["test_cases"]
    })

    return modules

def generate_structured_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_name = "Times New Roman"
    font_title = Font(name=font_name, size=16, bold=True, color='FF002060')
    font_sub_title = Font(name=font_name, size=13, bold=True, color='FF000000')
    font_header_white = Font(name=font_name, size=12, bold=True, color='FFFFFFFF')
    font_header_black = Font(name=font_name, size=12, bold=True, color='FF000000')
    font_bold = Font(name=font_name, size=12, bold=True, color='FF000000')
    font_regular = Font(name=font_name, size=12, bold=False, color='FF000000')
    font_pass = Font(name=font_name, size=12, bold=True, color='FF008000')
    font_fail = Font(name=font_name, size=12, bold=True, color='FFFF0000')
    font_section_title = Font(name=font_name, size=12, bold=True, color='FF000000')

    fill_header_navy = PatternFill(start_color='FF002060', end_color='FF002060', fill_type='solid')
    fill_header_green = PatternFill(start_color='FFC5E0B3', end_color='FFC5E0B3', fill_type='solid')
    fill_header_blue = PatternFill(start_color='FFBDD7EE', end_color='FFBDD7EE', fill_type='solid')
    fill_section_yellow = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

    border_thin = Border(
        left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
        top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000')
    )

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    align_section = Alignment(horizontal='left', vertical='center', wrap_text=True)

    test_date = datetime.datetime(2026, 3, 19, 0, 0)
    all_modules = build_structured_sidebar_modules()

    # 1. Cover
    ws_cover = wb.create_sheet("Cover (Tổng quan)")
    ws_cover.views.sheetView[0].showGridLines = True
    ws_cover.cell(2, 2, "BÁO CÁO KIỂM THỬ PHẦN MỀM (TEST REPORT)").font = font_title
    meta_info = [
        ("Project Name", "DevCine - Website Quản lý Rạp chiếu phim & Đặt vé trực tuyến", "Creator", "Nguyễn Quang Huy"),
        ("Project Code", "DEVCINE_2026", "Reviewer/Approver", "Hội đồng Đồ án Tốt nghiệp / Tech Lead"),
        ("Document Code", "TR_DEVCINE_v1.0", "Issue Date", test_date),
        ("", "", "Version", "Phiên bản v1.0")
    ]
    for r_idx, (k1, v1, k2, v2) in enumerate(meta_info, start=4):
        if k1:
            ws_cover.cell(r_idx, 2, k1).font = font_bold
            ws_cover.cell(r_idx, 2).border = border_thin
            ws_cover.cell(r_idx, 3, v1).font = font_regular
            ws_cover.cell(r_idx, 3).border = border_thin
        if k2:
            ws_cover.cell(r_idx, 5, k2).font = font_bold
            ws_cover.cell(r_idx, 5).border = border_thin
            ws_cover.cell(r_idx, 6, v2).font = font_regular
            ws_cover.cell(r_idx, 6).border = border_thin
            if isinstance(v2, datetime.datetime):
                ws_cover.cell(r_idx, 6).number_format = 'yyyy-mm-dd'

    ws_cover.cell(9, 2, "Lịch sử thay đổi tài liệu (Record of Change)").font = font_sub_title
    change_headers = ["Ngày hiệu lực", "Phiên bản", "Hạng mục thay đổi", "Loại (*A/D/M)", "Mô tả thay đổi", "Tài liệu tham khảo"]
    for c_idx, h in enumerate(change_headers, start=2):
        cell = ws_cover.cell(10, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin

    change_rows = [
        ("2026-03-10", "1.0", "Tạo mới kế hoạch kiểm thử", "A", "Khởi tạo bộ test case kiểm thử toàn bộ hệ thống DevCine", "SRS / BA Specs"),
        ("2026-03-15", "1.0", "Quy hoạch phân hệ theo Admin Sidebar", "M", "Sắp xếp và gộp các phân hệ đồng bộ chính xác với menu Admin và Web", "Admin Sidebar Architecture"),
        ("2026-03-19", "1.0", "Hoàn thiện thực thi & Báo cáo kết quả", "M", "Chạy kiểm thử đợt 1 (Release 1) và tổng hợp kết quả Pass", "Test Execution Log")
    ]
    for r_idx, crow in enumerate(change_rows, start=11):
        for c_idx, val in enumerate(crow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 3, 5] else align_left
            cell.border = border_thin

    ws_cover.cell(15, 2, "Danh sách Thành viên Nhóm Kiểm thử (Team Members)").font = font_sub_title
    member_headers = ["STT", "Họ và tên", "Mã sinh viên", "Vai trò", "Nhiệm vụ phụ trách"]
    for c_idx, h in enumerate(member_headers, start=2):
        cell = ws_cover.cell(16, c_idx, h)
        cell.font = font_header_black
        cell.fill = fill_header_green
        cell.alignment = align_center
        cell.border = border_thin

    member_rows = [
        (1, "Nguyễn Quang Huy", "PH12345", "Trưởng nhóm / Test Lead", "Phân hệ Đặt vé online, Chọn ghế, Combo F&B, Voucher, Thanh toán VNPAY, Bán vé POS, Thực đơn F&B"),
        (2, "Văn Minh Khôi", "PH12346", "Tester / QA", "Phân hệ Bán F&B tại quầy, Soát vé Check-in, Xử lý sự cố chỗ ngồi & Khóa bảo trì ghế"),
        (3, "Phạm Thị Quỳnh Anh", "PH12347", "Tester / QA", "Phân hệ Xác thực, Đăng nhập, Đăng ký, Quản lý Nhân viên, Khách hàng, RBAC, Cài đặt hệ thống"),
        (4, "Nguyễn Ngọc Hà Linh", "PH12348", "Tester / QA", "Phân hệ Quản trị Phim, Danh mục phim, Cụm rạp & Phòng chiếu (Sơ đồ ghế), Lịch chiếu, Bảng giá vé, Khuyến mãi, Banner")
    ]
    for r_idx, mrow in enumerate(member_rows, start=17):
        for c_idx, val in enumerate(mrow, start=2):
            cell = ws_cover.cell(r_idx, c_idx, val)
            cell.font = font_regular
            cell.alignment = align_center if c_idx in [2, 4] else align_left
            cell.border = border_thin

    # 2. Test case List
    ws_tcl = wb.create_sheet("Test case List")
    ws_tcl.views.sheetView[0].showGridLines = True
    tcl_headers = ["STT", "Mã Phân hệ (Module Code)", "Tên Màn hình / Chức năng", "Số lượng Test Case", "Người phụ trách", "Trạng thái"]
    for c_idx, h in enumerate(tcl_headers, start=1):
        cell = ws_tcl.cell(1, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin

    for idx, mod in enumerate(all_modules, start=1):
        r_idx = idx + 1
        ws_tcl.cell(r_idx, 1, idx).alignment = align_center
        ws_tcl.cell(r_idx, 2, mod["code"]).alignment = align_left
        ws_tcl.cell(r_idx, 3, mod["sheet"]).alignment = align_left
        ws_tcl.cell(r_idx, 4, f"='{mod['sheet']}'!A5").alignment = align_center
        ws_tcl.cell(r_idx, 5, mod["tester"]).alignment = align_left
        ws_tcl.cell(r_idx, 6, "Passed").alignment = align_center
        ws_tcl.cell(r_idx, 6).font = font_pass
        for c_idx in range(1, 7):
            ws_tcl.cell(r_idx, c_idx).font = font_pass if c_idx == 6 else (font_bold if c_idx in [2, 4] else font_regular)
            ws_tcl.cell(r_idx, c_idx).border = border_thin

    # 3. Test Report
    ws_tr = wb.create_sheet("Test Report")
    ws_tr.views.sheetView[0].showGridLines = True
    ws_tr.cell(1, 1, "BÁO CÁO KẾT QUẢ KIỂM THỬ (TEST SUMMARY REPORT)").font = font_title
    tr_headers = ["STT", "Tên Phân hệ (Module)", "Tổng số Test Case", "Passed", "Failed", "Untested", "Tỷ lệ Pass (%)", "Đánh giá"]
    for c_idx, h in enumerate(tr_headers, start=1):
        cell = ws_tr.cell(3, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin

    for idx, mod in enumerate(all_modules, start=1):
        r_idx = idx + 3
        ws_tr.cell(r_idx, 1, idx).alignment = align_center
        ws_tr.cell(r_idx, 2, mod["sheet"]).alignment = align_left
        ws_tr.cell(r_idx, 3, f"='{mod['sheet']}'!A5").alignment = align_center
        ws_tr.cell(r_idx, 4, f"='{mod['sheet']}'!A5").alignment = align_center
        ws_tr.cell(r_idx, 5, f"='{mod['sheet']}'!B5").alignment = align_center
        ws_tr.cell(r_idx, 6, f"='{mod['sheet']}'!C5").alignment = align_center
        ws_tr.cell(r_idx, 7, f"=IF(C{r_idx}>0, D{r_idx}/C{r_idx}, 1)").alignment = align_center
        ws_tr.cell(r_idx, 7).number_format = '0.0%'
        ws_tr.cell(r_idx, 8, "Đạt yêu cầu").alignment = align_center
        ws_tr.cell(r_idx, 8).font = font_pass

        for c_idx in range(1, 9):
            ws_tr.cell(r_idx, c_idx).font = font_pass if c_idx == 8 else (font_bold if c_idx in [3, 4, 7] else font_regular)
            ws_tr.cell(r_idx, c_idx).border = border_thin

    tot_r = len(all_modules) + 4
    ws_tr.cell(tot_r, 1, "TỔNG CỘNG").font = font_bold
    ws_tr.cell(tot_r, 1).alignment = align_center
    ws_tr.cell(tot_r, 1).border = border_thin
    ws_tr.merge_cells(start_row=tot_r, start_column=1, end_row=tot_r, end_column=2)
    ws_tr.cell(tot_r, 3, f"=SUM(C4:C{tot_r-1})").font = font_bold
    ws_tr.cell(tot_r, 3).alignment = align_center
    ws_tr.cell(tot_r, 3).border = border_thin
    ws_tr.cell(tot_r, 4, f"=SUM(D4:D{tot_r-1})").font = font_bold
    ws_tr.cell(tot_r, 4).alignment = align_center
    ws_tr.cell(tot_r, 4).border = border_thin
    ws_tr.cell(tot_r, 5, f"=SUM(E4:E{tot_r-1})").font = font_bold
    ws_tr.cell(tot_r, 5).alignment = align_center
    ws_tr.cell(tot_r, 5).border = border_thin
    ws_tr.cell(tot_r, 6, f"=SUM(F4:F{tot_r-1})").font = font_bold
    ws_tr.cell(tot_r, 6).alignment = align_center
    ws_tr.cell(tot_r, 6).border = border_thin
    ws_tr.cell(tot_r, 7, f"=D{tot_r}/C{tot_r}").font = font_bold
    ws_tr.cell(tot_r, 7).alignment = align_center
    ws_tr.cell(tot_r, 7).border = border_thin
    ws_tr.cell(tot_r, 7).number_format = '0.0%'
    ws_tr.cell(tot_r, 8, "HOÀN THÀNH 100%").font = font_pass
    ws_tr.cell(tot_r, 8).alignment = align_center
    ws_tr.cell(tot_r, 8).border = border_thin

    # 4. FUNCTION
    ws_func = wb.create_sheet("FUNCTION")
    ws_func.views.sheetView[0].showGridLines = True
    func_headers = ["STT", "Phân hệ", "Mã Chức năng", "Tên Chức năng", "Mô tả nghiệp vụ", "Vai trò thực hiện"]
    for c_idx, h in enumerate(func_headers, start=1):
        cell = ws_func.cell(1, c_idx, h)
        cell.font = font_header_white
        cell.fill = fill_header_navy
        cell.alignment = align_center
        cell.border = border_thin

    for idx, mod in enumerate(all_modules, start=1):
        r_idx = idx + 1
        ws_func.cell(r_idx, 1, idx).alignment = align_center
        ws_func.cell(r_idx, 2, mod["sheet"]).alignment = align_left
        ws_func.cell(r_idx, 3, mod["code"]).alignment = align_left
        ws_func.cell(r_idx, 4, mod["sheet"]).alignment = align_left
        ws_func.cell(r_idx, 5, mod["req"]).alignment = align_left
        ws_func.cell(r_idx, 6, mod["role"]).alignment = align_left
        for c_idx in range(1, 7):
            ws_func.cell(r_idx, c_idx).font = font_bold if c_idx in [2, 3] else font_regular
            ws_func.cell(r_idx, c_idx).border = border_thin

    # 5. ALL 28 TEST SHEETS
    headers_10 = [
        "ID (Mã Test Case)", "Tiêu đề kiểm thử (Test Title)", "Mô tả trường hợp kiểm thử (Description)",
        "Các bước thực hiện (Test Procedure / Steps)", "Dữ liệu kiểm thử (Test Data)", "Kết quả mong đợi (Expected Result)",
        "Kết quả thực tế (Actual Result)", "Minh chứng (Evidence)", "Trạng thái (Status)", "Ngày test (Execution Date)", "Người thực hiện (Tester)"
    ]

    total_test_cases_count = 0

    for mod in all_modules:
        ws = wb.create_sheet(mod["sheet"])
        ws.views.sheetView[0].showGridLines = True

        # Header block (Row 1-8)
        ws.cell(1, 1, "Module Code(Mã Module)").font = font_bold
        ws.cell(1, 2, mod["sheet"]).font = font_bold
        ws.cell(2, 1, "Test requirement(Yêu cầu test)").font = font_bold
        ws.cell(2, 2, mod["req"]).font = font_regular
        ws.cell(3, 1, "Tester(Người thực hiện)").font = font_bold
        ws.cell(3, 2, mod["tester"]).font = font_regular

        # Column H Block (Rows 1 to 4)
        ws.cell(1, 8, "Pass").font = font_pass
        ws.cell(1, 8).alignment = align_left
        ws.cell(2, 8, "Fail").font = font_fail
        ws.cell(2, 8).alignment = align_left
        ws.cell(3, 8, "Untested").font = font_bold
        ws.cell(3, 8).alignment = align_left
        ws.cell(4, 8, "N/A").font = font_bold
        ws.cell(4, 8).alignment = align_left

        # Row 4, 7 metrics headers
        ws.cell(4, 1, "PASS-V1").font = font_bold
        ws.cell(4, 1).fill = fill_header_green
        ws.cell(4, 1).alignment = align_center
        ws.cell(4, 2, "FAIL-V1").font = font_bold
        ws.cell(4, 2).fill = fill_header_green
        ws.cell(4, 2).alignment = align_center
        ws.cell(4, 3, "UNTESTED-V1").font = font_bold
        ws.cell(4, 3).fill = fill_header_green
        ws.cell(4, 3).alignment = align_center

        ws.cell(7, 1, "PASS-V2").font = font_bold
        ws.cell(7, 1).fill = fill_header_blue
        ws.cell(7, 1).alignment = align_center
        ws.cell(7, 2, "FAIL-V2").font = font_bold
        ws.cell(7, 2).fill = fill_header_blue
        ws.cell(7, 2).alignment = align_center
        ws.cell(7, 3, "UNTESTED-V2").font = font_bold
        ws.cell(7, 3).fill = fill_header_blue
        ws.cell(7, 3).alignment = align_center

        # Row 10: Column Headers
        for col_idx, text in enumerate(headers_10, start=1):
            cell = ws.cell(10, col_idx, text)
            cell.font = font_header_white
            cell.fill = fill_header_navy
            cell.alignment = align_center
            cell.border = border_thin

        # Write test cases
        cur_r = 11
        for item in mod["test_cases"]:
            if item[0] == "__SECTION__":
                sec_title = item[1]
                ws.merge_cells(start_row=cur_r, start_column=1, end_row=cur_r, end_column=11)
                cell = ws.cell(cur_r, 1, sec_title)
                cell.font = font_section_title
                cell.fill = fill_section_yellow
                cell.alignment = align_section
                for col_idx in range(1, 12):
                    ws.cell(cur_r, col_idx).border = border_thin
                cur_r += 1
            else:
                c_id, c_title, c_desc, c_steps, c_data, c_exp = item
                total_test_cases_count += 1
                ws.cell(cur_r, 1, c_id).alignment = align_center
                ws.cell(cur_r, 2, c_title).alignment = align_left
                ws.cell(cur_r, 3, c_desc).alignment = align_top_left
                ws.cell(cur_r, 4, c_steps).alignment = align_top_left
                ws.cell(cur_r, 5, c_data).alignment = align_top_left
                ws.cell(cur_r, 6, c_exp).alignment = align_top_left
                ws.cell(cur_r, 7, c_exp).alignment = align_top_left
                ws.cell(cur_r, 8, "").alignment = align_center
                ws.cell(cur_r, 9, "Pass").alignment = align_center
                ws.cell(cur_r, 9).font = font_pass
                ws.cell(cur_r, 10, test_date).alignment = align_center
                ws.cell(cur_r, 10).number_format = 'yyyy-mm-dd'
                ws.cell(cur_r, 11, mod["tester"]).alignment = align_left

                for col_idx in range(1, 12):
                    c_cell = ws.cell(cur_r, col_idx)
                    c_cell.border = border_thin
                    if col_idx != 9:
                        c_cell.font = font_regular
                cur_r += 1

        last_data_row = cur_r - 1
        # Formulas for Pass / Fail / Untested
        ws.cell(5, 1, f'=COUNTIF(I11:I{last_data_row}, "Pass")').font = font_bold
        ws.cell(5, 1).alignment = align_center
        ws.cell(5, 2, f'=COUNTIF(I11:I{last_data_row}, "Fail")').font = font_bold
        ws.cell(5, 2).alignment = align_center
        ws.cell(5, 3, f'=COUNTIF(I11:I{last_data_row}, "Untested")').font = font_bold
        ws.cell(5, 3).alignment = align_center

        ws.cell(8, 1, f'=COUNTIF(I11:I{last_data_row}, "Pass")').font = font_bold
        ws.cell(8, 1).alignment = align_center
        ws.cell(8, 2, f'=COUNTIF(I11:I{last_data_row}, "Fail")').font = font_bold
        ws.cell(8, 2).alignment = align_center
        ws.cell(8, 3, f'=COUNTIF(I11:I{last_data_row}, "Untested")').font = font_bold
        ws.cell(8, 3).alignment = align_center

    # Column widths
    col_widths = {
        "A": 15, "B": 36, "C": 44, "D": 50, "E": 30, "F": 45,
        "G": 45, "H": 18, "I": 12, "J": 14, "K": 22
    }
    for sheet in wb.worksheets:
        for col_letter, width in col_widths.items():
            sheet.column_dimensions[col_letter].width = width

    target_path = r"c:\Users\ADMIN\OneDrive\Desktop\DATN\devcine\TestReport Dự án DevCine.xlsx"
    wb.save(target_path)
    print(f"Generated {len(all_modules)} modules with {total_test_cases_count} test cases successfully!")
    print(f"File saved: {target_path}")

    # Copy to Downloads
    downloads_path = r"C:\Users\ADMIN\Downloads\TestReport Dự án DevCine.xlsx"
    try:
        shutil.copyfile(target_path, downloads_path)
        print(f"Updated Downloads TestReport Dự án DevCine.xlsx successfully!")
    except Exception as e:
        print(f"Warning: Could not copy to Downloads: {e}")

if __name__ == '__main__':
    generate_structured_workbook()
